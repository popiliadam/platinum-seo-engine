"""scripts/state/dump_workspace.py — Y-07 manager session state summary.

One-shot read of the 5 data sources a manager session needs to triage state at
a glance:

  1. _state/events.jsonl son N satır (parsed JSON)
  2. master.xlsx#master_task TODO satır sayısı (project ROOT — the live
     convention; init-project writes projects/<slug>/master.xlsx)
  3. _state/workflows/<run_id>.json içinde status="awaiting_approval" olanlar
  4. _state/backups/master-excel-<TIMESTAMP>.xlsx son N (mtime-sorted)
  5. Drift verdict (GREEN/AMBER/RED): _state/consistency-report.json if a report
     is persisted, else computed LIVE from the root master.xlsx via the invariant
     evaluator (no report is persisted today, so this avoids a null verdict)

CLI semantics:
  - --project <slug> opsiyonel. Eğer atlanırsa shared/active.json'daki
    `active_project` key okunur (legacy `slug` fallback + deprecation uyarısı).
    Active.json yoksa FileNotFoundError raise edilir.
  - --workspace-root  PSEO_WORKSPACE_ROOT environment variable'ından default
    alınır; --workspace-root flag explicit override.

All data sources are graceful: missing path -> None / empty list (manager
session use case'i için raise YERINE — partial state'te de iş görür).

Usage::

    python3 scripts/state/dump_workspace.py
    python3 scripts/state/dump_workspace.py --project vento
    PSEO_WORKSPACE_ROOT=/path/to/ws python3 scripts/state/dump_workspace.py

Authority: ADR-021 events.jsonl path, ADR-035 PSEO_WORKSPACE_ROOT,
  rules/append-only-state.md workflow paths, schemas/master-excel.schema.json
  master_task statusEnum, schemas/consistency-report.schema.json verdict enum.
Brief: docs/superpowers/plans/v1.5-audit-closure-brief.md Phase 3 Tier 2.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import List

REPO_ROOT_DEFAULT = Path(__file__).resolve().parents[2]


def _resolve_slug(workspace_root: Path, project_slug: str | None) -> str:
    if project_slug is not None:
        return project_slug
    active = workspace_root / "shared" / "active.json"
    if not active.exists():
        raise FileNotFoundError(
            f"no bound project — provide --project <slug> or create "
            f"{active} (PSEO bootstrap convention)"
        )
    try:
        data = json.loads(active.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise FileNotFoundError(
            f"shared/active.json is not valid JSON: {exc}"
        ) from exc
    slug = data.get("active_project") or data.get("slug")
    if not slug:
        raise FileNotFoundError(
            f"shared/active.json has no 'active_project' key: {active}"
        )
    if "active_project" not in data and "slug" in data:
        print(
            f"warning: shared/active.json uses legacy 'slug' key; rename to "
            f"'active_project' ({active})",
            file=sys.stderr,
        )
    return str(slug)


def _events_tail(events_path: Path, n: int) -> List[dict]:
    if not events_path.exists():
        return []
    out: List[dict] = []
    text = events_path.read_text(encoding="utf-8")
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out[-n:]


def _master_task_todo_count(xlsx_path: Path) -> int | None:
    if not xlsx_path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None
    if "master_task" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["master_task"]
    status_col: int | None = None
    header_row: int | None = None
    for hr in (1, 2, 3):
        try:
            row_cells = list(ws[hr])
        except (IndexError, ValueError):
            continue
        for cell in row_cells:
            if cell.value == "status":
                status_col = cell.column
                header_row = hr
                break
        if status_col is not None:
            break
    if status_col is None or header_row is None:
        wb.close()
        return None
    count = 0
    for row in ws.iter_rows(
        min_row=header_row + 1,
        min_col=status_col,
        max_col=status_col,
        values_only=True,
    ):
        if row[0] == "TODO":
            count += 1
    wb.close()
    return count


def _pending_approvals(workflows_dir: Path) -> List[str]:
    if not workflows_dir.exists():
        return []
    pending: List[str] = []
    for path in sorted(workflows_dir.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        if data.get("status") == "awaiting_approval":
            pending.append(path.stem)
    return pending


def _backups_recent(backups_dir: Path, n: int) -> List[str]:
    if not backups_dir.exists():
        return []
    # Recurse: the approved Excel writer rotates backups into a master/ subdir
    # (transaction.py writes _state/backups/master/master-<ISO>.xlsx), so a flat
    # iterdir saw only the subdirectory and reported no backups even when the
    # keep-7 rotation held real files.
    files = [p for p in backups_dir.rglob("*.xlsx") if p.is_file()]
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return [p.name for p in files[:n]]


def _drift_verdict(
    report_path: Path,
    workbook_path: Path | None = None,
    slug: str | None = None,
    workspace_root: Path | None = None,
) -> str | None:
    """Drift verdict for the manager summary.

    A persisted consistency-report.json wins if present (a future drift-check
    may write it). Otherwise — the live case today, where no report is persisted
    — compute the verdict directly from the root master.xlsx via the invariant
    evaluator so the summary shows a real GREEN/AMBER/RED instead of null.
    Graceful: any read/eval error returns None.
    """
    if report_path.exists():
        try:
            data = json.loads(report_path.read_text(encoding="utf-8"))
            verdict = data.get("verdict")
            if isinstance(verdict, str):
                return verdict
        except (json.JSONDecodeError, OSError):
            pass
    if workbook_path and workbook_path.exists() and slug:
        try:
            import importlib.util

            import openpyxl

            vi_path = Path(__file__).resolve().parents[1] / "validation" / "validate_invariants.py"
            spec = importlib.util.spec_from_file_location("vi_dump_gate", vi_path)
            vi = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(vi)
            wb = openpyxl.load_workbook(str(workbook_path), read_only=True, data_only=True)
            agg = vi.aggregate_verdicts(vi.evaluate_all(wb, slug, workspace_root=workspace_root))
            overall = agg.get("overall")
            return overall if isinstance(overall, str) else None
        except Exception:
            return None
    return None


def dump_workspace(
    workspace_root: Path,
    project_slug: str | None = None,
    events_tail_n: int = 5,
    backups_recent_n: int = 7,
) -> dict:
    workspace_root = Path(workspace_root)
    if not workspace_root.exists():
        raise FileNotFoundError(f"workspace not found: {workspace_root}")

    slug = _resolve_slug(workspace_root, project_slug)
    project_dir = workspace_root / "projects" / slug
    if not project_dir.exists():
        raise FileNotFoundError(
            f"unknown project slug {slug!r} under {workspace_root}/projects/"
        )

    state_dir = project_dir / "_state"

    return {
        "project": slug,
        "events_tail": _events_tail(state_dir / "events.jsonl", events_tail_n),
        "master_task_todo_count": _master_task_todo_count(
            project_dir / "master.xlsx"
        ),
        "workflow_pending_approvals": _pending_approvals(state_dir / "workflows"),
        "backups_recent": _backups_recent(state_dir / "backups", backups_recent_n),
        "drift_verdict": _drift_verdict(
            state_dir / "consistency-report.json",
            workbook_path=project_dir / "master.xlsx",
            slug=slug,
            workspace_root=workspace_root,
        ),
    }


def _format_summary(result: dict) -> str:
    lines: List[str] = []
    lines.append(f"project: {result['project']}")
    lines.append(f"drift verdict: {result['drift_verdict'] or '(no report)'}")
    todo = result["master_task_todo_count"]
    lines.append(
        f"master_task TODO count: "
        f"{todo if todo is not None else '(no master.xlsx)'}"
    )
    pending = result["workflow_pending_approvals"]
    if pending:
        lines.append(f"workflows awaiting approval ({len(pending)}):")
        for p in pending:
            lines.append(f"  - {p}")
    else:
        lines.append("workflows awaiting approval: none")
    backups = result["backups_recent"]
    if backups:
        lines.append(f"backups recent ({len(backups)}):")
        for b in backups:
            lines.append(f"  - {b}")
    else:
        lines.append("backups recent: none")
    tail = result["events_tail"]
    if tail:
        lines.append(f"events tail (last {len(tail)}):")
        for evt in tail:
            evt_id = evt.get("event_id", "?")
            evt_type = evt.get("event_type", "?")
            lines.append(f"  - {evt_id} [{evt_type}]")
    else:
        lines.append("events tail: empty")
    return "\n".join(lines)


def main(argv: List[str] | None = None) -> int:
    env_ws = os.environ.get("PSEO_WORKSPACE_ROOT", "")
    parser = argparse.ArgumentParser(
        prog="dump_workspace.py",
        description="Manager session state summary tek komut (Y-07).",
    )
    parser.add_argument(
        "--workspace-root",
        default=env_ws,
        help="Workspace root path (default: $PSEO_WORKSPACE_ROOT env var)",
    )
    parser.add_argument(
        "--project",
        default=None,
        dest="slug",
        help="Project slug (default: shared/active.json active_project)",
    )
    parser.add_argument(
        "--events-tail",
        type=int,
        default=5,
        help="Number of events.jsonl tail lines to include (default: 5)",
    )
    parser.add_argument(
        "--backups-recent",
        type=int,
        default=7,
        help="Number of recent backups to list (default: 7)",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Emit machine-readable JSON instead of human summary",
    )
    args = parser.parse_args(argv)

    if not args.workspace_root:
        print(
            "error: --workspace-root or $PSEO_WORKSPACE_ROOT required",
            file=sys.stderr,
        )
        return 2

    try:
        result = dump_workspace(
            workspace_root=Path(args.workspace_root),
            project_slug=args.slug,
            events_tail_n=args.events_tail,
            backups_recent_n=args.backups_recent,
        )
    except FileNotFoundError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(_format_summary(result), file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
