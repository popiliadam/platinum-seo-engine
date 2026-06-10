#!/usr/bin/env python3
"""
sheet_merge.py — idempotent, namespace-scoped row merge into a master.xlsx sheet.

sf-import projects six sheets via ``transaction.replace`` (whole-block snapshot
refresh): any rows another writer *appends* to ``robots_txt`` / ``redirect_404``
are WIPED on the next sf-import. The tech-SEO governance skills
(hreflang-audit / facet-nav-audit / robots-policy-audit / migration-map) must add
COMPUTED findings to those same sheets without (a) duplicating their own rows on
re-run, or (b) clobbering sf_projection's ``R-NNN`` rows and each other's
``HF-`` / ``FN-`` / ``RP-`` rows.

This module is the one merge path that reconciles those two facts. Each call
reads the current sheet, drops ONLY the rows in the caller's own namespace
(``id`` startswith ``id_prefix``, or ``key_column`` value in the new rows' keys),
appends the caller's new rows, then re-lands the UNION through the single
idempotent commit path ``scripts.orchestration.committer.commit`` (whole-block
``transaction.replace``). Re-running the same skill replaces its own namespace in
place — no growth, no foreign-row loss.

Two functions:
  - ``merge_prefixed_rows`` — id-namespaced sheets (``robots_txt``): generates
    sequential ids ``f"{id_prefix}{seq:03d}"`` for the new rows.
  - ``merge_keyed_rows``   — keyed sheets without an id column (``redirect_404``):
    replaces rows whose ``key_column`` value matches a new row's key.

Discipline (mirrors sf_projection / tech_audit_transform):
  - Pure read → compute union → single committer.commit; no slug literals.
  - Never mutates the caller's row dicts (immutability; new dicts only).
  - Schema validation + lock + backup are committer.commit's job — RowSchemaError
    / SchemaSheetMismatchError / LockHeldError propagate unchanged.
  - Append-only state preserved: committer.commit emits the provenance event.

Refs: scripts/orchestration/committer.py (single commit path),
scripts/excel/transaction.py (replace semantics + exceptions + data_start_row),
schemas/master-excel.schema.json (sheet shapes), scripts/ingestion/sf_projection.py
(the R-NNN rows this merge preserves), rules/tech-seo-governance.md (R-125..R-136).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# scripts is a namespace package; ensure repo root on sys.path for absolute imports.
_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.excel.transaction import (  # noqa: E402
    SchemaSheetMismatchError,
    WriteResult,
)
from scripts.orchestration import committer  # noqa: E402

_DEFAULT_SCHEMA_PATH = _REPO_ROOT / "schemas" / "master-excel.schema.json"


# ---------------------------------------------------------------------------
# Schema resolution (mirrors transaction._ordered_columns / _data_start_row)
# ---------------------------------------------------------------------------

def _load_schema(schema_path: Path | str | None) -> dict:
    p = Path(schema_path) if schema_path else _DEFAULT_SCHEMA_PATH
    with p.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _sheet_columns(schema: dict, sheet: str) -> list[str]:
    """Ordered column NAMES for ``sheet``; raise if the sheet is not declared.

    Raising here (not deep inside committer) gives the merge an early,
    explicit ``refuses unknown sheet`` contract independent of whether the
    workbook/sheet physically exists yet.
    """
    sheets = schema.get("sheets") or {}
    if sheet not in sheets:
        raise SchemaSheetMismatchError(
            f"sheet {sheet!r} not declared in master-excel.schema.json"
        )
    cols = sheets[sheet].get("required_columns") or []
    return [c["name"] for c in cols]


def _data_start_row(schema: dict, sheet: str) -> int:
    declared = (schema.get("sheets") or {}).get(sheet, {}).get("data_start_row")
    return declared if isinstance(declared, int) and declared >= 1 else 2


# ---------------------------------------------------------------------------
# Current-sheet read (values-only, from the schema's data_start_row)
# ---------------------------------------------------------------------------

def _read_existing_rows(
    xlsx_path: Path, sheet: str, columns: list[str], data_start_row: int
) -> list[dict]:
    """Read the current data block of ``sheet`` into dicts keyed by column name.

    Returns ``[]`` when the workbook or the sheet does not exist yet (first run)
    — committer.commit creates them. Fully-blank trailing rows are skipped so a
    re-landed union never accumulates empty rows.
    """
    if not Path(xlsx_path).exists():
        return []
    wb = load_workbook(str(xlsx_path), data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        rows: list[dict] = []
        for raw in ws.iter_rows(min_row=data_start_row, values_only=True):
            if raw is None or all(v is None for v in raw):
                continue
            rows.append({name: (raw[i] if i < len(raw) else None)
                         for i, name in enumerate(columns)})
        return rows
    finally:
        wb.close()


def _commit_union(
    xlsx_path: Path | str,
    sheet: str,
    union_rows: list[dict],
    *,
    run_id: str,
    project_slug: str,
    schema_path: Path | str | None,
    state_root: Path | str | None,
    writer: str,
) -> WriteResult:
    return committer.commit(
        xlsx_path,
        sheet,
        union_rows,
        run_id=run_id,
        project_slug=project_slug,
        schema_path=schema_path,
        state_root=state_root,
        writer=writer,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def merge_prefixed_rows(
    xlsx_path: Path | str,
    sheet: str,
    new_rows: list[dict],
    *,
    id_prefix: str,
    id_column: str = "id",
    run_id: str,
    project_slug: str,
    writer: str,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
) -> WriteResult:
    """Idempotently merge ``new_rows`` into an id-namespaced ``sheet``.

    Reads the current sheet, drops every row whose ``id_column`` value starts
    with ``id_prefix`` (the caller's own namespace), assigns the new rows
    sequential ids ``f"{id_prefix}{seq:03d}"`` (seq from 1), and re-lands the
    union (foreign rows + freshly-id'd new rows) via committer.commit. Re-running
    with the same ``new_rows`` reproduces the same ids → no duplication, no
    foreign-row loss. Caller dicts are never mutated.

    RowSchemaError / SchemaSheetMismatchError / LockHeldError propagate from the
    commit path unchanged.
    """
    schema = _load_schema(schema_path)
    columns = _sheet_columns(schema, sheet)  # raises on unknown sheet (early)
    data_start = _data_start_row(schema, sheet)

    existing = _read_existing_rows(Path(xlsx_path), sheet, columns, data_start)
    kept = [
        row for row in existing
        if not str(row.get(id_column) or "").startswith(id_prefix)
    ]

    renumbered: list[dict] = [
        {**row, id_column: f"{id_prefix}{seq:03d}"}
        for seq, row in enumerate(new_rows, start=1)
    ]

    union = kept + renumbered
    return _commit_union(
        xlsx_path, sheet, union,
        run_id=run_id, project_slug=project_slug,
        schema_path=schema_path, state_root=state_root, writer=writer,
    )


def merge_keyed_rows(
    xlsx_path: Path | str,
    sheet: str,
    new_rows: list[dict],
    *,
    key_column: str = "url",
    run_id: str,
    project_slug: str,
    writer: str,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
) -> WriteResult:
    """Idempotently merge ``new_rows`` into a keyed ``sheet`` with no id column.

    For sheets like ``redirect_404`` (key = ``url``): drops every existing row
    whose ``key_column`` value matches a key present in ``new_rows``, keeps the
    rest, appends ``new_rows`` verbatim, and re-lands the union via
    committer.commit. Re-running replaces the same keys in place (idempotent).
    Caller dicts are never mutated.
    """
    schema = _load_schema(schema_path)
    columns = _sheet_columns(schema, sheet)  # raises on unknown sheet (early)
    data_start = _data_start_row(schema, sheet)

    new_keys = {row.get(key_column) for row in new_rows}
    existing = _read_existing_rows(Path(xlsx_path), sheet, columns, data_start)
    kept = [row for row in existing if row.get(key_column) not in new_keys]

    union = kept + [dict(row) for row in new_rows]
    return _commit_union(
        xlsx_path, sheet, union,
        run_id=run_id, project_slug=project_slug,
        schema_path=schema_path, state_root=state_root, writer=writer,
    )


__all__ = ("merge_prefixed_rows", "merge_keyed_rows")
