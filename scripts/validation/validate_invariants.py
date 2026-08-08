#!/usr/bin/env python3
"""
validate_invariants.py — 25 hand-coded invariant rules for drift-check.

Owned by `skills/governance/drift-check/SKILL.md`. Read-only governance:
loads `master.xlsx` (data_only=True, read_only=True) plus _state side
files (events.jsonl, workflows/*.json, backups/) and evaluates the 25
invariants enumerated in §17.2 + the F-08 quick-wins URL subset rule.

Single exception to read-only: F-12 records/advances a monotonic high-water
line count to `_state/events.snapshot.json` (a derived governance baseline,
not the SSoT) so the append-only shrink check is not vacuous. No other rule
writes to the workspace.

Each rule is a small Python function. NO DSL — that is a Phase 6+
refactor candidate. The intent now is auditable, hand-verifiable
governance: every rule is one short function, every function returns
the same shape:

    {
        "id":        "F-XX",
        "severity":  "CRITICAL|HIGH|MEDIUM",
        "verdict":   "PASS|FAIL|SKIP",
        "evidence":  "human-readable one-liner",
        "rule":      "rule body verbatim",
        "category":  "csr_foundation|csr_data|...",
        "auto_repair_available": bool,
        "manual_triage": bool,                  # F-15-style routing
        "affected_sheets": [...],               # optional
        "sample_violations": [...],             # optional, ≤ 20
        "affected_rows": int,                   # optional
    }

Aggregation per §17.2:
  - any FAIL with severity CRITICAL  → overall RED
  - any FAIL with severity HIGH      → overall RED, UNLESS manual_triage
                                       routes it to AMBER
  - any FAIL with severity MEDIUM    → overall AMBER
  - any SKIP                         → overall AMBER (not RED)
  - all PASS                         → overall GREEN

The aggregator also produces a schema-valid `consistency-report.json`
ready to be persisted by drift-check step 5.

Refs:
  - schemas/master-excel.schema.json (definitions: statusEnum,
    severityEnum; per-sheet column shape)
  - schemas/events.schema.json (event_kind=audit + run_id integer)
  - schemas/workflow-run.schema.json (schema_version const "1.0",
    run_id pattern)
  - schemas/consistency-report.schema.json (output shape)
  - schemas/cross-sheet-invariants.json (rule registry)
  - rules/excel-discipline.md (master.xlsx never mutated outside
    transaction.py)
  - ADR-018 (statusEnum 7-value), ADR-019 (workflow schema_version),
    ADR-020 (event_kind=workflow), ADR-021 (_state path).
"""

from __future__ import annotations

import functools
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, quote, unquote, urlencode, urlsplit, urlunsplit

from jsonschema.exceptions import ValidationError as _JSValidationError

# AMO batch 3-gov-driftF (F-27): reuse 3a's declared-tool parser and IMPORT the
# 2b gate's single gated-MCP constant, so the drift rule can never diverge from
# the gate it audits — there is no second copy of "what MCP tool is gated".
# The check_excel_writer hook loads THIS file STANDALONE via
# spec_from_file_location (repo root NOT on sys.path), so bootstrap the repo
# root onto sys.path BEFORE these package imports — mirrors
# outward_action_gate.py's own bootstrap and the hook's file-path loader (the
# dual-context load pattern; the 0c bare-CLI lesson).
_BOOTSTRAP_ROOT = str(Path(__file__).resolve().parents[2])
if _BOOTSTRAP_ROOT not in sys.path:
    sys.path.insert(0, _BOOTSTRAP_ROOT)
from scripts.hooks.outward_action_gate import _MCP_SUBMIT_TOOL  # noqa: E402
from scripts.validation.skill_mcp_usage import (  # noqa: E402
    declared_tools,
    split_frontmatter_body,
)
# build_validator (not a raw Draft7Validator) so the consistency-report
# generated_at format:date-time is ENFORCED with the strict UTC '…Z' checker
# (P1-02 / time-discipline §8.10).
from scripts.validation.validate_schema import build_validator  # noqa: E402


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_REPO_ROOT = Path(__file__).resolve().parents[2]
_MASTER_EXCEL_SCHEMA = _REPO_ROOT / "schemas" / "master-excel.schema.json"
_CONSISTENCY_REPORT_SCHEMA = _REPO_ROOT / "schemas" / "consistency-report.schema.json"

_STATUS_ENUM_7: frozenset[str] = frozenset({
    "TODO", "ONGOING", "EXISTS", "DONE",
    "BLOCKED", "DEFERRED", "CANCELED",
})
_SEVERITY_ENUM_4: frozenset[str] = frozenset({
    "LOW", "MEDIUM", "HIGH", "CRITICAL",
})

_WORKFLOW_SCHEMA_VERSION = "1.0"
_WORKFLOW_RUN_ID_PATTERN = re.compile(
    r"^[a-z][a-z0-9-]*-\d{4}-\d{2}-\d{2}-[a-f0-9]{4}$"
)

_EVENTS_MAX_LINE_BYTES = 64 * 1024
_EXCEL_CELL_MAX_CHARS = 32_767
_BACKUP_KEEP_LIMIT = 7
_TRACKING_PARAMS = {"utm_source", "utm_medium", "utm_campaign",
                    "utm_term", "utm_content", "gclid", "fbclid"}

_SAMPLE_CAP = 20  # consistency-report.schema.json sample_violations.maxItems


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class DriftCheckError(Exception):
    """Base class for drift-check evaluator errors."""


class UnknownRuleError(DriftCheckError):
    """Aggregator received an unknown rule id."""


class ConsistencyReportInvalidError(DriftCheckError):
    """Built consistency-report failed schema validation."""


# ---------------------------------------------------------------------------
# Workbook helpers (read-only)
# ---------------------------------------------------------------------------

def _has_sheet(workbook: Any, name: str) -> bool:
    try:
        return name in workbook.sheetnames
    except Exception:
        return False


def _resolve_header_row(workbook: Any, sheet: str) -> int:
    """Phase 14 W3-W2-C-a: resolve the effective header row for `sheet` by
    consulting `schemas/master-excel.schema.json` first, falling back to row 1.

    Logic:
      1. If schema declares `sheets[sheet].header_row` as an int and the
         workbook's row at that index has at least one non-None cell whose
         text matches one of the schema's `required_columns`, trust schema.
      2. Otherwise (legacy / synthetic test workbook with header at row 1),
         return 1.

    This keeps existing tests (header at row 1) PASS while enabling W1
    bootstrap workbooks (header at row 3/4/5 per schema) to validate
    without 4 mechanical header-parse FAILs (F-01/F-05/F-17/F-18).
    """
    try:
        schema = _load_master_excel_schema()
    except Exception:
        return 1
    sdef = schema.get("sheets", {}).get(sheet) or {}
    hr = sdef.get("header_row")
    if not isinstance(hr, int) or hr <= 0:
        return 1
    required = sdef.get("required_columns") or []
    if not required:
        return 1
    if not _has_sheet(workbook, sheet):
        return 1
    ws = workbook[sheet]
    try:
        # Probe the schema-declared header row using openpyxl's read-only
        # iter_rows (min_row/max_row inclusive, 1-indexed).
        probe = next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True))
    except (StopIteration, Exception):
        return 1
    if probe is None:
        return 1
    # required_columns entries are dicts ({"col","name",...}); extract name.
    def _col_name(c: Any) -> str:
        if isinstance(c, dict):
            return str(c.get("name") or "").strip()
        return str(c).strip()

    required_set = {n for n in (_col_name(c) for c in required) if n}
    probe_set = {str(c).strip() for c in probe if c is not None}
    # Require at least 50% header match to trust schema header_row.
    if required_set and probe_set:
        overlap = len(required_set & probe_set)
        if overlap >= max(1, len(required_set) // 2):
            return hr
    return 1


def _iter_rows_as_dicts(workbook: Any, sheet: str) -> list[dict]:
    """Read sheet header row + data rows; return list of dicts keyed
    by header column name. Returns [] if sheet missing or has 0 data rows.
    Read-only friendly (uses iter_rows + values_only=True).

    Phase 14 W3-W2-C-a: header row is resolved via `_resolve_header_row`
    (schema authority dynamic, fallback to row 1)."""
    if not _has_sheet(workbook, sheet):
        return []
    ws = workbook[sheet]
    header_row_idx = _resolve_header_row(workbook, sheet)
    try:
        header_tuple = next(ws.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    except StopIteration:
        return []
    if header_tuple is None:
        return []
    headers = [str(h) if h is not None else "" for h in header_tuple]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None:
            continue
        if all(c is None for c in row):
            continue
        d: dict[str, Any] = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            d[key] = val
        out.append(d)
    return out


def _normalize_url(raw: Any) -> str:
    """D-03 idempotent URL normalize. Mirrors quickwins_transform.normalize_url
    spec (lowercase scheme+host, strip default ports, drop fragment, sort+
    filter query, trailing slash collapse except root). Inputs that aren't
    str-shaped are returned as the empty string so callers can detect drift."""
    if not isinstance(raw, str) or not raw:
        return ""
    parts = urlsplit(raw.strip())
    scheme = parts.scheme.lower()
    netloc = parts.netloc
    if "@" in netloc:
        netloc = netloc.split("@", 1)[1]
    host = netloc
    port = ""
    if ":" in netloc and not netloc.endswith("]"):
        host, _, port = netloc.rpartition(":")
    host = host.lower()
    try:
        host = host.encode("idna").decode("ascii")
    except (UnicodeError, UnicodeDecodeError):
        pass
    default_ports = {"http": "80", "https": "443"}
    if port and port == default_ports.get(scheme):
        port = ""
    netloc = host + (f":{port}" if port else "")
    path = parts.path or "/"
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    pairs = [
        (k, v) for k, v in parse_qsl(parts.query, keep_blank_values=True)
        if k not in _TRACKING_PARAMS
    ]
    pairs.sort()
    query = urlencode(pairs)
    return urlunsplit((scheme, netloc, path, query, ""))


# ---------------------------------------------------------------------------
# Workspace path helpers
# ---------------------------------------------------------------------------

def _resolve_workspace_root(workspace_root: Path | None) -> Path:
    if workspace_root is not None:
        return Path(workspace_root)
    env = os.getenv("PSEO_WORKSPACE_ROOT")
    if env:
        return Path(env)
    raise DriftCheckError(
        "workspace_root not provided and PSEO_WORKSPACE_ROOT not set"
    )


def _project_dir(project_slug: str, workspace_root: Path | None) -> Path:
    return _resolve_workspace_root(workspace_root) / "projects" / project_slug


# ---------------------------------------------------------------------------
# Rule helpers
# ---------------------------------------------------------------------------

def _make_result(
    *, id_: str, severity: str, verdict: str, evidence: str,
    rule: str, category: str,
    auto_repair_available: bool = False,
    manual_triage: bool = False,
    affected_sheets: list[str] | None = None,
    sample_violations: list[str] | None = None,
    affected_rows: int | None = None,
) -> dict:
    out: dict[str, Any] = {
        "id": id_,
        "severity": severity,
        "verdict": verdict,
        "evidence": evidence,
        "rule": rule,
        "category": category,
        "auto_repair_available": auto_repair_available,
        "manual_triage": manual_triage,
    }
    if affected_sheets:
        out["affected_sheets"] = affected_sheets
    if sample_violations is not None:
        out["sample_violations"] = sample_violations[:_SAMPLE_CAP]
    if affected_rows is not None:
        out["affected_rows"] = affected_rows
    return out


# ---------------------------------------------------------------------------
# CRITICAL (5)
# ---------------------------------------------------------------------------

def check_F_01(workbook: Any, project_slug: str, **_) -> dict:
    """master_task.status ⊆ statusEnum 7-value."""
    rule = "master_task.status ⊆ statusEnum {TODO,ONGOING,EXISTS,DONE,BLOCKED,DEFERRED,CANCELED}"
    if not _has_sheet(workbook, "master_task"):
        return _make_result(
            id_="F-01", severity="CRITICAL", verdict="SKIP",
            evidence="master_task sheet missing — pilot wb may be sparse",
            rule=rule, category="csr_foundation",
            affected_sheets=["master_task"],
        )
    rows = _iter_rows_as_dicts(workbook, "master_task")
    bad: list[str] = []
    for r in rows:
        s = r.get("status")
        if s is None:
            continue
        if str(s) not in _STATUS_ENUM_7:
            bad.append(str(s))
    if bad:
        return _make_result(
            id_="F-01", severity="CRITICAL", verdict="FAIL",
            evidence=f"{len(bad)} master_task.status values outside 7-value enum",
            rule=rule, category="csr_foundation",
            affected_sheets=["master_task"],
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-01", severity="CRITICAL", verdict="PASS",
        evidence=f"all {len(rows)} master_task.status values within 7-value enum",
        rule=rule, category="csr_foundation",
        affected_sheets=["master_task"],
    )


def _check_no_excel_formula(workbook: Any, project_slug: str, *,
                            id_: str, formula_token: str,
                            formula_workbook: Any = None) -> dict:
    """F-02/F-03/F-04: dashboard sheet holds no live Excel formula of the token.

    REQUIRES a workbook opened with ``data_only=False``. Every other rule wants
    the cached VALUES, so the shared handle is opened ``data_only=True`` — and in
    that mode openpyxl never yields a formula string at all, it yields the cached
    result (``None`` for a freshly written formula, a number for one Excel has
    saved). The previous implementation sniffed the shared handle for a leading
    ``'=' + token`` and so could not match anything: it reported PASS over a live
    formula for as long as it existed, and fired only on a harmless TEXT cell that
    merely looked like a formula — the inverse of the rule it states.

    Measured directly on one file::

        data_only=True  -> dashboard!A3 = None
        data_only=False -> dashboard!A3 = '=AVERAGEIF(A1:A2,1)'

    Without the formula view this reports SKIP, never PASS. A check that cannot
    see its subject has not measured it, and calling that a pass is precisely the
    failure this rule exists to prevent.
    """
    rule = f"dashboard sheet contains no live `={formula_token}(...)` formulas"
    if formula_workbook is None:
        return _make_result(
            id_=id_, severity="CRITICAL", verdict="SKIP",
            evidence=(
                "no data_only=False workbook view supplied — a formula is "
                "INVISIBLE to the shared data_only=True handle, so this rule was "
                "NOT MEASURED (this is not a pass)"
            ),
            rule=rule, category="csr_foundation",
            affected_sheets=["dashboard"],
        )
    if not _has_sheet(formula_workbook, "dashboard"):
        return _make_result(
            id_=id_, severity="CRITICAL", verdict="SKIP",
            evidence="dashboard sheet missing — pilot wb may be sparse",
            rule=rule, category="csr_foundation",
            affected_sheets=["dashboard"],
        )
    ws = formula_workbook["dashboard"]
    bad: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        for cell in row:
            if isinstance(cell, str) and cell.startswith("=" + formula_token):
                bad.append(cell)
                if len(bad) >= _SAMPLE_CAP:
                    break
        if len(bad) >= _SAMPLE_CAP:
            break
    if bad:
        return _make_result(
            id_=id_, severity="CRITICAL", verdict="FAIL",
            evidence=f"{len(bad)} `={formula_token}(...)` formula cells found in dashboard",
            rule=rule, category="csr_foundation",
            affected_sheets=["dashboard"],
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_=id_, severity="CRITICAL", verdict="PASS",
        evidence=f"no `={formula_token}(...)` formulas detected in dashboard",
        rule=rule, category="csr_foundation",
        affected_sheets=["dashboard"],
    )


def check_F_02(workbook: Any, project_slug: str, *,
               formula_workbook: Any = None, **_) -> dict:
    return _check_no_excel_formula(workbook, project_slug,
                                    id_="F-02", formula_token="COUNTIF",
                                    formula_workbook=formula_workbook)


def check_F_03(workbook: Any, project_slug: str, *,
               formula_workbook: Any = None, **_) -> dict:
    return _check_no_excel_formula(workbook, project_slug,
                                    id_="F-03", formula_token="SUMIF",
                                    formula_workbook=formula_workbook)


def check_F_04(workbook: Any, project_slug: str, *,
               formula_workbook: Any = None, **_) -> dict:
    return _check_no_excel_formula(workbook, project_slug,
                                    id_="F-04", formula_token="AVERAGEIF",
                                    formula_workbook=formula_workbook)


def check_F_05(workbook: Any, project_slug: str, **_) -> dict:
    """schema_version field per-sheet present: every present sheet has a
    header row whose column count matches the master-excel schema's
    required_columns.

    Phase 14 W3-W2-C-a: header row index resolved via schema authority
    (`_resolve_header_row`) — fallback to row 1 for legacy/synthetic
    workbooks. Eliminates the 4 mechanical header-parse FAILs
    (F-01/F-05/F-17/F-18) seen on W1 bootstrap workspace master.xlsx where
    sheets use header_row=3/4/5 per schema."""
    rule = "every sheet present in workbook matches its master-excel.schema header column count"
    schema = _load_master_excel_schema()
    sheets_def = schema.get("sheets", {})
    bad: list[str] = []
    checked = 0
    for sheet_name in workbook.sheetnames:
        sdef = sheets_def.get(sheet_name)
        if not sdef:
            continue  # unknown sheet — outside this check's scope
        required = sdef.get("required_columns")
        if not required:
            continue  # dashboard uses required_cells, not required_columns
        ws = workbook[sheet_name]
        header_row_idx = _resolve_header_row(workbook, sheet_name)
        try:
            header = next(ws.iter_rows(
                min_row=header_row_idx, max_row=header_row_idx,
                values_only=True))
        except StopIteration:
            bad.append(f"{sheet_name}: no header row at row {header_row_idx}")
            continue
        if header is None:
            bad.append(f"{sheet_name}: no header row at row {header_row_idx}")
            continue
        n_header = sum(1 for h in header if h is not None)
        if n_header != len(required):
            bad.append(f"{sheet_name}: header at row {header_row_idx} has "
                       f"{n_header} cols, schema requires {len(required)}")
        checked += 1
    if checked == 0:
        return _make_result(
            id_="F-05", severity="CRITICAL", verdict="SKIP",
            evidence="no schema-known sheets found (pilot wb sparse)",
            rule=rule, category="schema_validation",
        )
    if bad:
        return _make_result(
            id_="F-05", severity="CRITICAL", verdict="FAIL",
            evidence=f"{len(bad)}/{checked} sheets fail schema column count",
            rule=rule, category="schema_validation",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-05", severity="CRITICAL", verdict="PASS",
        evidence=f"all {checked} schema-known sheets match column count",
        rule=rule, category="schema_validation",
    )


# ---------------------------------------------------------------------------
# HIGH (14)
# ---------------------------------------------------------------------------

def check_F_08(workbook: Any, project_slug: str, **_) -> dict:
    """quick_wins.url ⊆ (crawl_sitemap.url ∪ gsc_performance.url).

    F2 flag: pilot wb is sparse (sf-import has not yet populated
    crawl_sitemap; gsc_performance also absent). When BOTH reference
    sheets are missing we promote this from RED to AMBER via the
    manual_triage routing. Wave 2 closeout will re-evaluate after
    sf-import (W-R) completes.
    """
    rule = "quick_wins.url ⊆ (crawl_sitemap.url ∪ gsc_performance.url)"
    if not _has_sheet(workbook, "quick_wins"):
        return _make_result(
            id_="F-08", severity="HIGH", verdict="SKIP",
            evidence="quick_wins sheet missing — nothing to check",
            rule=rule, category="csr_foundation",
            affected_sheets=["quick_wins"],
        )
    has_sitemap = _has_sheet(workbook, "crawl_sitemap")
    has_perf = _has_sheet(workbook, "gsc_performance")
    # F2 flag (sparse-pilot tolerance): if EITHER reference sheet is
    # missing, F-08 cannot be authoritatively evaluated. Route through
    # manual_triage so the aggregator surfaces AMBER (not RED) and the
    # Wave 2 closeout (sf-import + gsc_performance loader complete) can
    # re-run drift-check for an authoritative GREEN/RED.
    pilot_sparse = not (has_sitemap and has_perf)
    qw_rows = _iter_rows_as_dicts(workbook, "quick_wins")
    qw_urls = {_normalize_url(r.get("url")) for r in qw_rows
               if r.get("url") is not None}
    qw_urls.discard("")
    ref_urls: set[str] = set()
    if has_sitemap:
        for r in _iter_rows_as_dicts(workbook, "crawl_sitemap"):
            # crawl_sitemap layout uses "value" col for URL per schema (cat/metric/value).
            for key in ("url", "value"):
                v = r.get(key)
                if v is not None:
                    n = _normalize_url(v)
                    if n:
                        ref_urls.add(n)
                    break
    if has_perf:
        for r in _iter_rows_as_dicts(workbook, "gsc_performance"):
            v = r.get("url")
            if v is not None:
                n = _normalize_url(v)
                if n:
                    ref_urls.add(n)
    orphans = sorted(qw_urls - ref_urls)
    if orphans:
        evidence = (
            f"{len(orphans)} quick_wins URLs not in "
            f"(crawl_sitemap ∪ gsc_performance)"
        )
        if pilot_sparse:
            evidence += (
                " — pilot wb sparse (has_sitemap="
                f"{has_sitemap}, has_perf={has_perf}); routed to "
                "AMBER via manual_triage (F2 flag, re-eval after sf-import)"
            )
        return _make_result(
            id_="F-08", severity="HIGH", verdict="FAIL",
            evidence=evidence,
            rule=rule, category="csr_foundation",
            affected_sheets=["quick_wins", "crawl_sitemap", "gsc_performance"],
            manual_triage=pilot_sparse,
            sample_violations=orphans,
            affected_rows=len(orphans),
        )
    return _make_result(
        id_="F-08", severity="HIGH", verdict="PASS",
        evidence=f"all {len(qw_urls)} quick_wins URLs covered by reference sheets",
        rule=rule, category="csr_foundation",
        affected_sheets=["quick_wins", "crawl_sitemap", "gsc_performance"],
    )


def check_F_09(workbook: Any, project_slug: str, **_) -> dict:
    """master_task.task_id unique."""
    rule = "master_task.task_id is unique across all rows"
    if not _has_sheet(workbook, "master_task"):
        return _make_result(
            id_="F-09", severity="HIGH", verdict="SKIP",
            evidence="master_task sheet missing",
            rule=rule, category="csr_foundation",
            affected_sheets=["master_task"],
        )
    rows = _iter_rows_as_dicts(workbook, "master_task")
    seen: dict[str, int] = {}
    dups: list[str] = []
    for r in rows:
        tid = r.get("task_id")
        if tid is None:
            continue
        s = str(tid)
        seen[s] = seen.get(s, 0) + 1
    for k, v in seen.items():
        if v > 1:
            dups.append(f"{k}×{v}")
    if dups:
        return _make_result(
            id_="F-09", severity="HIGH", verdict="FAIL",
            evidence=f"{len(dups)} duplicate task_id values",
            rule=rule, category="csr_foundation",
            affected_sheets=["master_task"],
            sample_violations=dups,
            affected_rows=len(dups),
        )
    return _make_result(
        id_="F-09", severity="HIGH", verdict="PASS",
        evidence=f"{len(seen)} unique task_id values across {len(rows)} rows",
        rule=rule, category="csr_foundation",
        affected_sheets=["master_task"],
    )


def check_F_10(workbook: Any, project_slug: str, **_) -> dict:
    """quick_wins.url D-03 normalize idempotent."""
    rule = "url_normalizer(x) == x for every quick_wins.url (D-03 idempotent)"
    if not _has_sheet(workbook, "quick_wins"):
        return _make_result(
            id_="F-10", severity="HIGH", verdict="SKIP",
            evidence="quick_wins sheet missing",
            rule=rule, category="url_normalization",
            affected_sheets=["quick_wins"],
        )
    rows = _iter_rows_as_dicts(workbook, "quick_wins")
    drift: list[str] = []
    n_total = 0
    for r in rows:
        u = r.get("url")
        if not isinstance(u, str) or not u:
            continue
        n_total += 1
        n1 = _normalize_url(u)
        n2 = _normalize_url(n1)
        if u != n1 or n1 != n2:
            drift.append(f"{u} -> {n1}")
    if drift:
        return _make_result(
            id_="F-10", severity="HIGH", verdict="FAIL",
            evidence=f"{len(drift)} quick_wins URLs not idempotent under D-03",
            rule=rule, category="url_normalization",
            affected_sheets=["quick_wins"],
            sample_violations=drift,
            affected_rows=len(drift),
        )
    return _make_result(
        id_="F-10", severity="HIGH", verdict="PASS",
        evidence=f"all {n_total} quick_wins URLs are D-03 idempotent",
        rule=rule, category="url_normalization",
        affected_sheets=["quick_wins"],
    )


def check_F_11(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """workflow-run schema_version "1.0" across all _state/workflows/*.json."""
    rule = 'every _state/workflows/*.json has schema_version == "1.0"'
    pdir = _project_dir(project_slug, workspace_root)
    wf_dir = pdir / "_state" / "workflows"
    if not wf_dir.is_dir():
        return _make_result(
            id_="F-11", severity="HIGH", verdict="SKIP",
            evidence="_state/workflows/ directory missing — no workflow runs yet",
            rule=rule, category="schema_validation",
        )
    bad: list[str] = []
    n = 0
    for p in sorted(wf_dir.glob("*.json")):
        n += 1
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            bad.append(f"{p.name}: unreadable")
            continue
        sv = obj.get("schema_version")
        if sv != _WORKFLOW_SCHEMA_VERSION:
            bad.append(f"{p.name}: schema_version={sv!r}")
    if n == 0:
        return _make_result(
            id_="F-11", severity="HIGH", verdict="SKIP",
            evidence="no workflow run files present",
            rule=rule, category="schema_validation",
        )
    if bad:
        return _make_result(
            id_="F-11", severity="HIGH", verdict="FAIL",
            evidence=f"{len(bad)}/{n} workflow runs have wrong schema_version",
            rule=rule, category="schema_validation",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-11", severity="HIGH", verdict="PASS",
        evidence=f"all {n} workflow runs have schema_version 1.0",
        rule=rule, category="schema_validation",
    )


def _write_events_snapshot(snap_path: Path, lines: int) -> None:
    """Atomically record the events.jsonl high-water line count (F-12 baseline).

    Written via tmp + replace so a crash can never leave a half-written sidecar.
    The snapshot is a derived governance baseline, not the SSoT — events.jsonl
    itself remains the single source of truth.
    """
    tmp = snap_path.with_suffix(snap_path.suffix + ".tmp")
    tmp.write_text(json.dumps({"lines": int(lines)}), encoding="utf-8")
    tmp.replace(snap_path)


def check_F_12(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """events.jsonl append-only — compare the current line count against a
    monotonic high-water baseline in `_state/events.snapshot.json`, FAIL on a
    shrink, then record/advance the baseline for the next run."""
    rule = "events.jsonl line count is monotonically non-decreasing"
    pdir = _project_dir(project_slug, workspace_root)
    events_path = pdir / "_state" / "events.jsonl"
    if not events_path.exists():
        return _make_result(
            id_="F-12", severity="HIGH", verdict="SKIP",
            evidence="events.jsonl missing — no events emitted yet",
            rule=rule, category="row_count_integrity",
        )
    cur_lines = sum(1 for _ in events_path.open("r", encoding="utf-8")
                    if _.strip())
    snap_path = pdir / "_state" / "events.snapshot.json"
    prev_lines = 0
    if snap_path.exists():
        try:
            prev_lines = int(json.loads(snap_path.read_text("utf-8")).get("lines", 0))
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            prev_lines = 0
    # Record/advance the high-water baseline so the guard is not vacuous and a
    # later shrink stays detectable. Until this producer existed (deep-audit
    # 2026-06-04) nothing ever wrote the sidecar, so prev_lines was permanently
    # 0 and the FAIL branch was unreachable. High-water (max) means a shrink
    # keeps FAILing on subsequent runs until the rows are restored.
    _write_events_snapshot(snap_path, max(prev_lines, cur_lines))
    if cur_lines < prev_lines:
        return _make_result(
            id_="F-12", severity="HIGH", verdict="FAIL",
            evidence=f"events.jsonl shrank: {prev_lines} → {cur_lines}",
            rule=rule, category="row_count_integrity",
            affected_rows=prev_lines - cur_lines,
        )
    return _make_result(
        id_="F-12", severity="HIGH", verdict="PASS",
        evidence=f"events.jsonl line count {cur_lines} ≥ snapshot {prev_lines}",
        rule=rule, category="row_count_integrity",
    )


def check_F_13(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """provenance.run_id integer (per events.schema)."""
    rule = "every event_kind=provenance row has run_id of integer type"
    pdir = _project_dir(project_slug, workspace_root)
    events_path = pdir / "_state" / "events.jsonl"
    if not events_path.exists():
        return _make_result(
            id_="F-13", severity="HIGH", verdict="SKIP",
            evidence="events.jsonl missing",
            rule=rule, category="schema_validation",
        )
    bad: list[str] = []
    n_prov = 0
    for line in events_path.open("r", encoding="utf-8"):
        if not line.strip():
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if obj.get("event_kind") != "provenance":
            continue
        n_prov += 1
        rid = obj.get("run_id")
        if not isinstance(rid, int) or isinstance(rid, bool):
            bad.append(f"event_id={obj.get('event_id')!r} run_id={rid!r}")
    if bad:
        return _make_result(
            id_="F-13", severity="HIGH", verdict="FAIL",
            evidence=f"{len(bad)}/{n_prov} provenance events have non-int run_id",
            rule=rule, category="schema_validation",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-13", severity="HIGH", verdict="PASS",
        evidence=f"all {n_prov} provenance events have integer run_id",
        rule=rule, category="schema_validation",
    )


def check_F_14(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """workflow.workflow_run_id pattern (per events.schema)."""
    rule = "every event_kind=workflow row has workflow_run_id matching {slug}-YYYY-MM-DD-{hash4}"
    pdir = _project_dir(project_slug, workspace_root)
    events_path = pdir / "_state" / "events.jsonl"
    if not events_path.exists():
        return _make_result(
            id_="F-14", severity="HIGH", verdict="SKIP",
            evidence="events.jsonl missing",
            rule=rule, category="schema_validation",
        )
    bad: list[str] = []
    n_wf = 0
    for line in events_path.open("r", encoding="utf-8"):
        if not line.strip():
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if obj.get("event_kind") != "workflow":
            continue
        n_wf += 1
        wri = obj.get("workflow_run_id")
        if not isinstance(wri, str) or not _WORKFLOW_RUN_ID_PATTERN.match(wri):
            bad.append(f"event_id={obj.get('event_id')!r} workflow_run_id={wri!r}")
    if bad:
        return _make_result(
            id_="F-14", severity="HIGH", verdict="FAIL",
            evidence=f"{len(bad)}/{n_wf} workflow events have malformed workflow_run_id",
            rule=rule, category="schema_validation",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-14", severity="HIGH", verdict="PASS",
        evidence=f"all {n_wf} workflow events have well-formed workflow_run_id",
        rule=rule, category="schema_validation",
    )


def check_F_15(workbook: Any, project_slug: str, **_) -> dict:
    """Manual triage placeholder — populates manual_review_required[]
    when cannibalization sheet has rows AND any AMBER routing condition
    is detected. Always classified as HIGH+manual_triage so the
    aggregator routes to AMBER (NOT RED). Per §9.3 F-15 requires human
    decision (intent-matrix overlap)."""
    rule = "cannibalization intent çakışması intent-matrix.json ile tutarlı (manual triage)"
    if not _has_sheet(workbook, "cannibalization"):
        return _make_result(
            id_="F-15", severity="HIGH", verdict="SKIP",
            evidence="cannibalization sheet missing — no manual triage required",
            rule=rule, category="csr_foundation",
            affected_sheets=["cannibalization"],
            manual_triage=True,
        )
    rows = _iter_rows_as_dicts(workbook, "cannibalization")
    if not rows:
        return _make_result(
            id_="F-15", severity="HIGH", verdict="PASS",
            evidence="cannibalization sheet present but empty — nothing to triage",
            rule=rule, category="csr_foundation",
            affected_sheets=["cannibalization"],
            manual_triage=True,
        )
    samples = []
    for r in rows[:_SAMPLE_CAP]:
        pair = r.get("conflict_pair") or r.get("pair") or "?"
        samples.append(str(pair))
    return _make_result(
        id_="F-15", severity="HIGH", verdict="FAIL",
        evidence=f"{len(rows)} cannibalization rows require manual intent-matrix triage",
        rule=rule, category="csr_foundation",
        affected_sheets=["cannibalization"],
        manual_triage=True,
        sample_violations=samples,
        affected_rows=len(rows),
    )


def _extract_url_from_action_field(text: str) -> str:
    """Extract URL from opportunity.assigned_url_action field.

    Supports two formats (Q-V1.2-OPP-COVERAGE-01 defensive parsing,
    schema-first override 17'inci uygulama Phase B post-closeout):
    - Canonical "url | action" (Phase 8 quickwins_transform.py paterni)
    - Freeform "Optimize <url> for query 'X'" (manual/non-canonical
      content drift; workspace gerçek master.xlsx 2026-05-06 state)

    Returns first matched URL (post-strip), or empty string.
    """
    if not isinstance(text, str) or not text:
        return ""
    # Canonical first: split by '|', validate URL prefix
    if "|" in text:
        head = text.split("|", 1)[0].strip()
        if head.startswith(("http://", "https://")):
            return head
    # Freeform fallback: regex extract first URL in text
    m = re.search(r'https?://[^\s\'"|]+', text)
    return m.group(0) if m else ""


def check_F_16(workbook: Any, project_slug: str, **_) -> dict:
    """quick_wins.url ⊆ opportunity URL set (cross-sheet foreign key).

    opportunity.assigned_url_action defensive parse: supports canonical
    "url | action" (quickwins_transform paterni) + freeform "Optimize
    <url> for query 'X'" (workspace data drift Q-V1.2-OPP-COVERAGE-01,
    Phase B post-closeout schema-first override 17'inci uygulama).
    """
    rule = "quick_wins.url ⊆ opportunity.assigned_url_action URL set"
    if not (_has_sheet(workbook, "quick_wins") and _has_sheet(workbook, "opportunity")):
        return _make_result(
            id_="F-16", severity="HIGH", verdict="SKIP",
            evidence="quick_wins or opportunity sheet missing",
            rule=rule, category="csr_foundation",
            affected_sheets=["quick_wins", "opportunity"],
        )
    qw_urls = set()
    for r in _iter_rows_as_dicts(workbook, "quick_wins"):
        u = r.get("url")
        if isinstance(u, str) and u:
            qw_urls.add(_normalize_url(u))
    qw_urls.discard("")

    opp_urls = set()
    for r in _iter_rows_as_dicts(workbook, "opportunity"):
        au = r.get("assigned_url_action")
        if isinstance(au, str) and au:
            url_part = _extract_url_from_action_field(au)
            if url_part:
                opp_urls.add(_normalize_url(url_part))
    opp_urls.discard("")

    orphans = sorted(qw_urls - opp_urls)
    if orphans:
        return _make_result(
            id_="F-16", severity="HIGH", verdict="FAIL",
            evidence=f"{len(orphans)} quick_wins URLs not present in opportunity",
            rule=rule, category="csr_foundation",
            affected_sheets=["quick_wins", "opportunity"],
            sample_violations=orphans,
            affected_rows=len(orphans),
        )
    return _make_result(
        id_="F-16", severity="HIGH", verdict="PASS",
        evidence=f"all {len(qw_urls)} quick_wins URLs covered by opportunity",
        rule=rule, category="csr_foundation",
        affected_sheets=["quick_wins", "opportunity"],
    )


def check_F_23(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """SF MCP cross-sheet invariant (v1.8 Phase 4; v1.9 Phase 5 enhancement):
    if any _state/workflows/*.json shows a completed sf-crawl-orchestrator
    run, the repo-root mcp-tool-registry.json MUST list 'sf' under servers.

    Detection logic:
      1. Walk projects/{slug}/_state/workflows/*.json.
      2. Filter to skill=='sf-crawl-orchestrator' AND status=='done'.
      3. Read engine repo-root mcp-tool-registry.json (PRIMARY); check 'sf'
         key in servers.
      4. Non-empty evidence + missing 'sf' → FAIL HIGH (RED).
      5. Both registries absent (fresh engine checkout) → SKIP (AMBER).

    v1.9 Phase 5 (Q-V1.9-03 — ADDITIVE workspace-aware fallback, NOT a
    breaking refactor): when a workspace-level mcp-tool-registry.json exists
    at {workspace_root}/mcp-tool-registry.json (PSEO_WORKSPACE_ROOT
    deployments, spec FE-5), it is checked IN ADDITION to the engine
    registry. The invariant FAILs if EITHER existing registry is missing
    'sf' (broad enforcement — orphan crawl evidence OR missing inventory both
    surface). Engine-repo-only deployments (the default — no workspace
    registry) behave identically to v1.8.
    """
    rule = (
        "if any project's _state/workflows/ has a completed "
        "sf-crawl-orchestrator run, mcp-tool-registry.json MUST list 'sf' "
        "under servers"
    )
    pdir = _project_dir(project_slug, workspace_root)
    wf_dir = pdir / "_state" / "workflows"
    if not wf_dir.is_dir():
        return _make_result(
            id_="F-23", severity="HIGH", verdict="SKIP",
            evidence="_state/workflows/ missing — no SF crawl evidence yet",
            rule=rule, category="csr_mcp",
        )
    sf_runs: list[str] = []
    for p in sorted(wf_dir.glob("*.json")):
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if obj.get("skill") != "sf-crawl-orchestrator":
            continue
        if obj.get("status") != "done":
            continue
        rid = obj.get("run_id") or p.stem
        sf_runs.append(str(rid))
    if not sf_runs:
        return _make_result(
            id_="F-23", severity="HIGH", verdict="PASS",
            evidence="no completed sf-crawl-orchestrator runs found — invariant vacuous",
            rule=rule, category="csr_mcp",
        )
    # v1.9 Phase 5 (Q-V1.9-03): ADDITIVE dual-registry check. The engine
    # repo-root registry stays the PRIMARY source; a workspace-level
    # mcp-tool-registry.json (PSEO_WORKSPACE_ROOT deployments, spec FE-5) is
    # checked IN ADDITION when present. The workspace root is already
    # resolvable here (else _project_dir above would have raised).
    engine_registry_path = _REPO_ROOT / "mcp-tool-registry.json"
    workspace_registry_path = (
        _resolve_workspace_root(workspace_root) / "mcp-tool-registry.json"
    )
    candidate_registries = [("engine", engine_registry_path)]
    if (workspace_registry_path != engine_registry_path
            and workspace_registry_path.exists()):
        candidate_registries.append(("workspace", workspace_registry_path))

    present: list[tuple[str, dict]] = []
    for label, path in candidate_registries:
        if not path.exists():
            continue
        try:
            present.append(
                (label, json.loads(path.read_text(encoding="utf-8")))
            )
        except (OSError, json.JSONDecodeError) as exc:
            return _make_result(
                id_="F-23", severity="HIGH", verdict="FAIL",
                evidence=f"{label} mcp-tool-registry.json unparseable: {exc}",
                rule=rule, category="csr_mcp",
                sample_violations=sf_runs[:_SAMPLE_CAP],
                affected_rows=len(sf_runs),
            )

    if not present:
        # Neither engine nor workspace registry present — engine state
        # ambiguous (preserves the v1.8 SKIP/AMBER path verbatim).
        return _make_result(
            id_="F-23", severity="HIGH", verdict="SKIP",
            evidence=(
                f"mcp-tool-registry.json missing at {engine_registry_path} "
                f"but {len(sf_runs)} sf-crawl-orchestrator run(s) present — "
                "engine state ambiguous, surfaces AMBER"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=sf_runs[:_SAMPLE_CAP],
        )

    missing = [
        label for label, reg in present
        if "sf" not in (reg.get("servers") or {})
    ]
    if missing:
        return _make_result(
            id_="F-23", severity="HIGH", verdict="FAIL",
            evidence=(
                f"{len(sf_runs)} completed sf-crawl-orchestrator run(s) but "
                f"{', '.join(missing)} mcp-tool-registry.json servers "
                "missing 'sf' key"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=sf_runs[:_SAMPLE_CAP],
            affected_rows=len(sf_runs),
        )
    return _make_result(
        id_="F-23", severity="HIGH", verdict="PASS",
        evidence=(
            f"{len(sf_runs)} sf-crawl-orchestrator run(s) backed by "
            f"{', '.join(label for label, _ in present)} "
            "mcp-tool-registry.json servers['sf'] entry"
        ),
        rule=rule, category="csr_mcp",
    )


# D-SF-02 (naming.md) alignment: .mcp.json carries the legacy capitalized
# transport label "ScraplingServer" while mcp-tool-registry.json uses the
# canonical lowercase server key "scrapling". Normalize the .mcp.json side to
# the registry's canonical key before the set comparison. This is an EXPLICIT
# alias map, NOT a blanket .lower() — a generic casefold would yield
# "scraplingserver" != "scrapling" and raise a false-positive FAIL (spec R3).
_MCP_JSON_KEY_ALIASES = {"ScraplingServer": "scrapling"}


def check_F_24(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """`.mcp.json` ↔ `mcp-tool-registry.json` servers-key sync (v1.9 Phase 2):
    the set of `.mcp.json` mcpServers keys MUST equal the set of
    mcp-tool-registry.json servers keys, after the ScraplingServer→scrapling
    alias (D-SF-02). Engine-level invariant — ignores workspace_root; reads
    the two repo-root engine files only (never mutates them, F-16 safe).

    Detection logic:
      1. Read engine repo-root .mcp.json + mcp-tool-registry.json.
      2. Either file missing → SKIP (engine state ambiguous → AMBER).
      3. Either file unparseable → FAIL HIGH.
      4. Normalize .mcp.json keys via _MCP_JSON_KEY_ALIASES; compare sets.
      5. Server in .mcp.json but not registry (orphan transport) OR in
         registry but not .mcp.json (orphan inventory) → FAIL HIGH (RED).
      6. Sets equal → PASS.
    """
    rule = (
        ".mcp.json mcpServers keys MUST equal mcp-tool-registry.json servers "
        "keys (sets comparison; ScraplingServer↔scrapling case-fold per "
        "D-SF-02 alignment)"
    )
    mcp_path = _REPO_ROOT / ".mcp.json"
    registry_path = _REPO_ROOT / "mcp-tool-registry.json"
    missing = [p.name for p in (mcp_path, registry_path) if not p.exists()]
    if missing:
        return _make_result(
            id_="F-24", severity="HIGH", verdict="SKIP",
            evidence=(
                f"{', '.join(missing)} missing at engine repo root — "
                "engine state ambiguous, surfaces AMBER"
            ),
            rule=rule, category="csr_mcp",
        )
    try:
        mcp_config = json.loads(mcp_path.read_text(encoding="utf-8"))
        registry = json.loads(registry_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return _make_result(
            id_="F-24", severity="HIGH", verdict="FAIL",
            evidence=f".mcp.json or mcp-tool-registry.json unparseable: {exc}",
            rule=rule, category="csr_mcp",
        )
    # ScraplingServer→scrapling normalization applied to the transport side.
    transport_keys = {
        _MCP_JSON_KEY_ALIASES.get(k, k)
        for k in (mcp_config.get("mcpServers") or {})
    }
    inventory_keys = set(registry.get("servers") or {})
    orphan_transport = sorted(transport_keys - inventory_keys)
    orphan_inventory = sorted(inventory_keys - transport_keys)
    if orphan_transport or orphan_inventory:
        violations = (
            [f"orphan_transport:{k}" for k in orphan_transport]
            + [f"orphan_inventory:{k}" for k in orphan_inventory]
        )
        return _make_result(
            id_="F-24", severity="HIGH", verdict="FAIL",
            evidence=(
                ".mcp.json↔mcp-tool-registry.json server-key delta "
                "(post ScraplingServer→scrapling normalization): "
                f"orphan_transport={orphan_transport or '∅'} "
                f"orphan_inventory={orphan_inventory or '∅'}"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=violations,
            affected_rows=len(violations),
        )
    return _make_result(
        id_="F-24", severity="HIGH", verdict="PASS",
        evidence=(
            f"{len(transport_keys)} server keys aligned across .mcp.json + "
            f"mcp-tool-registry.json: {sorted(transport_keys)}"
        ),
        rule=rule, category="csr_mcp",
    )


# Migration 0005 (v1.4→v1.5) is what introduces the project.config `sf` block,
# so sf.mcp.enabled=true is only coherent on schema_version >= 1.5. The
# threshold is an int TUPLE, compared against _version_tuple(schema_version) —
# NOT a lexicographic string compare. For 1.4/1.5 the two happen to agree, but
# "1.10" >= "1.5" is FALSE under string ordering and TRUE under tuple ordering;
# we want the latter (D-V1.9-10).
_SF_MCP_MIN_SCHEMA_VERSION = (1, 5)


def _version_tuple(value: Any) -> tuple[int, ...]:
    """Parse a dotted version string into a comparable int tuple.
    '1.5' → (1, 5); '1.10' → (1, 10). Non-int segments degrade to 0 so a
    malformed version string can never raise inside the comparison."""
    parts: list[int] = []
    for seg in str(value).split("."):
        try:
            parts.append(int(seg))
        except (TypeError, ValueError):
            parts.append(0)
    return tuple(parts)


def check_F_25(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """`sf.mcp.enabled=true` requires `schema_version >= 1.5` (v1.9 Phase 3):
    Migration 0005 (v1.4→v1.5) is what adds the `sf` block, so a config that
    turns SF MCP on while still declaring schema_version 1.4 is incoherent.
    Per-project invariant — reads projects/{slug}/project.config.json (never the
    engine repo-root files; F-16 safe, .mcp.json untouched).

    Detection logic (D-V1.9-10):
      1. project.config.json missing → SKIP (state ambiguous → AMBER).
      2. schema_version field absent → SKIP (coupling cannot be evaluated).
      3. project.config.json unparseable → FAIL HIGH.
      4. sf.mcp.enabled not true → PASS (default behavior valid on any version).
      5. enabled=true AND schema_version >= 1.5 → PASS.
      6. enabled=true AND schema_version < 1.5 → FAIL HIGH (RED). Migration 0005
         makes this impossible in practice, but the defensive check catches a
         manual misconfiguration that bypassed the migration.
    """
    rule = (
        "if project.config.sf.mcp.enabled=true, project.config.schema_version "
        "MUST be >= '1.5' (Migration 0005 prerequisite)"
    )
    pdir = _project_dir(project_slug, workspace_root)
    cfg_path = pdir / "project.config.json"
    if not cfg_path.exists():
        return _make_result(
            id_="F-25", severity="HIGH", verdict="SKIP",
            evidence=f"{cfg_path.name} missing — no SF MCP coupling to check",
            rule=rule, category="csr_mcp",
        )
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return _make_result(
            id_="F-25", severity="HIGH", verdict="FAIL",
            evidence=f"project.config.json unparseable: {exc}",
            rule=rule, category="csr_mcp",
        )
    schema_version = cfg.get("schema_version")
    if not schema_version:
        return _make_result(
            id_="F-25", severity="HIGH", verdict="SKIP",
            evidence=(
                "project.config.json has no schema_version field — "
                "SF MCP coupling cannot be evaluated"
            ),
            rule=rule, category="csr_mcp",
        )
    mcp_block = (cfg.get("sf") or {}).get("mcp") or {}
    enabled = bool(mcp_block.get("enabled"))
    if not enabled:
        return _make_result(
            id_="F-25", severity="HIGH", verdict="PASS",
            evidence=(
                f"sf.mcp.enabled false/absent (schema_version="
                f"{schema_version!r}) — default behavior valid on any version"
            ),
            rule=rule, category="csr_mcp",
        )
    if _version_tuple(schema_version) >= _SF_MCP_MIN_SCHEMA_VERSION:
        return _make_result(
            id_="F-25", severity="HIGH", verdict="PASS",
            evidence=(
                f"sf.mcp.enabled=true backed by schema_version="
                f"{schema_version!r} (>= 1.5 Migration 0005 prerequisite)"
            ),
            rule=rule, category="csr_mcp",
        )
    return _make_result(
        id_="F-25", severity="HIGH", verdict="FAIL",
        evidence=(
            f"sf.mcp.enabled=true but schema_version={schema_version!r} < 1.5 — "
            "Migration 0005 prerequisite unmet (sf block requires v1.5)"
        ),
        rule=rule, category="csr_mcp",
        sample_violations=[
            f"schema_version={schema_version}<1.5 with sf.mcp.enabled=true"
        ],
        affected_rows=1,
    )


def check_F_17(workbook: Any, project_slug: str, **_) -> dict:
    """severity column ⊆ severityEnum 4-value (LOW/MEDIUM/HIGH/CRITICAL).

    Scans every sheet that the master-excel schema marks as carrying a
    severity-typed column (priority/level/impact with $ref severityEnum).
    """
    rule = "every severity-typed cell ∈ severityEnum {LOW,MEDIUM,HIGH,CRITICAL}"
    schema = _load_master_excel_schema()
    sheets_def = schema.get("sheets", {})
    severity_columns: list[tuple[str, str]] = []  # (sheet, col_name)
    for sheet, sdef in sheets_def.items():
        for col in sdef.get("required_columns", []):
            ref = col.get("ref")
            if ref and ref.endswith("/severityEnum"):
                severity_columns.append((sheet, col["name"]))

    bad: list[str] = []
    n_checked = 0
    affected_sheets: set[str] = set()
    for sheet, col_name in severity_columns:
        if not _has_sheet(workbook, sheet):
            continue
        for r in _iter_rows_as_dicts(workbook, sheet):
            v = r.get(col_name)
            if v is None or v == "":
                continue
            n_checked += 1
            if str(v).upper() not in _SEVERITY_ENUM_4:
                bad.append(f"{sheet}.{col_name}={v!r}")
                affected_sheets.add(sheet)
    if n_checked == 0:
        return _make_result(
            id_="F-17", severity="HIGH", verdict="SKIP",
            evidence="no severity-typed cells found across present sheets",
            rule=rule, category="csr_data",
        )
    if bad:
        return _make_result(
            id_="F-17", severity="HIGH", verdict="FAIL",
            evidence=f"{len(bad)}/{n_checked} severity cells outside 4-value enum",
            rule=rule, category="csr_data",
            affected_sheets=sorted(affected_sheets),
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-17", severity="HIGH", verdict="PASS",
        evidence=f"all {n_checked} severity cells within 4-value enum",
        rule=rule, category="csr_data",
    )


# ---------------------------------------------------------------------------
# MEDIUM (6)
# ---------------------------------------------------------------------------

def check_F_18(workbook: Any, project_slug: str, **_) -> dict:
    """master_task.created_at (col K) parses as ISO 8601."""
    rule = "every master_task.created_date is ISO 8601 parseable"
    if not _has_sheet(workbook, "master_task"):
        return _make_result(
            id_="F-18", severity="MEDIUM", verdict="SKIP",
            evidence="master_task sheet missing",
            rule=rule, category="csr_data",
            affected_sheets=["master_task"],
        )
    rows = _iter_rows_as_dicts(workbook, "master_task")
    bad: list[str] = []
    n = 0
    for r in rows:
        v = r.get("created_date")
        if v is None or v == "":
            continue
        n += 1
        if isinstance(v, datetime):
            continue  # openpyxl already parsed it
        try:
            datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        except (ValueError, TypeError):
            bad.append(f"row_value={v!r}")
    if bad:
        return _make_result(
            id_="F-18", severity="MEDIUM", verdict="FAIL",
            evidence=f"{len(bad)}/{n} master_task.created_date values not ISO 8601",
            rule=rule, category="csr_data",
            affected_sheets=["master_task"],
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-18", severity="MEDIUM", verdict="PASS",
        evidence=f"all {n} master_task.created_date values parse as ISO 8601",
        rule=rule, category="csr_data",
        affected_sheets=["master_task"],
    )


def check_F_19(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """project.config.json carries the canonical locale + market keys per
    schema 1.3 (ADR-030 / Migration 0003 / project-config.schema.json):
    nested ``language.content_locale`` (IETF BCP 47) + root ``market``.

    Wave 2 fix (Q-WAVE1-F19-VALIDATOR-01): pre-2026-05-06 implementation
    read root-level ``locale``/``defaults.locale`` (schema 1.2 paths);
    schema 1.3 made root ``locale`` forbidden via
    ``additionalProperties: false`` and moved the value to
    ``language.content_locale``. Only the 1.3 canonical path is read.
    """
    rule = (
        "project.config.json declares language.content_locale (nested) "
        "+ market (root) per schema 1.3"
    )
    pdir = _project_dir(project_slug, workspace_root)
    cfg_path = pdir / "project.config.json"
    if not cfg_path.exists():
        return _make_result(
            id_="F-19", severity="MEDIUM", verdict="SKIP",
            evidence=f"{cfg_path.name} missing",
            rule=rule, category="schema_validation",
        )
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return _make_result(
            id_="F-19", severity="MEDIUM", verdict="FAIL",
            evidence=f"project.config.json unparseable: {exc}",
            rule=rule, category="schema_validation",
        )
    language = cfg.get("language") or {}
    has_locale = bool(language.get("content_locale"))
    has_market = bool(cfg.get("market"))
    missing = [
        k for k, ok in (
            ("language.content_locale", has_locale),
            ("market", has_market),
        )
        if not ok
    ]
    if missing:
        return _make_result(
            id_="F-19", severity="MEDIUM", verdict="FAIL",
            evidence=f"project.config.json missing keys: {missing}",
            rule=rule, category="schema_validation",
            sample_violations=missing,
        )
    return _make_result(
        id_="F-19", severity="MEDIUM", verdict="PASS",
        evidence=(
            "project.config.json declares language.content_locale + market"
        ),
        rule=rule, category="schema_validation",
    )


def check_F_20(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """events.jsonl per-line size <64 KB cap."""
    rule = "every events.jsonl line is < 64 KiB (events_writer cap)"
    pdir = _project_dir(project_slug, workspace_root)
    events_path = pdir / "_state" / "events.jsonl"
    if not events_path.exists():
        return _make_result(
            id_="F-20", severity="MEDIUM", verdict="SKIP",
            evidence="events.jsonl missing",
            rule=rule, category="row_count_integrity",
        )
    bad: list[str] = []
    n = 0
    with events_path.open("rb") as fh:
        for i, line in enumerate(fh, start=1):
            if not line.strip():
                continue
            n += 1
            if len(line) > _EVENTS_MAX_LINE_BYTES:
                bad.append(f"line {i}: {len(line)} bytes")
    if bad:
        return _make_result(
            id_="F-20", severity="MEDIUM", verdict="FAIL",
            evidence=f"{len(bad)}/{n} events.jsonl lines exceed 64 KiB cap",
            rule=rule, category="row_count_integrity",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-20", severity="MEDIUM", verdict="PASS",
        evidence=f"all {n} events.jsonl lines under 64 KiB cap",
        rule=rule, category="row_count_integrity",
    )


def check_F_21(workbook: Any, project_slug: str, **_) -> dict:
    """Every cell value <32767 chars (Excel hard limit)."""
    rule = "every cell value < 32767 chars (Excel/openpyxl hard limit)"
    bad: list[str] = []
    n_cells = 0
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            for cell in row:
                if cell is None:
                    continue
                n_cells += 1
                if isinstance(cell, str) and len(cell) >= _EXCEL_CELL_MAX_CHARS:
                    bad.append(f"{sheet_name}: {len(cell)} chars")
                    if len(bad) >= _SAMPLE_CAP:
                        break
            if len(bad) >= _SAMPLE_CAP:
                break
        if len(bad) >= _SAMPLE_CAP:
            break
    if bad:
        return _make_result(
            id_="F-21", severity="MEDIUM", verdict="FAIL",
            evidence=f"{len(bad)} cells at/over 32767 char Excel limit",
            rule=rule, category="schema_validation",
            sample_violations=bad,
            affected_rows=len(bad),
        )
    return _make_result(
        id_="F-21", severity="MEDIUM", verdict="PASS",
        evidence=f"all {n_cells} non-null cells within 32767 char limit",
        rule=rule, category="schema_validation",
    )


def check_F_22(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None, **_) -> dict:
    """backup directory FIFO 7 (transaction.py keep-7 rotation).

    auto_repair_available=True — exceeding the cap is mechanically
    fixable by deleting oldest backups.
    """
    rule = "_state/backups/master keeps at most 7 .xlsx backups (FIFO)"
    pdir = _project_dir(project_slug, workspace_root)
    backups_dir = pdir / "_state" / "backups" / "master"
    if not backups_dir.is_dir():
        return _make_result(
            id_="F-22", severity="MEDIUM", verdict="SKIP",
            evidence="backups dir missing — no writes have occurred",
            rule=rule, category="file_hash_integrity",
            auto_repair_available=True,
        )
    backups = sorted(backups_dir.glob("master-*.xlsx"))
    n = len(backups)
    if n > _BACKUP_KEEP_LIMIT:
        return _make_result(
            id_="F-22", severity="MEDIUM", verdict="FAIL",
            evidence=f"{n} backups present, FIFO cap is {_BACKUP_KEEP_LIMIT}",
            rule=rule, category="file_hash_integrity",
            auto_repair_available=True,
            affected_rows=n - _BACKUP_KEEP_LIMIT,
            sample_violations=[p.name for p in backups[:_SAMPLE_CAP]],
        )
    return _make_result(
        id_="F-22", severity="MEDIUM", verdict="PASS",
        evidence=f"{n} backups within FIFO cap {_BACKUP_KEEP_LIMIT}",
        rule=rule, category="file_hash_integrity",
        auto_repair_available=True,
    )


# Canonical local SF MCP endpoint (D-SF-14 / .mcp.json `sf` server entry). F-26
# builds its OWN client against this default rather than reading `.mcp.json`, so
# the F-16 byte-for-byte `.mcp.json` guard is never touched. Overridable via the
# injected `mcp_client` arg (tests pass a stub; never a live socket).
_SF_MCP_DEFAULT_URL = "http://127.0.0.1:11435/mcp"
# Risk R2 mitigation: drift-check must never block on SF MCP downtime, so the
# health probe (and the follow-up progress calls on the same client) use a tight
# 1s timeout and the check SKIPs rather than hangs when the probe fails.
_SF_MCP_PROBE_TIMEOUT_S = 1.0
# Workflow run statuses that leave a crawl potentially orphaned: the orchestrator
# gave up (failed) or was paused, yet the SF GUI crawl may still be running.
_SF_CRAWL_ORPHAN_STATUSES: frozenset[str] = frozenset({"paused", "failed"})


def _extract_crawl_id(run_obj: dict) -> str | None:
    """Pull the SF crawl_id out of a workflow-run JSON.

    `workflow-run.schema.json` is additionalProperties:false at root, so the
    orchestrator records the crawl_id inside a step's `output_ref` as the string
    `crawl_id=<value>` (sf-crawl-orchestrator SKILL.md trigger step). We check a
    top-level `crawl_id` first (forward-compat / convenience), then scan
    `steps[].output_ref`. Returns None when no crawl_id is recorded — such a run
    cannot be probed and is therefore NOT counted as an orphan."""
    top = run_obj.get("crawl_id")
    if isinstance(top, str) and top.strip():
        return top.strip()
    for step in run_obj.get("steps") or []:
        if not isinstance(step, dict):
            continue
        ref = step.get("output_ref")
        if isinstance(ref, str) and ref.startswith("crawl_id="):
            val = ref.split("=", 1)[1].strip()
            if val:
                return val
    return None


def _progress_is_in_progress(raw: Any) -> bool:
    """True iff a `sf_crawl_progress` result reports an in-flight crawl.

    Mirrors scripts.ingestion.sf_crawl_orchestrator.parse_progress_response's
    shape contract (flat `{"status": ...}` OR nested `{"progress": {"status":
    ...}}`; canonical tokens IN_PROGRESS / DONE / FAILED) — kept inline to avoid
    a governance→ingestion module import. NEVER raises: a malformed/empty
    payload is treated as "not in progress" so a flaky GUI response cannot
    manufacture a false orphan (AMBER-only ethos)."""
    if not isinstance(raw, dict):
        return False
    nested = raw.get("progress") if isinstance(raw.get("progress"), dict) else None
    src = nested or raw
    status = src.get("status")
    if not isinstance(status, str):
        return False
    return status.strip().upper() == "IN_PROGRESS"


def check_F_26(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None,
               mcp_client: Any = None, **_) -> dict:
    """Orphan SF GUI crawl detection (v1.9 Phase 4, MCP-aware → AMBER).

    If a project's `_state/workflows/` has an sf-crawl-orchestrator run that is
    paused/failed, BUT the SF GUI still reports that crawl IN_PROGRESS, the
    workflow state and the GUI disagree — an orphan crawl the operator should
    clean up. This is an operator hint, NOT a data-integrity break, so it is
    severity MEDIUM → AMBER (never RED) per D-V1.9-11 / spec v2.2 line 210.

    Detection logic (Risk R2 — never hang drift-check on MCP downtime):
      1. Walk projects/{slug}/_state/workflows/*.json; collect
         skill=='sf-crawl-orchestrator' AND status ∈ {paused, failed}.
      2. No such runs → PASS (vacuous; the common case, NO MCP call made).
      3. 1s health probe via SfMcpClient.health(); probe False → SKIP (MCP down;
         surfaces AMBER but never blocks, and no progress call is made).
      4. For each paused/failed run with a recorded crawl_id, call
         sf_crawl_progress(crawl_id); a result reporting IN_PROGRESS is an
         orphan. call/parse failures are swallowed (not counted as orphans).
      5. Any orphan → FAIL MEDIUM (AMBER). No orphans → PASS.

    OPTIONAL / best-effort: never reads `.mcp.json` (F-16 safe — builds its own
    client at `_SF_MCP_DEFAULT_URL`, or uses the injected `mcp_client`) and never
    lets an MCP error escalate past AMBER. `workbook` is unused (this is a
    workflow + MCP check, like F-11/F-23)."""
    rule = (
        "if a project's _state/workflows/ has a paused/failed "
        "sf-crawl-orchestrator run whose crawl_id still reports IN_PROGRESS via "
        "sf_crawl_progress, surface AMBER (workflow state vs SF GUI disagree)"
    )
    pdir = _project_dir(project_slug, workspace_root)
    wf_dir = pdir / "_state" / "workflows"
    # Collect paused/failed sf-crawl-orchestrator runs (+ their crawl_ids).
    # glob on a missing dir yields nothing → vacuous PASS below (no is_dir guard
    # needed; keeps the "nothing to reconcile" path single).
    stalled: list[tuple[str, str | None]] = []  # (run_id, crawl_id|None)
    for p in sorted(wf_dir.glob("*.json")):
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if obj.get("skill") != "sf-crawl-orchestrator":
            continue
        if obj.get("status") not in _SF_CRAWL_ORPHAN_STATUSES:
            continue
        rid = str(obj.get("run_id") or p.stem)
        stalled.append((rid, _extract_crawl_id(obj)))
    if not stalled:
        return _make_result(
            id_="F-26", severity="MEDIUM", verdict="PASS",
            evidence=(
                "no paused/failed sf-crawl-orchestrator runs — "
                "orphan-crawl invariant vacuous"
            ),
            rule=rule, category="csr_mcp",
        )
    # MCP-aware path. Build a client if one was not injected; a missing
    # sf_mcp_client dependency must SKIP, never crash drift-check.
    client = mcp_client
    if client is None:
        try:
            from scripts.util.sf_mcp_client import SfMcpClient
            client = SfMcpClient(_SF_MCP_DEFAULT_URL,
                                 timeout_seconds=_SF_MCP_PROBE_TIMEOUT_S)
        except Exception as exc:  # noqa: BLE001 — optional dep; never block
            return _make_result(
                id_="F-26", severity="MEDIUM", verdict="SKIP",
                evidence=(
                    f"SF MCP client unavailable ({type(exc).__name__}) — "
                    f"{len(stalled)} paused/failed run(s) left unverified"
                ),
                rule=rule, category="csr_mcp",
                sample_violations=[rid for rid, _ in stalled[:_SAMPLE_CAP]],
            )
    # 1s health probe BEFORE any sf_crawl_progress call (Risk R2).
    try:
        healthy = bool(client.health())
    except Exception:  # noqa: BLE001 — a flaky probe is treated as "down"
        healthy = False
    if not healthy:
        return _make_result(
            id_="F-26", severity="MEDIUM", verdict="SKIP",
            evidence=(
                "SF MCP health probe failed within 1s — "
                f"{len(stalled)} paused/failed run(s) left unverified (MCP down)"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=[rid for rid, _ in stalled[:_SAMPLE_CAP]],
        )
    # MCP responding — reconcile each stalled run against live SF GUI state.
    orphans: list[str] = []
    for rid, crawl_id in stalled:
        if not crawl_id:
            continue  # no crawl_id recorded → cannot probe → not an orphan
        try:
            raw = client.call_tool("sf_crawl_progress", crawl_id=crawl_id)
        except Exception:  # noqa: BLE001 — query failure ≠ orphan; AMBER-only
            continue
        if _progress_is_in_progress(raw):
            orphans.append(f"{rid} (crawl_id={crawl_id} still IN_PROGRESS)")
    if orphans:
        return _make_result(
            id_="F-26", severity="MEDIUM", verdict="FAIL",
            evidence=(
                f"{len(orphans)} orphan SF crawl(s): workflow paused/failed but "
                "SF GUI still reports IN_PROGRESS — operator cleanup hint (AMBER)"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=orphans,
            affected_rows=len(orphans),
        )
    return _make_result(
        id_="F-26", severity="MEDIUM", verdict="PASS",
        evidence=(
            f"{len(stalled)} paused/failed sf-crawl-orchestrator run(s) "
            "reconciled — SF GUI reports none still IN_PROGRESS"
        ),
        rule=rule, category="csr_mcp",
    )


# ---------------------------------------------------------------------------
# F-27 (AMO batch 3-gov-driftF): declared OUTWARD MCP tool ⊆ gate matcher set
# ---------------------------------------------------------------------------

# The OUTWARD MCP tools the engine KNOWS perform an irreversible/outward
# submission, in registry-key form (server__tool — what skill_mcp_usage's
# declared_tools() yields). TODAY exactly one exists: gsc__submit_sitemap, the
# GSC sitemap-submit the batch-2b gate covers. Adding a NEW outward MCP tool (a
# future Indexing-API URL_UPDATED submit, a publish tool, …) means adding it
# BOTH here AND to an outward_action_gate matcher — F-27 FAILs until the gate
# catches up.
_OUTWARD_MCP_TOOLS: frozenset[str] = frozenset({"gsc__submit_sitemap"})

# A precise outward-VERB pattern for FUTURE tools the curated set has not caught
# yet (the AMBER tripwire) — a submit/publish/ping/indexnow/url_updated verb as
# a whole underscore-delimited segment of the TOOL name.
_OUTWARD_VERB_RE = re.compile(
    r"(?:^|_)(?:submit|publish|ping|indexnow|url_updated)(?:_|$)"
)
# Read/query verbs that are NEVER outward — an explicit exclusion guaranteeing
# ZERO false-positives against the real skills (gsc__index_inspect,
# gsc__list_sitemaps, gsc__get_sitemap, …search/detect/analytics/overview/…).
_READ_VERB_RE = re.compile(
    r"(?:^|_)(?:inspect|list|get|search|detect|analytics|overview|suggestions"
    r"|ideas|keywords|volume|parsing|lighthouse)(?:_|$)"
)


def _strip_mcp_prefix(qualified: str) -> str:
    """``mcp__gsc__submit_sitemap`` → ``gsc__submit_sitemap`` (registry-key form)
    so the gate's qualified constant compares against declared_tools() keys."""
    prefix = "mcp__"
    return qualified[len(prefix):] if qualified.startswith(prefix) else qualified


def _looks_outward(registry_tool: str) -> bool:
    """True iff a registry-key tool's NAME looks like an outward submission and
    is not a read/query tool — the tripwire for an UNCLASSIFIED outward tool."""
    name = registry_tool.split("__", 1)[-1]
    if _READ_VERB_RE.search(name):
        return False
    return bool(_OUTWARD_VERB_RE.search(name))


@functools.lru_cache(maxsize=1)
def _registry_outward_tools() -> frozenset[str]:
    """Outward-looking tools as declared by mcp-tool-registry.json.

    An INDEPENDENT source of truth for what the gate must cover. Deriving the
    expectation from the gate's own constant — as the rest of F-27 does — can
    only confirm the gate is consistent with itself.
    """
    for path in (_REPO_ROOT / "mcp-tool-registry.json",):
        try:
            doc = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        out: set[str] = set()
        for cfg in (doc.get("servers") or {}).values():
            if not isinstance(cfg, dict):
                continue
            for tool in cfg.get("tools") or []:
                name = tool.get("tool_name") if isinstance(tool, dict) else tool
                if isinstance(name, str) and _looks_outward(name):
                    out.add(name)
        return frozenset(out)
    return frozenset()


def _scan_declared_outward(
    skills_root: Path,
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    """Walk ``skills_root/**/SKILL.md``; partition each skill's DECLARED MCP
    tools (3a parser) into curated-outward (∈ _OUTWARD_MCP_TOOLS) and
    unclassified outward-looking, each mapped to the skill paths declaring it.
    Pure read — never mutates a skill or any input."""
    curated: dict[str, list[str]] = {}
    unclassified: dict[str, list[str]] = {}
    for skill in sorted(skills_root.glob("**/SKILL.md")):
        try:
            frontmatter, _ = split_frontmatter_body(
                skill.read_text(encoding="utf-8")
            )
        except OSError:
            continue
        rel = skill.relative_to(skills_root).as_posix()
        for tool in declared_tools(frontmatter):
            if tool in _OUTWARD_MCP_TOOLS:
                curated.setdefault(tool, []).append(rel)
            elif _looks_outward(tool):
                unclassified.setdefault(tool, []).append(rel)
    return curated, unclassified


def check_F_27(workbook: Any, project_slug: str, *,
               workspace_root: Path | None = None,
               skills_root: Path | None = None, **_) -> dict:
    """Every OUTWARD MCP tool a skill DECLARES ⊆ the outward_action_gate's
    gated-MCP set (AMO batch 3-gov-driftF; spec §7-2b). The batch-2b gate DENIES
    an outward MCP submission without per-session consent, but its only gated
    MCP tool is ``mcp__gsc__submit_sitemap``. If a skill later declares a NEW
    outward MCP tool (a future Indexing URL_UPDATED submit, a publish tool)
    WITHOUT a matching gate matcher, that action ships UNGATED — a silent hole
    in the consent wall (Süleyman's Indexing hard-constraint). Engine-level
    invariant — IGNORES workbook + workspace_root; reads the skills tree + the
    IMPORTED gate constant only (no drift; F-16 safe).

    Detection logic (mirrors F-24's sets-comparison shape):
      1. skills/ missing → SKIP (engine state ambiguous → AMBER).
      2. gate_matchers = { _MCP_SUBMIT_TOOL stripped to registry-key form }.
      3. Scan declared MCP tools across skills/**/SKILL.md (3a parser).
      4. A declared CURATED-outward tool the gate does NOT cover — or a curated
         outward tool the gate forgot entirely (_OUTWARD_MCP_TOOLS ⊄
         gate_matchers) → FAIL HIGH (RED): a hole in the consent wall.
      5. A declared tool that merely LOOKS outward but is unclassified → FAIL
         MEDIUM (→ AMBER per §17.2): classify it before it ships ungated.
      6. Else → PASS.
    """
    rule = (
        "every OUTWARD MCP tool a skill declares in mcp_tools "
        "(required|optional) MUST be covered by an outward_action_gate matcher "
        "(the gate's gated-MCP set); an ungated declared outward MCP tool is a "
        "hole in the consent wall"
    )
    root = skills_root if skills_root is not None else _REPO_ROOT / "skills"
    if not root.is_dir():
        return _make_result(
            id_="F-27", severity="HIGH", verdict="SKIP",
            evidence=(
                f"skills/ missing at {root} — engine state ambiguous, "
                "surfaces AMBER"
            ),
            rule=rule, category="csr_mcp",
        )
    gate_matchers = {_strip_mcp_prefix(_MCP_SUBMIT_TOOL)}
    curated, unclassified = _scan_declared_outward(root)
    # FAIL (RED): an outward tool the gate does NOT cover — a declared curated
    # tool that is ungated, OR a curated outward tool the gate forgot wholesale,
    # OR an outward-looking tool the REGISTRY carries that the gate never saw.
    #
    # The third source matters because the first two both flow through the gate's
    # own constant: the declared set is normalised with _strip_mcp_prefix and
    # compared against a set built from _MCP_SUBMIT_TOOL, so the comparison can
    # only ever confirm the gate agrees with itself about tools somebody wrote a
    # SKILL.md for. A tool that exists in mcp-tool-registry.json and is outward by
    # its own name was invisible here — it could ship ungated with this rule GREEN.
    ungated = sorted(
        {t for t in curated if t not in gate_matchers}
        | (set(_OUTWARD_MCP_TOOLS) - gate_matchers)
        | (_registry_outward_tools() - gate_matchers)
    )
    if ungated:
        violations = [
            f"{t}: declared by "
            f"{curated.get(t, ['(gate hole — no skill declares it)'])}, not gated"
            for t in ungated
        ]
        return _make_result(
            id_="F-27", severity="HIGH", verdict="FAIL",
            evidence=(
                "ungated OUTWARD MCP tool(s) — a hole in the consent wall: "
                f"{ungated} not in gate matchers {sorted(gate_matchers)}"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=violations, affected_rows=len(violations),
        )
    # AMBER (MEDIUM → AMBER per §17.2): an outward-LOOKING tool nobody classified.
    if unclassified:
        flagged = sorted(unclassified)
        violations = [
            f"{t}: declared by {unclassified[t]}, outward-looking but unclassified"
            for t in flagged
        ]
        return _make_result(
            id_="F-27", severity="MEDIUM", verdict="FAIL",
            evidence=(
                "review: declared MCP tool(s) look OUTWARD but are unclassified "
                "— if outward, add to _OUTWARD_MCP_TOOLS + an outward_action_gate "
                f"matcher: {flagged}"
            ),
            rule=rule, category="csr_mcp",
            sample_violations=violations, affected_rows=len(violations),
        )
    return _make_result(
        id_="F-27", severity="HIGH", verdict="PASS",
        evidence=(
            "all declared outward MCP tools are gated: declared_outward="
            f"{sorted(curated)}, gate_matchers={sorted(gate_matchers)}; "
            "no unclassified outward-looking tool"
        ),
        rule=rule, category="csr_mcp",
    )


# ---------------------------------------------------------------------------
# Rule registry
# ---------------------------------------------------------------------------

_RULE_FUNCTIONS = (
    # CRITICAL
    check_F_01, check_F_02, check_F_03, check_F_04, check_F_05,
    # HIGH
    check_F_08, check_F_09, check_F_10, check_F_11, check_F_12,
    check_F_13, check_F_14, check_F_15, check_F_16, check_F_17,
    check_F_23,  # v1.8 Phase 4 SF MCP cross-sheet
    check_F_24,  # v1.9 Phase 2 .mcp.json↔registry key sync
    check_F_25,  # v1.9 Phase 3 sf.mcp.enabled ⇒ schema_version >= 1.5
    check_F_27,  # AMO 3-gov-driftF declared OUTWARD MCP tool ⊆ gate matchers
    # MEDIUM
    check_F_18, check_F_19, check_F_20, check_F_21, check_F_22,
    check_F_26,  # v1.9 Phase 4 orphan SF crawl detection (MCP-aware AMBER)
)


def evaluate_all(workbook: Any, project_slug: str, *,
                 workspace_root: Path | None = None,
                 formula_workbook: Any = None) -> list[dict]:
    """Run every rule function, return list of result dicts in
    declaration order. Each function may raise; caller handles DURUR-3.

    ``formula_workbook`` is the same file opened ``data_only=False``. Only
    F-02/F-03/F-04 use it, and without it those three report SKIP rather than a
    PASS they cannot justify — see _check_no_excel_formula. Every other rule
    absorbs the kwarg through **_.
    """
    out: list[dict] = []
    for fn in _RULE_FUNCTIONS:
        result = fn(workbook, project_slug, workspace_root=workspace_root,
                    formula_workbook=formula_workbook)
        out.append(result)
    return out


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

_IMPLEMENTED_RULE_IDS: frozenset[str] = frozenset(
    fn.__name__.replace("check_", "").replace("_", "-")
    for fn in _RULE_FUNCTIONS
)


@functools.lru_cache(maxsize=1)
def _declared_rules() -> dict[str, dict]:
    """Every rule id declared in cross-sheet-invariants.json.

    The registry above is what RUNS; this is what the contract PROMISES. They
    are not the same set, and the difference is the point of the next function.
    """
    path = _REPO_ROOT / "schemas" / "cross-sheet-invariants.json"
    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {r["id"]: r for r in doc.get("rules", []) if isinstance(r, dict) and "id" in r}


def unimplemented_results() -> list[dict]:
    """One NOT_IMPLEMENTED row per declared rule that has no check function.

    Without this, a declared rule with no implementation emits NOTHING — not a
    FAIL, not even a SKIP — so a run reports a verdict over the rules that
    happen to exist as though it covered the whole contract. Two of the absent
    rules are CRITICAL, which means a GREEN report could be issued while a
    critical invariant was never evaluated by anything.
    """
    out: list[dict] = []
    for rid, spec in sorted(_declared_rules().items()):
        if rid in _IMPLEMENTED_RULE_IDS:
            continue
        out.append(_make_result(
            id_=rid,
            severity=str(spec.get("severity", "MEDIUM")),
            verdict="NOT_IMPLEMENTED",
            evidence=(
                "declared in cross-sheet-invariants.json but no check_%s function "
                "exists — this rule was NOT evaluated"
                % rid.replace("-", "_")
            ),
            rule=str(spec.get("rule", "")),
            category=str(spec.get("category", "unknown")),
        ))
    return out


# Accepts both what runs and what is merely declared, so aggregation can report
# the gap instead of refusing to look at it.
_VALID_RULE_IDS: frozenset[str] = _IMPLEMENTED_RULE_IDS | frozenset(_declared_rules())


def aggregate_verdicts(rule_results: list[dict]) -> dict:
    """Aggregate per §17.2 verdict logic. Returns a small summary dict.

    Raises UnknownRuleError if a result.id is outside the registered set
    — that would mean the caller injected a synthetic rule outside the
    25-rule contract.
    """
    pass_count = warn_count = fail_count = 0
    overall = "GREEN"
    manual_review_required: list[str] = []
    auto_repair_available: list[str] = []
    not_implemented: list[str] = []

    saw_red = False
    saw_amber = False

    for r in rule_results:
        rid = r.get("id")
        if rid not in _VALID_RULE_IDS:
            raise UnknownRuleError(
                f"unknown rule id {rid!r} — outside 25-rule registry"
            )
        verdict = r.get("verdict")
        severity = r.get("severity", "MEDIUM")
        manual = bool(r.get("manual_triage"))
        repair = bool(r.get("auto_repair_available"))
        if repair:
            auto_repair_available.append(rid)

        if verdict == "PASS":
            pass_count += 1
        elif verdict == "SKIP":
            warn_count += 1
            saw_amber = True
        elif verdict == "NOT_IMPLEMENTED":
            # A declared rule nobody wrote is a coverage gap, not a data fault.
            # It must never be silent and must never read as GREEN.
            warn_count += 1
            saw_amber = True
            not_implemented.append(rid)
        elif verdict == "FAIL":
            fail_count += 1
            if manual:
                # F-15-style routing: stay AMBER, surface for human triage.
                saw_amber = True
                manual_review_required.append(rid)
            elif severity in ("CRITICAL", "HIGH"):
                saw_red = True
            else:  # MEDIUM/LOW
                saw_amber = True
        else:
            # Defensive: unknown verdict is treated as AMBER skip.
            warn_count += 1
            saw_amber = True

    if saw_red:
        overall = "RED"
    elif saw_amber:
        overall = "AMBER"
    else:
        overall = "GREEN"

    return {
        "overall": overall,
        "pass_count": pass_count,
        "warn_count": warn_count,
        "fail_count": fail_count,
        "total": len(rule_results),
        "declared_total": len(_declared_rules()) or len(rule_results),
        "manual_review_required": manual_review_required,
        "auto_repair_available": auto_repair_available,
        "not_implemented": not_implemented,
    }


# ---------------------------------------------------------------------------
# Consistency-report builder
# ---------------------------------------------------------------------------

@functools.lru_cache(maxsize=2)
def _load_schema_cached(schema_path: str) -> dict:
    return json.loads(Path(schema_path).read_text(encoding="utf-8"))


def _load_master_excel_schema() -> dict:
    return _load_schema_cached(str(_MASTER_EXCEL_SCHEMA))


def _load_consistency_report_schema() -> dict:
    return _load_schema_cached(str(_CONSISTENCY_REPORT_SCHEMA))


def _verdict_to_check_verdict(verdict: str) -> str:
    """Map our internal PASS/FAIL/SKIP to the schema's PASS/WARN/FAIL."""
    if verdict == "PASS":
        return "PASS"
    if verdict == "FAIL":
        return "FAIL"
    return "WARN"  # SKIP and unknown -> WARN


def build_consistency_report(
    *,
    rule_results: list[dict],
    aggregation: dict,
    project_id: str,
    report_id: int,
    generated_at: str,
    run_id: int | None = None,
    notes: str | None = None,
) -> dict:
    """Assemble + schema-validate a consistency-report.json payload.

    Raises ConsistencyReportInvalidError on schema failure (DURUR-6).
    """
    by_category: dict[str, dict[str, int]] = {}
    checks: list[dict] = []
    for r in rule_results:
        check_verdict = _verdict_to_check_verdict(r.get("verdict", "SKIP"))
        # Re-route F-15-style manual triage so the schema-level check
        # records WARN (not FAIL) — keeps the per-row severity but the
        # verdict reflects the routing decision.
        if r.get("manual_triage") and check_verdict == "FAIL":
            check_verdict = "WARN"
        check: dict[str, Any] = {
            "check_id": r["id"],
            "category": r.get("category", "csr_foundation"),
            "verdict": check_verdict,
        }
        if "severity" in r:
            check["severity"] = r["severity"]
        if "rule" in r:
            check["rule"] = r["rule"]
        if "evidence" in r:
            check["details"] = r["evidence"]
        if "affected_sheets" in r:
            check["affected_sheets"] = r["affected_sheets"]
        if "sample_violations" in r:
            check["sample_violations"] = r["sample_violations"][:_SAMPLE_CAP]
        if "affected_rows" in r:
            check["affected_rows"] = r["affected_rows"]
        if r.get("auto_repair_available"):
            check["auto_repair_available"] = True
        checks.append(check)

        cat = check["category"]
        bucket = by_category.setdefault(cat, {"pass": 0, "warn": 0, "fail": 0})
        if check_verdict == "PASS":
            bucket["pass"] += 1
        elif check_verdict == "WARN":
            bucket["warn"] += 1
        else:
            bucket["fail"] += 1

    report: dict[str, Any] = {
        "schema_version": "1.0",
        "report_id": report_id,
        "generated_at": generated_at,
        "project_id": project_id,
        "verdict": aggregation["overall"],
        "checks": checks,
        "summary": {
            "total_checks": aggregation["total"],
            "pass_count": aggregation["pass_count"],
            "warn_count": aggregation["warn_count"],
            "fail_count": aggregation["fail_count"],
            "by_category": by_category,
        },
    }
    if run_id is not None:
        report["run_id"] = run_id
    if aggregation.get("manual_review_required"):
        report["manual_review_required"] = aggregation["manual_review_required"]
    if aggregation.get("auto_repair_available"):
        report["auto_repair_performed"] = []  # nothing performed; flag presence
    if notes:
        report["notes"] = notes

    schema = _load_consistency_report_schema()
    validator = build_validator(schema)
    errors = sorted(validator.iter_errors(report), key=lambda e: e.path)
    if errors:
        msgs = []
        for err in errors:
            loc = "/".join(str(p) for p in err.absolute_path) or "<root>"
            msgs.append(f"{loc}: {err.message}")
        raise ConsistencyReportInvalidError("; ".join(msgs))
    return report


# ---------------------------------------------------------------------------
# Module export list
# ---------------------------------------------------------------------------

__all__: Iterable[str] = (
    "DriftCheckError",
    "UnknownRuleError",
    "ConsistencyReportInvalidError",
    "evaluate_all",
    "aggregate_verdicts",
    "build_consistency_report",
    "check_F_01", "check_F_02", "check_F_03", "check_F_04", "check_F_05",
    "check_F_08", "check_F_09", "check_F_10", "check_F_11", "check_F_12",
    "check_F_13", "check_F_14", "check_F_15", "check_F_16", "check_F_17",
    "check_F_18", "check_F_19", "check_F_20", "check_F_21", "check_F_22",
    "check_F_23",
    "check_F_24",
    "check_F_25",
    "check_F_26",
    "check_F_27",
)


# ---------------------------------------------------------------------------
# Command-line entry point
# ---------------------------------------------------------------------------
#
# This module was 2000 lines with no __main__ block. Running it as a script did
# nothing and exited 0 — which is indistinguishable from running it and passing.
# Every "I ran the invariants and they were clean" was unverifiable until now.
#
# Three things this prints that the library alone never surfaced:
#   * NOT_IMPLEMENTED rows, so declared-but-absent rules stop being invisible
#   * the row count each rule saw, so a 3-row PASS cannot be read as a 3000-row one
#   * a non-zero exit on RED, so a caller can actually gate on it

def _sheet_row_counts(workbook: Any, sheets: list[str] | None) -> str:
    """Rows behind a verdict. A PASS over an empty sheet is not a measurement."""
    if not sheets:
        return ""
    parts = []
    for s in sheets:
        if not _has_sheet(workbook, s):
            parts.append("%s=absent" % s)
            continue
        try:
            parts.append("%s=%d" % (s, len(_iter_rows_as_dicts(workbook, s))))
        except Exception:
            parts.append("%s=?" % s)
    return " ".join(parts)


def main(argv: list[str] | None = None) -> int:
    import argparse

    ap = argparse.ArgumentParser(
        prog="validate_invariants.py",
        description="Evaluate the declared cross-sheet invariants for one project.",
    )
    ap.add_argument("--project", required=True, help="project slug under projects/")
    ap.add_argument("--workspace-root", default=None,
                    help="workspace root (defaults to the resolved workspace)")
    ap.add_argument("--json", action="store_true", help="emit the full result as JSON")
    args = ap.parse_args(argv)

    try:
        ws_root = _resolve_workspace_root(
            Path(args.workspace_root) if args.workspace_root else None)
        wb_path = _project_dir(args.project, ws_root) / "master.xlsx"
    except Exception as exc:
        print("ERROR resolving workspace: %s" % exc, file=sys.stderr)
        return 2
    if not wb_path.is_file():
        print("ERROR master.xlsx not found: %s" % wb_path, file=sys.stderr)
        return 2

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR openpyxl is not installed", file=sys.stderr)
        return 2

    wb = load_workbook(filename=str(wb_path), read_only=True, data_only=True)
    # Second view of the SAME file, formulas visible. F-02/F-03/F-04 cannot see
    # their subject without it and correctly report SKIP when it is absent.
    fwb = load_workbook(filename=str(wb_path), read_only=True, data_only=False)
    try:
        results = evaluate_all(wb, args.project, workspace_root=ws_root,
                               formula_workbook=fwb)
        rows = {r["id"]: _sheet_row_counts(wb, r.get("affected_sheets"))
                for r in results}
    finally:
        wb.close()
        fwb.close()

    results = results + unimplemented_results()
    agg = aggregate_verdicts(results)

    if args.json:
        print(json.dumps({"summary": agg, "checks": results}, indent=1, default=str))
    else:
        print("invariants — project=%s  workbook=%s" % (args.project, wb_path))
        for r in sorted(results, key=lambda x: x["id"]):
            print("  %-6s %-16s %-8s %s" % (
                r["id"], r["verdict"], r["severity"],
                rows.get(r["id"], "") or r["evidence"][:80]))
        print()
        print("  overall=%s   pass=%d warn=%d fail=%d   evaluated=%d of %d declared"
              % (agg["overall"], agg["pass_count"], agg["warn_count"],
                 agg["fail_count"], len(results) - len(agg["not_implemented"]),
                 agg["declared_total"]))
        if agg["not_implemented"]:
            print("  NOT IMPLEMENTED (declared, never evaluated): %s"
                  % ", ".join(agg["not_implemented"]))
        if agg["manual_review_required"]:
            print("  manual review: %s" % ", ".join(agg["manual_review_required"]))

    return 1 if agg["overall"] == "RED" else 0


if __name__ == "__main__":
    sys.exit(main())
