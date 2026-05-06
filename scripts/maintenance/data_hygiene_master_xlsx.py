#!/usr/bin/env python3
"""data_hygiene_master_xlsx.py — pilot data hygiene per ADR-037.

ADR-037 codifies hygiene policy: code-driven script + dry-run + audit
trail + Süleyman approval. This script handles the Wave 3 finding:

  F-17 priority code drift — `master_task.priority` cells using legacy
  P1/P2/P3 codes outside severityEnum. Mapping: P1→HIGH, P2→MEDIUM,
  P3→LOW (severityEnum 4-value canonical).

Note on dup-header artifact (`row 1 transaction-write` + `row N
bootstrap header_row=N`): validator's `_resolve_header_row` (Phase
14 W3-W2-C-a) already defends against this — no cleanup needed.

Usage::

  python -m scripts.maintenance.data_hygiene_master_xlsx \\
      --workspace /path/to/workspace --project dentnotion --dry-run

  # ... review audit at outputs/reports/{date}-data-hygiene-master.md ...
  # ... obtain Süleyman approval ...

  python -m scripts.maintenance.data_hygiene_master_xlsx \\
      --workspace /path/to/workspace --project dentnotion --apply

The script is idempotent: a clean workspace produces a no-op audit
report ("0 cells changed").
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from openpyxl import load_workbook  # noqa: E402
from scripts.excel import transaction  # noqa: E402
from scripts.validation.validate_invariants import (  # noqa: E402
    _iter_rows_as_dicts,
)

PRIORITY_MAPPING = {
    "P1": "HIGH",
    "P2": "MEDIUM",
    "P3": "LOW",
}


@dataclass(frozen=True)
class PriorityChange:
    row_index: int  # 1-based row index in master_task sheet (post-header)
    task_id: str | None
    old: str
    new: str


def _scan_priority_drift(workbook_path: Path) -> list[PriorityChange]:
    """Identify master_task.priority cells outside severityEnum.

    Reads via openpyxl + uses validate_invariants `_iter_rows_as_dicts`
    which respects schema `header_row`.  Returns the per-row change set.
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = _iter_rows_as_dicts(wb, "master_task")
    finally:
        wb.close()
    changes: list[PriorityChange] = []
    for i, row in enumerate(rows, start=1):
        prio = row.get("priority")
        if isinstance(prio, str) and prio in PRIORITY_MAPPING:
            changes.append(
                PriorityChange(
                    row_index=i,
                    task_id=row.get("task_id"),
                    old=prio,
                    new=PRIORITY_MAPPING[prio],
                )
            )
    return changes


def _format_audit(
    changes: list[PriorityChange],
    *,
    project_slug: str,
    apply_mode: bool,
) -> str:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    mode = "APPLY" if apply_mode else "DRY-RUN"
    out: list[str] = [
        f"# Data Hygiene Audit — {project_slug}",
        "",
        f"**Date:** {today}",
        f"**Mode:** {mode}",
        f"**Scope:** F-17 priority code normalize (master_task.priority)",
        f"**Mapping:** P1→HIGH, P2→MEDIUM, P3→LOW (ADR-037)",
        f"**Cells affected:** {len(changes)}",
        "",
    ]
    if not changes:
        out.append("0 cells outside severityEnum — no-op (idempotent re-run).")
        return "\n".join(out) + "\n"
    out.extend([
        "| # | task_id | old | new |",
        "|---|---|---|---|",
    ])
    for c in changes:
        tid = c.task_id or "(missing)"
        out.append(f"| {c.row_index} | `{tid}` | {c.old} | {c.new} |")
    out.append("")
    out.append("> Audit trail mandated by ADR-037.")
    return "\n".join(out) + "\n"


def _apply_priority_normalize(
    workbook_path: Path,
    changes: list[PriorityChange],
    project_slug: str,
) -> int:
    """Apply each P-code → severityEnum mapping via transaction.update.

    Each call mutates rows where priority == old to set priority = new.
    Returns total rows_affected (sums across the 3 mappings).
    """
    total = 0
    for old in ("P1", "P2", "P3"):
        if not any(c.old == old for c in changes):
            continue
        new = PRIORITY_MAPPING[old]
        try:
            result = transaction.update(
                workbook_path,
                "master_task",
                where={"priority": old},
                set_={"priority": new},
                project_slug=project_slug,
                writer="data_hygiene_master_xlsx",
            )
            total += result.rows_affected
        except Exception as exc:  # surfacing via stderr is fine for CLI
            print(f"WARN: update {old}→{new} raised {exc!r}", file=sys.stderr)
    return total


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workspace", required=True, type=Path)
    parser.add_argument("--project", required=True, type=str)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    parser.add_argument("--out", type=Path, default=None,
                        help="Audit report path (default: project's outputs/reports/)")
    args = parser.parse_args(argv)

    project_dir = args.workspace / "projects" / args.project
    workbook_path = project_dir / "master.xlsx"
    if not workbook_path.is_file():
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    changes = _scan_priority_drift(workbook_path)
    audit_text = _format_audit(changes, project_slug=args.project, apply_mode=args.apply)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if args.out is None:
        reports_dir = project_dir / "outputs" / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        suffix = "apply" if args.apply else "dry-run"
        args.out = reports_dir / f"{today}-data-hygiene-master-{suffix}.md"

    args.out.write_text(audit_text, encoding="utf-8")
    print(audit_text)
    print(f"\nAudit written to: {args.out}")

    if args.apply and changes:
        affected = _apply_priority_normalize(workbook_path, changes, args.project)
        print(f"\nApplied: {affected} rows updated via transaction.update")
    elif args.apply:
        print("\nApply mode: 0 changes (already clean)")
    else:
        print("\nDry-run: no workbook mutation. Re-run with --apply after Süleyman approval.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
