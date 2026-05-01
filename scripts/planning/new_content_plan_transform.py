#!/usr/bin/env python3
"""
new_content_plan_transform.py — pure transform: Phase 7 content-gaps
staging JSON (+ optional DFS keyword_overview / search_intent enrichment) →
schema-shaped rows for master.xlsx#new_content_plan.

Reads:
  - _state/staging/content_gaps_keyword_ideas_*.json   (REQUIRED;
        Phase 7 W-B1 producer, content-gaps skill)
  - inbox/dfs/{date}-keyword_overview-newplan-{slug}.json   (optional)
  - inbox/dfs/{date}-search_intent-newplan-{slug}.json      (optional)
  - master.xlsx#cluster_keywords snapshot                   (cross-ref)

Writes:
  - schema-shaped rows ready for transaction.append()
        master.xlsx#new_content_plan (11 cols)

Tolerates BOTH DFS response shapes via
``scripts.ingestion.dfs_pull._normalize_dfs_response`` (D-003 root-cause
fix — REST envelope ``tasks[0].result[0].items`` AND flat wrapper
``items``). The function is IMPORTED — never copied — so any future shape
fix in dfs_pull is inherited automatically. Same for ``StagingSchemaError``.

Pure function discipline (mirrors content_gaps_transform.py +
cannibalization_transform.py):
  - No state mutation.
  - No file write side-effects when imported as a module (CLI only).
  - Idempotent: same input → byte-identical output.

Schema target (master-excel.schema.json#new_content_plan, 11 cols):
  id, title, url_slug, primary_keyword, monthly_volume, assigned_cluster,
  target_word_count, priority, created_date, tivl_tag, lifecycle_status

CLI:
  python3 scripts/planning/new_content_plan_transform.py \\
      --staging _state/staging/content_gaps_keyword_ideas_{date}_{slug}.json \\
      [--cluster-map cluster_keywords.json] \\
      [--raw-keyword-overview inbox/dfs/{date}-keyword_overview-newplan-{slug}.json] \\
      [--raw-search-intent inbox/dfs/{date}-search_intent-newplan-{slug}.json] \\
      [--default-status TODO] \\
      [--default-tivl-tag I] \\
      [--target-word-count 1500] \\
      [--max-rows 50] \\
      [--date 2026-05-01] \\
      [--output-dir _state/transform/{run_id}/]

Refs: schemas/master-excel.schema.json (new_content_plan
required_columns + statusEnum + tivl_tag enum + lifecycle_status enum),
schemas/events.schema.json (target_excel_sheet=new_content_plan +
source.kind=dataforseo_mcp), schemas/skill-frontmatter.schema.json
(frontmatter contract). Pattern reference:
scripts/discovery/content_gaps_transform.py (D-003 staging consume +
budget pre-flight + KOPYA YASAK import).
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date as date_cls, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

# scripts/ingestion + scripts/budget are namespace packages; ensure repo
# root on sys.path so absolute imports resolve when invoked as a script.
_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# IMPORT discipline (NOT copy) — D-003 invariant: _normalize_dfs_response +
# StagingSchemaError are owned by scripts.ingestion.dfs_pull. We REUSE them
# so any future shape / error-class evolution is inherited automatically.
# KOPYA YASAK is enforced by the import-discipline identity test.
from scripts.ingestion.dfs_pull import (   # noqa: E402  (sys.path mutation above)
    _normalize_dfs_response,
    StagingSchemaError,
)


# ---------------------------------------------------------------------------
# Constants — schema-aligned column names
# (master-excel.schema.json#new_content_plan required_columns)
# ---------------------------------------------------------------------------

#: Schema-locked column tuple — 11 cols. ANY drift = DURUR #3.
NEW_CONTENT_PLAN_COLUMNS: tuple[str, ...] = (
    "id",
    "title",
    "url_slug",
    "primary_keyword",
    "monthly_volume",
    "assigned_cluster",
    "target_word_count",
    "priority",
    "created_date",
    "tivl_tag",
    "lifecycle_status",
)

#: statusEnum allowed values
#: (master-excel.schema.json#/definitions/statusEnum). Used to validate
#: callers' default_status input even though the new_content_plan sheet
#: itself does not carry a status column — the priority column is the
#: planner-facing status surface here.
_STATUS_ENUM: frozenset = frozenset({
    "TODO", "ONGOING", "EXISTS", "DONE",
    "BLOCKED", "DEFERRED", "CANCELED",
})

#: tivl_tag enum (master-excel.schema.json#new_content_plan col J).
#: T = Transactional, I = Informational, V = Video, L = Local.
_TIVL_ENUM: frozenset = frozenset({"T", "I", "V", "L"})

#: lifecycle_status enum (master-excel.schema.json#new_content_plan col K).
_LIFECYCLE_ENUM: frozenset = frozenset({"GREEN", "RED", "ON_HOLD", "REMOVED"})

#: Default lifecycle for newly-planned rows. GREEN = active in pipeline.
_DEFAULT_LIFECYCLE: str = "GREEN"

#: DoS guard — refuse to project more than this many rows even when the
#: staging table contains more candidates. Configurable via --max-rows.
DEFAULT_MAX_ROWS: int = 50

#: Default target word count for informational pages. Overridden by the
#: tivl_tag heuristic for T/L/V (lighter formats).
DEFAULT_TARGET_WORD_COUNT: int = 1500

#: Per-tag word count overrides (tivl_tag → target_word_count).
_WORD_COUNT_BY_TAG: dict = {
    "T": 1200,    # transactional landing pages
    "L": 1000,    # local listings — short, scannable
    "V":  800,    # video page — supporting copy only
    # "I" handled via target_word_count input (caller-provided default).
}

#: Path to the canonical budget pre-flight script (subprocess wrapper paterni
#: from W-A3/W-A4). Skill-orchestrator invokes this before any paid call.
BUDGET_CHECK_SCRIPT: Path = _REPO_ROOT / "scripts" / "budget" / "check_budget.py"

#: Glob pattern for content-gaps staging (Phase 7 W-B1 producer).
_CONTENT_GAPS_STAGING_GLOB: str = "content_gaps_keyword_ideas_*.json"

#: Turkish character → ASCII fold map for url_slug. Plugin-agnostik:
#: applies to every input keyword regardless of slug content.
_TR_FOLD: dict = {
    "ğ": "g", "Ğ": "g",
    "ş": "s", "Ş": "s",
    "ı": "i", "İ": "i",
    "ü": "u", "Ü": "u",
    "ö": "o", "Ö": "o",
    "ç": "c", "Ç": "c",
}


# ---------------------------------------------------------------------------
# Exceptions — DURUR signal classes
# ---------------------------------------------------------------------------

class NewContentPlanError(Exception):
    """Base class for new_content_plan_transform errors."""


class BudgetExceededError(NewContentPlanError):
    """Budget pre-flight DURUR — projected DFS spend exceeds the daily cap."""


class StagingNotFoundError(NewContentPlanError):
    """Phase 7 content-gaps staging consume DURUR — no staging file present
    AND no fallback fixture supplied."""


class ClusterKeywordsEmptyError(NewContentPlanError):
    """Cross-ref DURUR (CLUSTER_KEYWORDS_EMPTY) — master.xlsx#cluster_keywords
    sheet is empty when this skill runs. Workflow ordering issue: run
    cluster-map first."""


class RowSchemaError(NewContentPlanError):
    """A produced row drifts from the master-excel new_content_plan schema
    (column tuple mismatch, type mismatch, or enum violation)."""


class TivlTagDriftError(RowSchemaError):
    """Produced row's tivl_tag is not in the master-excel enum
    (T / I / V / L). Defensive: caller-input default_tivl_tag is also
    range-checked."""


class UrlSlugUniquenessError(RowSchemaError):
    """Two keywords in the same batch reduce to the same url_slug —
    DURUR #8. The caller picks a different max_rows or input set."""


# Re-export StagingSchemaError so callers can `except` it without
# importing dfs_pull directly. The IDENTITY (`is`) is preserved — see
# the import-discipline test.
__all_errors__ = (
    "NewContentPlanError",
    "BudgetExceededError",
    "StagingNotFoundError",
    "ClusterKeywordsEmptyError",
    "RowSchemaError",
    "TivlTagDriftError",
    "UrlSlugUniquenessError",
    "StagingSchemaError",
)


# ---------------------------------------------------------------------------
# Helpers — type coercion (mirrors content_gaps_transform / dfs_pull patterns)
# ---------------------------------------------------------------------------

def _safe_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def _safe_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _utc_today_iso() -> str:
    """ISO 8601 date (no time) for the created_date column. UTC-pinned."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# Budget pre-flight (subprocess wrapper paterni — W-A3/W-A4)
# ---------------------------------------------------------------------------

def preflight_budget(
    *,
    project_config_path: str | os.PathLike[str],
    events_path: str | os.PathLike[str],
    script_path: str | os.PathLike[str] = BUDGET_CHECK_SCRIPT,
) -> dict:
    """Run scripts/budget/check_budget.py as a subprocess and DURUR if exit 1.

    Returns the parsed envelope on PASS:
        {"budget_per_day": int, "used_24h": ..., "remaining": ..., "exceeded": bool}

    Raises BudgetExceededError when the subprocess exits non-zero — that is
    the canonical "budget would be exceeded by this run" signal per
    §16.8 + W-A3/W-A4. Stderr is preserved on the exception for manager
    diagnostics.
    """
    cmd = [
        sys.executable,
        str(script_path),
        "--project-config", str(project_config_path),
        "--events", str(events_path),
    ]
    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise BudgetExceededError(
            "budget pre-flight FAIL "
            f"(exit={proc.returncode}): stdout={proc.stdout.strip()!r} "
            f"stderr={proc.stderr.strip()!r}"
        )
    try:
        return json.loads(proc.stdout.strip())
    except json.JSONDecodeError as exc:
        raise BudgetExceededError(
            f"budget pre-flight stdout was not JSON: {proc.stdout!r}"
        ) from exc


# ---------------------------------------------------------------------------
# url_slug generation — idempotent, plugin-agnostik
# ---------------------------------------------------------------------------

def make_url_slug(keyword: str) -> str:
    """
    Turn a keyword into a URL slug. Idempotent; plugin-agnostik.

    Rules (deterministic):
      1. Lowercase.
      2. Turkish character fold (ğ→g, ş→s, ı→i, ü→u, ö→o, ç→c).
      3. Replace any non-alphanum with single "-".
      4. Collapse runs of "-" into one.
      5. Strip leading/trailing "-".

    Returns "" for an empty / whitespace-only input. Raises ValueError
    if the input is not a string.

    Idempotency: ``make_url_slug(make_url_slug(k)) == make_url_slug(k)``.
    """
    if not isinstance(keyword, str):
        raise ValueError(
            f"keyword must be a string, got {type(keyword).__name__}"
        )
    s = keyword.strip().lower()
    if not s:
        return ""
    # Turkish fold (deterministic char-by-char).
    s = "".join(_TR_FOLD.get(ch, ch) for ch in s)
    # Replace non-alphanum (ASCII) with "-".
    s = re.sub(r"[^a-z0-9]+", "-", s)
    # Collapse runs and strip edges.
    s = re.sub(r"-+", "-", s).strip("-")
    return s


# ---------------------------------------------------------------------------
# Staging file locator — Phase 7 W-B1 consume
# ---------------------------------------------------------------------------

def locate_content_gaps_staging(
    *,
    project_root: Path | None = None,
    explicit_path: str | os.PathLike[str] | None = None,
    fallback_payload: dict | None = None,
) -> Path | dict:
    """Locate the newest content_gaps_keyword_ideas_*.json staging file.

    Priority order:
      1. ``explicit_path`` if provided AND file exists.
      2. ``project_root/_state/staging/content_gaps_keyword_ideas_*.json``
         — newest by mtime.
      3. ``fallback_payload`` (dict) — TEST-ONLY fallback fixture; returned
         as-is. Skill runtime SHOULD NOT pass this; tests use it to
         exercise the transform without a real staging file on disk.
      4. Raises StagingNotFoundError when none of the above resolve.

    Returns:
        Path: when (1) or (2) succeeded.
        dict: when (3) — caller must handle inline payload mode.

    Raises:
        StagingNotFoundError: nothing on disk, no fallback.
    """
    if explicit_path is not None:
        p = Path(explicit_path)
        if p.exists():
            return p
        # Explicit path provided but missing → DURUR (caller picked a
        # specific file, do not silently fall through).
        raise StagingNotFoundError(
            f"explicit content_gaps_staging_path does not exist: {p}"
        )

    if project_root is not None:
        staging_dir = Path(project_root) / "_state" / "staging"
        if staging_dir.is_dir():
            matches = sorted(
                staging_dir.glob(_CONTENT_GAPS_STAGING_GLOB),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            if matches:
                return matches[0]

    if fallback_payload is not None:
        return fallback_payload

    raise StagingNotFoundError(
        "no content_gaps_keyword_ideas_*.json staging file found under "
        f"{project_root}/_state/staging/ (and no fallback fixture). "
        "Run content-gaps skill (Phase 7 W-B1) first."
    )


# ---------------------------------------------------------------------------
# cluster_keywords snapshot reader (cross-ref at workflow_runner gate)
# ---------------------------------------------------------------------------

def read_cluster_keywords_snapshot(
    *,
    workbook_path: str | os.PathLike[str] | None = None,
    rows: Sequence[Mapping[str, Any]] | None = None,
) -> dict[str, str]:
    """Read master.xlsx#cluster_keywords sheet and build a {keyword: cluster}
    snapshot map. The map is consumed by transform() to fill the
    `assigned_cluster` column (col F).

    Two input modes:
      - ``workbook_path``: read live from master.xlsx (uses openpyxl).
      - ``rows``: TEST-ONLY mode; pass a list of dicts with at least
        the ``keyword`` and ``cluster`` columns (matching the
        cluster_keywords sheet shape).

    Returns:
        dict[str, str]: lowercased-keyword → cluster name.

    Raises:
        ClusterKeywordsEmptyError: sheet is empty (no data rows).
            Workflow ordering issue, not a parallel-write race.

    Notes (concurrent write awareness):
        This is a SNAPSHOT-AT-DISPATCH read. The skill body invokes this
        function right before transaction.append(), so the cluster_keywords
        snapshot is consistent with the time of the planning write.
        scripts/excel/transaction.py provides atomic file locking that
        serialises concurrent writes from sibling planning skills (e.g.
        cluster-map running in parallel) — see ADR-021.
    """
    if rows is not None:
        out: dict[str, str] = {}
        for r in rows:
            if not isinstance(r, dict):
                continue
            kw = _safe_str(r.get("keyword"))
            cl = _safe_str(r.get("cluster"))
            if kw and cl:
                out[kw.lower()] = cl
        if not out:
            raise ClusterKeywordsEmptyError(
                "cluster_keywords rows produced an empty {keyword: cluster} "
                "map — run cluster-map first before new-content-plan "
                "(workflow ordering issue, not a parallel-write race)."
            )
        return out

    if workbook_path is None:
        raise ValueError("either workbook_path or rows must be provided")

    wb_path = Path(workbook_path)
    if not wb_path.exists():
        raise ClusterKeywordsEmptyError(
            f"workbook does not exist: {wb_path}"
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:    # pragma: no cover — tested via rows= path
        raise NewContentPlanError(
            "openpyxl is required to read master.xlsx#cluster_keywords; "
            "install the project requirements"
        ) from exc

    wb = load_workbook(wb_path, read_only=True, data_only=True)
    if "cluster_keywords" not in wb.sheetnames:
        raise ClusterKeywordsEmptyError(
            f"workbook has no 'cluster_keywords' sheet: {wb_path}"
        )
    ws = wb["cluster_keywords"]
    # Per master-excel.schema.json: header_row=3, data_start_row=4.
    # Col A=cluster, Col B=keyword.
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row is None:
            continue
        if len(row) < 2:
            continue
        cluster = _safe_str(row[0])
        keyword = _safe_str(row[1])
        if keyword and cluster:
            out[keyword.lower()] = cluster
    wb.close()
    if not out:
        raise ClusterKeywordsEmptyError(
            f"cluster_keywords sheet has no data rows in {wb_path} — "
            "run cluster-map first before new-content-plan "
            "(workflow ordering issue, not a parallel-write race)."
        )
    return out


# ---------------------------------------------------------------------------
# Search intent → tivl_tag heuristic
# ---------------------------------------------------------------------------

#: Canonical DFS search_intent → tivl_tag map.
_INTENT_TO_TIVL: dict = {
    "informational": "I",
    "transactional": "T",
    "commercial":    "T",   # transactional umbrella in TIVL
    "navigational":  "I",   # informational fallback (no nav tag in TIVL)
    "local":         "L",
    "video":         "V",
}


def _intent_to_tivl(intent_label: Any, *, default: str) -> str:
    """Map a DFS search_intent label to TIVL tag (T/I/V/L).

    Unknown / missing → default (caller-provided, frontmatter input).
    Defensive: validates the result against _TIVL_ENUM before return.
    """
    if isinstance(intent_label, str) and intent_label.strip():
        token = intent_label.strip().lower()
        tag = _INTENT_TO_TIVL.get(token)
        if tag is not None:
            return tag
    if default not in _TIVL_ENUM:
        raise TivlTagDriftError(
            f"default_tivl_tag must be one of {sorted(_TIVL_ENUM)}, "
            f"got {default!r}"
        )
    return default


def _build_intent_map(
    raw_search_intent: dict | None,
) -> dict[str, str]:
    """Extract a {keyword_lower: intent_label} map from a search_intent payload.

    Tolerates both DFS shapes via the imported _normalize_dfs_response.
    Returns {} when raw_search_intent is None or empty — caller falls back
    to default_tivl_tag.
    """
    if raw_search_intent is None:
        return {}
    items = _normalize_dfs_response(
        raw_search_intent,
        expected_endpoint="dataforseo_labs_search_intent",
    )
    out: dict[str, str] = {}
    for it in items:
        if not isinstance(it, dict):
            continue
        kw = _safe_str(it.get("keyword"))
        if not kw:
            continue
        # search_intent shapes: {keyword_intent: {label}} or
        # {main_intent: ...} or top-level intent.
        intent: Any = None
        ki = it.get("keyword_intent")
        if isinstance(ki, dict):
            intent = ki.get("label") or ki.get("main_intent")
        if intent is None:
            intent = it.get("main_intent")
        if intent is None:
            intent = it.get("intent")
        if isinstance(intent, str):
            out[kw.lower()] = intent.strip().lower()
    return out


def _build_overview_volume_map(
    raw_keyword_overview: dict | None,
) -> dict[str, int]:
    """Extract {keyword_lower: search_volume} from a keyword_overview payload.

    Used to override / refresh monthly_volume on the projected subset
    (the staging table's volume comes from keyword_ideas; overview is
    a separate per-keyword refresh). Returns {} when payload absent.
    """
    if raw_keyword_overview is None:
        return {}
    items = _normalize_dfs_response(
        raw_keyword_overview,
        expected_endpoint="dataforseo_labs_google_keyword_overview",
    )
    out: dict[str, int] = {}
    for it in items:
        if not isinstance(it, dict):
            continue
        kw = _safe_str(it.get("keyword"))
        if not kw:
            continue
        vol = _safe_int((it.get("keyword_info") or {}).get("search_volume"))
        if vol is None:
            vol = _safe_int(it.get("search_volume"))
        if vol is not None:
            out[kw.lower()] = int(vol)
    return out


# ---------------------------------------------------------------------------
# Priority + word count heuristics
# ---------------------------------------------------------------------------

def _priority_label(*, gap_score: float, monthly_volume: int) -> str:
    """Priority tier — gap_score-driven with monthly_volume fallback.

    gap_score >= 1000   → P1
    gap_score >=  100   → P2
    otherwise           → P3
    Volume-only fallback when gap_score==0:
        volume >= 1000  → P1
        volume >=  100  → P2
        else            → P3
    """
    if gap_score >= 1000:
        return "P1"
    if gap_score >= 100:
        return "P2"
    if gap_score > 0:
        return "P3"
    # gap_score==0 → volume fallback.
    if monthly_volume >= 1000:
        return "P1"
    if monthly_volume >= 100:
        return "P2"
    return "P3"


def _target_word_count(
    *,
    tivl_tag: str,
    target_word_count_default: int,
) -> int:
    """Per-tag word count override; falls back to caller default."""
    if tivl_tag in _WORD_COUNT_BY_TAG:
        return int(_WORD_COUNT_BY_TAG[tivl_tag])
    return int(target_word_count_default)


# ---------------------------------------------------------------------------
# Core transform
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class _StagingRow:
    """A single content-gaps staging row after extraction."""
    keyword: str
    monthly_volume: int
    gap_score: float


def _extract_staging_rows(staging_payload: Mapping[str, Any]) -> list[_StagingRow]:
    """Extract staging rows from a content-gaps keyword_ideas staging dict.

    Expected shape (mirrors content_gaps_transform._staging_table_dict):
        {"schema_version": "1.0", "tool": "keyword_ideas", "rows": [
            {"id": ..., "keyword": ..., "search_volume": ...,
             "gap_score": ..., ...}, ...
        ], ...}

    Tolerates a bare list of rows (test-only convenience).
    """
    if isinstance(staging_payload, list):
        rows = staging_payload
    elif isinstance(staging_payload, dict):
        rows = staging_payload.get("rows") or []
    else:
        raise NewContentPlanError(
            "staging_payload must be a dict (with 'rows') or a list, "
            f"got {type(staging_payload).__name__}"
        )

    out: list[_StagingRow] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        kw = _safe_str(r.get("keyword"))
        if not kw:
            continue
        vol = _safe_int(r.get("search_volume")) or 0
        gscore = _safe_float(r.get("gap_score")) or 0.0
        out.append(_StagingRow(
            keyword=kw,
            monthly_volume=int(vol),
            gap_score=float(gscore),
        ))
    return out


def transform(
    staging_payload: Mapping[str, Any] | Sequence[Mapping[str, Any]],
    *,
    cluster_map: Mapping[str, str],
    raw_keyword_overview: dict | None = None,
    raw_search_intent: dict | None = None,
    default_status: str = "TODO",
    default_tivl_tag: str = "I",
    target_word_count: int = DEFAULT_TARGET_WORD_COUNT,
    max_rows: int = DEFAULT_MAX_ROWS,
    created_date: str | None = None,
) -> dict:
    """
    Transform Phase 7 content-gaps staging + optional DFS enrichment +
    cluster_keywords snapshot into schema-shaped new_content_plan rows.

    Args:
        staging_payload: Phase 7 W-B1 producer output (content-gaps
            keyword_ideas staging dict OR a bare list of rows).
        cluster_map: {keyword_lower: cluster_name} from
            master.xlsx#cluster_keywords (read at workflow_runner gate
            via read_cluster_keywords_snapshot()). MUST be non-empty —
            empty map is the CLUSTER_KEYWORDS_EMPTY DURUR (caller
            handles before invoking transform()).
        raw_keyword_overview: optional parsed JSON from
            mcp__dataforseo__dataforseo_labs_google_keyword_overview
            for volume refresh on the projected subset.
        raw_search_intent: optional parsed JSON from
            mcp__dataforseo__dataforseo_labs_search_intent for tivl_tag
            assignment. Absent → every row uses default_tivl_tag.
        default_status: statusEnum value (defensive — new_content_plan
            sheet has no status column, but worker brief required this
            input for cross-skill consistency).
        default_tivl_tag: TIVL tag fallback for keywords without intent
            data. Must be one of T/I/V/L.
        target_word_count: default word count for tivl_tag=I rows.
            Other tags use _WORD_COUNT_BY_TAG override.
        max_rows: DoS guard; only the top max_rows by gap_score are
            projected.
        created_date: ISO 8601 date string (YYYY-MM-DD) for the
            created_date column. Defaults to today UTC.

    Returns:
        {"new_content_plan": [<row>, ...], "meta": {...}}.

    new_content_plan row shape (master-excel.schema.json#new_content_plan):
      id                  int
      title               str
      url_slug            str   (idempotent slug — see make_url_slug)
      primary_keyword     str
      monthly_volume      int
      assigned_cluster    str   (from cluster_map cross-ref)
      target_word_count   int
      priority            str   "P1" / "P2" / "P3"
      created_date        str   ISO date YYYY-MM-DD
      tivl_tag            str   "T" / "I" / "V" / "L" (master-excel enum)
      lifecycle_status    str   "GREEN" (default for new rows)

    DURUR triggers (raise NewContentPlanError or subclass):
      - default_status not in statusEnum
      - default_tivl_tag not in TIVL enum (TivlTagDriftError)
      - cluster_map empty (ClusterKeywordsEmptyError)
      - max_rows <= 0
      - row column tuple drifts from NEW_CONTENT_PLAN_COLUMNS (RowSchemaError)
      - tivl_tag value drifts (TivlTagDriftError)
      - lifecycle_status drift (RowSchemaError)
      - url_slug uniqueness violation within batch (UrlSlugUniquenessError)
    """
    if default_status not in _STATUS_ENUM:
        raise NewContentPlanError(
            f"default_status must be one of {sorted(_STATUS_ENUM)}, "
            f"got {default_status!r}"
        )
    if default_tivl_tag not in _TIVL_ENUM:
        raise TivlTagDriftError(
            f"default_tivl_tag must be one of {sorted(_TIVL_ENUM)}, "
            f"got {default_tivl_tag!r}"
        )
    if not isinstance(max_rows, int) or max_rows <= 0:
        raise NewContentPlanError(
            f"max_rows must be a positive int, got {max_rows!r}"
        )
    if not isinstance(target_word_count, int) or target_word_count <= 0:
        raise NewContentPlanError(
            f"target_word_count must be a positive int, got {target_word_count!r}"
        )
    if not cluster_map:
        raise ClusterKeywordsEmptyError(
            "cluster_map is empty — run cluster-map first before "
            "new-content-plan (workflow ordering issue, not a "
            "parallel-write race)."
        )

    created = created_date or _utc_today_iso()

    # 1. Extract staging rows + sort by gap_score desc (deterministic).
    staging_rows = _extract_staging_rows(staging_payload)
    staging_rows.sort(key=lambda r: (-r.gap_score, r.keyword))
    # 2. Cap to max_rows.
    projected = staging_rows[:max_rows]

    # 3. Build enrichment maps (search_intent → tivl, keyword_overview → volume).
    intent_map = _build_intent_map(raw_search_intent)
    overview_volume_map = _build_overview_volume_map(raw_keyword_overview)

    # 4. Project rows.
    rows: list[dict] = []
    seen_slugs: dict[str, str] = {}    # slug → first-seen keyword
    for next_id, sr in enumerate(projected, start=1):
        kw_lower = sr.keyword.lower()

        # tivl_tag.
        intent_label = intent_map.get(kw_lower)
        tivl_tag = _intent_to_tivl(intent_label, default=default_tivl_tag)
        if tivl_tag not in _TIVL_ENUM:
            raise TivlTagDriftError(
                f"tivl_tag drift for keyword {sr.keyword!r}: got {tivl_tag!r} "
                f"(expected one of {sorted(_TIVL_ENUM)})"
            )

        # cluster cross-ref.
        cluster = cluster_map.get(kw_lower, "")

        # url_slug — idempotent.
        slug = make_url_slug(sr.keyword)
        if slug and slug in seen_slugs and seen_slugs[slug] != sr.keyword:
            raise UrlSlugUniquenessError(
                f"url_slug uniqueness violation within batch: {slug!r} "
                f"produced by both {seen_slugs[slug]!r} and {sr.keyword!r}"
            )
        seen_slugs[slug] = sr.keyword

        # monthly_volume — overview override > staging value.
        volume = overview_volume_map.get(kw_lower, sr.monthly_volume)
        volume = int(volume)

        # target_word_count — per-tag override.
        twc = _target_word_count(
            tivl_tag=tivl_tag,
            target_word_count_default=target_word_count,
        )

        # priority.
        priority = _priority_label(
            gap_score=sr.gap_score,
            monthly_volume=volume,
        )

        # title — Title Case projection.
        title = sr.keyword.strip()
        # Capitalise each whitespace-separated word; preserve original
        # casing of acronyms is out-of-scope (planner refines manually).
        title = " ".join(w[:1].upper() + w[1:] for w in title.split() if w)

        row = {
            "id": int(next_id),
            "title": title,
            "url_slug": slug,
            "primary_keyword": sr.keyword,
            "monthly_volume": int(volume),
            "assigned_cluster": cluster,
            "target_word_count": int(twc),
            "priority": priority,
            "created_date": created,
            "tivl_tag": tivl_tag,
            "lifecycle_status": _DEFAULT_LIFECYCLE,
        }

        # Schema-shape self-check (defensive — catches any future drift in
        # the row literal above before it lands in the workbook).
        if tuple(row.keys()) != NEW_CONTENT_PLAN_COLUMNS:
            raise RowSchemaError(
                f"row column drift: got {tuple(row.keys())}, "
                f"expected {NEW_CONTENT_PLAN_COLUMNS}"
            )
        # Type / enum invariants.
        if not isinstance(row["id"], int):
            raise RowSchemaError(
                f"id must be int, got {type(row['id']).__name__}"
            )
        if not isinstance(row["monthly_volume"], int):
            raise RowSchemaError(
                f"monthly_volume must be int, got {type(row['monthly_volume']).__name__}"
            )
        if not isinstance(row["target_word_count"], int):
            raise RowSchemaError(
                f"target_word_count must be int, got "
                f"{type(row['target_word_count']).__name__}"
            )
        if row["tivl_tag"] not in _TIVL_ENUM:
            raise TivlTagDriftError(
                f"tivl_tag {row['tivl_tag']!r} not in master-excel enum "
                f"{sorted(_TIVL_ENUM)}"
            )
        if row["lifecycle_status"] not in _LIFECYCLE_ENUM:
            raise RowSchemaError(
                f"lifecycle_status {row['lifecycle_status']!r} not in "
                f"master-excel enum {sorted(_LIFECYCLE_ENUM)}"
            )
        # created_date sanity — ISO 8601 date string.
        try:
            datetime.strptime(row["created_date"], "%Y-%m-%d")
        except (TypeError, ValueError):
            raise RowSchemaError(
                f"created_date must be ISO 8601 YYYY-MM-DD, "
                f"got {row['created_date']!r}"
            )

        rows.append(row)

    return {
        "new_content_plan": rows,
        "meta": {
            "input_staging_rows": len(staging_rows),
            "row_count": len(rows),
            "max_rows": int(max_rows),
            "default_status": default_status,
            "default_tivl_tag": default_tivl_tag,
            "target_word_count": int(target_word_count),
            "created_date": created,
            "intent_enrichment_present": raw_search_intent is not None,
            "overview_enrichment_present": raw_keyword_overview is not None,
            "cluster_map_size": len(cluster_map),
        },
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args(argv: Iterable[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="new_content_plan_transform.py",
        description=(
            "Transform Phase 7 content-gaps staging JSON (+ optional DFS "
            "keyword_overview / search_intent enrichment + cluster_keywords "
            "snapshot) → schema-shaped master.xlsx#new_content_plan rows."
        ),
    )
    p.add_argument("--staging", required=True,
                   help="Path to content_gaps_keyword_ideas_*.json staging file.")
    p.add_argument("--cluster-map", default=None,
                   help="Path to cluster_keywords snapshot JSON ({keyword: cluster}).")
    p.add_argument("--raw-keyword-overview", default=None,
                   help="Optional path to keyword_overview JSON (volume refresh).")
    p.add_argument("--raw-search-intent", default=None,
                   help="Optional path to search_intent JSON (tivl_tag assignment).")
    p.add_argument("--default-status", default="TODO",
                   help="statusEnum seed (default: TODO).")
    p.add_argument("--default-tivl-tag", default="I",
                   choices=sorted(_TIVL_ENUM),
                   help="TIVL tag fallback (default: I).")
    p.add_argument("--target-word-count", type=int, default=DEFAULT_TARGET_WORD_COUNT,
                   help=f"Default target_word_count for tivl_tag=I "
                        f"(default: {DEFAULT_TARGET_WORD_COUNT}).")
    p.add_argument("--max-rows", type=int, default=DEFAULT_MAX_ROWS,
                   help=f"DoS guard: top-N rows by gap_score "
                        f"(default: {DEFAULT_MAX_ROWS}).")
    p.add_argument("--date", default=None,
                   help="ISO date (YYYY-MM-DD) for created_date column; defaults to today UTC.")
    p.add_argument("--output-dir", default=None,
                   help="If set, write new_content_plan.json into this directory.")
    return p.parse_args(list(argv))


def _read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def main(argv: list[str]) -> int:
    args = _parse_args(argv)

    staging_path = Path(args.staging)
    if not staging_path.exists():
        print(f"staging JSON not found: {staging_path}", file=sys.stderr)
        return 2
    staging_payload = _read_json(staging_path)

    if args.cluster_map is None:
        print(
            "--cluster-map required: pass a JSON file shaped as "
            '{"keyword_lower": "cluster_name", ...}',
            file=sys.stderr,
        )
        return 2
    cluster_map_path = Path(args.cluster_map)
    if not cluster_map_path.exists():
        print(f"cluster_map JSON not found: {cluster_map_path}", file=sys.stderr)
        return 2
    cluster_map_raw = _read_json(cluster_map_path)
    if not isinstance(cluster_map_raw, dict):
        print(
            f"cluster_map must be a JSON object, got "
            f"{type(cluster_map_raw).__name__}",
            file=sys.stderr,
        )
        return 2
    cluster_map = {str(k).lower(): str(v) for k, v in cluster_map_raw.items()}

    raw_overview: dict | None = None
    if args.raw_keyword_overview:
        ov_path = Path(args.raw_keyword_overview)
        if ov_path.exists():
            raw_overview = _read_json(ov_path)

    raw_intent: dict | None = None
    if args.raw_search_intent:
        intent_path = Path(args.raw_search_intent)
        if intent_path.exists():
            raw_intent = _read_json(intent_path)

    try:
        result = transform(
            staging_payload,
            cluster_map=cluster_map,
            raw_keyword_overview=raw_overview,
            raw_search_intent=raw_intent,
            default_status=args.default_status,
            default_tivl_tag=args.default_tivl_tag,
            target_word_count=args.target_word_count,
            max_rows=args.max_rows,
            created_date=args.date,
        )
    except NewContentPlanError as exc:
        print(f"transform failed: {exc}", file=sys.stderr)
        return 1
    except ValueError as exc:
        # _normalize_dfs_response raises plain ValueError on shape drift.
        print(f"transform failed (DFS shape): {exc}", file=sys.stderr)
        return 1

    if args.output_dir:
        out_dir = Path(args.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / "new_content_plan.json"
        out_path.write_text(
            json.dumps(result["new_content_plan"], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(json.dumps({
            "new_content_plan_path": str(out_path.resolve()),
            "meta": result["meta"],
        }, ensure_ascii=False, indent=2))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))


__all__ = (
    "NEW_CONTENT_PLAN_COLUMNS",
    "DEFAULT_MAX_ROWS",
    "DEFAULT_TARGET_WORD_COUNT",
    "BUDGET_CHECK_SCRIPT",
    "NewContentPlanError",
    "BudgetExceededError",
    "StagingNotFoundError",
    "ClusterKeywordsEmptyError",
    "RowSchemaError",
    "TivlTagDriftError",
    "UrlSlugUniquenessError",
    "StagingSchemaError",          # re-exported from scripts.ingestion.dfs_pull
    "_normalize_dfs_response",     # re-exported (IDENTITY preserved) for callers
    "preflight_budget",
    "make_url_slug",
    "locate_content_gaps_staging",
    "read_cluster_keywords_snapshot",
    "transform",
    "main",
)
