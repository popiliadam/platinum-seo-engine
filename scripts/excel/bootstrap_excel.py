#!/usr/bin/env python3
"""bootstrap_excel.py - schema-driven Excel template generator.

Reads ``schemas/master-excel.schema.json`` and emits ``templates/master-excel.xlsx``.
Plugin-agnostic: no project slug, no domain, no hardcoded paths beyond the
repo-relative schema input and template output. Idempotent: identical schema
input produces an identical workbook (deterministic core properties).

Disciplines (ADR-008/-009/-010):
- formula_policy = values_only -> only literal header strings are written
- header rows honor each sheet's ``header_row`` from the schema
- dashboard has no ``required_columns`` (KPI cells populated by runtime tools)
- determinism: ZIP entry mtimes pinned to 1980-01-01, core.xml dates pinned
  to 1970-01-01 (UNIX epoch); re-runs produce byte-identical xlsx

Run:
    python3 scripts/excel/bootstrap_excel.py
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMA_PATH = REPO_ROOT / "schemas" / "master-excel.schema.json"
OUTPUT_PATH = REPO_ROOT / "templates" / "master-excel.xlsx"

# Pinned timestamp for deterministic core properties (UNIX epoch in W3CDTF form).
EPOCH_W3CDTF = "1970-01-01T00:00:00Z"
# ZIP/PKZIP cannot represent dates before 1980, so use 1980-01-01 for archive entries.
ZIP_FIXED_DATE = (1980, 1, 1, 0, 0, 0)
# Match any W3CDTF dcterms:created / dcterms:modified value.
_W3CDTF_RE = re.compile(
    r'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)'
)


def load_schema(path: Path = SCHEMA_PATH) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"schema not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def write_headers(ws: Worksheet, sheet_spec: dict) -> None:
    """Write column headers literally; no formulas, no defaults beyond names.

    Sheets without ``required_columns`` (e.g. dashboard) get an empty grid.
    """
    columns = sheet_spec.get("required_columns")
    if not columns:
        return
    header_row = sheet_spec.get("header_row", 1)
    for col_def in columns:
        col_letter = col_def["col"]
        name = col_def["name"]
        if isinstance(name, str) and name.startswith("="):
            raise ValueError(
                f"formula_policy violation: column '{col_letter}' has formula-like name '{name}'"
            )
        ws[f"{col_letter}{header_row}"] = name


def build_workbook(schema: dict) -> Workbook:
    sheets = schema.get("sheets")
    if not isinstance(sheets, dict) or not sheets:
        raise ValueError("schema.sheets is empty or malformed")
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet seeded by Workbook()
    for sheet_id, sheet_spec in sheets.items():
        ws = wb.create_sheet(title=sheet_id)
        write_headers(ws, sheet_spec)
    # Best-effort core property pin; openpyxl rewrites ``modified`` on save,
    # so the canonical fix is in ``normalize_zip_timestamps`` below.
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    return wb


def _pin_core_xml_dates(payload: bytes) -> bytes:
    """Rewrite ``dcterms:created`` and ``dcterms:modified`` values to EPOCH."""
    text = payload.decode("utf-8")
    pinned = _W3CDTF_RE.sub(rf"\g<1>{EPOCH_W3CDTF}\g<2>", text)
    return pinned.encode("utf-8")


def normalize_zip_timestamps(xlsx_path: Path) -> None:
    """Rewrite the xlsx ZIP container for byte-level determinism.

    openpyxl's ``Workbook.save`` stamps:
      1. each ZIP entry with the current time (breaks file mtimes), and
      2. ``docProps/core.xml`` with a fresh ``dcterms:modified`` value.

    Stream the archive into a fresh ZipFile with a fixed ``date_time`` and
    pin core.xml's W3CDTF dates so re-runs produce identical bytes.
    """
    buffer = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, "r") as src:
        entries = sorted(src.infolist(), key=lambda i: i.filename)
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as dst:
            for info in entries:
                payload = src.read(info.filename)
                if info.filename == "docProps/core.xml":
                    payload = _pin_core_xml_dates(payload)
                fixed = zipfile.ZipInfo(filename=info.filename, date_time=ZIP_FIXED_DATE)
                fixed.compress_type = zipfile.ZIP_DEFLATED
                fixed.external_attr = info.external_attr
                dst.writestr(fixed, payload)
    xlsx_path.write_bytes(buffer.getvalue())


def main() -> None:
    schema = load_schema()
    wb = build_workbook(schema)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    normalize_zip_timestamps(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
