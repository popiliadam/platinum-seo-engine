#!/usr/bin/env python3
"""
transaction.py — Atomic, schema-validated, backup-protected Excel writer.

Single approved write path for `master.xlsx` (rules/excel-discipline.md +
ADR-009: master excel is schema-generated, never hand-edited). 19 sheets
defined in schemas/master-excel.schema.json.

Discipline: atomic write (tempfile + os.fsync + os.replace + dir fsync);
backup-then-write to `{ws}/projects/{slug}/_state/backups/master/
master-{ISO}.xlsx` with FIFO keep-7 rotation; lock on sentinel
`_state/excel.lock` via fcntl.flock(LOCK_EX|LOCK_NB) (NOT on the xlsx
itself — openpyxl re-opens it); 3-layer schema validation (sheet exists
→ per-row jsonschema synthesized from required_columns+definitions
ADR-018 → formula_policy '=' prefix); cell length cap 32767; master_task
gated by allowed_writers; AFTER write emit provenance event via
events_writer.append_provenance(operation="project_excel"). Plugin-
agnostik (no slug literals).

Public API: write/append/replace/update (no delete — rules/append-only-state.md).
`replace` is an idempotent in-sheet refresh (clears the data block below the
schema header_row, re-lands at data_start_row); it is NOT a row delete — the
append-only state ledger (events.jsonl) is untouched and a provenance event is
still emitted for the refresh.

Refs: spec §3 + §8.5; schemas/master-excel.schema.json (definitions,
formula_policy, allowed_writers); rules/excel-discipline, schema-first,
append-only-state, single-source-of-truth; ADR-009, ADR-012, ADR-018,
ADR-021.
Imports events_writer + anomaly_recorder one-way; never imported by them. On a
provenance emit failure the write is KEPT and a durable anomaly is recorded to
the _state/anomalies.jsonl sidecar (operator decision D-B: no roll-back, no
hard-block); reconciliation is a follow-up.
"""

from __future__ import annotations

import fcntl
import functools
import json
import os
import shutil
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from jsonschema import Draft7Validator
from jsonschema.exceptions import ValidationError as _JSValidationError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook

from scripts.state import anomaly_recorder, events_writer
from scripts.validation.validate_schema import build_validator

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_REPO_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_SCHEMA_PATH = _REPO_ROOT / "schemas" / "master-excel.schema.json"

_BACKUPS_KEEP = 7
_EXCEL_CELL_MAX_CHARS = 32_767  # openpyxl/Excel hard limit
_LOCK_STALE_SECONDS = 5 * 60  # 5 minutes
# acquire_blocking=True bounds: wait at most _LOCK_BLOCKING_DEADLINE_SECONDS for a
# concurrent holder to release, polling every _LOCK_POLL_INTERVAL_SECONDS. Fixed
# cadence (NOT exponential backoff) — adequate for a single-operator tool, and the
# deadline keeps a wedged peer from hanging a writer forever.
_LOCK_BLOCKING_DEADLINE_SECONDS = 30.0
_LOCK_POLL_INTERVAL_SECONDS = 0.1

# JSON-schema "type" → Python type tuple acceptable by openpyxl-bound rows.
_TYPE_MAP: dict[str, tuple[type, ...]] = {
    "string": (str,),
    "integer": (int,),         # bool is subclass of int — handled separately
    "number": (int, float),
    "boolean": (bool,),
    "date": (str,),            # we store ISO 8601 strings (rules/time-discipline.md)
}

# ---------------------------------------------------------------------------
# WRITER_REGISTRY — per-sheet allowed-writer convention (LC-1, Q-W3W2B-WRITER-01)
# ---------------------------------------------------------------------------
#
# Codifies which writer identity is *expected* to author each logical sheet.
# `master_task` mirrors schemas/master-excel.schema.json `allowed_writers`
# (the only sheet with a HARD schema-level gate); every other sheet carries
# allowed_writers=None in the schema (no hard gate) and historically accepted
# arbitrary writer strings — `_check_writer_scope` is vacuous when
# allowed_writers is None — so the producing-skill convention was lost.
#
# v1.9 = OPT-IN / ADVISORY ONLY. WRITER_REGISTRY does NOT gate writes; the
# write/append/update path is intentionally UNCHANGED (spec Risk R6 — must not
# break existing transaction.py callers). `writer_registry_status()` is the
# opt-in surface: a caller (or v2.0 enforcement) may consult it to emit a
# warning when a writer is off-convention. Enforcement is reserved for v2.0.
WRITER_REGISTRY_ENFORCEMENT: bool = False  # v1.9 OPT-IN (advisory); v2.0 flips

WRITER_REGISTRY: dict[str, frozenset[str]] = {
    # Hard-gated sheet — mirrors master-excel.schema.json allowed_writers.
    "master_task":      frozenset({"human", "dashboard_refresh", "done_protocol", "master_task_sync"}),
    # KPI projection sheet — refreshed by the dashboard projector.
    "dashboard":        frozenset({"dashboard_refresh"}),
    # Planning sheets.
    "topical_map":      frozenset({"topical-map"}),
    "cluster_keywords": frozenset({"cluster-map"}),
    "cannibalization":  frozenset({"cannibalization"}),
    "quick_wins":       frozenset({"quick-wins"}),
    "opportunity":      frozenset({"quick-wins"}),
    "new_content_plan": frozenset({"new-content-plan"}),
    "content_improve":  frozenset({"content-decay"}),
    "content_decay":    frozenset({"content-decay"}),
    # Ingestion / audit sheets.
    "gsc_performance":  frozenset({"gsc-pull"}),
    "on_page_audit":    frozenset({"on-page-audit"}),
    "tech_seo":         frozenset({"tech-audit"}),
    "gbp_audit":        frozenset({"gbp-audit"}),
    "schema":           frozenset({"schema-audit"}),
    "crawl_sitemap":    frozenset({"sf-import"}),
    "robots_txt":       frozenset({"tech-audit"}),
    "redirect_404":     frozenset({"sf-import"}),
    "completed_work":   frozenset({"done-protocol"}),
}

# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class TransactionError(Exception):
    """Base class for all transaction.py errors."""


class SchemaSheetMismatchError(TransactionError):
    """The requested sheet is not declared in master-excel.schema.json."""


class RowSchemaError(TransactionError):
    """A row failed per-column schema validation."""


class FormulaPolicyViolation(TransactionError):
    """A cell value starts with '=' — formulas are forbidden (formula_policy)."""


class CellValueTooLongError(TransactionError):
    """A cell value exceeds Excel's 32,767-character cap."""


class LockHeldError(TransactionError):
    """Another writer currently holds the excel.lock sentinel (fail-fast path)."""


class LockTimeout(TransactionError):
    """acquire_blocking=True waited past its bounded deadline for excel.lock.

    Distinct from LockHeldError so callers can tell "waited and gave up" from
    "fail-fast refused immediately". A sibling under TransactionError (not a
    LockHeldError subclass), so `except LockHeldError` never swallows a timeout,
    yet a broad `except TransactionError` still catches both.
    """


class BackupDirUnwritableError(TransactionError):
    """Cannot create or write into the backups/master/ directory."""


class WorkbookLockedByExcelError(TransactionError):
    """The workbook is open in Excel/LibreOffice (~$ lockfile present)."""


class WorkbookCorruptError(TransactionError):
    """openpyxl could not parse the workbook on read."""


class WriterScopeError(TransactionError):
    """master_task write attempted with missing/disallowed writer identity."""


# ---------------------------------------------------------------------------
# Result type
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class WriteResult:
    workbook_path: Path
    sheet: str
    rows_affected: int
    backup_path: Path
    event_id: str  # provenance event_id emitted post-write


# ---------------------------------------------------------------------------
# Internals — schema loading + per-sheet validator synthesis
# ---------------------------------------------------------------------------

def _log(msg: str) -> None:
    print(msg, file=sys.stderr)


@functools.lru_cache(maxsize=4)
def _load_schema(schema_path: str) -> dict:
    p = Path(schema_path)
    with p.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _resolve_definition_ref(schema: dict, ref: str) -> dict:
    """Resolve '#/definitions/<name>' against the master-excel schema."""
    if not ref.startswith("#/definitions/"):
        raise RowSchemaError(f"unsupported $ref form: {ref!r}")
    name = ref.split("/", 2)[2]
    defs = schema.get("definitions") or {}
    if name not in defs:
        raise RowSchemaError(f"definition not found: {name!r} (ADR-018)")
    return defs[name]


def _build_row_validator(schema: dict, sheet: str, *, allow_extra: bool = False) -> Draft7Validator:
    """Synthesize a Draft7 validator for a single row of the named sheet.

    allow_extra (P1-09): when False (default), unknown row keys are REJECTED
    (additionalProperties:false) so a typo'd column name surfaces instead of
    being silently dropped at write time. Pass True to keep the legacy behaviour
    (extras validate, then get skipped because they have no column)."""
    sheets = schema.get("sheets") or {}
    if sheet not in sheets:
        raise SchemaSheetMismatchError(
            f"sheet {sheet!r} not declared in master-excel.schema.json"
        )
    sdef = sheets[sheet]
    cols = sdef.get("required_columns") or []
    if not cols:
        # Sheets like 'dashboard' use required_cells (KPI projection) — not
        # supported by row-style write/append/update. Caller must use a
        # different code path for KPI projection writes.
        raise SchemaSheetMismatchError(
            f"sheet {sheet!r} has no required_columns (use KPI projection)"
        )

    properties: dict[str, Any] = {}
    required: list[str] = []
    for col in cols:
        name = col["name"]
        required.append(name)
        prop: dict[str, Any] = {}
        ctype = col.get("type")
        cenum = col.get("enum")
        cref = col.get("ref")
        if cref:
            ref_def = _resolve_definition_ref(schema, cref)
            # Nullable, consistent with typed/string columns below: a blank cell
            # (None) is a valid sparse value. Without this, update()'s full-row
            # re-validation rejects unrelated edits to any row that carries a
            # blank enum/ref cell (deep-audit 2026-06-04 HIGH). A non-blank value
            # is still constrained to the enum.
            prop = {"anyOf": [ref_def, {"type": "null"}]}
        elif cenum:
            prop["type"] = ["string", "null"]
            prop["enum"] = list(cenum) + [None]
        elif ctype:
            # Allow None as well (sparse columns are common in Excel land).
            json_types = {
                "integer": "integer",
                "number":  "number",
                "boolean": "boolean",
                "date":    "string",
            }.get(ctype)
            if json_types is None:
                prop["type"] = "string"
            else:
                prop["type"] = [json_types, "null"]
        else:
            prop["type"] = ["string", "null"]
        properties[name] = prop

    row_schema = {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        # P1-09: closed by default — an unknown key is almost always a typo and
        # would otherwise be silently dropped at write time. allow_extra=True
        # reopens it for callers that intentionally carry non-column metadata.
        "additionalProperties": bool(allow_extra),
        "required": required,
        "properties": properties,
    }
    # build_validator (not a raw Draft7Validator) so the engine's single strict
    # FormatChecker governs row validation too — today the dynamic row_schema
    # carries only `type` (no `format`), so this is behavior-preserving, but any
    # future date-time/uri column would then be enforced consistently (P1-02).
    return build_validator(row_schema)


# ---------------------------------------------------------------------------
# Internals — formula policy + cell length checks
# ---------------------------------------------------------------------------

def _check_formula_policy(rows: list[dict]) -> None:
    for idx, row in enumerate(rows):
        for k, v in row.items():
            if isinstance(v, str) and v.lstrip().startswith("="):
                raise FormulaPolicyViolation(
                    f"row {idx} field {k!r}: formula prefix '=' forbidden"
                )


def _check_cell_lengths(rows: list[dict]) -> None:
    for idx, row in enumerate(rows):
        for k, v in row.items():
            if isinstance(v, str) and len(v) > _EXCEL_CELL_MAX_CHARS:
                raise CellValueTooLongError(
                    f"row {idx} field {k!r}: {len(v)} chars > {_EXCEL_CELL_MAX_CHARS}"
                )


def _validate_rows(rows: list[dict], validator: Draft7Validator) -> None:
    for idx, row in enumerate(rows):
        errors = sorted(validator.iter_errors(row), key=lambda e: e.path)
        if errors:
            err: _JSValidationError = errors[0]
            loc = "/".join(str(p) for p in err.absolute_path) or "<row>"
            raise RowSchemaError(f"row {idx} {loc}: {err.message}")


# ---------------------------------------------------------------------------
# Internals — lock sentinel
# ---------------------------------------------------------------------------

def _state_dir(state_root: Path | None, workbook_path: Path, project_slug: str) -> Path:
    """
    Resolve {ws}/projects/{slug}/_state/. If state_root explicitly given,
    use it directly (test convenience). Else derive from workbook_path:
    walk up to find projects/{slug}/_state/.
    """
    if state_root is not None:
        sd = Path(state_root)
        sd.mkdir(parents=True, exist_ok=True)
        return sd
    # Default: workbook_path is .../projects/{slug}/{filename}.xlsx
    sd = workbook_path.parent / "_state"
    sd.mkdir(parents=True, exist_ok=True)
    return sd


def _read_sentinel(lock_path: Path) -> tuple[int | None, str | None]:
    if not lock_path.exists():
        return None, None
    try:
        raw = lock_path.read_text(encoding="utf-8").strip()
        obj = json.loads(raw)
        return obj.get("pid"), obj.get("ts")
    except (OSError, json.JSONDecodeError, ValueError):
        return None, None


def _is_stale(pid: int | None, ts: str | None) -> bool:
    if pid is None or ts is None:
        # Corrupt sentinel = stale.
        return True
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        # No process with that PID → stale.
        return True
    except PermissionError:
        # Process exists but is not ours; not stale.
        pid_alive = True
    else:
        pid_alive = True
    if not pid_alive:  # pragma: no cover (ProcessLookupError handled above)
        return True
    # Age check.
    try:
        when = datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except ValueError:
        return True
    age = (datetime.now(timezone.utc) - when).total_seconds()
    return age > _LOCK_STALE_SECONDS


def _flock_or_raise(
    fd: int,
    lock_path: Path,
    *,
    acquire_blocking: bool,
    deadline_seconds: float,
) -> None:
    """flock(LOCK_EX) on `fd`. Closes `fd` before raising so it never leaks.

    Default (acquire_blocking=False): ONE LOCK_NB attempt — a concurrent holder
    raises LockHeldError immediately (byte-identical to the legacy fail-fast).

    acquire_blocking=True: retry LOCK_NB on a fixed cadence until it succeeds or
    `deadline_seconds` elapses, then raise LockTimeout. Bounded (never an
    unbounded blocking LOCK_EX) so a wedged peer can't hang a writer forever.
    Only EWOULDBLOCK (BlockingIOError = lock held) is retried; any other OSError
    (e.g. ENOLCK) is a non-contention failure and raises at once even when
    blocking.
    """
    deadline = time.monotonic() + deadline_seconds
    while True:
        try:
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            return
        except BlockingIOError as exc:
            if not acquire_blocking:
                os.close(fd)
                raise LockHeldError(f"another writer holds {lock_path}") from exc
            if time.monotonic() >= deadline:
                os.close(fd)
                raise LockTimeout(
                    f"waited {deadline_seconds:g}s for {lock_path}, still held"
                ) from exc
            time.sleep(_LOCK_POLL_INTERVAL_SECONDS)
        except OSError as exc:
            os.close(fd)
            raise LockHeldError(f"flock failed on {lock_path}: {exc}") from exc


def _acquire_lock(
    lock_path: Path,
    *,
    acquire_blocking: bool = False,
    deadline_seconds: float = _LOCK_BLOCKING_DEADLINE_SECONDS,
) -> int:
    """Acquire excel.lock; return the fd to hold.

    acquire_blocking=False (default): fail-fast — raise LockHeldError if another
    writer holds the lock (unchanged behaviour; no regression for any caller).
    acquire_blocking=True: wait up to `deadline_seconds` for the holder to
    release, then raise LockTimeout. Opt-in parallel-write safety so a concurrent
    second writer waits its turn instead of silently losing its work.
    """
    # First, stale reclaim. Sentinel content is informational; flock provides
    # the actual mutual exclusion below.
    pid, ts = _read_sentinel(lock_path)
    if pid is not None and _is_stale(pid, ts):
        try:
            lock_path.unlink()
        except FileNotFoundError:  # pragma: no cover
            pass

    fd = os.open(str(lock_path), os.O_WRONLY | os.O_CREAT, 0o644)
    _flock_or_raise(
        fd, lock_path,
        acquire_blocking=acquire_blocking, deadline_seconds=deadline_seconds,
    )

    # Refresh sentinel content for the holder. Truncate + rewrite is OK:
    # this file is informational, not the durability target.
    payload = json.dumps({
        "pid": os.getpid(),
        "ts": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
    }, separators=(",", ":")).encode("utf-8") + b"\n"
    try:
        os.ftruncate(fd, 0)
        os.lseek(fd, 0, os.SEEK_SET)
        os.write(fd, payload)
        os.fsync(fd)
    except OSError:  # pragma: no cover (informational, not blocking)
        pass
    return fd


def _release_lock(fd: int, lock_path: Path) -> None:
    try:
        fcntl.flock(fd, fcntl.LOCK_UN)
    except OSError:  # pragma: no cover
        pass
    try:
        os.close(fd)
    except OSError:  # pragma: no cover
        pass
    # Best-effort sentinel cleanup; OK if it lingers (next writer reclaims).
    try:
        lock_path.unlink()
    except FileNotFoundError:  # pragma: no cover
        pass


# ---------------------------------------------------------------------------
# Internals — backup + atomic save
# ---------------------------------------------------------------------------

def _iso_compact() -> str:
    """Backup-friendly UTC timestamp: 2026-04-30T12-00-00-123456Z (lex-sortable)."""
    now = datetime.now(timezone.utc)
    # Substitute ':' with '-' so it's a valid filename on every OS.
    return now.strftime("%Y-%m-%dT%H-%M-%S-%fZ")


def _backup(target: Path, state_dir: Path) -> Path:
    backup_dir = state_dir / "backups" / "master"
    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise BackupDirUnwritableError(f"cannot create {backup_dir}: {exc}") from exc

    if not target.exists():
        # First-write case: there is nothing to back up. Touch a marker
        # so the rotation invariant ("a backup precedes every write") is
        # satisfied. Marker has the same name shape so it sorts naturally.
        marker = backup_dir / f"master-{_iso_compact()}.empty"
        marker.write_bytes(b"")
        return marker

    bp = backup_dir / f"master-{_iso_compact()}.xlsx"
    try:
        shutil.copy2(target, bp)
    except OSError as exc:
        raise BackupDirUnwritableError(f"cannot copy {target} → {bp}: {exc}") from exc
    return bp


def _unlink_quietly(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:  # pragma: no cover
        pass


def _rotate_backups(state_dir: Path) -> None:
    """Prune backups to the FIFO keep-7 budget, scoped to real snapshots.

    The budget counts ONLY `master-*.xlsx` — the SAME glob the F-22 invariant
    (validate_invariants.check_F_22) and dump_workspace._backups_recent use. The
    previous all-files `iterdir()` rotation let first-write `.empty` markers (and
    any stray tempfile) count against the keep-7 budget, so a newer marker could
    evict a real `master-*.xlsx` snapshot while F-22 (which only counts xlsx)
    still reported compliant — a silently lost recoverable backup.

    `.empty` markers exist only to satisfy "a backup precedes the first write"
    when there is nothing to copy; once any real snapshot exists they are
    vestigial, so they are pruned to zero and never share the xlsx budget.
    """
    backup_dir = state_dir / "backups" / "master"
    if not backup_dir.is_dir():
        return
    # Lex-sort by name; the embedded ISO timestamps are lex-sortable (oldest 1st).
    snapshots = sorted(backup_dir.glob("master-*.xlsx"))
    excess = len(snapshots) - _BACKUPS_KEEP
    for old in snapshots[:max(0, excess)]:
        _unlink_quietly(old)
    # Vestigial once a real snapshot exists; the lone first-write marker (no
    # snapshot yet) is left intact as the only proof a backup preceded it.
    if snapshots:
        for marker in backup_dir.glob("master-*.empty"):
            _unlink_quietly(marker)


def _atomic_save(wb: Workbook, target: Path) -> None:
    """Save wb to a sibling tempfile, fsync, then os.replace target."""
    parent = target.parent
    parent.mkdir(parents=True, exist_ok=True)
    tmp_fd, tmp_name = tempfile.mkstemp(
        prefix=".master-", suffix=".xlsx.tmp", dir=str(parent)
    )
    tmp_path = Path(tmp_name)
    try:
        os.close(tmp_fd)  # openpyxl re-opens the path
        wb.save(str(tmp_path))
        # Re-open the file just to fsync the data we wrote.
        with tmp_path.open("rb") as fh:
            os.fsync(fh.fileno())
        os.replace(str(tmp_path), str(target))
    except Exception:
        # Cleanup leftover tempfile on failure (do NOT swallow the error).
        try:
            tmp_path.unlink()
        except FileNotFoundError:  # pragma: no cover
            pass
        raise

    # Parent dir fsync — durability of the rename.
    try:
        dir_fd = os.open(str(parent), os.O_RDONLY)
        try:
            os.fsync(dir_fd)
        finally:
            os.close(dir_fd)
    except OSError:  # pragma: no cover (best-effort, not all FS support it)
        pass


def _excel_lockfile_present(target: Path) -> bool:
    """Detect Excel/LibreOffice owner-lock sidecar (~$Filename.xlsx)."""
    sidecar = target.parent / f"~${target.name}"
    return sidecar.exists()


def _load_or_new(target: Path) -> Workbook:
    """Load an existing wb or create a new one. Wraps openpyxl exceptions."""
    if not target.exists():
        return Workbook()
    if _excel_lockfile_present(target):
        raise WorkbookLockedByExcelError(
            f"workbook {target} appears open in Excel/LibreOffice "
            f"(found ~${target.name})"
        )
    try:
        return load_workbook(str(target))
    except InvalidFileException as exc:
        raise WorkbookCorruptError(f"cannot parse {target}: {exc}") from exc
    except OSError as exc:
        raise WorkbookCorruptError(f"cannot read {target}: {exc}") from exc


def _check_writer_scope(schema: dict, sheet: str, writer: str | None) -> None:
    sdef = schema.get("sheets", {}).get(sheet, {})
    allowed = sdef.get("allowed_writers")
    if allowed is None:
        return
    if writer is None:
        raise WriterScopeError(
            f"sheet {sheet!r} requires writer arg (allowed: {allowed})"
        )
    if writer not in allowed:
        raise WriterScopeError(
            f"writer {writer!r} not in allowed_writers for {sheet!r}: {allowed}"
        )


def writer_registry_status(
    sheet: str,
    writer: str | None,
    *,
    enforce: bool = WRITER_REGISTRY_ENFORCEMENT,
) -> str:
    """Advisory per-sheet writer-convention check (LC-1, Q-W3W2B-WRITER-01).

    Returns:
        "ok"           — writer is registered for the sheet, OR the sheet is
                         not in WRITER_REGISTRY, OR writer is None (vacuous).
        "unregistered" — the sheet IS in WRITER_REGISTRY but `writer` is not
                         an allowed writer for it. A warning is logged.

    v1.9 contract: ADVISORY ONLY — NEVER raises and NEVER gates a write (spec
    Risk R6). `enforce` defaults to WRITER_REGISTRY_ENFORCEMENT (False in
    v1.9) and is reserved for v2.0; even when True it currently only escalates
    the log marker. The sole HARD write gate remains `_check_writer_scope`
    (schema allowed_writers). transaction.write/append/update do NOT call this
    helper — opting in is the caller's choice.
    """
    allowed = WRITER_REGISTRY.get(sheet)
    if allowed is None or writer is None or writer in allowed:
        return "ok"
    msg = (
        f"writer {writer!r} off-convention for sheet {sheet!r} "
        f"(WRITER_REGISTRY allows {sorted(allowed)}) — advisory "
        f"(LC-1 OPT-IN; enforcement reserved v2.0)"
    )
    _log(f"WARNING{' [enforce]' if enforce else ''}: {msg}")
    return "unregistered"


# ---------------------------------------------------------------------------
# Internals — sheet I/O (header detection + row write)
# ---------------------------------------------------------------------------

def _ordered_columns(schema: dict, sheet: str) -> list[dict]:
    return list(schema["sheets"][sheet]["required_columns"])


def _ensure_sheet_with_header(
    wb: Workbook, sheet: str, columns: list[dict], header_row: int = 1
) -> None:
    """Create the sheet if missing; ensure the header lands at `header_row`.

    P1-09: previously the header was always written to row 1, which disagreed
    with the schema header_row (>=3 for 18/19 sheets) and the data_start_row
    (=header_row+1) used by replace() and every reader — so a freshly-created
    sheet had its header at row 1 but data was expected at row 5. Writing the
    header at the declared header_row keeps creation, replace(), append's
    _next_data_row landing, and all readers consistent, and stops the spurious
    row-1 header that clobbered template title blocks on bootstrap-seeded sheets.
    """
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        # Drop a bare "Sheet" if it's the only one and we're adding the first.
        ws = wb.create_sheet(sheet)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"]["A1"].value is None:
            del wb["Sheet"]
    # Always (re)write the header row in canonical order at the schema header_row.
    header_names = [c["name"] for c in columns]
    for col_idx, name in enumerate(header_names, start=1):
        ws.cell(row=header_row, column=col_idx, value=name)


def _row_to_tuple(row: dict, columns: list[dict]) -> tuple:
    return tuple(row.get(col["name"]) for col in columns)


def _next_data_row(ws) -> int:
    """First empty row after the existing data block."""
    return (ws.max_row or 1) + 1


def _data_start_row(schema: dict, sheet: str) -> int:
    """First physical row where DATA lives for `sheet`.

    Template-seeded sheets (bootstrap_excel.py honors each sheet's header_row)
    place data below a multi-row header block: a header_row=3 sheet has data at
    row 4, a header_row=4 sheet at row 5. The schema sheet dict carries this as
    `data_start_row`. Readers (weekly_summary, portfolio_heatmap,
    validate_invariants) all read from there, so the replace writer must land
    data there too.

    Fallback: a sheet that declares no `data_start_row` keeps the legacy
    header@1/data@2 layout (return 2), matching `_ensure_sheet_with_header`.
    """
    sdef = schema.get("sheets", {}).get(sheet, {})
    declared = sdef.get("data_start_row")
    if isinstance(declared, int) and declared >= 1:
        return declared
    return 2


def _header_row(schema: dict, sheet: str) -> int:
    """Physical row where the column HEADER lives for `sheet` (P1-09).

    Mirrors _data_start_row: returns the schema's declared header_row (3/4/5 for
    template-seeded sheets) or 1 for legacy sheets that declare none. The invariant
    data_start_row == header_row + 1 holds across the schema, so writing the header
    here keeps append/replace landing + readers aligned.
    """
    sdef = schema.get("sheets", {}).get(sheet, {})
    declared = sdef.get("header_row")
    if isinstance(declared, int) and declared >= 1:
        return declared
    return 1


# ---------------------------------------------------------------------------
# Public API — write / append / update
# ---------------------------------------------------------------------------

def write(
    workbook_path: Path | str,
    sheet: str,
    rows: list[dict],
    project_slug: str,
    *,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
    writer: str | None = None,
    allow_extra: bool = False,
    acquire_blocking: bool = False,
) -> WriteResult:
    """
    Append-and-replace write: appends `rows` to `sheet`, atomically.

    The 'replace' aspect is not destructive: existing rows are preserved.
    Use update() for in-place mutation. The function name 'write' is
    retained for back-compat with the brief. allow_extra (P1-09): keep False to
    reject unknown row keys (typo guard); True restores the legacy skip.
    acquire_blocking (default False = fail-fast): True waits a bounded deadline
    for a concurrent holder instead of raising LockHeldError immediately.
    """
    return _write_or_append(
        workbook_path, sheet, rows, project_slug,
        schema_path=schema_path, state_root=state_root, writer=writer,
        mode="append", allow_extra=allow_extra, acquire_blocking=acquire_blocking,
    )


def append(
    workbook_path: Path | str,
    sheet: str,
    rows: list[dict],
    project_slug: str,
    *,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
    writer: str | None = None,
    allow_extra: bool = False,
    acquire_blocking: bool = False,
) -> WriteResult:
    """Append rows to the bottom of `sheet`. Alias for write(mode='append').
    allow_extra (P1-09): keep False to reject unknown row keys; True skips them.
    acquire_blocking (default False = fail-fast): True waits a bounded deadline
    for a concurrent holder instead of raising LockHeldError immediately."""
    return _write_or_append(
        workbook_path, sheet, rows, project_slug,
        schema_path=schema_path, state_root=state_root, writer=writer,
        mode="append", allow_extra=allow_extra, acquire_blocking=acquire_blocking,
    )


def replace(
    workbook_path: Path | str,
    sheet: str,
    rows: list[dict],
    project_slug: str,
    *,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
    writer: str | None = None,
    allow_extra: bool = False,
    acquire_blocking: bool = False,
) -> WriteResult:
    """Idempotent refresh: clear `sheet`'s data block, then write `rows`.

    Unlike append(), re-importing the same source does NOT duplicate rows: the
    prior data block (everything from the schema's data_start_row down) is
    cleared first, then `rows` are written starting at data_start_row. The
    header block (rows 1..data_start_row-1) is preserved verbatim. Schema
    validation runs BEFORE the lock/clear, so an invalid payload leaves the
    sheet untouched. Same atomic envelope as append (backup → atomic_save →
    rotate, provenance emit). allow_extra (P1-09): keep False to reject unknown
    row keys; True skips them. acquire_blocking (default False = fail-fast): True
    waits a bounded deadline for a concurrent holder instead of raising at once.
    """
    return _write_or_append(
        workbook_path, sheet, rows, project_slug,
        schema_path=schema_path, state_root=state_root, writer=writer,
        mode="replace", allow_extra=allow_extra, acquire_blocking=acquire_blocking,
    )


def update(
    workbook_path: Path | str,
    sheet: str,
    where: dict,
    set_: dict,
    project_slug: str,
    *,
    schema_path: Path | str | None = None,
    state_root: Path | str | None = None,
    writer: str | None = None,
    acquire_blocking: bool = False,
) -> WriteResult:
    """
    Update existing rows where every key/value in `where` matches; apply
    `set_` to those rows. Returns rows_affected = number of matched rows.

    Raises if no rows match (callers should append() instead of silently
    failing). Schema-validates the resulting row(s) end-to-end. acquire_blocking
    (default False = fail-fast): True waits a bounded deadline for a concurrent
    holder instead of raising LockHeldError immediately.
    """
    workbook_path = Path(workbook_path)
    schema_p = Path(schema_path) if schema_path else _DEFAULT_SCHEMA_PATH
    schema = _load_schema(str(schema_p))

    if sheet not in (schema.get("sheets") or {}):
        raise SchemaSheetMismatchError(f"sheet {sheet!r} not in schema")
    columns = _ordered_columns(schema, sheet)
    _check_writer_scope(schema, sheet, writer)

    # Lock + load.
    state_dir = _state_dir(
        Path(state_root) if state_root else None, workbook_path, project_slug
    )
    lock_path = state_dir / "excel.lock"
    fd = _acquire_lock(lock_path, acquire_blocking=acquire_blocking)
    try:
        wb = _load_or_new(workbook_path)
        _ensure_sheet_with_header(wb, sheet, columns, _header_row(schema, sheet))
        ws = wb[sheet]

        col_index = {c["name"]: idx for idx, c in enumerate(columns, start=1)}
        affected_rows: list[dict] = []
        # Walk data rows from the schema's data_start_row down — never a hardcoded
        # row 2. Template-seeded sheets (header_row>1) keep header/blank rows above
        # data_start_row; scanning from row 2 would match and mutate them. Mirrors
        # the data_start_row contract _write_or_append/replace already honor.
        data_start = _data_start_row(schema, sheet)
        for row_idx in range(data_start, (ws.max_row or 1) + 1):
            current = {
                c["name"]: ws.cell(row=row_idx, column=col_index[c["name"]]).value
                for c in columns
            }
            if all(current.get(k) == v for k, v in where.items()):
                merged = {**current, **set_}
                affected_rows.append(merged)
                # Validate the merged row before flushing.
                validator = _build_row_validator(schema, sheet)
                _validate_rows([merged], validator)
                _check_formula_policy([merged])
                _check_cell_lengths([merged])
                # Apply mutation to ws cells.
                for k, v in set_.items():
                    if k not in col_index:
                        raise RowSchemaError(f"unknown column in set_: {k!r}")
                    ws.cell(row=row_idx, column=col_index[k], value=v)

        if not affected_rows:
            raise RowSchemaError(
                f"update matched 0 rows in {sheet!r} for where={where!r}"
            )

        backup_path = _backup(workbook_path, state_dir)
        _atomic_save(wb, workbook_path)
        _rotate_backups(state_dir)
    finally:
        _release_lock(fd, lock_path)

    # D-B: a provenance emit failure records a durable anomaly + keeps the
    # workbook (no roll-back, no hard-block) — never propagates. state_dir is the
    # project's _state dir (anomalies.jsonl lands beside events.jsonl).
    event_id = _emit_provenance_or_anomaly(
        project_slug=project_slug,
        sheet=sheet,
        rows_affected=len(affected_rows),
        state_dir=state_dir,
        operation="project_excel",
    )
    return WriteResult(
        workbook_path=workbook_path,
        sheet=sheet,
        rows_affected=len(affected_rows),
        backup_path=backup_path,
        event_id=event_id,
    )


# ---------------------------------------------------------------------------
# Internals — core write/append driver
# ---------------------------------------------------------------------------

def _write_or_append(
    workbook_path: Path | str,
    sheet: str,
    rows: list[dict],
    project_slug: str,
    *,
    schema_path: Path | str | None,
    state_root: Path | str | None,
    writer: str | None,
    mode: str,
    allow_extra: bool = False,
    acquire_blocking: bool = False,
) -> WriteResult:
    if mode not in ("append", "replace"):
        raise TransactionError(f"unsupported mode: {mode!r}")

    workbook_path = Path(workbook_path)
    schema_p = Path(schema_path) if schema_path else _DEFAULT_SCHEMA_PATH
    schema = _load_schema(str(schema_p))

    # 1. Schema layer: sheet exists.
    if sheet not in (schema.get("sheets") or {}):
        raise SchemaSheetMismatchError(f"sheet {sheet!r} not in schema")
    columns = _ordered_columns(schema, sheet)

    # 2. Writer scope (master_task only, currently).
    _check_writer_scope(schema, sheet, writer)

    # 3. Per-row schema validation. P1-09: unknown keys rejected unless allow_extra.
    validator = _build_row_validator(schema, sheet, allow_extra=allow_extra)
    if not isinstance(rows, list):
        raise RowSchemaError("rows must be a list of dicts")
    _validate_rows(rows, validator)

    # 4. Formula policy + cell length cap.
    _check_formula_policy(rows)
    _check_cell_lengths(rows)

    # 5. Lock + load + write under flock.
    state_dir = _state_dir(
        Path(state_root) if state_root else None, workbook_path, project_slug
    )
    lock_path = state_dir / "excel.lock"
    fd = _acquire_lock(lock_path, acquire_blocking=acquire_blocking)
    try:
        wb = _load_or_new(workbook_path)
        _ensure_sheet_with_header(wb, sheet, columns, _header_row(schema, sheet))
        ws = wb[sheet]

        # Resolve the write cursor by mode. append: bottom of existing data
        # (byte-identical legacy behavior). replace: clear the prior data block
        # then re-land at the schema's data_start_row (idempotent refresh).
        # Rows are already schema-validated above (step 3), so a clear here is
        # only reached for a valid payload — a bad row raised before the lock.
        if mode == "replace":
            data_start = _data_start_row(schema, sheet)
            if (ws.max_row or 0) >= data_start:
                ws.delete_rows(data_start, ws.max_row - data_start + 1)
            write_row = data_start
        else:
            write_row = _next_data_row(ws)

        col_index = {c["name"]: idx for idx, c in enumerate(columns, start=1)}
        for row in rows:
            for k, v in row.items():
                if k not in col_index:
                    # Validator allowed this key (additionalProperties:true);
                    # we still skip writing it because it has no column.
                    continue
                ws.cell(row=write_row, column=col_index[k], value=v)
            write_row += 1

        # 6. Backup → atomic save → rotation. Order matters: backup must
        # complete before save, save must complete before rotation prunes.
        backup_path = _backup(workbook_path, state_dir)
        _atomic_save(wb, workbook_path)
        _rotate_backups(state_dir)
    finally:
        _release_lock(fd, lock_path)

    # 7. Emit provenance event into events.jsonl. Failure to emit does NOT roll
    # back the write (Excel is on disk) and does NOT hard-block the caller —
    # operator decision D-B. Instead a DURABLE ANOMALY is recorded into the
    # _state/anomalies.jsonl sidecar (a SEPARATE path from the events.jsonl whose
    # emit just failed) so the missing provenance event stays reconcilable.
    event_id = _emit_provenance_or_anomaly(
        project_slug=project_slug,
        sheet=sheet,
        rows_affected=len(rows),
        state_dir=state_dir,
        operation="project_excel",
    )

    return WriteResult(
        workbook_path=workbook_path,
        sheet=sheet,
        rows_affected=len(rows),
        backup_path=backup_path,
        event_id=event_id,
    )


# ---------------------------------------------------------------------------
# Internals — event emission helper
# ---------------------------------------------------------------------------

def _emit_provenance(
    *,
    project_slug: str,
    sheet: str,
    rows_affected: int,
    state_root: Path,
    operation: str,
) -> str:
    """
    Emit a provenance event recording the write. The events_writer module
    resolves the events.jsonl path from {workspace_root}/projects/{slug}/_state/.
    state_root here is the project directory — its parent is `projects/`,
    grandparent is the workspace root expected by events_writer.
    """
    workspace_root = state_root.parent.parent  # projects/<slug> → projects → workspace
    # source.kind enum (events.schema): manual | tool_computed | sf_csv | *_mcp.
    # An Excel projection by transaction.py is internally tool_computed —
    # it's the engine writing structured output, not a human key-by-key edit.
    # run_id auto-allocated race-free under the append flock (P1-11).
    result = events_writer.append_provenance(
        project_id=project_slug,
        source={"kind": "tool_computed"},
        operation=operation,
        target_excel_sheet=sheet,
        rows_written=rows_affected,
        workspace_root=workspace_root,
    )
    return result.event_id


def _emit_provenance_or_anomaly(
    *,
    project_slug: str,
    sheet: str,
    rows_affected: int,
    state_dir: Path,
    operation: str,
) -> str:
    """Emit provenance; on emit failure record a durable anomaly (operator D-B).

    The workbook is already saved + the lock released by the time this runs. D-B
    rules that an audit/event emit hiccup must NOT roll back the workbook and must
    NOT hard-block the caller. So on ANY emit failure we:
      1. write a DURABLE ANOMALY into _state/anomalies.jsonl — a SEPARATE path
         from the events.jsonl whose emit just failed (so a schema-load/lock/disk
         failure on events.jsonl can't take the fallback down with it), and
      2. return "" (no event_id) instead of raising.
    A failure of the anomaly write itself is downgraded to a WARNING — a fallback
    failure must never re-brick an already-committed write. Reconciliation
    (drift-check surfacing anomalies.jsonl) is a follow-up.
    """
    try:
        return _emit_provenance(
            project_slug=project_slug,
            sheet=sheet,
            rows_affected=rows_affected,
            state_root=state_dir.parent,  # _state lives at projects/{slug}/_state
            operation=operation,
        )
    except Exception as exc:
        try:
            anomaly_recorder.record_anomaly(
                state_dir,
                kind="provenance_emit_failed",
                detail={
                    "sheet": sheet,
                    "rows_written": rows_affected,
                    "operation": operation,
                },
                error=f"{type(exc).__name__}: {exc}",
            )
        except Exception as rec_exc:  # never re-brick on a fallback failure
            _log(
                f"WARNING: anomaly record ALSO failed for sheet {sheet!r} "
                f"(provenance emit lost, unreconcilable): {rec_exc}"
            )
        _log(
            f"WARNING: provenance emit failed for sheet {sheet!r} — anomaly "
            f"recorded, workbook kept (D-B: no roll-back, no hard-block): {exc}"
        )
        return ""


# ---------------------------------------------------------------------------
# Module export list
# ---------------------------------------------------------------------------

__all__: Iterable[str] = (
    "WriteResult",
    "TransactionError",
    "SchemaSheetMismatchError",
    "RowSchemaError",
    "FormulaPolicyViolation",
    "CellValueTooLongError",
    "LockHeldError",
    "LockTimeout",
    "BackupDirUnwritableError",
    "WorkbookLockedByExcelError",
    "WorkbookCorruptError",
    "WriterScopeError",
    "WRITER_REGISTRY",
    "WRITER_REGISTRY_ENFORCEMENT",
    "writer_registry_status",
    "write",
    "append",
    "replace",
    "update",
)
