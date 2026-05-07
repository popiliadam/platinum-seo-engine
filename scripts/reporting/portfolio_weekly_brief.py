#!/usr/bin/env python3
"""portfolio_weekly_brief.py — Phase 9 Wave 1 reporting transform.

Multi-project LOCAL aggregator across active_projects in
`projects/_portfolio/portfolio.config.json`. READ-ONLY: per-project
events.jsonl (work-kind, 7-day window) + master.xlsx#master_task
(created_date / done_date in window) + last_sync_at vs SLA freshness.

Outputs (frontmatter):
  - "master.xlsx#none" (READ-ONLY confirm)
  - "outputs/reports/{date}-portfolio-weekly-brief.md"
  - "inbox/local/{date}-portfolio-weekly-brief.json"

events.jsonl NOT written (REVIZE 3; Q-RP-01 defer Phase 14). Mirrors
W-D1 master_task_sync discipline (pure-fn first, plugin agnostik,
idempotent, openpyxl read_only=True, no D-003 import).

Schema authority: schemas/portfolio-config.schema.json v1.1.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import date as _date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence


_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


# ---- Constants ------------------------------------------------------------

WEEKDAY_ENUM: tuple[str, ...] = (
    "Monday", "Tuesday", "Wednesday", "Thursday",
    "Friday", "Saturday", "Sunday",
)
_WEEKDAY_TO_ISO: Mapping[str, int] = {n: i + 1 for i, n in enumerate(WEEKDAY_ENUM)}
HOUR_MIN, HOUR_MAX = 0, 23
SLA_MIN_DAYS, SLA_MAX_DAYS = 1, 30
FRESHNESS_FRESH, FRESHNESS_STALE = "fresh", "stale"
DELTA_WINDOW_DAYS: int = 7
WRITER_NAME: str = "portfolio_weekly_brief"


# ---- DURUR exceptions -----------------------------------------------------

class PortfolioWeeklyBriefError(Exception):
    """Base error (no silent fallback)."""


class PortfolioConfigSchemaError(PortfolioWeeklyBriefError):
    """portfolio.config.json missing keys or out-of-range values."""


class WeekEndParseError(PortfolioWeeklyBriefError):
    """Malformed week_end ISO date or cadence-derived week_end."""


class WorkspaceRootUnsetError(PortfolioWeeklyBriefError):
    """workspace_root unresolvable (env + arg both absent)."""


class FreshnessFlagInputError(PortfolioWeeklyBriefError):
    """compute_freshness_flag got malformed last_sync_at or bad sla_days."""


# ---- Dataclasses ----------------------------------------------------------

@dataclass(frozen=True)
class ProjectDelta:
    slug: str
    tasks_added: int
    tasks_done: int
    work_events: int
    last_sync_at: str  # "" when null/never-synced
    freshness_flag: str  # FRESHNESS_FRESH | FRESHNESS_STALE


@dataclass(frozen=True)
class PortfolioBrief:
    week_start: str
    week_end: str
    generated_at: str  # UTC ISO with 'Z'
    sla_weekly_sync_max_days: int
    projects: tuple[ProjectDelta, ...]
    totals_tasks_added: int
    totals_tasks_done: int
    totals_work_events: int
    totals_stale_projects: int
    totals_fresh_projects: int


# ---- Schema-light validators (full Draft7 lives in tests) -----------------

def _validate_cadence(cadence: Mapping[str, Any]) -> None:
    if not isinstance(cadence, Mapping):
        raise PortfolioConfigSchemaError("cadence must be an object")
    wb = cadence.get("weekly_brief")
    if not isinstance(wb, Mapping):
        raise PortfolioConfigSchemaError("cadence.weekly_brief required (object)")
    day, hour = wb.get("day"), wb.get("hour")
    if day not in WEEKDAY_ENUM:
        raise PortfolioConfigSchemaError(
            f"cadence.weekly_brief.day={day!r} not in {WEEKDAY_ENUM}"
        )
    if not isinstance(hour, int) or hour < HOUR_MIN or hour > HOUR_MAX:
        raise PortfolioConfigSchemaError(
            f"cadence.weekly_brief.hour={hour!r} not in [{HOUR_MIN}, {HOUR_MAX}]"
        )


def _validate_sla(sla: Mapping[str, Any]) -> int:
    if not isinstance(sla, Mapping):
        raise PortfolioConfigSchemaError("slas must be an object")
    days = sla.get("weekly_sync_max_days")
    if (
        not isinstance(days, int)
        or days < SLA_MIN_DAYS or days > SLA_MAX_DAYS
    ):
        raise PortfolioConfigSchemaError(
            f"slas.weekly_sync_max_days={days!r} not in "
            f"[{SLA_MIN_DAYS}, {SLA_MAX_DAYS}]"
        )
    return days


# ---- Date / freshness helpers --------------------------------------------

def _parse_iso_date(value: str) -> _date:
    if not isinstance(value, str) or not value:
        raise WeekEndParseError(f"week_end must be non-empty ISO date, got {value!r}")
    try:
        return _date.fromisoformat(value)
    except ValueError as exc:
        raise WeekEndParseError(f"week_end {value!r} not valid ISO date: {exc}") from exc


def _parse_iso_timestamp(value: str) -> datetime:
    if not isinstance(value, str) or not value:
        raise FreshnessFlagInputError(f"timestamp must be non-empty ISO, got {value!r}")
    raw = value.strip()
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError as exc:
        raise FreshnessFlagInputError(f"timestamp {value!r} not ISO-8601: {exc}") from exc
    return dt if dt.tzinfo is not None else dt.replace(tzinfo=timezone.utc)


def resolve_week_end(
    cadence: Mapping[str, Any],
    *,
    today: _date,
    explicit: str | None = None,
) -> _date:
    """Resolve week_end: explicit override OR cadence.weekly_brief.day's
    most recent occurrence on or before `today`."""
    if explicit is not None:
        return _parse_iso_date(explicit)
    _validate_cadence(cadence)
    target_iso = _WEEKDAY_TO_ISO[cadence["weekly_brief"]["day"]]
    delta = (today.isoweekday() - target_iso) % 7
    return today - timedelta(days=delta)


def compute_freshness_flag(
    *, last_sync_at: str | None, sla_days: int, now: datetime,
) -> str:
    """'fresh' or 'stale' for one project. None/'' → stale."""
    if (
        not isinstance(sla_days, int)
        or sla_days < SLA_MIN_DAYS or sla_days > SLA_MAX_DAYS
    ):
        raise FreshnessFlagInputError(
            f"sla_days={sla_days!r} not in [{SLA_MIN_DAYS}, {SLA_MAX_DAYS}]"
        )
    if now.tzinfo is None:
        raise FreshnessFlagInputError("now must be timezone-aware (UTC expected)")
    if last_sync_at is None or last_sync_at == "":
        return FRESHNESS_STALE
    if (now - _parse_iso_timestamp(last_sync_at)) > timedelta(days=sla_days):
        return FRESHNESS_STALE
    return FRESHNESS_FRESH


# ---- events.jsonl reader (READ-ONLY, fault-tolerant) ---------------------

def read_events_in_window(
    events_path: Path,
    *,
    window_start: datetime,
    window_end: datetime,
    event_kind: str = "work",
) -> list[dict[str, Any]]:
    """Read events.jsonl, return matching entries. Empty/missing → []."""
    if not events_path.exists():
        return []
    out: list[dict[str, Any]] = []
    raw = events_path.read_text(encoding="utf-8")
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            evt = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(evt, dict):
            continue
        if evt.get("event_kind") != event_kind:
            continue
        ts_raw = evt.get("timestamp")
        if not isinstance(ts_raw, str):
            continue
        try:
            ts = _parse_iso_timestamp(ts_raw)
        except FreshnessFlagInputError:
            continue
        if window_start <= ts <= window_end:
            out.append(evt)
    return out


# ---- master.xlsx#master_task delta (READ-ONLY) ---------------------------

def read_master_task_delta(
    workbook_path: Path,
    *,
    window_start: _date,
    window_end: _date,
) -> tuple[int, int]:
    """Return (tasks_added, tasks_done) in [window_start, window_end].
    READ-ONLY (load_workbook read_only=True). Missing wb/sheet → (0, 0)."""
    if not workbook_path.exists():
        return 0, 0
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise PortfolioWeeklyBriefError("openpyxl required to read master.xlsx") from exc
    try:
        wb = load_workbook(
            filename=str(workbook_path), read_only=True, data_only=True,
        )
    except Exception as exc:  # noqa: BLE001
        raise PortfolioWeeklyBriefError(f"could not open {workbook_path}: {exc}") from exc
    try:
        if "master_task" not in wb.sheetnames:
            return 0, 0
        rows_iter = wb["master_task"].iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return 0, 0
        if not header:
            return 0, 0
        col_idx = {n: i for i, n in enumerate(header) if isinstance(n, str)}
        i_status, i_created, i_done = (
            col_idx.get("status"), col_idx.get("created_date"),
            col_idx.get("done_date"),
        )
        if i_status is None or i_created is None or i_done is None:
            return 0, 0
        added = done = 0
        for row in rows_iter:
            if row is None:
                continue
            created = (
                _coerce_date_cell(row[i_created])
                if i_created < len(row) else None
            )
            if created is not None and window_start <= created <= window_end:
                added += 1
            sv = row[i_status] if i_status < len(row) else None
            dv = _coerce_date_cell(row[i_done]) if i_done < len(row) else None
            if (
                isinstance(sv, str) and sv.strip().upper() == "DONE"
                and dv is not None and window_start <= dv <= window_end
            ):
                done += 1
        return added, done
    finally:
        wb.close()


def _coerce_date_cell(value: Any) -> _date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    if not isinstance(value, str):
        return None
    try:
        return _date.fromisoformat(value.strip())
    except ValueError:
        return None


# ---- Workspace root resolver ---------------------------------------------

def _resolve_workspace_root(workspace_root: Path | str | None) -> Path:
    if workspace_root is not None:
        return Path(workspace_root).expanduser().resolve()
    env_val = os.environ.get("PSEO_WORKSPACE_ROOT")
    if env_val:
        return Path(env_val).expanduser().resolve()
    raise WorkspaceRootUnsetError(
        "PSEO_WORKSPACE_ROOT env var unset and no workspace_root passed"
    )


# ---- Per-project aggregator ----------------------------------------------

def aggregate_project(
    *,
    slug: str,
    workspace_root: Path,
    last_sync_at: str | None,
    sla_days: int,
    week_start: _date,
    week_end: _date,
    now: datetime,
) -> ProjectDelta:
    """Compute one project's 7-day delta. Reads filesystem; never writes."""
    project_root = workspace_root / "projects" / slug
    ws_dt = datetime.combine(week_start, datetime.min.time(), tzinfo=timezone.utc)
    we_dt = datetime.combine(week_end, datetime.max.time(), tzinfo=timezone.utc)
    events = read_events_in_window(
        project_root / "_state" / "events.jsonl",
        window_start=ws_dt, window_end=we_dt, event_kind="work",
    )
    tasks_added, tasks_done = read_master_task_delta(
        project_root / "master.xlsx",
        window_start=week_start, window_end=week_end,
    )
    flag = compute_freshness_flag(
        last_sync_at=last_sync_at, sla_days=sla_days, now=now,
    )
    return ProjectDelta(
        slug=slug, tasks_added=tasks_added, tasks_done=tasks_done,
        work_events=len(events),
        last_sync_at=last_sync_at if isinstance(last_sync_at, str) else "",
        freshness_flag=flag,
    )


# ---- Top-level brief builder ---------------------------------------------

def build_portfolio_brief(
    *,
    portfolio_config: Mapping[str, Any],
    workspace_root: Path,
    now: datetime,
    week_end: str | None = None,
) -> PortfolioBrief:
    """Top-level pure aggregator. Idempotent: same inputs → same brief."""
    if not isinstance(portfolio_config, Mapping):
        raise PortfolioConfigSchemaError("portfolio_config must be a JSON object")
    cadence = portfolio_config.get("cadence")
    sla = portfolio_config.get("slas")
    if cadence is None:
        raise PortfolioConfigSchemaError("portfolio.config.json missing 'cadence'")
    if sla is None:
        raise PortfolioConfigSchemaError("portfolio.config.json missing 'slas'")
    _validate_cadence(cadence)
    sla_days = _validate_sla(sla)
    projects_in = portfolio_config.get("active_projects")
    if not isinstance(projects_in, list) or not projects_in:
        raise PortfolioConfigSchemaError(
            "portfolio.config.json active_projects must be a non-empty list"
        )
    today = now.date()
    week_end_d = resolve_week_end(cadence, today=today, explicit=week_end)
    week_start_d = week_end_d - timedelta(days=DELTA_WINDOW_DAYS - 1)

    deltas: list[ProjectDelta] = []
    for entry in projects_in:
        if not isinstance(entry, Mapping):
            raise PortfolioConfigSchemaError(
                f"active_projects entry must be object, got {type(entry).__name__}"
            )
        slug = entry.get("slug")
        if not isinstance(slug, str) or not slug:
            raise PortfolioConfigSchemaError(
                f"active_projects[].slug must be non-empty string, got {slug!r}"
            )
        # Per-project SLA override (editorial_overrides.sla_days).
        eff_sla_days = sla_days
        eo = entry.get("editorial_overrides")
        if isinstance(eo, Mapping):
            od = eo.get("sla_days")
            if isinstance(od, int):
                if od < SLA_MIN_DAYS or od > SLA_MAX_DAYS:
                    raise PortfolioConfigSchemaError(
                        f"editorial_overrides.sla_days={od!r} for slug={slug!r} "
                        f"not in [{SLA_MIN_DAYS}, {SLA_MAX_DAYS}]"
                    )
                eff_sla_days = od
        last_sync = entry.get("last_sync_at")
        if last_sync is not None and not isinstance(last_sync, str):
            raise PortfolioConfigSchemaError(
                f"active_projects[slug={slug!r}].last_sync_at must be string "
                f"or null, got {type(last_sync).__name__}"
            )
        deltas.append(aggregate_project(
            slug=slug, workspace_root=workspace_root,
            last_sync_at=last_sync, sla_days=eff_sla_days,
            week_start=week_start_d, week_end=week_end_d, now=now,
        ))

    deltas.sort(key=lambda d: d.slug)
    return PortfolioBrief(
        week_start=week_start_d.isoformat(),
        week_end=week_end_d.isoformat(),
        generated_at=now.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        sla_weekly_sync_max_days=sla_days,
        projects=tuple(deltas),
        totals_tasks_added=sum(d.tasks_added for d in deltas),
        totals_tasks_done=sum(d.tasks_done for d in deltas),
        totals_work_events=sum(d.work_events for d in deltas),
        totals_stale_projects=sum(
            1 for d in deltas if d.freshness_flag == FRESHNESS_STALE
        ),
        totals_fresh_projects=sum(
            1 for d in deltas if d.freshness_flag == FRESHNESS_FRESH
        ),
    )


# ---- Render helpers -------------------------------------------------------

def render_projects_delta_table(brief: PortfolioBrief) -> str:
    """Markdown table for $projects_delta_table."""
    header = (
        "| slug | tasks_added | tasks_done | last_sync_at | freshness_flag |\n"
        "|---|---:|---:|---|---|"
    )
    if not brief.projects:
        return header + "\n| _(no active projects)_ | 0 | 0 | — | — |"
    lines = [header]
    for p in brief.projects:
        sync = p.last_sync_at or "—"
        lines.append(
            f"| {p.slug} | {p.tasks_added} | {p.tasks_done} | "
            f"{sync} | {p.freshness_flag} |"
        )
    return "\n".join(lines)


def render_totals_summary(brief: PortfolioBrief) -> str:
    return (
        f"Toplam tasks_added={brief.totals_tasks_added}, "
        f"tasks_done={brief.totals_tasks_done}, "
        f"work_events={brief.totals_work_events}, "
        f"fresh={brief.totals_fresh_projects}, "
        f"stale={brief.totals_stale_projects} "
        f"(SLA={brief.sla_weekly_sync_max_days}d, "
        f"projects={len(brief.projects)})"
    )


def build_template_vars(
    brief: PortfolioBrief, *, scope: str = "portfolio",
) -> dict[str, str]:
    """Flat dict for string.Template $var substitution."""
    return {
        "week_start": brief.week_start,
        "week_end": brief.week_end,
        "generated_at": brief.generated_at,
        "scope": scope,
        "sla_weekly_sync_max_days": str(brief.sla_weekly_sync_max_days),
        "projects_count": str(len(brief.projects)),
        "projects_delta_table": render_projects_delta_table(brief),
        "totals_summary": render_totals_summary(brief),
        "totals_tasks_added": str(brief.totals_tasks_added),
        "totals_tasks_done": str(brief.totals_tasks_done),
        "totals_work_events": str(brief.totals_work_events),
        "totals_fresh_projects": str(brief.totals_fresh_projects),
        "totals_stale_projects": str(brief.totals_stale_projects),
    }


def build_snapshot(brief: PortfolioBrief) -> dict[str, Any]:
    """Drift-recovery snapshot — JSON-serialisable PortfolioBrief."""
    return {
        "writer": WRITER_NAME,
        "week_start": brief.week_start, "week_end": brief.week_end,
        "generated_at": brief.generated_at,
        "sla_weekly_sync_max_days": brief.sla_weekly_sync_max_days,
        "totals": {
            "tasks_added": brief.totals_tasks_added,
            "tasks_done": brief.totals_tasks_done,
            "work_events": brief.totals_work_events,
            "fresh_projects": brief.totals_fresh_projects,
            "stale_projects": brief.totals_stale_projects,
        },
        "projects": [
            {
                "slug": p.slug, "tasks_added": p.tasks_added,
                "tasks_done": p.tasks_done, "work_events": p.work_events,
                "last_sync_at": p.last_sync_at,
                "freshness_flag": p.freshness_flag,
            } for p in brief.projects
        ],
    }


def render_brief_markdown(
    *, brief: PortfolioBrief, template_path: Path,
) -> str:
    """Render Markdown via string.Template (matches render_template.py)."""
    from string import Template
    if not template_path.exists():
        raise PortfolioWeeklyBriefError(f"template not found: {template_path}")
    tpl = Template(template_path.read_text(encoding="utf-8"))
    try:
        return tpl.safe_substitute(build_template_vars(brief))
    except KeyError as exc:
        raise PortfolioWeeklyBriefError(
            f"template references unknown variable: {exc.args[0]}"
        ) from exc


# ---- CLI orchestrator -----------------------------------------------------

def _now_utc() -> datetime:
    return datetime.now(tz=timezone.utc).replace(microsecond=0)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build portfolio weekly brief (READ-ONLY aggregator).",
    )
    parser.add_argument("--portfolio-config", required=True)
    parser.add_argument("--workspace-root", default=None)
    parser.add_argument("--week-end", default=None)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--template", default=None)
    parser.add_argument("--now", default=None)
    args = parser.parse_args(argv)

    cfg_path = Path(args.portfolio_config)
    if not cfg_path.exists():
        print(f"portfolio.config.json not found: {cfg_path}", file=sys.stderr)
        return 1
    portfolio_config = json.loads(cfg_path.read_text(encoding="utf-8"))
    workspace_root = _resolve_workspace_root(args.workspace_root)
    now = _parse_iso_timestamp(args.now) if args.now else _now_utc()

    brief = build_portfolio_brief(
        portfolio_config=portfolio_config, workspace_root=workspace_root,
        now=now, week_end=args.week_end,
    )

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    snap_path = out_dir / f"{brief.week_end}-portfolio-weekly-brief.json"
    snap_path.write_text(
        json.dumps(
            build_snapshot(brief), ensure_ascii=False, indent=2, sort_keys=True,
        ),
        encoding="utf-8",
    )
    template_path = (
        Path(args.template) if args.template
        else _REPO_ROOT / "templates" / "reports"
        / "portfolio-weekly-brief.template.md"
    )
    report_path = out_dir / f"{brief.week_end}-portfolio-weekly-brief.md"
    report_path.write_text(
        render_brief_markdown(brief=brief, template_path=template_path),
        encoding="utf-8",
    )

    print(json.dumps({
        "writer": WRITER_NAME,
        "week_start": brief.week_start, "week_end": brief.week_end,
        "report_path": str(report_path), "snapshot_path": str(snap_path),
        "totals_tasks_added": brief.totals_tasks_added,
        "totals_tasks_done": brief.totals_tasks_done,
        "totals_work_events": brief.totals_work_events,
        "totals_fresh_projects": brief.totals_fresh_projects,
        "totals_stale_projects": brief.totals_stale_projects,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
