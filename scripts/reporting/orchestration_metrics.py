#!/usr/bin/env python3
"""orchestration_metrics.py — the AMO INDEPENDENT correctness oracle (spec G5).

AMO claims "<=5% error on structured workflows" (spec G5). That number is
worthless if it is read back from the run's OWN self-reported status: a run can
write ``verdict="pass"`` while silently committing wrong/short data. This module
is the independent auditor. For each run it RE-DERIVES the truth from the actual
on-disk artifacts — the raw provenance drop the model wrote, the transform
output, and the COMMITTED ``master.xlsx`` — and reports whether they reconcile,
WITHOUT trusting the coverage record's ``verdict`` / ``scored_count``.

Headline number: the structured-error rate = (runs whose committed workbook does
NOT reconcile with their provenance) / (total reconcilable runs). The dangerous
catch is ``fake_green``: a run whose self-reported ``verdict == "pass"`` but whose
committed data does NOT reconcile — success was reported while the data is wrong.

The expected committed-row count is the length of the transform OUTPUT, NOT the
raw row count: steps like ``quick_wins`` legitimately FILTER their input, so
comparing committed-vs-raw would false-flag every filtering step. Reconcile
compares committed-vs-output (the transform output is the bridge); the raw->output
drop is surfaced separately as an advisory silent-skip signal (reusing
``verify.silent_skip_exceeds`` so the threshold cannot drift).

Workflow-AGNOSTIC: a run's workflow (``monthly`` / ``audit`` / …) is resolved from
its coverage record by which workflow's STEPS-names are a subset of the record's
step names (the name sets are disjoint), so EVERY workflow's runs are reconciled —
including audit's model_attested sheet-writing steps, whose only independent ≤5%
backstop is this oracle. No ``workflow`` field is added to the frozen record.

READ-ONLY: this module opens artifacts + the workbook (openpyxl read_only) and
writes NO state — it adds no hook, schema, or command and never blocks anything.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook

from scripts.orchestration.coverage import coverage_path
from scripts.orchestration.verify import silent_skip_exceeds
from scripts.orchestration.workflows import audit_suite, monthly_maintenance

_REPO_ROOT = Path(__file__).resolve().parents[2]
_DEFAULT_SCHEMA_PATH = _REPO_ROOT / "schemas" / "master-excel.schema.json"

# Workflow registry — the oracle is workflow-AGNOSTIC: it reconciles whichever
# workflow a run belongs to. A run's workflow is resolved from its coverage record
# by which workflow's STEPS-names are a subset of the record's step names. The two
# name sets are DISJOINT (monthly={gsc_pull,quick_wins,content_decay};
# audit={tech_audit,schema_audit,on_page_audit,cannibalization}), so a real run
# matches exactly one — resolution is unambiguous WITHOUT adding a `workflow` field
# to the frozen coverage record. A 3rd workflow is a one-line append here; the
# spine + the coverage schema + the workflow modules stay UNTOUCHED.
_WORKFLOWS: tuple[tuple[str, Sequence[dict]], ...] = (
    ("monthly", monthly_maintenance.STEPS),
    ("audit", audit_suite.STEPS),
)

# Stable reason codes — a report keys on these strings, so a mismatch is always
# explainable (which of R1-R5 failed).
RAW_MISSING = "raw_missing"               # R1: raw drop absent / unparseable
IDENTITY_MISMATCH = "identity_mismatch"   # R1: provenance run_id/slug disagree
TRUNCATED = "truncated"                   # R2: declared_count != len(raw.rows)
OUTPUT_MISSING = "output_missing"         # R3: transform output absent / not a list
WORKBOOK_MISMATCH = "workbook_mismatch"   # R4: committed rows != len(output)
SCORED_COUNT_MISMATCH = "scored_count_mismatch"  # R5: scored_count != committed
WORKBOOK_ABSENT = "workbook_absent"       # no committed workbook to check R4/R5


# ---------------------------------------------------------------------------
# Artifact readers (thin IO; never raise for an expected-missing artifact)
# ---------------------------------------------------------------------------

def _read_json(path: Path | str) -> Any:
    """Parse JSON at ``path``; return the object, or None if missing/unreadable."""
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _output_rows(payload: Any) -> list | None:
    """Extract the row list from a transform-output payload.

    The transform output is a bare JSON list OR ``{"rows": [...]}`` (the two
    shapes monthly_maintenance._output_loader accepts). Returns the list, or None
    if the shape is neither.
    """
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return payload["rows"]
    return None


# ---------------------------------------------------------------------------
# Workflow resolution + artifact-path derivation
#
# The oracle knows INDEPENDENTLY where each workflow's artifacts live — it does
# NOT import a workflow's ``output_path`` (monthly's is SHEET-keyed, which is wrong
# for audit's explicit ``output_file``). Knowing the layout itself is correct for
# an independent auditor.
# ---------------------------------------------------------------------------

def _inbox_path(workspace_root: Path | str, run_id: str, slug: str, step: str) -> Path:
    """``projects/{slug}/_state/inbox/{run_id}/{step}.json`` (raw drop, keyed by
    step NAME). This inbox convention is IDENTICAL across every workflow."""
    return (
        Path(workspace_root) / "projects" / slug / "_state" / "inbox" / run_id
        / f"{step}.json"
    )


def _output_file_path(
    workspace_root: Path | str, run_id: str, slug: str, output_file: str
) -> Path:
    """``projects/{slug}/_state/transform/{run_id}/{output_file}`` (transform
    output, keyed by the step's OUTPUT FILE). Only the filename differs across
    workflows — see ``_step_output_file``."""
    return (
        Path(workspace_root) / "projects" / slug / "_state" / "transform" / run_id
        / output_file
    )


def _step_output_file(entry: dict) -> str:
    """The transform-output FILENAME for a STEPS entry.

    Audit entries carry an EXPLICIT ``output_file`` (schema_audit -> schema_audit.json
    though its sheet is ``schema``); monthly entries carry none, so the file is
    ``{sheet}.json``. ``entry.get("output_file") or {sheet}.json`` is correct for
    BOTH workflow shapes and touches neither workflow module."""
    return entry.get("output_file") or f"{entry['sheet']}.json"


def _resolve_workflow_steps(
    coverage_record: dict,
) -> tuple[str, Sequence[dict]] | None:
    """Resolve which workflow a coverage record belongs to.

    Returns ``(workflow_name, STEPS)`` for the registered workflow whose
    STEPS-names are ALL present in the record's step names (a subset match), or
    None if no workflow matches (an unknown/legacy record). The workflow name sets
    are DISJOINT, so a real run matches at most one workflow — resolution is
    unambiguous WITHOUT a ``workflow`` field on the frozen coverage record. Pure:
    reads only ``steps[].name``."""
    names = {
        s.get("name")
        for s in (coverage_record.get("steps") or [])
        if isinstance(s, dict)
    }
    for workflow_name, steps in _WORKFLOWS:
        if all(entry["name"] in names for entry in steps):
            return workflow_name, steps
    return None


# ---------------------------------------------------------------------------
# Workbook reads (openpyxl read_only; degrade to None/False, never crash)
# ---------------------------------------------------------------------------

def committed_row_count(
    workbook_path: Path | str, sheet: str, data_start_row: int
) -> int | None:
    """Count CONTIGUOUS non-empty rows from ``data_start_row`` down.

    Reads the workbook read_only/data_only and, for worksheet ``sheet``, counts
    rows from ``data_start_row`` that have at least one non-None cell, stopping at
    the first all-empty row. This mirrors exactly how ``transaction.replace`` lands
    a single contiguous data block from ``data_start_row``, so the count is an
    independent re-derivation of what was committed (not a guess). Returns None if
    the workbook or sheet is missing/unreadable — the caller treats None as
    ``workbook_absent`` for that sheet (a reporting tool must never crash a build).
    """
    path = Path(workbook_path)
    if not path.exists():
        return None
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — any read defect ⇒ "absent", never raise
        return None
    try:
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        count = 0
        for row in ws.iter_rows(min_row=data_start_row):
            if all(cell.value is None for cell in row):
                break
            count += 1
        return count
    finally:
        wb.close()


def _workbook_present(workbook_path: Path | str) -> bool:
    """True iff the workbook file exists and openpyxl can open it (read_only)."""
    path = Path(workbook_path)
    if not path.exists():
        return False
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — unreadable ⇒ treated as absent
        return False
    wb.close()
    return True


def data_start_row_for(sheet: str, schema_path: Path | str | None = None) -> int:
    """Return ``sheets[sheet].data_start_row`` from master-excel.schema.json.

    Raises ``KeyError`` (clear message) if the sheet is unknown or declares no
    integer ``data_start_row`` — fail loud rather than silently miscount.
    """
    path = Path(schema_path) if schema_path else _DEFAULT_SCHEMA_PATH
    sheets = (json.loads(path.read_text(encoding="utf-8")).get("sheets") or {})
    if sheet not in sheets:
        raise KeyError(f"sheet {sheet!r} not in master-excel.schema.json")
    dsr = sheets[sheet].get("data_start_row")
    if not isinstance(dsr, int):
        raise KeyError(f"sheet {sheet!r} declares no integer data_start_row")
    return dsr


# ---------------------------------------------------------------------------
# Reconcile core (pure) + IO wrapper
# ---------------------------------------------------------------------------

def _reconcile_step_data(
    *,
    run_id: str,
    slug: str,
    step_name: str,
    sheet: str,
    raw: Any,
    output_rows: list | None,
    committed_count: int | None,
    scored_count: int | None,
    has_workbook: bool,
) -> dict:
    """Apply R1-R5 to already-read data; build a NEW result dict (no mutation).

    R4/R5 are evaluated ONLY when ``has_workbook`` is True; otherwise they are
    skipped (committed_count stays None) and the run layer marks the run absent.
    ``high_silent_skip`` is ADVISORY ONLY (reuses verify.silent_skip_exceeds) and
    never affects ``independent_ok`` — a step may legitimately drop >50% of its
    raw input (filtering) yet still reconcile.
    """
    failed: list[str] = []

    prov = raw.get("provenance") if isinstance(raw, dict) else None
    raw_rows = raw.get("rows") if isinstance(raw, dict) else None
    raw_count = len(raw_rows) if isinstance(raw_rows, list) else None

    # R1 identity + R2 untruncated — only when the raw drop is structurally sound.
    if not isinstance(prov, dict) or not isinstance(raw_rows, list):
        failed.append(RAW_MISSING)
    else:
        if prov.get("run_id") != run_id or prov.get("slug") != slug:
            failed.append(IDENTITY_MISMATCH)
        if prov.get("declared_count") != len(raw_rows):
            failed.append(TRUNCATED)

    # R3 output present + a list.
    output_count = len(output_rows) if isinstance(output_rows, list) else None
    if output_count is None:
        failed.append(OUTPUT_MISSING)

    # R4 committed-vs-OUTPUT + R5 self-report-vs-committed (workbook only).
    if has_workbook:
        if committed_count is None:
            failed.append(WORKBOOK_ABSENT)
        else:
            if output_count is not None and committed_count != output_count:
                failed.append(WORKBOOK_MISMATCH)
            if scored_count != committed_count:
                failed.append(SCORED_COUNT_MISMATCH)

    high_silent_skip = bool(
        raw_count is not None
        and committed_count is not None
        and silent_skip_exceeds(raw_count, committed_count)
    )
    return {
        "step": step_name,
        "sheet": sheet,
        "independent_ok": not failed,
        "failed_checks": failed,
        "raw_count": raw_count,
        "output_count": output_count,
        "committed_count": committed_count,
        "scored_count": scored_count,
        "high_silent_skip": high_silent_skip,
    }


def reconcile_step(
    *,
    workspace_root: Path | str,
    slug: str,
    run_id: str,
    step_name: str,
    sheet: str,
    coverage_step: dict,
    output_file: str | None = None,
    workbook_path: Path | str | None = None,
    schema_path: Path | str | None = None,
) -> dict:
    """Read this step's artifacts (raw drop, transform output, committed rows) and
    reconcile them via R1-R5. Returns a NEW result dict.

    ``output_file`` is the transform-output filename; when None it defaults to
    ``{sheet}.json`` (monthly's convention), so existing sheet-keyed callers are
    unchanged. Audit passes the step's EXPLICIT ``output_file`` (e.g.
    ``schema_audit.json`` for sheet ``schema``). The raw drop is ALWAYS keyed by
    ``step_name`` (the shared inbox convention)."""
    resolved_output_file = output_file if output_file is not None else f"{sheet}.json"
    raw = _read_json(_inbox_path(workspace_root, run_id, slug, step_name))
    output_rows = _output_rows(
        _read_json(
            _output_file_path(workspace_root, run_id, slug, resolved_output_file)
        )
    )
    scored_count = (
        coverage_step.get("scored_count") if isinstance(coverage_step, dict) else None
    )
    has_workbook = workbook_path is not None
    committed_count = None
    if has_workbook:
        committed_count = committed_row_count(
            workbook_path, sheet, data_start_row_for(sheet, schema_path)
        )
    return _reconcile_step_data(
        run_id=run_id, slug=slug, step_name=step_name, sheet=sheet, raw=raw,
        output_rows=output_rows, committed_count=committed_count,
        scored_count=scored_count, has_workbook=has_workbook,
    )


def _run_result(run_id: str, slug: str, verdict: str, self_verdict, steps: list) -> dict:
    """Assemble a reconcile_run result — ONE shape for every verdict. ``fake_green``
    (pass claimed but committed data mismatches) is DERIVED, so non-mismatch
    verdicts (reconciled / workbook_absent / unresolved_workflow) are never
    fake_green."""
    return {
        "run_id": run_id,
        "slug": slug,
        "independent_verdict": verdict,
        "self_reported_verdict": self_verdict,
        "steps": steps,
        "fake_green": self_verdict == "pass" and verdict == "mismatch",
    }


def reconcile_run(
    *,
    workspace_root: Path | str,
    slug: str,
    run_id: str,
    workbook_path: Path | str | None = None,
    schema_path: Path | str | None = None,
    steps: Sequence[dict] | None = None,
    coverage_record: dict | None = None,
) -> dict:
    """Reconcile one run (workflow-AGNOSTIC); derive the independent verdict.

    Loads the coverage record (coverage_path) if not passed. The run's workflow is
    resolved from the record (``_resolve_workflow_steps``) unless ``steps`` is
    given explicitly (back-compat). Reconciles EVERY resolved STEPS entry that has
    a matching coverage step — regardless of ``verification_class``, because every
    STEPS entry commits a sheet (audit's model_attested steps write sheets too, so
    they get the same committed-vs-output R4/R5 backstop). A coverage step NOT in
    the STEPS table (e.g. a synthetic ``monthly_report``) is never reconciled.
    The run is:
      - ``unresolved_workflow`` if ``steps`` is None and the record matches no
        workflow (reported, EXCLUDED from the error-rate denominator — never a
        silent pass),
      - ``workbook_absent`` if no readable workbook was given (reported, NOT scored),
      - else ``reconciled`` if EVERY reconciled step independent_ok, else ``mismatch``.
    ``fake_green`` = self-reported verdict == "pass" AND independent verdict ==
    "mismatch" (success claimed while the committed data disagrees).
    """
    if coverage_record is None:
        coverage_record = _read_json(coverage_path(workspace_root, slug, run_id)) or {}
    self_verdict = coverage_record.get("verdict")

    if steps is None:
        resolved = _resolve_workflow_steps(coverage_record)
        if resolved is None:
            # Unknown/legacy record — not about the workbook; reported, not scored.
            return _run_result(run_id, slug, "unresolved_workflow", self_verdict, [])
        _, steps = resolved

    cov_by_name = {
        s.get("name"): s
        for s in (coverage_record.get("steps") or [])
        if isinstance(s, dict)
    }
    wb_present = workbook_path is not None and _workbook_present(workbook_path)
    effective_wb = workbook_path if wb_present else None

    results: list[dict] = []
    for entry in steps:
        cov_step = cov_by_name.get(entry["name"])
        if not isinstance(cov_step, dict):
            continue
        results.append(
            reconcile_step(
                workspace_root=workspace_root, slug=slug, run_id=run_id,
                step_name=entry["name"], sheet=entry["sheet"],
                output_file=_step_output_file(entry), coverage_step=cov_step,
                workbook_path=effective_wb, schema_path=schema_path,
            )
        )

    if not wb_present or not results:
        # No readable workbook (or no STEPS step present) ⇒ not scored.
        independent_verdict = "workbook_absent"
    elif all(r["independent_ok"] for r in results):
        independent_verdict = "reconciled"
    else:
        independent_verdict = "mismatch"

    return _run_result(run_id, slug, independent_verdict, self_verdict, results)


# ---------------------------------------------------------------------------
# Aggregation + project-wide enumeration
# ---------------------------------------------------------------------------

def structured_error_rate(run_reconciliations: Iterable[dict]) -> dict:
    """Aggregate reconcile_run dicts into the headline structured-error metrics.

    ``reconcilable`` EXCLUDES both ``workbook_absent`` and ``unresolved_workflow``
    runs (neither was scored). ``error_rate`` = mismatched / reconcilable (0.0 when
    reconcilable == 0). ``workbook_absent`` and ``unresolved_workflow`` are each
    surfaced as their own count — NEVER silently dropped or treated as a pass.
    """
    runs = list(run_reconciliations)
    verdicts = [r.get("independent_verdict") for r in runs]
    reconciled = verdicts.count("reconciled")
    mismatched = verdicts.count("mismatch")
    workbook_absent = verdicts.count("workbook_absent")
    unresolved_workflow = verdicts.count("unresolved_workflow")
    reconcilable = reconciled + mismatched
    return {
        "total": len(runs),
        "reconcilable": reconcilable,
        "reconciled": reconciled,
        "mismatched": mismatched,
        "workbook_absent": workbook_absent,
        "unresolved_workflow": unresolved_workflow,
        "error_rate": (mismatched / reconcilable) if reconcilable else 0.0,
        "fake_green_count": sum(1 for r in runs if r.get("fake_green")),
        "mismatched_run_ids": [
            r.get("run_id") for r in runs
            if r.get("independent_verdict") == "mismatch"
        ],
    }


def oracle_report(
    *,
    workspace_root: Path | str,
    slug: str,
    workbook_path: Path | str | None = None,
    schema_path: Path | str | None = None,
) -> dict:
    """Reconcile EVERY coverage run under ``slug`` and aggregate the metrics.

    Enumerates ``{run_id}.json`` under the slug's coverage dir (the parent of
    ``coverage_path``), reconcile_run each, and returns
    ``{slug, runs:[...], metrics: structured_error_rate(...)}``.
    """
    coverage_dir = coverage_path(workspace_root, slug, "_").parent
    runs: list[dict] = []
    if coverage_dir.is_dir():
        for record_path in sorted(coverage_dir.glob("*.json")):
            runs.append(
                reconcile_run(
                    workspace_root=workspace_root, slug=slug, run_id=record_path.stem,
                    workbook_path=workbook_path, schema_path=schema_path,
                )
            )
    return {"slug": slug, "runs": runs, "metrics": structured_error_rate(runs)}


# ---------------------------------------------------------------------------
# CLI (READ-ONLY report; exit 0 always — it never fails the build)
# ---------------------------------------------------------------------------

def _format_report(report: dict) -> str:
    """Render the compact human report (pure: returns the string)."""
    m = report["metrics"]
    lines = [
        f"AMO oracle — slug={report['slug']} "
        f"(READ-ONLY; independent of coverage.verdict)",
        f"  structured error rate: {m['error_rate'] * 100:.1f}%  "
        f"({m['mismatched']}/{m['reconcilable']} reconcilable runs mismatched)",
        f"  reconciled={m['reconciled']}  mismatched={m['mismatched']}  "
        f"workbook_absent={m['workbook_absent']}  "
        f"unresolved_workflow={m['unresolved_workflow']}  total={m['total']}",
    ]
    if m["fake_green_count"]:
        loud = ", ".join(r["run_id"] for r in report["runs"] if r["fake_green"])
        lines.append(
            f"  *** FAKE-GREEN ({m['fake_green_count']}): {loud} — "
            f"verdict=pass but committed data does NOT reconcile ***"
        )
    else:
        lines.append("  fake-green: none")
    for r in report["runs"]:
        flag = "  [FAKE-GREEN]" if r["fake_green"] else ""
        lines.append(
            f"    {r['run_id']}: {r['independent_verdict']} "
            f"(self={r['self_reported_verdict']}){flag}"
        )
    return "\n".join(lines)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python3 -m scripts.reporting.orchestration_metrics",
        description="AMO independent oracle: reconcile the committed master.xlsx "
                    "against raw provenance for every run of a project (READ-ONLY).",
    )
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--slug", required=True)
    parser.add_argument(
        "--workbook", default=None,
        help="committed master.xlsx; omit to skip R4/R5 (runs marked "
             "workbook_absent — no silent degradation)",
    )
    parser.add_argument(
        "--schema", default=None,
        help="override path to master-excel.schema.json (data_start_row source)",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI boundary: enumerate + reconcile + print the report. Exit 0 always."""
    args = _build_arg_parser().parse_args(argv)
    if args.workbook is None:
        print(
            "note: --workbook not given → R4/R5 skipped, runs marked "
            "workbook_absent (no silent degradation)."
        )
    report = oracle_report(
        workspace_root=args.workspace_root, slug=args.slug,
        workbook_path=args.workbook, schema_path=args.schema,
    )
    print(_format_report(report))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
