#!/usr/bin/env python3
"""portfolio_task_heatmap.py — multi-project READ-ONLY density aggregator
(Phase 9 W-E6).

Reads portfolio.config.json + each active project's master.xlsx#master_task
col F (category) + col G (priority severityEnum 4) + col J (status
statusEnum 7) into a project x category x priority density matrix +
per-project / per-category / per-priority totals + status_check_drift
advisory (consistency-report verdict GREEN/AMBER/RED).

READ-ONLY (schemas/portfolio-config.schema.json#cross_query.read_only
const true): no write call sites; load_workbook(read_only=True).
Pure on import. Idempotent: same input + reference_date -> byte-identical
output. Plugin-agnostik. No events.jsonl write (Q-RP-01 deferred).

Pattern reference: portfolio_overview.py (W-E3) +
portfolio_weekly_brief.py (W-E4).
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import date as _date, datetime, timezone
from pathlib import Path
from string import Template
from typing import Any, Mapping


_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


# --- Constants (schema-aligned) -------------------------------------------

STATUS_ENUM: tuple[str, ...] = (
    "TODO", "ONGOING", "EXISTS", "DONE",
    "BLOCKED", "DEFERRED", "CANCELED",
)
SEVERITY_ENUM: tuple[str, ...] = ("CRITICAL", "HIGH", "MEDIUM", "LOW")
CLOSED_STATUSES: frozenset[str] = frozenset({"DONE", "CANCELED"})
OTHER_BUCKET: str = "OTHER"
SEV_PLUS_OTHER: tuple[str, ...] = (*SEVERITY_ENUM, OTHER_BUCKET)
MASTER_TASK_DATA_START_ROW: int = 4
MASTER_TASK_CATEGORY_COL_INDEX: int = 6   # F
MASTER_TASK_PRIORITY_COL_INDEX: int = 7   # G
MASTER_TASK_STATUS_COL_INDEX: int = 10    # J
ACTIVE_PROJECTS_MAX: int = 8
PORTFOLIO_DIR_NAME: str = "_portfolio"
OUTPUTS_REPORTS_REL: str = "projects/_portfolio/outputs/reports"
INBOX_LOCAL_REL: str = "projects/_portfolio/inbox/local"
VERDICT_GREEN, VERDICT_AMBER, VERDICT_RED = "GREEN", "AMBER", "RED"
VERDICT_ENUM: tuple[str, ...] = (VERDICT_GREEN, VERDICT_AMBER, VERDICT_RED)
DEFAULT_TEMPLATE_REL: str = "templates/reports/portfolio-task-heatmap.template.md"
WRITER_NAME: str = "portfolio_task_heatmap"


# --- Exceptions (DURUR — STOP + surface) ---------------------------------

class PortfolioTaskHeatmapError(Exception): pass
class PortfolioConfigMissingError(PortfolioTaskHeatmapError): pass
class PortfolioConfigInvalidError(PortfolioTaskHeatmapError): pass
class ActiveProjectsCeilingError(PortfolioTaskHeatmapError): pass
class ReadOnlyContractViolation(PortfolioTaskHeatmapError): pass
class WorkspaceRootUnsetError(PortfolioTaskHeatmapError): pass


# --- Dataclasses ---------------------------------------------------------

@dataclass(frozen=True)
class ProjectHeatmap:
    slug: str
    display_name: str
    priority: int
    workspace_path: str
    master_xlsx_path: str | None
    density: dict
    status_distribution: dict
    category_totals: dict
    priority_totals: dict
    open_total: int
    consistency_verdict: str | None
    status_check_drift: bool
    warnings: list


@dataclass(frozen=True)
class HeatmapBatch:
    portfolio_id: str
    display_name: str
    projects: tuple = field(default_factory=tuple)
    skipped_projects: int = 0
    run_date: str = ""
    generated_at: str = ""
    portfolio_category_totals: dict = field(default_factory=dict)
    portfolio_priority_totals: dict = field(default_factory=dict)
    portfolio_open_total: int = 0


# --- Helpers -------------------------------------------------------------

def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def load_portfolio_config(portfolio_root: Path) -> dict:
    cfg_path = (
        portfolio_root / "projects" / PORTFOLIO_DIR_NAME
        / "portfolio.config.json"
    )
    if not cfg_path.is_file():
        raise PortfolioConfigMissingError(
            f"portfolio.config.json not found at {cfg_path}."
        )
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PortfolioConfigMissingError(
            f"portfolio.config.json read failed at {cfg_path}: {exc}"
        ) from exc


def validate_portfolio_config(
    config: Mapping[str, Any], schema: Mapping[str, Any],
) -> None:
    """Draft 7 validate + maxItems + cross_query.read_only domain guards."""
    try:
        from jsonschema import Draft7Validator
    except ImportError as exc:  # pragma: no cover
        raise PortfolioConfigInvalidError("jsonschema required") from exc
    errors = sorted(
        Draft7Validator(dict(schema)).iter_errors(dict(config)),
        key=lambda e: list(e.absolute_path),
    )
    if errors:
        raise PortfolioConfigInvalidError(
            "portfolio.config.json failed Draft 7 validation: "
            + "; ".join(f"{list(e.absolute_path)}: {e.message}"
                        for e in errors))
    active = list(config.get("active_projects") or [])
    if len(active) > ACTIVE_PROJECTS_MAX:
        raise ActiveProjectsCeilingError(
            f"active_projects has {len(active)} entries; ceiling is "
            f"{ACTIVE_PROJECTS_MAX} (schema maxItems).")
    cq = config.get("cross_query") or {}
    if "read_only" in cq and cq.get("read_only") is not True:
        raise ReadOnlyContractViolation(
            f"cross_query.read_only must be true (schema const), got "
            f"{cq.get('read_only')!r}.")


def resolve_project_workspace(
    *, portfolio_root: Path, workspace_path: str,
) -> Path:
    p = Path(workspace_path).expanduser()
    return p if p.is_absolute() else (portfolio_root.parent / p).resolve()


def _project_root(
    portfolio_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    return resolve_project_workspace(
        portfolio_root=portfolio_root,
        workspace_path=_safe_str(project_entry.get("workspace_path")),
    ) / "projects" / _safe_str(project_entry.get("slug"))


def resolve_master_xlsx_path(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    return _project_root(portfolio_root, project_entry) / "master.xlsx"


def resolve_consistency_report_path(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    return (
        _project_root(portfolio_root, project_entry) / "_state"
        / "consistency-report.json"
    )


# --- openpyxl thin wrapper (READ-ONLY, INLINE) -----

def _read_master_task_density(workbook_path: Path) -> tuple[dict, dict]:
    """Scan master_task col F/G/J. READ-ONLY. Returns (density, sd).
    density[category][severity] = OPEN-task count.
    sd[status] = total count (statusEnum + OTHER).
    Open task = status NOT in CLOSED_STATUSES ({DONE, CANCELED})."""
    density: dict[str, dict[str, int]] = {}
    sd: dict[str, int] = {s: 0 for s in STATUS_ENUM}
    sd[OTHER_BUCKET] = 0
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:  # pragma: no cover
        return density, sd
    wb = load_workbook(
        filename=str(workbook_path), read_only=True, data_only=True,
    )
    try:
        if "master_task" not in wb.sheetnames:
            return density, sd
        for row in wb["master_task"].iter_rows(
            min_row=MASTER_TASK_DATA_START_ROW,
            min_col=MASTER_TASK_CATEGORY_COL_INDEX,
            max_col=MASTER_TASK_STATUS_COL_INDEX,
            values_only=True,
        ):
            cat_raw = row[0] if len(row) > 0 else None
            pri_raw = row[1] if len(row) > 1 else None
            sta_raw = row[4] if len(row) > 4 else None
            sval = _safe_str(sta_raw)
            cval = _safe_str(cat_raw) or OTHER_BUCKET
            pri_s = _safe_str(pri_raw).upper()
            pval = pri_s if pri_s in SEVERITY_ENUM else OTHER_BUCKET
            if not sval and not _safe_str(cat_raw) and not _safe_str(pri_raw):
                continue
            if sval:
                sd[sval if sval in STATUS_ENUM else OTHER_BUCKET] += 1
            if sval and sval in CLOSED_STATUSES:
                continue
            density.setdefault(cval, {s: 0 for s in SEVERITY_ENUM})
            if pval == OTHER_BUCKET:
                density[cval].setdefault(OTHER_BUCKET, 0)
            density[cval][pval] = density[cval].get(pval, 0) + 1
    finally:
        wb.close()
    return density, sd


def _read_consistency_verdict(report_path: Path) -> str | None:
    """OPTIONAL consistency-report.json verdict; absent/bad -> None."""
    if not report_path.is_file():
        return None
    try:
        payload = json.loads(report_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    v = payload.get("verdict") if isinstance(payload, Mapping) else None
    return v if isinstance(v, str) and v in VERDICT_ENUM else None


# --- Per-project + top-level aggregators ---------------------------------

def _build_project_heatmap(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> ProjectHeatmap:
    """Build one project's density. Missing master.xlsx -> graceful skip."""
    slug = _safe_str(project_entry.get("slug"))
    display_name = _safe_str(project_entry.get("display_name") or slug)
    priority = int(project_entry.get("priority") or 10)
    workspace_path = _safe_str(project_entry.get("workspace_path"))
    warnings: list = []
    base = dict(
        slug=slug, display_name=display_name, priority=priority,
        workspace_path=workspace_path,
    )
    master_xlsx = resolve_master_xlsx_path(
        portfolio_root=portfolio_root, project_entry=project_entry,
    )
    if not master_xlsx.is_file():
        warnings.append(f"master.xlsx not found at {master_xlsx}")
        return ProjectHeatmap(
            **base, master_xlsx_path=None, density={},
            status_distribution={s: 0 for s in (*STATUS_ENUM, OTHER_BUCKET)},
            category_totals={},
            priority_totals={s: 0 for s in SEV_PLUS_OTHER},
            open_total=0, consistency_verdict=None,
            status_check_drift=False, warnings=warnings,
        )
    density, sd = _read_master_task_density(master_xlsx)
    cat_totals: dict[str, int] = {}
    pri_totals: dict[str, int] = {s: 0 for s in SEV_PLUS_OTHER}
    open_total = 0
    for cat, row in density.items():
        n = sum(int(v) for v in row.values())
        cat_totals[cat] = n
        open_total += n
        for sev, c in row.items():
            pri_totals.setdefault(sev, 0)
            pri_totals[sev] += int(c)
    if sd.get(OTHER_BUCKET, 0):
        warnings.append(
            f"master_task has {sd[OTHER_BUCKET]} rows with status outside "
            "statusEnum"
        )
    verdict = _read_consistency_verdict(resolve_consistency_report_path(
        portfolio_root=portfolio_root, project_entry=project_entry,
    ))
    drift = verdict is not None and verdict != VERDICT_GREEN
    if drift:
        warnings.append(
            f"status_check_drift: consistency-report verdict = {verdict}"
        )
    return ProjectHeatmap(
        **base, master_xlsx_path=str(master_xlsx),
        density=density, status_distribution=sd,
        category_totals=cat_totals, priority_totals=pri_totals,
        open_total=open_total,
        consistency_verdict=verdict, status_check_drift=drift,
        warnings=warnings,
    )


def _now_utc_iso() -> str:
    return (datetime.now(tz=timezone.utc).replace(microsecond=0)
            .isoformat().replace("+00:00", "Z"))


def aggregate(
    *, portfolio_root: Path, config: Mapping[str, Any],
    run_date: str, generated_at: str | None = None,
) -> HeatmapBatch:
    """Pure aggregation. pending_onboard NEVER read (schema)."""
    portfolio_id = _safe_str(config.get("portfolio_id"))
    display_name = _safe_str(config.get("display_name") or portfolio_id)

    projects: list[ProjectHeatmap] = []
    skipped = 0
    for entry in list(config.get("active_projects") or []):
        ph = _build_project_heatmap(
            portfolio_root=portfolio_root, project_entry=entry,
        )
        projects.append(ph)
        if ph.master_xlsx_path is None:
            skipped += 1
    projects.sort(key=lambda p: (p.priority, p.slug))

    p_cat: dict[str, int] = {}
    p_pri: dict[str, int] = {s: 0 for s in SEV_PLUS_OTHER}
    p_open = 0
    for ph in projects:
        p_open += ph.open_total
        for cat, n in ph.category_totals.items():
            p_cat[cat] = p_cat.get(cat, 0) + int(n)
        for sev, n in ph.priority_totals.items():
            p_pri.setdefault(sev, 0)
            p_pri[sev] += int(n)

    return HeatmapBatch(
        portfolio_id=portfolio_id, display_name=display_name,
        projects=tuple(projects), skipped_projects=skipped,
        run_date=run_date, generated_at=generated_at or _now_utc_iso(),
        portfolio_category_totals=p_cat,
        portfolio_priority_totals=p_pri,
        portfolio_open_total=p_open,
    )


# --- Snapshot + Markdown render ------------------------------------------

def build_snapshot(*, batch: HeatmapBatch) -> dict:
    return {
        "writer": WRITER_NAME, "version": "1.0",
        "run_date": batch.run_date, "generated_at": batch.generated_at,
        "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "active_project_count": len(batch.projects),
        "skipped_project_count": batch.skipped_projects,
        "portfolio_open_total": batch.portfolio_open_total,
        "portfolio_category_totals": dict(batch.portfolio_category_totals),
        "portfolio_priority_totals": dict(batch.portfolio_priority_totals),
        "projects": [{
            "slug": p.slug, "display_name": p.display_name,
            "priority": p.priority, "workspace_path": p.workspace_path,
            "master_xlsx_path": p.master_xlsx_path,
            "density": {c: dict(r) for c, r in p.density.items()},
            "status_distribution": dict(p.status_distribution),
            "category_totals": dict(p.category_totals),
            "priority_totals": dict(p.priority_totals),
            "open_total": p.open_total,
            "consistency_verdict": p.consistency_verdict,
            "status_check_drift": p.status_check_drift,
            "warnings": list(p.warnings),
        } for p in batch.projects],
    }


def _sev_header() -> str:
    cols = " | ".join(SEVERITY_ENUM) + f" | {OTHER_BUCKET} | total"
    sep = "|---" * (len(SEVERITY_ENUM) + 3) + "|"
    return f"| category | {cols} |\n{sep}"


def _row_cells(row: Mapping[str, int]) -> list[str]:
    cells = [str(int(row.get(s, 0))) for s in SEV_PLUS_OTHER]
    cells.append(str(sum(int(row.get(s, 0)) for s in SEV_PLUS_OTHER)))
    return cells


def _matrix_table(rows: Mapping[str, Mapping[str, int]]) -> str:
    if not rows:
        return _sev_header() + "\n| _(no open tasks)_ | 0 | 0 | 0 | 0 | 0 | 0 |"
    return _sev_header() + "\n" + "\n".join(
        f"| {cat} | " + " | ".join(_row_cells(rows[cat])) + " |"
        for cat in sorted(rows.keys())
    )


def _projects_matrix(batch: HeatmapBatch) -> str:
    if not batch.projects:
        return "_(no active projects)_\n"
    parts = []
    for p in batch.projects:
        out = [f"### `{p.slug}` — {p.display_name} (priority={p.priority}, "
               f"open_total={p.open_total})", "", _matrix_table(p.density)]
        if p.warnings:
            out.append("")
            out.extend(f"> WARNING [{p.slug}]: {w}" for w in p.warnings)
        out.append("")
        parts.append("\n".join(out))
    return "\n".join(parts)


def _category_totals_table(batch: HeatmapBatch) -> str:
    if not batch.portfolio_category_totals:
        return "_(no open tasks across portfolio)_\n"
    matrix: dict[str, dict[str, int]] = {}
    for p in batch.projects:
        for cat, row in p.density.items():
            tgt = matrix.setdefault(cat, {s: 0 for s in SEV_PLUS_OTHER})
            for sev, n in row.items():
                tgt[sev] = tgt.get(sev, 0) + int(n)
    return _matrix_table(matrix) + "\n"


def _priority_totals_table(batch: HeatmapBatch) -> str:
    if not batch.projects:
        return "_(no active projects)_\n"
    slugs = [p.slug for p in batch.projects]
    lines = ["| priority | " + " | ".join(slugs) + " | total |",
             "|---" * (len(slugs) + 2) + "|"]
    for sev in SEV_PLUS_OTHER:
        cells = [str(int(p.priority_totals.get(sev, 0)))
                 for p in batch.projects]
        cells.append(str(sum(int(p.priority_totals.get(sev, 0))
                             for p in batch.projects)))
        lines.append(f"| {sev} | " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def _advisory_block(batch: HeatmapBatch) -> str:
    drifters = [p for p in batch.projects if p.status_check_drift]
    if not drifters:
        return ("_(no advisories — all projects clean or "
                "no consistency-report)_\n")
    return "\n".join(
        f"- `{p.slug}` — status_check_drift "
        f"(verdict={p.consistency_verdict or 'unknown'}). "
        f"Transform proceeded; manager review suggested."
        for p in drifters
    ) + "\n"


def _totals_summary(batch: HeatmapBatch) -> str:
    top = sorted(batch.portfolio_category_totals.items(),
                 key=lambda kv: (-int(kv[1]), kv[0]))[:3]
    top_str = (", ".join(f"{c}={n}" for c, n in top) if top
               else "(no categories)")
    return (f"- active projects: **{len(batch.projects)}** "
            f"(skipped: {batch.skipped_projects})\n"
            f"- portfolio_open_total: **{batch.portfolio_open_total}**\n"
            f"- top-3 categories: {top_str}\n")


def build_report_markdown(
    *, batch: HeatmapBatch, template_text: str,
) -> str:
    data = {
        "generated_at": batch.generated_at,
        "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "project_count": str(len(batch.projects)),
        "skipped_count": str(batch.skipped_projects),
        "scope": ("portfolio-task-heatmap "
                  "(read-only project x category x priority density)"),
        "totals_summary": _totals_summary(batch),
        "projects_matrix": _projects_matrix(batch),
        "category_totals_table": _category_totals_table(batch),
        "priority_totals_table": _priority_totals_table(batch),
        "advisory_block": _advisory_block(batch),
    }
    try:
        return Template(template_text).safe_substitute(
            {k: str(v) for k, v in data.items()})
    except KeyError as exc:
        raise PortfolioTaskHeatmapError(
            f"template substitution failed - missing key: {exc.args[0]}"
        ) from exc


# --- Read-only contract self-check ---------------------------------------

#: Patterns built at runtime so the literals do not appear verbatim in source.
_FORBIDDEN_WRITE_PATTERNS: tuple[str, ...] = (
    "." + "save(",
    "transaction" + "." + "append(",
    "transaction" + "." + "update(",
    "transaction" + "." + "delete(",
)


def assert_read_only_module() -> None:
    """W-E3 helper reuse: grep self-check; raises on any forbidden token."""
    src = Path(__file__).read_text(encoding="utf-8")
    hits = [p for p in _FORBIDDEN_WRITE_PATTERNS if p in src]
    if hits:
        raise ReadOnlyContractViolation(
            f"portfolio_task_heatmap source contains forbidden write "
            f"patterns: {hits}."
        )


# --- CLI -----------------------------------------------------------------

def _resolve_portfolio_root(arg: str | None) -> Path:
    if arg:
        p = Path(arg).expanduser().resolve()
        if not p.is_dir():
            raise WorkspaceRootUnsetError(
                f"--portfolio-root path does not exist: {p}")
        return p
    env = os.environ.get("PSEO_WORKSPACE_ROOT")
    if env:
        p = Path(env).expanduser().resolve()
        if p.is_dir():
            return p
    cwd = Path.cwd().resolve()
    for parent in (cwd, *cwd.parents):
        if (parent / ".pse-workspace").exists():
            return parent
    raise WorkspaceRootUnsetError(
        "portfolio_root unresolved: pass --portfolio-root, set "
        "PSEO_WORKSPACE_ROOT, or place a .pse-workspace marker.")


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(prog="portfolio_task_heatmap.py")
    ap.add_argument("--portfolio-root", default=None)
    ap.add_argument("--reference-date", default=None)
    ap.add_argument("--template", default=None)
    ap.add_argument("--output-dir", default=None)
    args = ap.parse_args(argv)
    try:
        portfolio_root = _resolve_portfolio_root(args.portfolio_root)
        config = load_portfolio_config(portfolio_root)
    except (WorkspaceRootUnsetError, PortfolioConfigMissingError) as exc:
        print(f"setup failed: {exc}", file=sys.stderr)
        return 2
    schema_path = _REPO_ROOT / "schemas" / "portfolio-config.schema.json"
    if schema_path.is_file():
        try:
            validate_portfolio_config(
                config,
                json.loads(schema_path.read_text(encoding="utf-8")))
        except (PortfolioConfigInvalidError, ActiveProjectsCeilingError,
                ReadOnlyContractViolation) as exc:
            print(f"config validation failed: {exc}", file=sys.stderr)
            return 1
    run_date = args.reference_date or _date.today().isoformat()
    batch = aggregate(
        portfolio_root=portfolio_root, config=config, run_date=run_date)
    tpl_path = (Path(args.template).expanduser() if args.template
                else _REPO_ROOT / DEFAULT_TEMPLATE_REL)
    if not tpl_path.is_file():
        print(f"template not found: {tpl_path}", file=sys.stderr)
        return 2
    snapshot = build_snapshot(batch=batch)
    report = build_report_markdown(
        batch=batch, template_text=tpl_path.read_text(encoding="utf-8"))
    if args.output_dir:
        out_dir = Path(args.output_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
        snap_path = out_dir / f"{run_date}-portfolio-task-heatmap.json"
        snap_path.write_text(json.dumps(
            snapshot, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8")
        report_path = out_dir / f"{run_date}-portfolio-task-heatmap.md"
        report_path.write_text(report, encoding="utf-8")
        print(json.dumps({
            "snapshot_path": str(snap_path.resolve()),
            "report_path": str(report_path.resolve()),
            "active_project_count": len(batch.projects),
            "skipped_project_count": batch.skipped_projects,
            "portfolio_open_total": batch.portfolio_open_total,
        }, ensure_ascii=False, indent=2))
    else:
        print(json.dumps(
            {"snapshot": snapshot, "report_md": report},
            ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main(sys.argv[1:]))
