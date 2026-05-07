#!/usr/bin/env python3
"""portfolio_kpi_trend.py — Phase 9 W2 reporting transform (W-E7).

Multi-project READ-ONLY daily time-series aggregator. period_days range
[7,90], default 30. Per active project: tasks_added (master_task col K
created_date), tasks_done (col L done_date status=DONE), work_events
density per event_type 10-enum (events.jsonl event_kind=work). LOCAL
approximation paterni (W-E1+W-E2): gscTotals stub all keys=0 until GSC
longitudinal Phase 6+. Schema authority: portfolio-config v1.1,
master-excel master_task K/L, events event_type 10 enum,
monthly-report gscTotals stub. Outputs:
projects/_portfolio/outputs/reports/{date}-portfolio-kpi-trend.md +
inbox/local/{date}-portfolio-kpi-trend.json + master.xlsx#none. Q-RP-01
defer (no events.jsonl write).
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import date as _date, datetime, timedelta, timezone
from pathlib import Path
from string import Template
from typing import Any, Mapping, Sequence

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

EVENT_TYPE_ENUM: tuple[str, ...] = (
    "content_new", "content_revise", "content_remove", "tech_fix",
    "quickwin_applied", "pillar_launch", "schema_fix",
    "redirect_deployed", "backlink_outreach", "manual",
)
STATUS_ENUM: tuple[str, ...] = (
    "TODO", "ONGOING", "EXISTS", "DONE", "BLOCKED", "DEFERRED", "CANCELED",
)
PERIOD_DAYS_MIN, PERIOD_DAYS_MAX, PERIOD_DAYS_DEFAULT = 7, 90, 30
ACTIVE_PROJECTS_MAX = 8
WRITER_NAME = "portfolio_kpi_trend"
DEFAULT_TEMPLATE_REL = "templates/reports/portfolio-kpi-trend.template.md"

class PortfolioKpiTrendError(Exception): pass
class PortfolioConfigMissingError(PortfolioKpiTrendError): pass
class PortfolioConfigInvalidError(PortfolioKpiTrendError): pass
class ActiveProjectsCeilingError(PortfolioKpiTrendError): pass
class ReadOnlyContractViolation(PortfolioKpiTrendError): pass
class WorkspaceRootUnsetError(PortfolioKpiTrendError): pass
class PeriodDaysOutOfRangeError(PortfolioKpiTrendError): pass

@dataclass(frozen=True)
class DailyBucket:
    day: str
    tasks_added: int
    tasks_done: int
    work_events_total: int

@dataclass(frozen=True)
class ProjectTrend:
    slug: str
    priority: int
    workspace_path: str
    master_xlsx_path: str | None
    events_jsonl_path: str | None
    daily: tuple[DailyBucket, ...]
    event_type_buckets: dict
    gsc_totals_stub: dict
    warnings: list

@dataclass(frozen=True)
class PortfolioKpiTrendBatch:
    portfolio_id: str
    display_name: str
    period_start: str
    period_end: str
    period_days: int
    generated_at: str
    projects: tuple[ProjectTrend, ...]
    totals_tasks_added: int
    totals_tasks_done: int
    totals_work_events: int
    totals_event_type_buckets: dict

def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()

def _coerce_date_cell(value: Any) -> _date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    if not isinstance(value, str):
        return None
    try:
        return _date.fromisoformat(value.strip()[:10])
    except ValueError:
        return None

def _parse_iso_timestamp(value: str) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    raw = value.strip()
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return None
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)

def _validate_period_days(value: Any) -> int:
    try:
        n = int(value)
    except (TypeError, ValueError) as exc:
        raise PeriodDaysOutOfRangeError(
            f"period_days={value!r} not int-coercible") from exc
    if n < PERIOD_DAYS_MIN or n > PERIOD_DAYS_MAX:
        raise PeriodDaysOutOfRangeError(
            f"period_days={n} not in [{PERIOD_DAYS_MIN},{PERIOD_DAYS_MAX}]")
    return n

def load_portfolio_config(portfolio_root: Path) -> dict:
    """Read portfolio.config.json. DURUR PortfolioConfigMissingError."""
    cfg_path = (portfolio_root / "projects" / "_portfolio"
                / "portfolio.config.json")
    if not cfg_path.is_file():
        raise PortfolioConfigMissingError(
            f"portfolio.config.json not found at {cfg_path}.")
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PortfolioConfigMissingError(
            f"portfolio.config.json read failed: {exc}") from exc

def validate_portfolio_config(
    config: Mapping[str, Any], schema: Mapping[str, Any],
) -> None:
    """Draft 7 validate + ActiveProjectsCeiling + ReadOnlyContract DURUR."""
    try:
        from jsonschema import Draft7Validator
    except ImportError as exc:  # pragma: no cover
        raise PortfolioConfigInvalidError("jsonschema required") from exc
    errors = sorted(Draft7Validator(dict(schema)).iter_errors(dict(config)),
                    key=lambda e: list(e.absolute_path))
    if errors:
        raise PortfolioConfigInvalidError(
            "portfolio.config.json failed Draft 7 validation: " + "; ".join(
                f"{list(e.absolute_path)}: {e.message}" for e in errors))
    active = list(config.get("active_projects") or [])
    if len(active) > ACTIVE_PROJECTS_MAX:
        raise ActiveProjectsCeilingError(
            f"active_projects has {len(active)} entries; ceiling is "
            f"{ACTIVE_PROJECTS_MAX} (schema maxItems).")
    cq = config.get("cross_query") or {}
    if "read_only" in cq and cq.get("read_only") is not True:
        raise ReadOnlyContractViolation(
            f"cross_query.read_only must be true (schema const), got "
            f"{cq.get('read_only')!r}.")

def _resolve_workspace_root(workspace_root: Path | str | None) -> Path:
    if workspace_root is not None:
        return Path(workspace_root).expanduser().resolve()
    env_val = os.environ.get("PSEO_WORKSPACE_ROOT")
    if env_val:
        return Path(env_val).expanduser().resolve()
    raise WorkspaceRootUnsetError(
        "PSEO_WORKSPACE_ROOT env var unset and no workspace_root passed")

def _resolve_project_root(
    *, workspace_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    """workspace_path override OR {workspace_root}/projects/{slug}."""
    slug = _safe_str(project_entry.get("slug"))
    explicit = _safe_str(project_entry.get("workspace_path"))
    if explicit:
        p = Path(explicit).expanduser()
        return p if p.is_absolute() else (workspace_root / p).resolve()
    return (workspace_root / "projects" / slug).resolve()

def read_work_events_in_window(
    events_path: Path, *, window_start: _date, window_end: _date,
) -> list[dict[str, Any]]:
    """Return event_kind=work events whose timestamp date ∈ window."""
    if not events_path.exists():
        return []
    out: list[dict[str, Any]] = []
    for line in events_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            evt = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(evt, dict) or evt.get("event_kind") != "work":
            continue
        ts_dt = _parse_iso_timestamp(evt.get("timestamp") or "")
        if ts_dt is None:
            continue
        if window_start <= ts_dt.date() <= window_end:
            out.append(evt)
    return out

def bucket_events_by_type(
    events: Sequence[Mapping[str, Any]],
) -> dict[str, int]:
    """Bucket by event_type. Every EVENT_TYPE_ENUM key seeded; unknown
    types accumulate under '__other__' (schema drift signal)."""
    buckets: dict[str, int] = {k: 0 for k in EVENT_TYPE_ENUM}
    buckets["__other__"] = 0
    for evt in events:
        et = evt.get("event_type")
        key = et if isinstance(et, str) and et in buckets else "__other__"
        buckets[key] += 1
    return buckets

def bucket_events_by_day(
    events: Sequence[Mapping[str, Any]],
) -> dict[str, int]:
    """Bucket by ISO date. Daily-axis builder fills gaps."""
    out: dict[str, int] = {}
    for evt in events:
        ts_dt = _parse_iso_timestamp(evt.get("timestamp") or "")
        if ts_dt is None:
            continue
        key = ts_dt.date().isoformat()
        out[key] = out.get(key, 0) + 1
    return out

def read_master_task_dated_rows(
    workbook_path: Path, *, window_start: _date, window_end: _date,
) -> tuple[dict[str, int], dict[str, int]]:
    """Return (added_by_day, done_by_day) ISO-date → count. READ-ONLY."""
    added: dict[str, int] = {}
    done: dict[str, int] = {}
    if not workbook_path.exists():
        return added, done
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise PortfolioKpiTrendError("openpyxl required") from exc
    try:
        wb = load_workbook(filename=str(workbook_path),
                           read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise PortfolioKpiTrendError(
            f"could not open {workbook_path}: {exc}") from exc
    try:
        if "master_task" not in wb.sheetnames:
            return added, done
        rows_iter = wb["master_task"].iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return added, done
        if not header:
            return added, done
        col = {n: i for i, n in enumerate(header) if isinstance(n, str)}
        i_st, i_c, i_d = (col.get("status"), col.get("created_date"),
                          col.get("done_date"))
        if i_st is None or i_c is None or i_d is None:
            return added, done
        for row in rows_iter:
            if row is None:
                continue
            cr = _coerce_date_cell(row[i_c]) if i_c < len(row) else None
            if cr is not None and window_start <= cr <= window_end:
                added[cr.isoformat()] = added.get(cr.isoformat(), 0) + 1
            sv = row[i_st] if i_st < len(row) else None
            dv = _coerce_date_cell(row[i_d]) if i_d < len(row) else None
            if (isinstance(sv, str) and sv.strip().upper() == "DONE"
                    and dv is not None
                    and window_start <= dv <= window_end):
                done[dv.isoformat()] = done.get(dv.isoformat(), 0) + 1
    finally:
        wb.close()
    return added, done

def _build_daily_axis(
    *, period_start: _date, period_end: _date,
    added_by_day: Mapping[str, int], done_by_day: Mapping[str, int],
    events_by_day: Mapping[str, int],
) -> tuple[DailyBucket, ...]:
    """Contiguous daily axis (no gaps) covering the period."""
    out: list[DailyBucket] = []
    cur = period_start
    while cur <= period_end:
        k = cur.isoformat()
        out.append(DailyBucket(
            day=k, tasks_added=int(added_by_day.get(k, 0)),
            tasks_done=int(done_by_day.get(k, 0)),
            work_events_total=int(events_by_day.get(k, 0))))
        cur += timedelta(days=1)
    return tuple(out)

def aggregate_project(
    *, project_entry: Mapping[str, Any], workspace_root: Path,
    period_start: _date, period_end: _date,
) -> ProjectTrend:
    """One project's period trend. Reads filesystem; never writes."""
    slug = _safe_str(project_entry.get("slug"))
    priority = int(project_entry.get("priority") or 10)
    workspace_path = _safe_str(project_entry.get("workspace_path"))
    project_root = _resolve_project_root(
        workspace_root=workspace_root, project_entry=project_entry)
    master_xlsx = project_root / "master.xlsx"
    events_jsonl = project_root / "_state" / "events.jsonl"
    warnings: list = []
    mp, ep = master_xlsx.is_file(), events_jsonl.exists()
    if not mp:
        warnings.append(f"master.xlsx not found at {master_xlsx}")
    if not ep:
        warnings.append(f"events.jsonl not found at {events_jsonl}")
    abd, dbd = read_master_task_dated_rows(
        master_xlsx, window_start=period_start, window_end=period_end)
    events = read_work_events_in_window(
        events_jsonl, window_start=period_start, window_end=period_end)
    etb = bucket_events_by_type(events)
    if etb.get("__other__", 0):
        warnings.append(
            f"events.jsonl has {etb['__other__']} entries with event_type "
            "outside the schema enum")
    # LOCAL approximation (W-E1+W-E2 reuse): gscTotals stub all keys=0
    # until GSC longitudinal Phase 6+ entegrasyonu.
    return ProjectTrend(
        slug=slug, priority=priority, workspace_path=workspace_path,
        master_xlsx_path=str(master_xlsx) if mp else None,
        events_jsonl_path=str(events_jsonl) if ep else None,
        daily=_build_daily_axis(
            period_start=period_start, period_end=period_end,
            added_by_day=abd, done_by_day=dbd,
            events_by_day=bucket_events_by_day(events)),
        event_type_buckets=etb,
        gsc_totals_stub={
            "clicks": 0, "impressions": 0, "avg_position": 0.0,
            "ctr": 0.0,
            "approximation_note": ("LOCAL approximation — GSC longitudinal "
                                   "Phase 6+ entegrasyonu öncesinde sentinel değer."),
        },
        warnings=warnings)

def build_kpi_trend_batch(
    *, portfolio_config: Mapping[str, Any], workspace_root: Path,
    now: datetime, period_days: int = PERIOD_DAYS_DEFAULT,
) -> PortfolioKpiTrendBatch:
    """Top-level pure aggregator. Idempotent: same inputs → same batch."""
    if not isinstance(portfolio_config, Mapping):
        raise PortfolioConfigInvalidError(
            "portfolio_config must be a JSON object")
    pd = _validate_period_days(period_days)
    if now.tzinfo is None:
        raise PortfolioKpiTrendError(
            "now must be timezone-aware (UTC expected)")
    portfolio_id = _safe_str(portfolio_config.get("portfolio_id"))
    display_name = _safe_str(
        portfolio_config.get("display_name") or portfolio_id)
    active = list(portfolio_config.get("active_projects") or [])
    if len(active) > ACTIVE_PROJECTS_MAX:
        raise ActiveProjectsCeilingError(
            f"active_projects has {len(active)} entries; ceiling is "
            f"{ACTIVE_PROJECTS_MAX}.")
    period_end = now.date()
    period_start = period_end - timedelta(days=pd - 1)
    trends: list[ProjectTrend] = []
    for entry in active:
        if not isinstance(entry, Mapping):
            raise PortfolioConfigInvalidError(
                f"active_projects entry must be object, got "
                f"{type(entry).__name__}")
        slug = entry.get("slug")
        if not isinstance(slug, str) or not slug:
            raise PortfolioConfigInvalidError(
                f"active_projects[].slug must be non-empty string, got "
                f"{slug!r}")
        trends.append(aggregate_project(
            project_entry=entry, workspace_root=workspace_root,
            period_start=period_start, period_end=period_end))
    trends.sort(key=lambda t: (t.priority, t.slug))
    tet: dict[str, int] = {k: 0 for k in EVENT_TYPE_ENUM}
    for t in trends:
        for k in EVENT_TYPE_ENUM:
            tet[k] += int(t.event_type_buckets.get(k, 0))
    return PortfolioKpiTrendBatch(
        portfolio_id=portfolio_id, display_name=display_name,
        period_start=period_start.isoformat(),
        period_end=period_end.isoformat(), period_days=pd,
        generated_at=now.replace(microsecond=0).isoformat().replace(
            "+00:00", "Z"),
        projects=tuple(trends),
        totals_tasks_added=sum(
            sum(b.tasks_added for b in t.daily) for t in trends),
        totals_tasks_done=sum(
            sum(b.tasks_done for b in t.daily) for t in trends),
        totals_work_events=sum(
            sum(b.work_events_total for b in t.daily) for t in trends),
        totals_event_type_buckets=tet)

def build_snapshot(batch: PortfolioKpiTrendBatch) -> dict:
    """Build inbox/local/{date}-portfolio-kpi-trend.json payload."""
    return {
        "writer": WRITER_NAME, "version": "1.0",
        "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "period_start": batch.period_start,
        "period_end": batch.period_end, "period_days": batch.period_days,
        "generated_at": batch.generated_at,
        "totals": {"tasks_added": batch.totals_tasks_added,
                   "tasks_done": batch.totals_tasks_done,
                   "work_events": batch.totals_work_events,
                   "event_type_buckets": dict(
                       batch.totals_event_type_buckets)},
        "projects": [{
            "slug": t.slug, "priority": t.priority,
            "workspace_path": t.workspace_path,
            "master_xlsx_path": t.master_xlsx_path,
            "events_jsonl_path": t.events_jsonl_path,
            "daily": [{"day": b.day, "tasks_added": b.tasks_added,
                       "tasks_done": b.tasks_done,
                       "work_events_total": b.work_events_total}
                      for b in t.daily],
            "event_type_buckets": dict(t.event_type_buckets),
            "gsc_totals_stub": dict(t.gsc_totals_stub),
            "warnings": list(t.warnings),
        } for t in batch.projects],
    }

def _render_project_trend_section(t: ProjectTrend) -> str:
    """Render one project's trend block (summary + event_type table)."""
    lines: list[str] = [f"### `{t.slug}` (priority={t.priority})"]
    if t.master_xlsx_path is None:
        lines.append("- master.xlsx eksik (graceful skip)")
    if t.events_jsonl_path is None:
        lines.append("- events.jsonl eksik (graceful skip)")
    pa = sum(b.tasks_added for b in t.daily)
    pd_ = sum(b.tasks_done for b in t.daily)
    pe = sum(b.work_events_total for b in t.daily)
    lines.append(
        f"- toplam tasks_added={pa}, tasks_done={pd_}, work_events={pe}")
    lines.extend(("", "| event_type | count |", "|---|---:|"))
    for k in EVENT_TYPE_ENUM:
        lines.append(f"| `{k}` | {int(t.event_type_buckets.get(k, 0))} |")
    if t.event_type_buckets.get("__other__", 0):
        lines.append(
            f"| `__other__` (schema drift) | "
            f"{int(t.event_type_buckets['__other__'])} |")
    lines.extend(("", (
        "> LOCAL approximation — gscTotals stub: clicks=0, impressions=0, "
        "avg_position=0.0, ctr=0.0 (GSC longitudinal Phase 6+ "
        "entegrasyonu öncesinde sentinel değer).")))
    return "\n".join(lines)

def build_template_vars(
    batch: PortfolioKpiTrendBatch, *, scope: str = "portfolio-kpi-trend",
) -> dict[str, str]:
    """Flat dict for string.Template $var substitution."""
    projects_block = ("_(no active projects)_\n" if not batch.projects
                      else "\n\n".join(_render_project_trend_section(t)
                                       for t in batch.projects) + "\n")
    et_lines = ["| event_type | count (portföy toplam) |", "|---|---:|"]
    for k in EVENT_TYPE_ENUM:
        et_lines.append(
            f"| `{k}` | {int(batch.totals_event_type_buckets.get(k, 0))} |")
    return {
        "scope": scope, "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "period_start": batch.period_start,
        "period_end": batch.period_end,
        "period_days": str(batch.period_days),
        "generated_at": batch.generated_at,
        "projects_count": str(len(batch.projects)),
        "totals_tasks_added": str(batch.totals_tasks_added),
        "totals_tasks_done": str(batch.totals_tasks_done),
        "totals_work_events": str(batch.totals_work_events),
        "totals_summary": (
            f"Toplam tasks_added={batch.totals_tasks_added}, "
            f"tasks_done={batch.totals_tasks_done}, "
            f"work_events={batch.totals_work_events} "
            f"(period={batch.period_days}d, projects={len(batch.projects)})"),
        "event_type_totals_table": "\n".join(et_lines),
        "projects_block": projects_block,
    }

def render_report_markdown(
    *, batch: PortfolioKpiTrendBatch, template_path: Path,
) -> str:
    """Render Markdown via string.Template (matches render_template.py)."""
    if not template_path.exists():
        raise PortfolioKpiTrendError(f"template not found: {template_path}")
    try:
        return Template(template_path.read_text(encoding="utf-8")
                        ).safe_substitute(build_template_vars(batch))
    except KeyError as exc:
        raise PortfolioKpiTrendError(
            f"template references unknown variable: {exc.args[0]}") from exc

#: Patterns built at runtime so the literals do not appear verbatim
#: in source (the self-check below greps the file).
_FORBIDDEN_WRITE_PATTERNS: tuple[str, ...] = (
    "." + "save(",
    "transaction" + "." + "append(",
    "transaction" + "." + "update(",
    "transaction" + "." + "delete(",
    "events_writer" + "." + "append_",
)

def assert_read_only_module() -> None:
    """Verify this module's source has no write call sites."""
    src = Path(__file__).read_text(encoding="utf-8")
    hits = [p for p in _FORBIDDEN_WRITE_PATTERNS if p in src]
    if hits:
        raise ReadOnlyContractViolation(
            f"portfolio_kpi_trend source contains forbidden write "
            f"patterns: {hits}.")

def main(argv: Sequence[str] | None = None) -> int:  # pragma: no cover
    p = argparse.ArgumentParser(
        prog="portfolio_kpi_trend.py",
        description="Multi-project daily KPI trend (READ-ONLY).")
    p.add_argument("--portfolio-root", default=None)
    p.add_argument("--period-days", type=int, default=PERIOD_DAYS_DEFAULT)
    p.add_argument("--output-dir", default=None)
    p.add_argument("--template", default=None)
    p.add_argument("--now", default=None)
    args = p.parse_args(argv)
    try:
        workspace_root = _resolve_workspace_root(args.portfolio_root)
        config = load_portfolio_config(workspace_root)
    except (WorkspaceRootUnsetError, PortfolioConfigMissingError) as exc:
        print(f"setup failed: {exc}", file=sys.stderr)
        return 2
    schema_path = _REPO_ROOT / "schemas" / "portfolio-config.schema.json"
    if schema_path.is_file():
        try:
            validate_portfolio_config(
                config, json.loads(schema_path.read_text(encoding="utf-8")))
        except (PortfolioConfigInvalidError, ActiveProjectsCeilingError,
                ReadOnlyContractViolation) as exc:
            print(f"config validation failed: {exc}", file=sys.stderr)
            return 1
    try:
        period_days = _validate_period_days(args.period_days)
    except PeriodDaysOutOfRangeError as exc:
        print(f"period_days invalid: {exc}", file=sys.stderr)
        return 1
    now = (_parse_iso_timestamp(args.now) if args.now
           else datetime.now(tz=timezone.utc).replace(microsecond=0))
    if now is None:
        print(f"--now invalid: {args.now!r}", file=sys.stderr)
        return 1
    batch = build_kpi_trend_batch(
        portfolio_config=config, workspace_root=workspace_root,
        now=now, period_days=period_days)
    template_path = (Path(args.template).expanduser() if args.template
                     else _REPO_ROOT / DEFAULT_TEMPLATE_REL)
    out_dir = (Path(args.output_dir).expanduser() if args.output_dir
               else workspace_root / "projects" / "_portfolio"
               / "outputs" / "reports")
    inbox_dir = (workspace_root / "projects" / "_portfolio"
                 / "inbox" / "local")
    out_dir.mkdir(parents=True, exist_ok=True)
    inbox_dir.mkdir(parents=True, exist_ok=True)
    snap_path = inbox_dir / f"{batch.period_end}-portfolio-kpi-trend.json"
    report_path = out_dir / f"{batch.period_end}-portfolio-kpi-trend.md"
    snap_path.write_text(json.dumps(
        build_snapshot(batch), ensure_ascii=False,
        indent=2, sort_keys=True), encoding="utf-8")
    report_path.write_text(render_report_markdown(
        batch=batch, template_path=template_path), encoding="utf-8")
    print(json.dumps({
        "writer": WRITER_NAME, "period_start": batch.period_start,
        "period_end": batch.period_end, "period_days": batch.period_days,
        "report_path": str(report_path), "snapshot_path": str(snap_path),
        "totals_tasks_added": batch.totals_tasks_added,
        "totals_tasks_done": batch.totals_tasks_done,
        "totals_work_events": batch.totals_work_events,
        "active_project_count": len(batch.projects),
    }, ensure_ascii=False))
    return 0

if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())

__all__ = (
    "EVENT_TYPE_ENUM", "STATUS_ENUM", "PERIOD_DAYS_MIN", "PERIOD_DAYS_MAX",
    "PERIOD_DAYS_DEFAULT", "ACTIVE_PROJECTS_MAX", "WRITER_NAME",
    "PortfolioKpiTrendError", "PortfolioConfigMissingError",
    "PortfolioConfigInvalidError", "ActiveProjectsCeilingError",
    "ReadOnlyContractViolation", "WorkspaceRootUnsetError",
    "PeriodDaysOutOfRangeError", "DailyBucket", "ProjectTrend",
    "PortfolioKpiTrendBatch", "load_portfolio_config",
    "validate_portfolio_config", "read_work_events_in_window",
    "bucket_events_by_type", "bucket_events_by_day",
    "read_master_task_dated_rows", "aggregate_project",
    "build_kpi_trend_batch", "build_snapshot", "build_template_vars",
    "render_report_markdown", "assert_read_only_module", "main")
