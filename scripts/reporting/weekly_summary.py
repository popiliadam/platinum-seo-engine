#!/usr/bin/env python3
"""
weekly_summary.py — pure aggregator: master.xlsx + events.jsonl → weekly
summary (LOCAL aggregation, NO MCP, NO DFS, NO budget pre-flight).

Phase 9 Wave 1 reporting skill (W-E2). 7-day rolling window per project.

Output shape (5 sections — monthly-report SUBSET, NO schema this Wave;
manual structural assert in test):
  - exec_summary, gsc_weekly_delta, tasks_done, tasks_added, drift_signals

Read sources (READ-ONLY):
  - master.xlsx#master_task / opportunity / quick_wins / content_decay
  - _state/events.jsonl (event_kind=work, in window)

Write sinks (3 outputs, REVIZE 1 — events.jsonl write removed):
  - outputs/reports/{date}-weekly-summary.md
  - inbox/local/{date}-weekly-{slug}.json

Discipline: pure functions; no MCP/DFS imports; idempotent; 0 slug literals;
0 events.jsonl writes (REVIZE 3 — Q-RP-01 deferred to Phase 14).

Refs: schemas/master-excel.schema.json,
scripts/reporting/render_template.py (string.Template $var compatible),
scripts/planning/master_task_sync.py (pattern reference).
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from datetime import date as _date, datetime, timedelta
from pathlib import Path
from string import Template
from typing import Any, Iterable, Mapping, Sequence


_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


# Constants
WINDOW_DAYS = 7
WEEKLY_SUMMARY_SECTIONS: tuple[str, ...] = (
    "exec_summary",
    "gsc_weekly_delta",
    "tasks_done",
    "tasks_added",
    "drift_signals",
)
STATUS_DONE = "DONE"
EVENT_KIND_WORK = "work"


# Exceptions
class WeeklySummaryError(Exception):
    """Base class for weekly-summary DURUR conditions."""


class WorkbookMissingError(WeeklySummaryError):
    """master.xlsx not found at projects/{slug}/. Run init-project first."""


class WorkspaceRootUnsetError(WeeklySummaryError):
    """workspace_root unresolvable; cannot reach projects/{slug}/."""


class TemplateMissingError(WeeklySummaryError):
    """templates/reports/weekly-summary.template.md not found."""


# Helpers
def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def snap_to_last_sunday(reference: _date) -> _date:
    """Snap to the most recent Sunday (inclusive). Sunday→Sunday is no-op.
    Python weekday(): Mon=0..Sun=6 → days_back = (weekday + 1) % 7."""
    days_back = (reference.weekday() + 1) % 7
    return reference - timedelta(days=days_back)


def compute_window(week_end: _date | None) -> tuple[_date, _date]:
    """7-day inclusive window ending on week_end (snapped to Sunday)."""
    if week_end is None:
        week_end = snap_to_last_sunday(_date.today())
    elif week_end.weekday() != 6:
        week_end = snap_to_last_sunday(week_end)
    week_start = week_end - timedelta(days=WINDOW_DAYS - 1)
    return week_start, week_end


def _parse_iso_date(value: Any) -> _date | None:
    """Parse ISO-8601 date or datetime; return None on failure / empty."""
    s = _safe_str(value)
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
        except ValueError:
            return None


def _in_window(value: Any, window_start: _date, window_end: _date) -> bool:
    d = _parse_iso_date(value)
    return d is not None and window_start <= d <= window_end


# Excel reader (INLINE openpyxl thin wrapper — Q-MTS-1: no shared helper)
def _read_sheet_rows(
    workbook_path: Path, sheet_name: str, header_row: int, data_start_row: int,
) -> list[dict]:
    """READ-ONLY: load sheet into list of header-keyed row-dicts. [] on
    missing file / missing sheet / empty header."""
    if not workbook_path.is_file():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as exc:    # pragma: no cover
        raise WeeklySummaryError(
            "openpyxl is required to read master.xlsx; install requirements"
        ) from exc

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        header_cells = next(
            ws.iter_rows(
                min_row=header_row, max_row=header_row, values_only=True,
            ),
            (),
        )
        headers = [_safe_str(h) for h in header_cells]
        if not any(headers):
            return []
        rows: list[dict] = []
        for raw in ws.iter_rows(min_row=data_start_row, values_only=True):
            if raw is None or all(v is None or _safe_str(v) == "" for v in raw):
                continue
            row = {}
            for idx, name in enumerate(headers):
                if not name:
                    continue
                row[name] = raw[idx] if idx < len(raw) else None
            rows.append(row)
        return rows
    finally:
        wb.close()


# Per-sheet header / data offsets (mirrors master-excel.schema.json).
_SHEET_OFFSETS: dict[str, tuple[int, int]] = {
    "master_task":   (3, 4),
    "opportunity":   (4, 5),
    "quick_wins":    (4, 5),
    "content_decay": (5, 6),
}


def read_master_sheets(
    workbook_path: Path,
    sheets: Iterable[str] = (
        "master_task", "opportunity", "quick_wins", "content_decay",
    ),
) -> dict[str, list[dict]]:
    """READ-ONLY: load multiple sheets in one workbook open."""
    out: dict[str, list[dict]] = {}
    for s in sheets:
        if s not in _SHEET_OFFSETS:
            out[s] = []
            continue
        header_row, data_start_row = _SHEET_OFFSETS[s]
        out[s] = _read_sheet_rows(
            workbook_path, s, header_row, data_start_row,
        )
    return out


# events.jsonl reader (READ-ONLY — last 7-day filter; NO writes)
def read_events_jsonl(
    events_path: Path,
    window_start: _date,
    window_end: _date,
    event_kind: str = EVENT_KIND_WORK,
) -> list[dict]:
    """READ-ONLY: stream events.jsonl, return entries with matching
    event_kind AND timestamp inside the inclusive window. Missing file
    → []; malformed lines → silently skipped."""
    if not events_path.is_file():
        return []
    out: list[dict] = []
    for line in events_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            evt = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(evt, dict):
            continue
        if _safe_str(evt.get("event_kind")) != event_kind:
            continue
        if not _in_window(evt.get("timestamp"), window_start, window_end):
            continue
        out.append(evt)
    return out


# Aggregation
@dataclass(frozen=True)
class WeeklySummaryBatch:
    """Pure-data weekly summary. No I/O implied."""

    project_slug: str
    week_start: _date
    week_end: _date
    exec_summary: str
    gsc_weekly_delta: dict
    tasks_done: list[dict] = field(default_factory=list)
    tasks_added: list[dict] = field(default_factory=list)
    drift_signals: dict = field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            "project_slug":     self.project_slug,
            "week_start":       self.week_start.isoformat(),
            "week_end":         self.week_end.isoformat(),
            "exec_summary":     self.exec_summary,
            "gsc_weekly_delta": dict(self.gsc_weekly_delta),
            "tasks_done":       list(self.tasks_done),
            "tasks_added":      list(self.tasks_added),
            "drift_signals":    dict(self.drift_signals),
        }


def _filter_tasks_done(
    rows: Sequence[Mapping[str, Any]],
    window_start: _date, window_end: _date,
) -> list[dict]:
    out = [
        dict(r) for r in rows
        if _safe_str(r.get("status")).upper() == STATUS_DONE
        and _in_window(r.get("done_date"), window_start, window_end)
    ]
    out.sort(key=lambda r: (_safe_str(r.get("done_date")), _safe_str(r.get("task_id"))))
    return out


def _filter_tasks_added(
    rows: Sequence[Mapping[str, Any]],
    window_start: _date, window_end: _date,
) -> list[dict]:
    out = [
        dict(r) for r in rows
        if _in_window(r.get("created_date"), window_start, window_end)
    ]
    out.sort(key=lambda r: (_safe_str(r.get("created_date")), _safe_str(r.get("task_id"))))
    return out


def _build_drift_signals(
    sheets: Mapping[str, Sequence[Mapping[str, Any]]],
) -> dict:
    """Counts from the three discovery sheets (proxy for open drift items;
    true drift-check is governance/drift-check skill scope)."""
    return {
        "opportunity_rows":   len(sheets.get("opportunity", []) or []),
        "quick_wins_rows":    len(sheets.get("quick_wins", []) or []),
        "content_decay_rows": len(sheets.get("content_decay", []) or []),
    }


def _build_gsc_weekly_delta(
    work_events: Sequence[Mapping[str, Any]],
    drift_signals: Mapping[str, int],
) -> dict:
    """LOCAL approximation — Wave 1 does NOT call MCP. Full GSC delta is
    monthly-report scope."""
    return {
        "work_events_in_window": len(list(work_events)),
        "open_quick_wins":       int(drift_signals.get("quick_wins_rows", 0)),
        "open_opportunity":      int(drift_signals.get("opportunity_rows", 0)),
        "open_content_decay":    int(drift_signals.get("content_decay_rows", 0)),
    }


def _build_exec_summary(
    *, project_slug: str, week_start: _date, week_end: _date,
    n_done: int, n_added: int, drift_signals: Mapping[str, int],
) -> str:
    drift_total = (
        int(drift_signals.get("opportunity_rows", 0))
        + int(drift_signals.get("quick_wins_rows", 0))
        + int(drift_signals.get("content_decay_rows", 0))
    )
    return (
        f"`{project_slug}` projesi {week_start.isoformat()}–{week_end.isoformat()} "
        f"haftasında {n_done} task DONE, {n_added} yeni task açıldı. "
        f"Açık drift sinyali toplamı: {drift_total}."
    )


def aggregate(
    *,
    project_slug: str,
    sheets: Mapping[str, Sequence[Mapping[str, Any]]],
    work_events: Sequence[Mapping[str, Any]],
    week_end: _date | None = None,
) -> WeeklySummaryBatch:
    """Pure aggregation: sheet snapshot + work events → WeeklySummaryBatch.

    Args:
        project_slug: active project slug.
        sheets: dict mapping sheet_name → row-dict list. Expected keys:
            master_task, opportunity, quick_wins, content_decay. Missing
            keys treated as empty.
        work_events: events.jsonl entries (event_kind=work, in window).
        week_end: window end ISO date. Default: today snapped to Sunday.

    Raises:
        WeeklySummaryError on bad inputs.
    """
    if not isinstance(project_slug, str) or not project_slug.strip():
        raise WeeklySummaryError(
            f"project_slug must be a non-empty string, got {project_slug!r}"
        )

    week_start, week_end_resolved = compute_window(week_end)

    master_rows = list(sheets.get("master_task", []) or [])
    tasks_done = _filter_tasks_done(master_rows, week_start, week_end_resolved)
    tasks_added = _filter_tasks_added(master_rows, week_start, week_end_resolved)
    drift_signals = _build_drift_signals(sheets)
    gsc_weekly_delta = _build_gsc_weekly_delta(work_events, drift_signals)
    exec_summary = _build_exec_summary(
        project_slug=project_slug,
        week_start=week_start, week_end=week_end_resolved,
        n_done=len(tasks_done), n_added=len(tasks_added),
        drift_signals=drift_signals,
    )

    return WeeklySummaryBatch(
        project_slug=project_slug,
        week_start=week_start, week_end=week_end_resolved,
        exec_summary=exec_summary,
        gsc_weekly_delta=gsc_weekly_delta,
        tasks_done=tasks_done, tasks_added=tasks_added,
        drift_signals=drift_signals,
    )


# Snapshot
def build_snapshot(*, batch: WeeklySummaryBatch, generated_at: str) -> dict:
    """JSON payload for inbox/local/{date}-weekly-{slug}.json."""
    return {
        "version":      "1.0",
        "report_kind":  "weekly_summary",
        "generated_at": generated_at,
        **batch.as_dict(),
    }


# Report rendering (string.Template — render_template.py compatible)
def _render_tasks_table(tasks: Sequence[Mapping[str, Any]]) -> str:
    """Compact markdown table (cap 50 rows; surplus collapsed)."""
    if not tasks:
        return "_Bu hafta hiç kayıt yok._"
    lines = ["| task_id | primary_source | url | status |", "|---|---|---|---|"]
    for r in tasks[:50]:
        tid = _safe_str(r.get("task_id"))
        src = _safe_str(r.get("primary_source"))
        url = _safe_str(r.get("url")) or "_(none)_"
        st = _safe_str(r.get("status"))
        lines.append(f"| `{tid}` | {src} | {url} | {st} |")
    if len(tasks) > 50:
        lines.append(f"| … | … | … | _+{len(tasks) - 50} satır daha_ |")
    return "\n".join(lines)


def _render_drift_signals(signals: Mapping[str, int]) -> str:
    return (
        f"- opportunity satırları: **{signals.get('opportunity_rows', 0)}**\n"
        f"- quick_wins satırları: **{signals.get('quick_wins_rows', 0)}**\n"
        f"- content_decay satırları: **{signals.get('content_decay_rows', 0)}**"
    )


def _render_gsc_delta(delta: Mapping[str, Any]) -> str:
    return (
        f"- bu haftaki work events: **{delta.get('work_events_in_window', 0)}**\n"
        f"- açık quick_wins: **{delta.get('open_quick_wins', 0)}**\n"
        f"- açık opportunity: **{delta.get('open_opportunity', 0)}**\n"
        f"- açık content_decay: **{delta.get('open_content_decay', 0)}**"
    )


def build_report_markdown(
    *, batch: WeeklySummaryBatch, generated_at: str, template_path: Path,
) -> str:
    """Render templates/reports/weekly-summary.template.md with $var subs."""
    if not template_path.is_file():
        raise TemplateMissingError(
            f"weekly-summary template not found: {template_path}"
        )
    template_text = template_path.read_text(encoding="utf-8")
    data = {
        "project_slug":      batch.project_slug,
        "week_start":        batch.week_start.isoformat(),
        "week_end":          batch.week_end.isoformat(),
        "generated_at":      generated_at,
        "exec_summary":      batch.exec_summary,
        "gsc_weekly_delta":  _render_gsc_delta(batch.gsc_weekly_delta),
        "tasks_done_table":  _render_tasks_table(batch.tasks_done),
        "tasks_added_table": _render_tasks_table(batch.tasks_added),
        "tasks_done_count":  str(len(batch.tasks_done)),
        "tasks_added_count": str(len(batch.tasks_added)),
        "drift_signals":     _render_drift_signals(batch.drift_signals),
    }
    return Template(template_text).substitute(
        {k: str(v) for k, v in data.items()}
    )


# CLI
def _parse_args(argv: Iterable[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="weekly_summary.py",
        description=(
            "Phase 9 W-E2 weekly-summary aggregator. LOCAL aggregation: "
            "master.xlsx + events.jsonl → 7-day rolling summary."
        ),
    )
    p.add_argument("--project-slug", required=True,
                   help="Active project slug.")
    p.add_argument("--workspace-root", required=True,
                   help="Workspace root path (resolves projects/{slug}/).")
    p.add_argument("--week-end", default=None,
                   help="Inclusive ISO date. Default: today snapped to Sunday.")
    p.add_argument("--output-dir", default=None,
                   help="If set, writes report.md + snapshot.json here.")
    return p.parse_args(list(argv))


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    week_end = None
    if args.week_end:
        try:
            week_end = datetime.strptime(args.week_end, "%Y-%m-%d").date()
        except ValueError as exc:
            print(f"--week-end invalid: {exc}", file=sys.stderr)
            return 2

    workspace_root = Path(args.workspace_root)
    project_root = workspace_root / "projects" / args.project_slug
    workbook_path = project_root / "master.xlsx"
    if not workbook_path.is_file():
        raise WorkbookMissingError(
            f"master.xlsx not found at {workbook_path}; run init-project first"
        )

    sheets = read_master_sheets(workbook_path)
    week_start, week_end_resolved = compute_window(week_end)
    events_path = project_root / "_state" / "events.jsonl"
    work_events = read_events_jsonl(events_path, week_start, week_end_resolved)

    batch = aggregate(
        project_slug=args.project_slug,
        sheets=sheets,
        work_events=work_events,
        week_end=week_end_resolved,
    )

    generated_at = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    template_path = (
        _REPO_ROOT / "templates" / "reports" / "weekly-summary.template.md"
    )
    report_md = build_report_markdown(
        batch=batch, generated_at=generated_at, template_path=template_path,
    )
    snapshot = build_snapshot(batch=batch, generated_at=generated_at)

    if args.output_dir:
        report_dir = Path(args.output_dir)
        inbox_dir = Path(args.output_dir)
    else:
        report_dir = project_root / "outputs" / "reports"
        inbox_dir = project_root / "inbox" / "local"
    report_dir.mkdir(parents=True, exist_ok=True)
    inbox_dir.mkdir(parents=True, exist_ok=True)

    report_path = report_dir / f"{week_end_resolved.isoformat()}-weekly-summary.md"
    snapshot_path = (
        inbox_dir / f"{week_end_resolved.isoformat()}-weekly-{args.project_slug}.json"
    )
    report_path.write_text(report_md, encoding="utf-8")
    snapshot_path.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )

    print(json.dumps({
        "report_path":       str(report_path.resolve()),
        "snapshot_path":     str(snapshot_path.resolve()),
        "tasks_done_count":  len(batch.tasks_done),
        "tasks_added_count": len(batch.tasks_added),
        "week_start":        batch.week_start.isoformat(),
        "week_end":          batch.week_end.isoformat(),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))


__all__ = (
    "WINDOW_DAYS",
    "WEEKLY_SUMMARY_SECTIONS",
    "STATUS_DONE",
    "EVENT_KIND_WORK",
    "WeeklySummaryError",
    "WorkbookMissingError",
    "WorkspaceRootUnsetError",
    "TemplateMissingError",
    "WeeklySummaryBatch",
    "snap_to_last_sunday",
    "compute_window",
    "read_master_sheets",
    "read_events_jsonl",
    "aggregate",
    "build_snapshot",
    "build_report_markdown",
    "main",
)
