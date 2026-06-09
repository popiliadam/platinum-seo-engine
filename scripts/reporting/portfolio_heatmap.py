#!/usr/bin/env python3
"""portfolio_heatmap.py — Phase 9 W-E8 reporting transform.

Multi-project LOCAL aggregator. READ-ONLY across portfolio.config.json
v1.1 active_projects[]: per project pulls 4 opportunity-style sheets
(opportunity, quick_wins, content_decay, cannibalization) row counts +
master_task col B/C/F/G into a density matrix + configurable
aggregate_dimension breakdown (category | priority | primary_source).

Outputs: master.xlsx#none + outputs/reports + inbox/local. events.jsonl
NOT written (Phase 9 W2; Q-RP-01 defer). Mirrors W-E3 + W-E4 + W-E1.
Schema authority: schemas/portfolio-config.schema.json v1.1 +
schemas/master-excel.schema.json.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass, field, replace as _dc_replace
from datetime import date as _date, datetime, timezone
from pathlib import Path
from string import Template
from typing import Any, Mapping, Sequence

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from scripts.state.active_projects import ACTIVE_PROJECTS_MAX  # noqa: E402

# --- Constants (schema-aligned) -------------------------------------------
PRIMARY_SOURCE_ENUM: tuple[str, ...] = (
    "content_decay", "quickwin", "tech_fix", "schema", "pillar",
    "manual", "sxo", "cannibalization", "redirect_404", "internal_links",
    "new_content_plan",
)  #: master_task col C primary_source 11-enum (Q-IL-1 internal_links;
#: +new_content_plan v1.2 audit followup — lockstep with
#: master-excel.schema.json#master_task col C, additive per ADR-018).
PRIORITY_ENUM: tuple[str, ...] = ("CRITICAL", "HIGH", "MEDIUM", "LOW")
SHEET_READ_SPECS: tuple[tuple[str, str, int], ...] = (
    ("opportunity", "B", 5), ("quick_wins", "J", 5),
    ("content_decay", "E", 6), ("cannibalization", "F", 5),
)
SHEET_NAMES: tuple[str, ...] = tuple(s[0] for s in SHEET_READ_SPECS)
MASTER_TASK_DATA_START_ROW = 4
AGGREGATE_DIMENSION_ENUM: frozenset[str] = frozenset(
    {"category", "priority", "primary_source"})
DEFAULT_AGGREGATE_DIMENSION = "category"
DEFAULT_TEMPLATE_REL = "templates/reports/portfolio-heatmap.template.md"
WRITER_NAME = "portfolio_heatmap"
_DENSITY_BARS: tuple[str, ...] = (" ", "▁", "▂", "▃", "▄", "▅", "▆", "▇")

# ---Exceptions (DURUR conditions) ----------------------------------------
class PortfolioHeatmapError(Exception): pass
class PortfolioConfigMissingError(PortfolioHeatmapError): pass
class PortfolioConfigInvalidError(PortfolioHeatmapError): pass
class ActiveProjectsCeilingError(PortfolioHeatmapError): pass
class ReadOnlyContractViolation(PortfolioHeatmapError): pass
class WorkspaceRootUnsetError(PortfolioHeatmapError): pass
class AggregateDimensionInvalidError(PortfolioHeatmapError): pass

# ---Dataclasses (frozen, read-only result containers) -------------------
@dataclass(frozen=True)
class ProjectHeatmapRow:
    slug: str
    display_name: str
    priority: int
    workspace_path: str
    master_xlsx_path: str | None
    sheet_counts: dict
    sheet_densities: dict
    master_task_total: int
    dimension_buckets: dict
    warnings: list

@dataclass(frozen=True)
class PortfolioHeatmap:
    portfolio_id: str
    display_name: str
    run_date: str
    generated_at: str
    aggregate_dimension: str
    projects: tuple = field(default_factory=tuple)
    sheet_max_counts: dict = field(default_factory=dict)
    dimension_keys: tuple = field(default_factory=tuple)
    skipped_projects: int = 0

# ---Helpers --------------------------------------------------------------
def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()

def _ensure_aggregate_dimension(value: str) -> str:
    """Validate aggregate_dimension input; raise DURUR on drift."""
    if value not in AGGREGATE_DIMENSION_ENUM:
        raise AggregateDimensionInvalidError(
            f"aggregate_dimension {value!r} not in "
            f"{sorted(AGGREGATE_DIMENSION_ENUM)}")
    return value

def _col_letter_to_index(letter: str) -> int:
    """Excel column letter → 1-based index."""
    s = letter.strip().upper()
    n = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"invalid column letter {letter!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def _now_utc_iso() -> str:
    return datetime.now(tz=timezone.utc).replace(
        microsecond=0).isoformat().replace("+00:00", "Z")

# ---portfolio.config.json read + validate --------------------------------

def load_portfolio_config(portfolio_root: Path) -> dict:
    """Read projects/_portfolio/portfolio.config.json. DURUR on error."""
    cfg_path = (
        portfolio_root / "projects" / "_portfolio" / "portfolio.config.json"
    )
    if not cfg_path.is_file():
        raise PortfolioConfigMissingError(
            f"portfolio.config.json not found at {cfg_path}.")
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PortfolioConfigMissingError(
            f"portfolio.config.json read failed at {cfg_path}: {exc}",
        ) from exc

def validate_portfolio_config(
    config: Mapping[str, Any], schema: Mapping[str, Any],
) -> None:
    """Validate against portfolio-config.schema.json (Draft 7) + surface
    ActiveProjectsCeilingError + ReadOnlyContractViolation."""
    try:
        from scripts.validation.validate_schema import build_validator
    except ImportError as exc:  # pragma: no cover
        raise PortfolioConfigInvalidError("jsonschema required") from exc
    errors = sorted(build_validator(dict(schema)).iter_errors(dict(config)),
                    key=lambda e: list(e.absolute_path))
    if errors:
        raise PortfolioConfigInvalidError(
            "portfolio.config.json failed Draft 7 validation: "
            + "; ".join(f"{list(e.absolute_path)}: {e.message}"
                        for e in errors))
    active = list(config.get("active_projects") or [])
    if len(active) > ACTIVE_PROJECTS_MAX:
        raise ActiveProjectsCeilingError(
            f"active_projects has {len(active)} entries; ceiling is "
            f"{ACTIVE_PROJECTS_MAX} (schema maxItems).")
    cq = config.get("cross_query") or {}
    if "read_only" in cq and cq.get("read_only") is not True:
        raise ReadOnlyContractViolation(
            f"cross_query.read_only must be true (schema const), got "
            f"{cq.get('read_only')!r}.")

def resolve_project_workspace(
    *, portfolio_root: Path, workspace_path: str,
) -> Path:
    """Resolve workspace_path. Absolute → as-is; relative → portfolio_root."""
    p = Path(workspace_path).expanduser()
    return p if p.is_absolute() else (portfolio_root / p).resolve()

def resolve_master_xlsx_path(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    """Compute {workspace_root}/projects/{slug}/master.xlsx."""
    slug = _safe_str(project_entry.get("slug"))
    return resolve_project_workspace(
        portfolio_root=portfolio_root,
        workspace_path=_safe_str(project_entry.get("workspace_path")),
    ) / "projects" / slug / "master.xlsx"

# ---openpyxl thin wrapper (READ-ONLY, INLINE) ---------------------------

def _read_workbook_sheets(
    workbook_path: Path,
) -> tuple[dict[str, list[tuple]] | None, list[str]]:
    """READ-ONLY: open workbook, extract row tuples for the 4 opportunity
    sheets + master_task. Missing sheet → [] + warning (W-E1 paterni)."""
    if not workbook_path.is_file():
        return None, [f"workbook missing: {workbook_path}"]
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return None, ["openpyxl not installed"]
    try:
        wb = load_workbook(filename=str(workbook_path),
                           read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return None, [f"workbook open failed ({exc})"]
    warnings: list[str] = []
    out: dict[str, list[tuple]] = {}
    try:
        for sheet_name, col_letter, data_start_row in SHEET_READ_SPECS:
            if sheet_name not in wb.sheetnames:
                warnings.append(f"{sheet_name}: sheet missing (graceful skip)")
                out[sheet_name] = []
                continue
            col_idx = _col_letter_to_index(col_letter)
            out[sheet_name] = list(wb[sheet_name].iter_rows(
                min_row=data_start_row, min_col=col_idx,
                max_col=col_idx, values_only=True))
        if "master_task" not in wb.sheetnames:
            warnings.append("master_task: sheet missing (graceful skip)")
            out["master_task"] = []
        else:
            out["master_task"] = list(wb["master_task"].iter_rows(
                min_row=MASTER_TASK_DATA_START_ROW,
                min_col=_col_letter_to_index("B"),
                max_col=_col_letter_to_index("G"), values_only=True))
        return out, warnings
    finally:
        try:
            wb.close()
        except Exception:
            pass

def _count_non_empty(rows: list[tuple]) -> int:
    """Count non-empty single-cell rows."""
    n = 0
    for row in rows:
        cell = row[0] if row else None
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            continue
        n += 1
    return n

def _master_task_to_dicts(rows: list[tuple]) -> list[dict]:
    """master_task row tuples (B..G; 0=B,1=C,4=F,5=G) → row dicts."""
    out: list[dict] = []
    for raw in rows:
        if raw is None or all(c in (None, "") for c in raw):
            continue
        task = _safe_str(raw[0]) if len(raw) > 0 else ""
        if not task or task.lower() == "task":
            continue
        out.append({
            "task": task,
            "primary_source": _safe_str(raw[1]) if len(raw) > 1 else "",
            "category": _safe_str(raw[4]) if len(raw) > 4 else "",
            "priority": _safe_str(raw[5]).upper() if len(raw) > 5 else "",
        })
    return out

# ---Bucketing + per-project row builder ---------------------------------

def _bucket_master_task(
    rows: Sequence[Mapping[str, Any]], *, aggregate_dimension: str,
) -> dict[str, int]:
    """Bucket master_task rows. primary_source → 11-enum (every value
    gets a bucket); priority → severityEnum 4-bucket; category →
    free-text cardinality."""
    if aggregate_dimension == "primary_source":
        out, key = {k: 0 for k in PRIMARY_SOURCE_ENUM}, "primary_source"
    elif aggregate_dimension == "priority":
        out, key = {k: 0 for k in PRIORITY_ENUM}, "priority"
    else:
        out = {}
        for r in rows:
            v = _safe_str(r.get("category"))
            if v:
                out[v] = out.get(v, 0) + 1
        return out
    for r in rows:
        v = _safe_str(r.get(key))
        if key == "priority":
            v = v.upper()
        if v in out:
            out[v] += 1
    return out

def _build_project_row(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
    aggregate_dimension: str,
) -> ProjectHeatmapRow:
    """Build one heatmap row (raw counts; densities filled portfolio-
    side once max_count per sheet known). Missing workbook → schema-
    valid empty shape (W-E1 paterni)."""
    slug = _safe_str(project_entry.get("slug"))
    display_name = _safe_str(project_entry.get("display_name") or slug)
    priority = int(project_entry.get("priority") or 10)
    workspace_path = _safe_str(project_entry.get("workspace_path"))
    master_xlsx = resolve_master_xlsx_path(
        portfolio_root=portfolio_root, project_entry=project_entry)
    sheets, warnings = _read_workbook_sheets(master_xlsx)
    if sheets is None:
        sheet_counts = {n: 0 for n in SHEET_NAMES}
        mt_rows: list[dict] = []
        mt_path: str | None = None
    else:
        mt_path = str(master_xlsx)
        sheet_counts = {n: _count_non_empty(sheets.get(n, []))
                        for n in SHEET_NAMES}
        mt_rows = _master_task_to_dicts(sheets.get("master_task", []))
    return ProjectHeatmapRow(
        slug=slug, display_name=display_name, priority=priority,
        workspace_path=workspace_path, master_xlsx_path=mt_path,
        sheet_counts=sheet_counts,
        sheet_densities={n: 0.0 for n in SHEET_NAMES},
        master_task_total=len(mt_rows),
        dimension_buckets=_bucket_master_task(
            mt_rows, aggregate_dimension=aggregate_dimension),
        warnings=warnings,
    )

# ---Top-level aggregator ------------------------------------------------

def aggregate(
    *, portfolio_root: Path, config: Mapping[str, Any], run_date: str,
    aggregate_dimension: str = DEFAULT_AGGREGATE_DIMENSION,
    generated_at: str | None = None,
) -> PortfolioHeatmap:
    """Pure aggregation: portfolio.config + per-project workbooks →
    PortfolioHeatmap. Idempotent given pinned generated_at."""
    aggregate_dimension = _ensure_aggregate_dimension(aggregate_dimension)
    portfolio_id = _safe_str(config.get("portfolio_id"))
    display_name = _safe_str(config.get("display_name") or portfolio_id)
    rows: list[ProjectHeatmapRow] = [
        _build_project_row(
            portfolio_root=portfolio_root, project_entry=entry,
            aggregate_dimension=aggregate_dimension)
        for entry in (config.get("active_projects") or [])
        if isinstance(entry, Mapping)
    ]
    rows.sort(key=lambda r: (r.priority, r.slug))
    skipped = sum(1 for r in rows if r.master_xlsx_path is None)
    sheet_max = {n: max((r.sheet_counts.get(n, 0) for r in rows),
                        default=0) for n in SHEET_NAMES}
    rows_normalized = [
        _dc_replace(r, sheet_densities={
            n: (round(r.sheet_counts[n] / sheet_max[n], 4)
                if sheet_max[n] > 0 else 0.0) for n in SHEET_NAMES})
        for r in rows
    ]
    if aggregate_dimension == "primary_source":
        dim_keys = PRIMARY_SOURCE_ENUM
    elif aggregate_dimension == "priority":
        dim_keys = PRIORITY_ENUM
    else:
        seen: dict[str, None] = {}
        for r in rows_normalized:
            for k in r.dimension_buckets.keys():
                seen.setdefault(k, None)
        dim_keys = tuple(sorted(seen.keys()))
    return PortfolioHeatmap(
        portfolio_id=portfolio_id, display_name=display_name,
        run_date=run_date, generated_at=generated_at or _now_utc_iso(),
        aggregate_dimension=aggregate_dimension,
        projects=tuple(rows_normalized), sheet_max_counts=sheet_max,
        dimension_keys=dim_keys, skipped_projects=skipped,
    )

# ---Snapshot / report builders ------------------------------------------

def build_snapshot(*, heatmap: PortfolioHeatmap) -> dict:
    """Build inbox/local/{date}-portfolio-heatmap.json payload."""
    return {
        "writer": WRITER_NAME, "version": "1.0",
        "run_date": heatmap.run_date, "generated_at": heatmap.generated_at,
        "portfolio_id": heatmap.portfolio_id,
        "display_name": heatmap.display_name,
        "aggregate_dimension": heatmap.aggregate_dimension,
        "active_project_count": len(heatmap.projects),
        "skipped_project_count": heatmap.skipped_projects,
        "sheet_max_counts": dict(heatmap.sheet_max_counts),
        "dimension_keys": list(heatmap.dimension_keys),
        "projects": [{
            "slug": r.slug, "display_name": r.display_name,
            "priority": r.priority, "workspace_path": r.workspace_path,
            "master_xlsx_path": r.master_xlsx_path,
            "sheet_counts": dict(r.sheet_counts),
            "sheet_densities": dict(r.sheet_densities),
            "master_task_total": r.master_task_total,
            "dimension_buckets": dict(r.dimension_buckets),
            "warnings": list(r.warnings),
        } for r in heatmap.projects],
    }

def _density_bar(d: float) -> str:
    if d <= 0:
        return _DENSITY_BARS[0]
    idx = min(int(d * (len(_DENSITY_BARS) - 1)) + 1,
              len(_DENSITY_BARS) - 1)
    return _DENSITY_BARS[idx]

def _build_density_table(heatmap: PortfolioHeatmap) -> str:
    """Render project × sheet density matrix as a markdown table."""
    if not heatmap.projects:
        return "_(no active projects)_\n"
    lines = [
        "| slug | priority | " + " | ".join(SHEET_NAMES) + " |",
        "|---|---:|" + "|".join(["---:"] * len(SHEET_NAMES)) + "|",
    ]
    for r in heatmap.projects:
        cells = [
            f"{r.sheet_counts.get(n, 0)} {_density_bar(r.sheet_densities.get(n, 0.0))}"
            for n in SHEET_NAMES
        ]
        lines.append(
            f"| `{r.slug}` | {r.priority} | " + " | ".join(cells) + " |"
        )
    return "\n".join(lines) + "\n"

def _build_dimension_table(heatmap: PortfolioHeatmap) -> str:
    """Render project × dimension_bucket count matrix."""
    if not heatmap.projects or not heatmap.dimension_keys:
        return "_(no buckets — empty data)_\n"
    lines = [
        "| slug | " + " | ".join(heatmap.dimension_keys) + " | total |",
        "|---|" + "|".join(["---:"] * (len(heatmap.dimension_keys) + 1))
        + "|",
    ]
    for r in heatmap.projects:
        cells = [str(r.dimension_buckets.get(k, 0))
                 for k in heatmap.dimension_keys]
        lines.append(
            f"| `{r.slug}` | " + " | ".join(cells)
            + f" | {r.master_task_total} |"
        )
    return "\n".join(lines) + "\n"

def _build_totals_summary(heatmap: PortfolioHeatmap) -> str:
    """Single-line totals rollup."""
    sheet_totals = {
        n: sum(r.sheet_counts.get(n, 0) for r in heatmap.projects)
        for n in SHEET_NAMES
    }
    parts = [f"{n}={sheet_totals[n]}" for n in SHEET_NAMES]
    parts.extend([
        f"master_task_total={sum(r.master_task_total for r in heatmap.projects)}",
        f"projects={len(heatmap.projects)}",
        f"skipped={heatmap.skipped_projects}",
    ])
    return "Toplam: " + ", ".join(parts) + "."

def build_report_markdown(
    *, heatmap: PortfolioHeatmap, template_text: str,
) -> str:
    """Render markdown report via string.Template (render_template.py
    convention reused inline). Idempotent / pure."""
    data = {
        "generated_at": heatmap.generated_at, "run_date": heatmap.run_date,
        "portfolio_id": heatmap.portfolio_id,
        "display_name": heatmap.display_name,
        "aggregate_dimension": heatmap.aggregate_dimension,
        "projects_count": str(len(heatmap.projects)),
        "skipped_count": str(heatmap.skipped_projects),
        "scope": "portfolio-heatmap (read-only multi-project density matrix)",
        "density_table": _build_density_table(heatmap),
        "dimension_table": _build_dimension_table(heatmap),
        "totals_summary": _build_totals_summary(heatmap),
    }
    try:
        return Template(template_text).safe_substitute(
            {k: str(v) for k, v in data.items()}
        )
    except KeyError as exc:
        raise PortfolioHeatmapError(
            f"template substitution failed - missing key: {exc.args[0]}"
        ) from exc
    except ValueError as exc:
        raise PortfolioHeatmapError(
            f"template syntax error: {exc}"
        ) from exc

# ---Read-only contract self-check ---------------------------------------

#: Built at runtime so literals don't appear verbatim. Tokens spell
#: openpyxl workbook write + transaction layer + events_writer write
#: shapes — NOT bare `.append(` (legitimate Python list mutation).
_FORBIDDEN_WRITE_PATTERNS: tuple[str, ...] = (
    "wb" + "." + "save(",
    "workbook" + "." + "save(",
    "transaction" + "." + "append(",
    "transaction" + "." + "update(",
    "transaction" + "." + "delete(",
    "events_writer" + "." + "append_",
    "from " + "scripts.excel.transaction",
    "from " + "scripts.state.events_writer",
)

def assert_read_only_module() -> None:
    """Verify this module's source has no write call sites. Per schema
    cross_query.read_only=true (const) — strict read-only. W-E3 helper
    paterni."""
    src = Path(__file__).read_text(encoding="utf-8")
    marker = "_FORBIDDEN_WRITE_PATTERNS"
    if marker in src:
        before, _sep, _rest = src.partition(marker)
        scan_src = before
    else:
        scan_src = src
    hits = [p for p in _FORBIDDEN_WRITE_PATTERNS if p in scan_src]
    if hits:
        raise ReadOnlyContractViolation(
            f"portfolio_heatmap source contains forbidden write patterns: "
            f"{hits}."
        )

# ---Workspace root resolver ---------------------------------------------

def _resolve_portfolio_root(arg: str | None) -> Path:
    """Resolve portfolio_root: arg → PSEO_WORKSPACE_ROOT env (W-E4)."""
    if arg:
        p = Path(arg).expanduser().resolve()
        if not p.is_dir():
            raise WorkspaceRootUnsetError(
                f"--portfolio-root path does not exist: {p}")
        return p
    env = os.environ.get("PSEO_WORKSPACE_ROOT")
    if env:
        p = Path(env).expanduser().resolve()
        if p.is_dir():
            return p
    raise WorkspaceRootUnsetError(
        "portfolio_root unresolved: pass --portfolio-root or set "
        "PSEO_WORKSPACE_ROOT.")

# ---CLI ------------------------------------------------------------------

def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(prog="portfolio_heatmap.py", description=(
        "Aggregate portfolio.config + per-project master.xlsx 4 sheet "
        "density + master_task aggregate_dimension breakdown. READ-ONLY."))
    p.add_argument("--portfolio-root", default=None)
    p.add_argument("--run-date", default=None)
    p.add_argument("--aggregate-dimension",
                   default=DEFAULT_AGGREGATE_DIMENSION,
                   choices=sorted(AGGREGATE_DIMENSION_ENUM))
    p.add_argument("--template", default=None)
    p.add_argument("--output-dir", default=None)
    p.add_argument("--generated-at", default=None)
    args = p.parse_args(argv)
    try:
        portfolio_root = _resolve_portfolio_root(args.portfolio_root)
        config = load_portfolio_config(portfolio_root)
    except (WorkspaceRootUnsetError, PortfolioConfigMissingError) as exc:
        print(f"setup failed: {exc}", file=sys.stderr)
        return 2
    schema_path = _REPO_ROOT / "schemas" / "portfolio-config.schema.json"
    if schema_path.is_file():
        try:
            validate_portfolio_config(
                config,
                json.loads(schema_path.read_text(encoding="utf-8")))
        except (PortfolioConfigInvalidError, ActiveProjectsCeilingError,
                ReadOnlyContractViolation) as exc:
            print(f"config validation failed: {exc}", file=sys.stderr)
            return 1
    run_date = args.run_date or _date.today().isoformat()
    try:
        heatmap = aggregate(
            portfolio_root=portfolio_root, config=config,
            run_date=run_date,
            aggregate_dimension=args.aggregate_dimension,
            generated_at=args.generated_at)
    except AggregateDimensionInvalidError as exc:
        print(f"aggregate failed: {exc}", file=sys.stderr)
        return 1
    tpl_path = (Path(args.template).expanduser() if args.template
                else _REPO_ROOT / DEFAULT_TEMPLATE_REL)
    if not tpl_path.is_file():
        print(f"template not found: {tpl_path}", file=sys.stderr)
        return 2
    snapshot = build_snapshot(heatmap=heatmap)
    report = build_report_markdown(
        heatmap=heatmap,
        template_text=tpl_path.read_text(encoding="utf-8"))
    if args.output_dir:
        out_dir = Path(args.output_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
        snap_path = out_dir / f"{run_date}-portfolio-heatmap.json"
        snap_path.write_text(json.dumps(
            snapshot, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8")
        report_path = out_dir / f"{run_date}-portfolio-heatmap.md"
        report_path.write_text(report, encoding="utf-8")
        print(json.dumps({
            "snapshot_path": str(snap_path.resolve()),
            "report_path": str(report_path.resolve()),
            "active_project_count": len(heatmap.projects),
            "skipped_project_count": heatmap.skipped_projects,
            "aggregate_dimension": heatmap.aggregate_dimension,
        }, ensure_ascii=False, indent=2))
    else:
        print(json.dumps({"snapshot": snapshot, "report_md": report},
                         ensure_ascii=False, indent=2))
    return 0

if __name__ == "__main__":  # pragma: no cover
    sys.exit(main(sys.argv[1:]))

# __all__ intentionally omitted — all public names module-scope already.
