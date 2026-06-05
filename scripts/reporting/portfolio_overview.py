#!/usr/bin/env python3
"""portfolio_overview.py — multi-project LOCAL aggregator (Phase 9 W1).

Reads portfolio.config.json + each active project's master.xlsx#dashboard
KPI cells + master_task status counts → snapshot + markdown report.

READ-ONLY contract (schemas/portfolio-config.schema.json#cross_query
.read_only = true const): no openpyxl write call sites; load_workbook
always uses read_only=True. Pure: no I/O when imported as a module.
Idempotent: same input → byte-identical output. Plugin-agnostik: 0
hardcoded slug values; iteration is driven by active_projects entirely.
Does NOT write events.jsonl (Phase 9 W1 convention; Q-RP-01 deferred).

Refs: schemas/portfolio-config.schema.json (v1.1, maxItems=12,
cross_query.read_only=true), schemas/master-excel.schema.json (dashboard
required_cells + master_task statusEnum), scripts/reporting/render_template.py.
Pattern reference: scripts/planning/master_task_sync.py (Phase 8 W-D1).
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import date as _date
from pathlib import Path
from string import Template
from typing import Any, Mapping


_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


# --- Constants (schema-aligned) ---------------------------------------------

#: Dashboard KPI cells (master-excel.schema.json#dashboard required_cells).
DASHBOARD_KPI_CELLS: tuple[tuple[str, str], ...] = (
    ("R10", "total_pillars"),
    ("R47", "master_task_total"),
    ("R48", "master_task_done"),
    ("R49", "master_task_todo"),
    ("R50", "master_task_ongoing"),
    ("R51", "cannibalization_count"),
    ("R52", "quick_wins_count"),
    ("R59", "last_audit_date"),
)

#: master_task statusEnum (master-excel.schema.json#/definitions/statusEnum).
STATUS_ENUM: tuple[str, ...] = (
    "TODO", "ONGOING", "EXISTS", "DONE",
    "BLOCKED", "DEFERRED", "CANCELED",
)

MASTER_TASK_DATA_START_ROW = 4
MASTER_TASK_STATUS_COL_INDEX = 10  # J = 10 (1-based)
ACTIVE_PROJECTS_MAX = 12
DEFAULT_TEMPLATE_REL = "templates/reports/portfolio-overview.template.md"

_SUMMABLE_KPIS: tuple[str, ...] = (
    "total_pillars",
    "master_task_total",
    "master_task_done",
    "master_task_todo",
    "master_task_ongoing",
    "cannibalization_count",
    "quick_wins_count",
)


# --- Exceptions (DURUR conditions; STOP + surface to manager) --------------

class PortfolioOverviewError(Exception): pass
class PortfolioConfigMissingError(PortfolioOverviewError): pass
class PortfolioConfigInvalidError(PortfolioOverviewError): pass
class ActiveProjectsCeilingError(PortfolioOverviewError): pass
class ReadOnlyContractViolation(PortfolioOverviewError): pass
class WorkspaceRootUnsetError(PortfolioOverviewError): pass


# --- Data classes (frozen, read-only result containers) --------------------

@dataclass(frozen=True)
class ProjectSnapshot:
    slug: str
    display_name: str
    priority: int
    workspace_path: str
    master_xlsx_path: str | None
    kpis: dict
    status_counts: dict
    status_check_drift: bool
    warnings: list


@dataclass(frozen=True)
class PortfolioOverviewBatch:
    portfolio_id: str
    display_name: str
    snapshots: list = field(default_factory=list)
    totals: dict = field(default_factory=dict)
    skipped_projects: int = 0
    run_date: str = ""


# --- Helpers ----------------------------------------------------------------

def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _safe_int(v: Any) -> int:
    """Coerce a workbook cell value to int. Blank / non-numeric → 0."""
    if v is None or isinstance(v, bool):
        return int(v) if isinstance(v, bool) else 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return 0


def load_portfolio_config(portfolio_root: Path) -> dict:
    """Read projects/_portfolio/portfolio.config.json."""
    cfg_path = (
        portfolio_root / "projects" / "_portfolio" / "portfolio.config.json"
    )
    if not cfg_path.is_file():
        raise PortfolioConfigMissingError(
            f"portfolio.config.json not found at {cfg_path}."
        )
    try:
        return json.loads(cfg_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise PortfolioConfigMissingError(
            f"portfolio.config.json read failed at {cfg_path}: {exc}"
        ) from exc


def validate_portfolio_config(
    config: Mapping[str, Any], schema: Mapping[str, Any],
) -> None:
    """Validate config against portfolio-config.schema.json (Draft 7).

    Also surfaces ActiveProjectsCeilingError + ReadOnlyContractViolation
    domain exceptions.
    """
    try:
        from jsonschema import Draft7Validator
    except ImportError as exc:  # pragma: no cover
        raise PortfolioConfigInvalidError("jsonschema required") from exc
    validator = Draft7Validator(dict(schema))
    errors = sorted(validator.iter_errors(dict(config)),
                    key=lambda e: list(e.absolute_path))
    if errors:
        raise PortfolioConfigInvalidError(
            "portfolio.config.json failed Draft 7 validation: "
            + "; ".join(f"{list(e.absolute_path)}: {e.message}" for e in errors)
        )
    active = list(config.get("active_projects") or [])
    if len(active) > ACTIVE_PROJECTS_MAX:
        raise ActiveProjectsCeilingError(
            f"active_projects has {len(active)} entries; ceiling is "
            f"{ACTIVE_PROJECTS_MAX} (schema maxItems)."
        )
    cq = config.get("cross_query") or {}
    if "read_only" in cq and cq.get("read_only") is not True:
        raise ReadOnlyContractViolation(
            f"cross_query.read_only must be true (schema const), got "
            f"{cq.get('read_only')!r}."
        )


def resolve_project_workspace(
    *, portfolio_root: Path, workspace_path: str,
) -> Path:
    """Resolve workspace_path entry. Absolute → as-is; ~ → expanduser;
    relative → relative to portfolio_root (W-E4 convention authority,
    5/6 majority + W-E8 explicit tercih, Phase 9 closeout resolution).
    """
    p = Path(workspace_path).expanduser()
    return p if p.is_absolute() else (portfolio_root / p).resolve()


def resolve_master_xlsx_path(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> Path:
    """Compute {workspace_path}/projects/{slug}/master.xlsx."""
    slug = _safe_str(project_entry.get("slug"))
    workspace_path = _safe_str(project_entry.get("workspace_path"))
    project_root = resolve_project_workspace(
        portfolio_root=portfolio_root, workspace_path=workspace_path,
    )
    return project_root / "projects" / slug / "master.xlsx"


# --- openpyxl thin wrapper (READ-ONLY) --------------------------------------

def _read_dashboard_kpis(workbook_path: Path) -> dict[str, Any]:
    """Read 8 dashboard KPI cells. READ-ONLY (load_workbook read_only=True)."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:  # pragma: no cover
        return {}
    out: dict[str, Any] = {}
    wb = load_workbook(filename=str(workbook_path),
                       read_only=True, data_only=True)
    try:
        if "dashboard" not in wb.sheetnames:
            return {kpi: "" for _cell, kpi in DASHBOARD_KPI_CELLS}
        ws = wb["dashboard"]
        for cell, kpi in DASHBOARD_KPI_CELLS:
            value = ws[cell].value
            if value is None:
                out[kpi] = ""
            elif isinstance(value, _date):
                out[kpi] = value.isoformat()
            else:
                out[kpi] = str(value).strip()
    finally:
        wb.close()
    return out


def _read_master_task_status_counts(workbook_path: Path) -> dict[str, int]:
    """Compute master_task status counts via column J scan. READ-ONLY."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:  # pragma: no cover
        return {s: 0 for s in STATUS_ENUM}
    counts: dict[str, int] = {s: 0 for s in STATUS_ENUM}
    counts["OTHER"] = 0
    wb = load_workbook(filename=str(workbook_path),
                       read_only=True, data_only=True)
    try:
        if "master_task" not in wb.sheetnames:
            return counts
        ws = wb["master_task"]
        for row in ws.iter_rows(
            min_row=MASTER_TASK_DATA_START_ROW,
            min_col=MASTER_TASK_STATUS_COL_INDEX,
            max_col=MASTER_TASK_STATUS_COL_INDEX,
            values_only=True,
        ):
            sval = _safe_str(row[0])
            if not sval:
                continue
            counts[sval if sval in counts else "OTHER"] += 1
    finally:
        wb.close()
    return counts


# --- Per-project snapshot ---------------------------------------------------

def _build_project_snapshot(
    *, portfolio_root: Path, project_entry: Mapping[str, Any],
) -> ProjectSnapshot:
    """Build one snapshot. READ-ONLY. Missing master.xlsx → warning."""
    slug = _safe_str(project_entry.get("slug"))
    display_name = _safe_str(project_entry.get("display_name") or slug)
    priority = int(project_entry.get("priority") or 10)
    workspace_path = _safe_str(project_entry.get("workspace_path"))
    warnings: list = []

    master_xlsx = resolve_master_xlsx_path(
        portfolio_root=portfolio_root, project_entry=project_entry,
    )
    if not master_xlsx.is_file():
        warnings.append(f"master.xlsx not found at {master_xlsx}")
        return ProjectSnapshot(
            slug=slug, display_name=display_name, priority=priority,
            workspace_path=workspace_path, master_xlsx_path=None,
            kpis={kpi: "" for _cell, kpi in DASHBOARD_KPI_CELLS},
            status_counts={s: 0 for s in STATUS_ENUM},
            status_check_drift=False, warnings=warnings,
        )

    kpis = _read_dashboard_kpis(master_xlsx)
    status_counts = _read_master_task_status_counts(master_xlsx)

    drift = False
    pairs = (
        ("master_task_total",
         sum(status_counts.get(s, 0) for s in STATUS_ENUM)),
        ("master_task_done",    status_counts.get("DONE", 0)),
        ("master_task_todo",    status_counts.get("TODO", 0)),
        ("master_task_ongoing", status_counts.get("ONGOING", 0)),
    )
    for kpi_name, local_value in pairs:
        cell_value = kpis.get(kpi_name, "")
        if cell_value == "":
            warnings.append(f"dashboard {kpi_name} blank")
            continue
        if _safe_int(cell_value) != local_value:
            drift = True
            warnings.append(
                f"status_check_drift: dashboard {kpi_name}={cell_value} "
                f"vs local={local_value}"
            )
    if status_counts.get("OTHER", 0):
        warnings.append(
            f"master_task has {status_counts['OTHER']} rows with status "
            "outside statusEnum"
        )

    return ProjectSnapshot(
        slug=slug, display_name=display_name, priority=priority,
        workspace_path=workspace_path,
        master_xlsx_path=str(master_xlsx),
        kpis=kpis, status_counts=status_counts,
        status_check_drift=drift, warnings=warnings,
    )


# --- Top-level aggregator ---------------------------------------------------

def aggregate(
    *, portfolio_root: Path, config: Mapping[str, Any],
    run_date: str, filter_active: bool = True,
) -> PortfolioOverviewBatch:
    """Pure aggregation: portfolio.config + per-project workbooks → batch.
    pending_onboard NEVER read (schema). filter_active forward-compat only.
    """
    portfolio_id = _safe_str(config.get("portfolio_id"))
    display_name = _safe_str(config.get("display_name") or portfolio_id)

    active = list(config.get("active_projects") or [])
    _ = filter_active  # forward-compat placeholder

    snapshots: list[ProjectSnapshot] = []
    skipped = 0
    for entry in active:
        snap = _build_project_snapshot(
            portfolio_root=portfolio_root, project_entry=entry,
        )
        snapshots.append(snap)
        if snap.master_xlsx_path is None:
            skipped += 1
    snapshots.sort(key=lambda s: (s.priority, s.slug))

    totals: dict[str, int] = {k: 0 for k in _SUMMABLE_KPIS}
    for snap in snapshots:
        for k in _SUMMABLE_KPIS:
            totals[k] += _safe_int(snap.kpis.get(k))
    totals["active_project_count"] = len(snapshots)
    totals["skipped_project_count"] = skipped

    return PortfolioOverviewBatch(
        portfolio_id=portfolio_id, display_name=display_name,
        snapshots=snapshots, totals=totals,
        skipped_projects=skipped, run_date=run_date,
    )


# --- Snapshot / report builders --------------------------------------------

def build_snapshot(*, batch: PortfolioOverviewBatch) -> dict:
    """Build inbox/local/{date}-portfolio-overview.json payload."""
    return {
        "version": "1.0",
        "run_date": batch.run_date,
        "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "active_project_count": len(batch.snapshots),
        "skipped_project_count": batch.skipped_projects,
        "totals": dict(batch.totals),
        "snapshots": [
            {
                "slug": s.slug,
                "display_name": s.display_name,
                "priority": s.priority,
                "workspace_path": s.workspace_path,
                "master_xlsx_path": s.master_xlsx_path,
                "kpis": dict(s.kpis),
                "status_counts": dict(s.status_counts),
                "status_check_drift": s.status_check_drift,
                "warnings": list(s.warnings),
            }
            for s in batch.snapshots
        ],
    }


def _build_projects_table(batch: PortfolioOverviewBatch) -> str:
    """Render snapshots as a markdown table."""
    if not batch.snapshots:
        return "_(no active projects)_\n"
    lines = [
        "| slug | name | priority | tasks_total | done | todo | ongoing | "
        "quick_wins | last_audit | drift |",
        "|---|---|---|---|---|---|---|---|---|---|",
    ]
    for s in batch.snapshots:
        drift = "YES" if s.status_check_drift else "no"
        lines.append(
            f"| `{s.slug}` | {s.display_name} | {s.priority} | "
            f"{_safe_int(s.kpis.get('master_task_total'))} | "
            f"{_safe_int(s.kpis.get('master_task_done'))} | "
            f"{_safe_int(s.kpis.get('master_task_todo'))} | "
            f"{_safe_int(s.kpis.get('master_task_ongoing'))} | "
            f"{_safe_int(s.kpis.get('quick_wins_count'))} | "
            f"{_safe_str(s.kpis.get('last_audit_date')) or '-'} | "
            f"{drift} |"
        )
    return "\n".join(lines) + "\n"


_TOTALS_LABELS: tuple[tuple[str, str], ...] = (
    ("active projects",               "active_project_count"),
    ("skipped (master.xlsx missing)", "skipped_project_count"),
    ("total pillars",                 "total_pillars"),
    ("master_task_total (sum)",       "master_task_total"),
    ("master_task_done (sum)",        "master_task_done"),
    ("master_task_todo (sum)",        "master_task_todo"),
    ("master_task_ongoing (sum)",     "master_task_ongoing"),
    ("cannibalization_count (sum)",   "cannibalization_count"),
    ("quick_wins_count (sum)",        "quick_wins_count"),
)


def _build_totals_summary(batch: PortfolioOverviewBatch) -> str:
    """Render aggregate totals as markdown bullets."""
    t = batch.totals
    return "\n".join(
        f"- {label}: **{t.get(key, 0)}**" for label, key in _TOTALS_LABELS
    ) + "\n"


def build_report_markdown(
    *, batch: PortfolioOverviewBatch, template_text: str,
) -> str:
    """Render markdown report via strict string.Template.substitute()
    (the render_template.py convention — a missing key raises, never a
    silently-leaked $placeholder). `$run_id` is the K-06 audit-trail
    placeholder a render-time hook injects (events.jsonl.next_run_id), so
    it defaults to the literal `$run_id` and survives the render intact —
    matching monthly_report.py. Template owns the layout; this only
    injects $variables.
    """
    data = {
        "generated_at": batch.run_date,
        "portfolio_id": batch.portfolio_id,
        "display_name": batch.display_name,
        "project_count": str(len(batch.snapshots)),
        "skipped_count": str(batch.skipped_projects),
        "scope": "portfolio-overview (read-only multi-project aggregation)",
        "projects_table": _build_projects_table(batch),
        "totals_summary": _build_totals_summary(batch),
    }
    # K-06 audit-trail placeholder: keep $run_id for the render-time hook.
    data.setdefault("run_id", "$run_id")
    try:
        return Template(template_text).substitute(
            {k: str(v) for k, v in data.items()}
        )
    except KeyError as exc:
        raise PortfolioOverviewError(
            f"template substitution failed - missing key: {exc.args[0]}"
        ) from exc
    except ValueError as exc:
        raise PortfolioOverviewError(
            f"template syntax error: {exc}"
        ) from exc


# --- Read-only contract self-check ------------------------------------------

#: Patterns built at runtime so the literals do not appear verbatim in
#: source (the self-check below greps the file). The tokens spell openpyxl
#: + transaction write call shapes.
_FORBIDDEN_WRITE_PATTERNS: tuple[str, ...] = (
    "." + "save(",
    "transaction" + "." + "append(",
    "transaction" + "." + "update(",
    "transaction" + "." + "delete(",
)


def assert_read_only_module() -> None:
    """Verify this module's source has no write call sites. Per schema
    cross_query.read_only=true (const) — strict read-only. Used by tests.
    """
    src = Path(__file__).read_text(encoding="utf-8")
    hits = [p for p in _FORBIDDEN_WRITE_PATTERNS if p in src]
    if hits:
        raise ReadOnlyContractViolation(
            f"portfolio_overview source contains forbidden write patterns: "
            f"{hits}."
        )


# --- CLI --------------------------------------------------------------------

def _resolve_portfolio_root(arg: str | None) -> Path:
    """Resolve portfolio_root: arg → env → .pse-workspace marker."""
    if arg:
        p = Path(arg).expanduser().resolve()
        if not p.is_dir():
            raise WorkspaceRootUnsetError(
                f"--portfolio-root path does not exist: {p}"
            )
        return p
    env = os.environ.get("PSEO_PORTFOLIO_ROOT") or os.environ.get(
        "PSEO_WORKSPACE_ROOT"
    )
    if env:
        p = Path(env).expanduser().resolve()
        if p.is_dir():
            return p
    cwd = Path.cwd().resolve()
    for parent in (cwd, *cwd.parents):
        if (parent / ".pse-workspace").exists():
            return parent
    raise WorkspaceRootUnsetError(
        "portfolio_root unresolved: pass --portfolio-root, set "
        "PSEO_PORTFOLIO_ROOT or PSEO_WORKSPACE_ROOT, or place a "
        ".pse-workspace marker in the workspace root."
    )


def _parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(prog="portfolio_overview.py", description=(
        "Aggregate portfolio.config + per-project master.xlsx KPIs. READ-ONLY."
    ))
    p.add_argument("--portfolio-root", default=None)
    p.add_argument("--run-date", default=None, help="ISO YYYY-MM-DD.")
    p.add_argument("--filter-active", action="store_true", default=True)
    p.add_argument("--template", default=None)
    p.add_argument("--output-dir", default=None)
    return p.parse_args(argv)


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    try:
        portfolio_root = _resolve_portfolio_root(args.portfolio_root)
        config = load_portfolio_config(portfolio_root)
    except (WorkspaceRootUnsetError, PortfolioConfigMissingError) as exc:
        print(f"setup failed: {exc}", file=sys.stderr)
        return 2

    schema_path = _REPO_ROOT / "schemas" / "portfolio-config.schema.json"
    if schema_path.is_file():
        try:
            validate_portfolio_config(
                config, json.loads(schema_path.read_text(encoding="utf-8")),
            )
        except (PortfolioConfigInvalidError, ActiveProjectsCeilingError,
                ReadOnlyContractViolation) as exc:
            print(f"config validation failed: {exc}", file=sys.stderr)
            return 1

    run_date = args.run_date or _date.today().isoformat()
    batch = aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date=run_date, filter_active=args.filter_active,
    )
    tpl_path = (Path(args.template).expanduser() if args.template
                else _REPO_ROOT / DEFAULT_TEMPLATE_REL)
    if not tpl_path.is_file():
        print(f"template not found: {tpl_path}", file=sys.stderr)
        return 2
    snapshot = build_snapshot(batch=batch)
    report = build_report_markdown(
        batch=batch, template_text=tpl_path.read_text(encoding="utf-8"),
    )

    if args.output_dir:
        out_dir = Path(args.output_dir).expanduser()
        out_dir.mkdir(parents=True, exist_ok=True)
        snap_path = out_dir / f"{run_date}-portfolio-overview.json"
        snap_path.write_text(json.dumps(
            snapshot, ensure_ascii=False, indent=2, sort_keys=True,
        ), encoding="utf-8")
        report_path = out_dir / f"{run_date}-portfolio-overview.md"
        report_path.write_text(report, encoding="utf-8")
        print(json.dumps({
            "snapshot_path": str(snap_path.resolve()),
            "report_path": str(report_path.resolve()),
            "active_project_count": len(batch.snapshots),
            "skipped_project_count": batch.skipped_projects,
        }, ensure_ascii=False, indent=2))
    else:
        print(json.dumps({"snapshot": snapshot, "report_md": report},
                         ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))


__all__ = (
    "DASHBOARD_KPI_CELLS", "STATUS_ENUM", "ACTIVE_PROJECTS_MAX",
    "PortfolioOverviewError", "PortfolioConfigMissingError",
    "PortfolioConfigInvalidError", "ActiveProjectsCeilingError",
    "ReadOnlyContractViolation", "WorkspaceRootUnsetError",
    "ProjectSnapshot", "PortfolioOverviewBatch",
    "load_portfolio_config", "validate_portfolio_config",
    "resolve_project_workspace", "resolve_master_xlsx_path",
    "aggregate", "build_snapshot", "build_report_markdown",
    "assert_read_only_module", "main",
)
