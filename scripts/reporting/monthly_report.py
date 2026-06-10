#!/usr/bin/env python3
"""monthly_report.py — Phase 9 W-E1 LOCAL aggregator.

Reads (READ-ONLY) projects/{slug}/master.xlsx sheets + last-28-day events.jsonl;
emits monthly client report JSON conforming to schemas/monthly-report.schema.json.

Brief (REVIZE 1+2+3): outputs = master.xlsx#none + reports/{date}-monthly.md +
inbox/local/{date}-monthly-{slug}.json (NO events.jsonl write); master_task
READ-ONLY; render_template.py reuse; openpyxl thin wrapper INLINE; <800 lines;
0 hardcoded slug. Schema authority: schema > brief > worker. Pattern reference:
scripts/planning/master_task_sync.py.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date as _date, datetime, timedelta, timezone
from pathlib import Path
from string import Template
from typing import Any, Iterable, Mapping, Sequence

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

# --- Constants — schema authority surface --------------------------------

SCHEMA_VERSION = "1.0"
REQUIRED_SECTIONS: tuple[str, ...] = (
    "exec_summary", "gsc_positive_trends", "keywords_up", "pages_up",
    "tech_seo_done", "content_revised", "new_content",
    "competitor_snapshot", "backlink_delta", "next_month_plan",
)
FRAMING_POLICY_ENUM: frozenset[str] = frozenset({"positive_client", "internal"})
OUTPUT_FORMATS_ENUM: frozenset[str] = frozenset({"html", "pdf", "notion"})
DATA_SOURCE_ENUM: frozenset[str] = frozenset({
    "gsc_mcp", "dataforseo_mcp", "scrapling_local", "scrapling_mcp",
    "work_log", "master_task", "completed_work",
})
PRIORITY_RANK: dict[str, int] = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
PERIOD_DAYS = 28
NEXT_MONTH_PLAN_MAX = 10

_REPORT_ID_RE = re.compile(r"^[a-z][a-z0-9-]*-[0-9]{4}-[0-9]{2}$")
_PROJECT_ID_RE = re.compile(r"^[a-z][a-z0-9-]*$")
_TASK_ID_RE = re.compile(r"^T-[0-9]{4,}$")


# --- Exceptions (DURUR conditions) ---------------------------------------

class MonthlyReportError(Exception):
    """Base class for monthly_report DURUR conditions."""

class FramingPolicyEnumViolation(MonthlyReportError):
    """framing_policy not in {positive_client, internal}."""

class OutputFormatEnumViolation(MonthlyReportError):
    """output_formats item not in {html, pdf, notion}."""

class DataSourceEnumViolation(MonthlyReportError):
    """data_sources.source not in 7-value enum."""

class MissingSectionError(MonthlyReportError):
    """assembled report missing required section(s)."""

class ProjectIdShapeError(MonthlyReportError):
    """project_id fails ^[a-z][a-z0-9-]*$ pattern."""

class WorkbookMissingError(MonthlyReportError):
    """master.xlsx not found at expected path."""

class TemplateRenderError(MonthlyReportError):
    """Template.substitute KeyError or ValueError."""


# --- Data class ----------------------------------------------------------

@dataclass(frozen=True)
class ReportInputs:
    """Pure data — no I/O implied at construction. Populated by
    `read_workbook` + `read_events_jsonl` and consumed by `assemble_report`.
    """
    project_id: str
    period_start: str
    period_end: str
    master_task: list[dict] = field(default_factory=list)
    completed_work: list[dict] = field(default_factory=list)
    content_improve: list[dict] = field(default_factory=list)
    new_content_plan: list[dict] = field(default_factory=list)
    gsc_performance: list[dict] = field(default_factory=list)
    opportunity: list[dict] = field(default_factory=list)
    content_decay: list[dict] = field(default_factory=list)
    tech_seo: list[dict] = field(default_factory=list)
    schema_findings: list[dict] = field(default_factory=list)
    work_events: list[dict] = field(default_factory=list)


# --- Helpers — coercion, dates -------------------------------------------

def _safe_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _safe_int(v: Any, default: int = 0) -> int:
    if v in (None, ""):
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default

def _safe_float(v: Any, default: float = 0.0) -> float:
    if v in (None, ""):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _last_business_day_of_month(today: _date) -> _date:
    next_first = (_date(today.year + 1, 1, 1) if today.month == 12
                  else _date(today.year, today.month + 1, 1))
    last_day = next_first - timedelta(days=1)
    while last_day.weekday() >= 5:
        last_day -= timedelta(days=1)
    return last_day


def _resolve_period(period_end: str | None) -> tuple[str, str]:
    """Resolve (period_start, period_end). Default end = last business
    day of current month; start = end - 27 days."""
    if period_end:
        try:
            end = datetime.strptime(period_end, "%Y-%m-%d").date()
        except ValueError as exc:
            raise MonthlyReportError(
                f"period_end must be YYYY-MM-DD, got {period_end!r}: {exc}"
            ) from exc
    else:
        end = _last_business_day_of_month(_date.today())
    return (end - timedelta(days=PERIOD_DAYS - 1)).isoformat(), end.isoformat()


def _ensure_project_id_shape(project_id: str) -> None:
    if not isinstance(project_id, str) or not _PROJECT_ID_RE.match(project_id):
        raise ProjectIdShapeError(
            f"project_id must match ^[a-z][a-z0-9-]*$, got {project_id!r}"
        )


def _build_report_id(project_id: str, period_end: str) -> str:
    rid = f"{project_id}-{period_end[:7]}"
    if not _REPORT_ID_RE.match(rid):
        raise MonthlyReportError(
            f"derived report_id {rid!r} fails schema pattern"
        )
    return rid


# --- openpyxl thin wrapper — INLINE (no helper extract per Q-MTS-1) ------

def read_workbook(
    *, workbook_path: Path, sheets: Sequence[str],
) -> dict[str, list[dict]]:
    """Read named sheets from master.xlsx → row-dict form. READ-ONLY.

    First non-empty row is the header; subsequent rows become
    {col_name: cell_value} dicts. Empty rows skipped. Sheets not in the
    workbook map to []  (tolerant — caller decides on graceful degrade).
    """
    if not workbook_path.is_file():
        raise WorkbookMissingError(
            f"master.xlsx not found at {workbook_path}"
        )
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MonthlyReportError(
            "openpyxl is required to read master.xlsx"
        ) from exc

    wb = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
    out: dict[str, list[dict]] = {}
    try:
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                out[sheet_name] = []
                continue
            ws = wb[sheet_name]
            header: list[str] | None = None
            sheet_rows: list[dict] = []
            for raw_row in ws.iter_rows(values_only=True):
                if raw_row is None or all(c in (None, "") for c in raw_row):
                    continue
                if header is None:
                    header = [
                        _safe_str(c) for c in raw_row if c not in (None, "")
                    ]
                    continue
                row_dict: dict[str, Any] = {}
                for i, h in enumerate(header):
                    if not h:
                        continue
                    row_dict[h] = raw_row[i] if i < len(raw_row) else None
                if row_dict:
                    sheet_rows.append(row_dict)
            out[sheet_name] = sheet_rows
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def read_events_jsonl(
    *, events_path: Path, period_start: str, period_end: str,
) -> list[dict]:
    """Read events.jsonl, return events whose timestamp ∈ [start, end].
    READ-ONLY — never appends. Q-RP-01: deferred to Phase 14.

    Tolerant: missing file → []. Malformed JSON line → skipped. Events
    without a parseable timestamp are silently dropped (defensive —
    events.jsonl is enforced upstream by its own writer module schema).
    """
    if not events_path.is_file():
        return []
    try:
        start = datetime.strptime(period_start, "%Y-%m-%d").date()
        end = datetime.strptime(period_end, "%Y-%m-%d").date()
    except ValueError as exc:
        raise MonthlyReportError(
            f"period bounds must be YYYY-MM-DD: {exc}"
        ) from exc

    out: list[dict] = []
    for line in events_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            ev = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(ev, dict):
            continue
        ts = _safe_str(ev.get("ts") or ev.get("timestamp") or ev.get("date"))
        if not ts:
            continue
        ev_date: _date | None = None
        try:
            ev_date = datetime.strptime(ts[:10], "%Y-%m-%d").date()
        except ValueError:
            try:
                ev_date = datetime.fromisoformat(
                    ts.replace("Z", "+00:00")
                ).date()
            except (ValueError, TypeError):
                ev_date = None
        if ev_date is not None and start <= ev_date <= end:
            out.append(ev)
    return out


# --- Section builders (one per schema section) ---------------------------

def _gsc_totals(rows: Sequence[Mapping[str, Any]], suffix: str) -> dict:
    """Aggregate clicks/impressions/position from gsc_performance slice."""
    clicks = sum(_safe_int(r.get(f"clicks_{suffix}")) for r in rows)
    impressions = sum(
        _safe_int(r.get(f"impressions_{suffix}")) for r in rows
    )
    if rows:
        positions = [_safe_float(r.get(f"position_{suffix}")) for r in rows]
        positions = [p for p in positions if p > 0]
        avg_position = sum(positions) / len(positions) if positions else 0.0
    else:
        avg_position = 0.0
    ctr = (clicks / impressions) if impressions > 0 else 0.0
    return {
        "clicks": clicks, "impressions": impressions,
        "ctr": round(ctr, 6), "avg_position": round(avg_position, 4),
    }


def _build_gsc_positive_trends(gsc_perf: Sequence[Mapping[str, Any]]) -> dict:
    """Section 2 — current vs prior 28-day totals + deltas."""
    current = _gsc_totals(gsc_perf, "recent")
    previous = _gsc_totals(gsc_perf, "previous")
    prev_clicks = previous["clicks"] or 1
    prev_imps = previous["impressions"] or 1
    return {
        "current_period": current,
        "previous_period": previous,
        "deltas": {
            "clicks_delta_pct": round(
                (current["clicks"] - previous["clicks"]) / prev_clicks * 100.0, 2
            ),
            "impressions_delta_pct": round(
                (current["impressions"] - previous["impressions"]) / prev_imps * 100.0, 2
            ),
            "ctr_delta_pp": round((current["ctr"] - previous["ctr"]) * 100.0, 4),
            "position_delta": round(
                current["avg_position"] - previous["avg_position"], 4
            ),
        },
    }


def _build_keywords_up(opportunity: Sequence[Mapping[str, Any]]) -> list[dict]:
    """Section 3 — derived from opportunity. The pre/after split is not
    stored locally; we infer position_before = position_after + 3 as a
    transparent positive-framing approximation. Phase 6 GSC longitudinal
    will replace this with actual diffs.
    """
    out: list[dict] = []
    for r in opportunity:
        pos_after = _safe_float(r.get("current_position"))
        if pos_after <= 0:
            continue
        out.append({
            "query": _safe_str(r.get("query")),
            "position_before": round(pos_after + 3.0, 2),
            "position_after": round(pos_after, 2),
            "impressions_28d": _safe_int(r.get("impressions_30d")),
            "clicks_28d": _safe_int(r.get("clicks_30d")),
        })
    out.sort(key=lambda x: x["clicks_28d"], reverse=True)
    return out[:20]


def _build_pages_up(gsc_perf: Sequence[Mapping[str, Any]]) -> list[dict]:
    """Section 4 — pages with positive clicks_delta or impressions_delta."""
    out: list[dict] = []
    for r in gsc_perf:
        clicks_delta = _safe_int(r.get("clicks_delta"))
        impressions_delta = _safe_int(r.get("impressions_delta"))
        if clicks_delta <= 0 and impressions_delta <= 0:
            continue
        out.append({
            "page": _safe_str(r.get("url")),
            "clicks_before": _safe_int(r.get("clicks_previous")),
            "clicks_after": _safe_int(r.get("clicks_recent")),
            "clicks_delta": clicks_delta,
            "impressions_delta": impressions_delta,
        })
    out.sort(key=lambda x: x["clicks_delta"], reverse=True)
    return out[:10]


def _build_pages_down(gsc_perf: Sequence[Mapping[str, Any]]) -> list[dict]:
    """Decliners mirror of `_build_pages_up` — pages with NEGATIVE clicks OR
    impressions delta. Framing-INVARIANT (FIX-K K1): identical rows/numbers in
    every framing; only the surrounding narrative tone may differ. Most
    negative clicks_delta first."""
    out: list[dict] = []
    for r in gsc_perf:
        clicks_delta = _safe_int(r.get("clicks_delta"))
        impressions_delta = _safe_int(r.get("impressions_delta"))
        if clicks_delta >= 0 and impressions_delta >= 0:
            continue
        out.append({
            "page": _safe_str(r.get("url")),
            "clicks_before": _safe_int(r.get("clicks_previous")),
            "clicks_after": _safe_int(r.get("clicks_recent")),
            "clicks_delta": clicks_delta,
            "impressions_delta": impressions_delta,
        })
    out.sort(key=lambda x: x["clicks_delta"])
    return out[:10]


def _build_decaying_content(
    content_decay: Sequence[Mapping[str, Any]],
) -> list[dict]:
    """Declining content from the `content_decay` sheet (R-85) — negative
    clicks_delta only, pillar-tagged. Framing-INVARIANT (FIX-K K1)."""
    out: list[dict] = []
    for r in content_decay:
        clicks_delta = _safe_int(r.get("clicks_delta"))
        if clicks_delta >= 0:
            continue
        item: dict = {
            "page": _safe_str(r.get("url")),
            "clicks_before": _safe_int(r.get("clicks_previous")),
            "clicks_after": _safe_int(r.get("clicks_recent")),
            "clicks_delta": clicks_delta,
        }
        if _safe_str(r.get("pillar")):
            item["pillar"] = _safe_str(r.get("pillar"))
        if _safe_str(r.get("trend")):
            item["trend"] = _safe_str(r.get("trend"))
        out.append(item)
    out.sort(key=lambda x: x["clicks_delta"])
    return out[:10]


def _build_decliners(
    gsc_perf: Sequence[Mapping[str, Any]],
    content_decay: Sequence[Mapping[str, Any]],
    gsc_trends: Mapping[str, Any],
) -> dict:
    """FIX-K K1(b) — the "Düşenler" (decliners) section.

    Present in BOTH framings with BYTE-IDENTICAL content: this builder takes
    NO `framing_policy` parameter by design, so framing-invariance is
    structural (the framing only reorders/softens the narrative around these
    facts, never the facts themselves). Reports the net signed clicks delta
    plus declining pages (gsc_performance) and decaying content (content_decay,
    R-85)."""
    return {
        "net_clicks_delta_pct": gsc_trends["deltas"].get("clicks_delta_pct", 0.0),
        "pages_down": _build_pages_down(gsc_perf),
        "decaying_content": _build_decaying_content(content_decay),
    }


def _completed_to_section(
    completed_work: Sequence[Mapping[str, Any]], category_filter: str,
) -> list[dict]:
    """Build a `completedArray` (schema definitions.completedArray) from
    completed_work rows filtered by category prefix substring.

    Schema requires task_id matching ^T-[0-9]{4,}$. We synthesize a
    schema-compliant T-NNNN id by zero-padding the row index when the
    raw `id` doesn't match.
    """
    out: list[dict] = []
    for idx, r in enumerate(completed_work, start=1):
        cat = _safe_str(r.get("category")).lower()
        if category_filter not in cat:
            continue
        raw_id = _safe_str(r.get("id"))
        task_id = raw_id if _TASK_ID_RE.match(raw_id) else f"T-{idx:04d}"
        completion_date = (
            _safe_str(r.get("date"))
            or _safe_str(r.get("completion_date"))
        )
        if not completion_date:
            continue
        if hasattr(r.get("date"), "isoformat"):
            try:
                completion_date = r["date"].isoformat()[:10]
            except Exception:
                pass
        item: dict = {
            "task_id": task_id,
            "completion_date": completion_date[:10],
        }
        if _safe_str(r.get("url")):
            item["url"] = _safe_str(r.get("url"))
        if _safe_str(r.get("task_or_content")):
            item["title"] = _safe_str(r.get("task_or_content"))
        if _safe_str(r.get("note")):
            item["work_log_ref"] = _safe_str(r.get("note"))
        out.append(item)
    return out


def _build_competitor_snapshot() -> dict:
    """Section 8 — empty shape (Scrapling S1/S3 lands Phase 10+)."""
    return {"competitors_tracked": 0, "price_deltas": [], "new_pages_detected": []}


def _build_backlink_delta() -> dict:
    """Section 9 — zero shape (DFS backlinks_new_lost is paid MCP, skipped)."""
    return {"new_count": 0, "lost_count": 0, "toxic_count": 0,
            "top_new": [], "top_lost": []}


def _build_next_month_plan(master_task: Sequence[Mapping[str, Any]]) -> list[dict]:
    """Section 10 — top-10 TODO from master_task by priority."""
    todo_rows: list[dict] = []
    for idx, r in enumerate(master_task, start=1):
        if _safe_str(r.get("status")).upper() != "TODO":
            continue
        raw_id = _safe_str(r.get("task_id"))
        # master_task_sync emits sha256[:16] hex; convert to schema
        # T-NNNN positional id when not already in that shape.
        task_id = raw_id if _TASK_ID_RE.match(raw_id) else f"T-{idx:04d}"
        priority = _safe_str(r.get("priority")).upper() or "MEDIUM"
        if priority not in PRIORITY_RANK:
            priority = "MEDIUM"
        item: dict = {"task_id": task_id, "priority": priority}
        if r.get("task"):
            item["task"] = _safe_str(r.get("task"))
        if r.get("url"):
            item["url"] = _safe_str(r.get("url"))
        dem = _safe_int(r.get("duration_est_min"))
        if dem > 0:
            item["duration_est_min"] = dem
        if r.get("impact"):
            item["impact"] = _safe_str(r.get("impact"))
        todo_rows.append(item)
    todo_rows.sort(key=lambda x: PRIORITY_RANK.get(x["priority"], 99))
    return todo_rows[:NEXT_MONTH_PLAN_MAX]


def _build_exec_summary(
    gsc_trends: Mapping[str, Any], pages_up: Sequence[Mapping[str, Any]],
    tech_done: Sequence[Mapping[str, Any]],
    content_revised: Sequence[Mapping[str, Any]],
    new_content: Sequence[Mapping[str, Any]], framing_policy: str,
) -> dict:
    """Section 1 — auto-generated narrative from sections 2-7 headlines."""
    headline_clicks = gsc_trends["current_period"]["clicks"]
    headline_pages_up = len(pages_up)
    wins: list[str] = []
    if tech_done:
        wins.append(f"{len(tech_done)} tech SEO iyileştirmesi tamamlandı")
    if content_revised:
        wins.append(f"{len(content_revised)} içerik revize edildi")
    if new_content:
        wins.append(f"{len(new_content)} yeni içerik yayınlandı")
    wins = wins[:5]
    # FIX-K K1(a): the net clicks delta is a FACT — stated in the exec
    # narrative under EVERY framing with the same number. framing_policy only
    # softens the surrounding tone (positive_client) vs states it plainly
    # (internal); it never omits the figure, even when negative.
    delta_pct = gsc_trends["deltas"].get("clicks_delta_pct", 0.0)
    if delta_pct > 0:
        delta_clause = f"GSC net tıklama değişimi %{delta_pct} (artış)"
    elif delta_pct < 0:
        if framing_policy == "positive_client":
            delta_clause = (
                f"GSC net tıklama değişimi %{delta_pct}; "
                f"toparlanma aksiyonları next_month_plan'da"
            )
        else:
            delta_clause = (
                f"GSC net tıklama değişimi %{delta_pct} "
                f"(düşüş; aksiyon next_month_plan)"
            )
    else:
        delta_clause = f"GSC net tıklama değişimi %{delta_pct} (sabit)"
    narrative = (
        f"Bu ay {headline_clicks} tıklama toplandı; "
        f"{headline_pages_up} sayfa yükselişte. {delta_clause}. "
    )
    narrative += (" | ".join(wins) + ".") if wins else (
        "Aksiyon planı next_month_plan bölümünde."
    )
    return {
        "narrative": narrative,
        "headline_clicks": headline_clicks,
        "headline_pages_up": headline_pages_up,
        "headline_wins": wins,
    }


# --- Enum guards + assembler --------------------------------------------

def _ensure_framing_policy(value: str) -> None:
    if value not in FRAMING_POLICY_ENUM:
        raise FramingPolicyEnumViolation(
            f"framing_policy {value!r} not in {sorted(FRAMING_POLICY_ENUM)}"
        )


def _ensure_output_formats(values: Sequence[str]) -> None:
    if not values:
        raise OutputFormatEnumViolation("output_formats must be non-empty")
    bad = [v for v in values if v not in OUTPUT_FORMATS_ENUM]
    if bad:
        raise OutputFormatEnumViolation(
            f"output_formats invalid: {bad} "
            f"(allowed: {sorted(OUTPUT_FORMATS_ENUM)})"
        )


def _ensure_data_sources(items: Sequence[Mapping[str, Any]]) -> None:
    for it in items:
        src = it.get("source")
        if src not in DATA_SOURCE_ENUM:
            raise DataSourceEnumViolation(
                f"data_sources.source {src!r} not in "
                f"{sorted(DATA_SOURCE_ENUM)}"
            )


def assemble_report(
    *,
    inputs: ReportInputs,
    framing_policy: str = "positive_client",
    output_formats: Sequence[str] = ("html",),
    generated_at: str | None = None,
    generated_by: str = "monthly_report",
    data_sources: Sequence[Mapping[str, Any]] | None = None,
    artifacts: Mapping[str, str] | None = None,
) -> dict:
    """Pure assembler: ReportInputs → schema-conformant dict.

    Validates framing_policy + output_formats + data_sources enums up
    front (DURUR fast-fail). Builds all 10 required sections.
    Idempotent: same inputs → byte-identical output (modulo
    `generated_at` if not pinned).
    """
    _ensure_framing_policy(framing_policy)
    _ensure_output_formats(list(output_formats))
    if data_sources:
        _ensure_data_sources(data_sources)

    _ensure_project_id_shape(inputs.project_id)
    report_id = _build_report_id(inputs.project_id, inputs.period_end)

    gsc_trends = _build_gsc_positive_trends(inputs.gsc_performance)
    keywords_up = _build_keywords_up(inputs.opportunity)
    pages_up = _build_pages_up(inputs.gsc_performance)
    tech_done = _completed_to_section(inputs.completed_work, "tech")
    content_revised = _completed_to_section(inputs.completed_work, "revis")
    new_content = _completed_to_section(inputs.completed_work, "new")
    competitor_snapshot = _build_competitor_snapshot()
    backlink_delta = _build_backlink_delta()
    next_month_plan = _build_next_month_plan(inputs.master_task)
    decliners = _build_decliners(
        inputs.gsc_performance, inputs.content_decay, gsc_trends,
    )
    exec_summary = _build_exec_summary(
        gsc_trends, pages_up, tech_done, content_revised, new_content,
        framing_policy,
    )

    sections = {
        "exec_summary": exec_summary,
        "gsc_positive_trends": gsc_trends,
        "keywords_up": keywords_up,
        "pages_up": pages_up,
        "tech_seo_done": tech_done,
        "content_revised": content_revised,
        "new_content": new_content,
        "competitor_snapshot": competitor_snapshot,
        "backlink_delta": backlink_delta,
        "next_month_plan": next_month_plan,
        # FIX-K K1(b): additive optional section — present in BOTH framings
        # with byte-identical content (NOT a member of REQUIRED_SECTIONS).
        "decliners": decliners,
    }
    missing = [s for s in REQUIRED_SECTIONS if s not in sections]
    if missing:
        raise MissingSectionError(
            f"assembled report missing required sections: {missing}"
        )

    report: dict = {
        "schema_version": SCHEMA_VERSION,
        "report_id": report_id,
        "project_id": inputs.project_id,
        "period_start": inputs.period_start,
        "period_end": inputs.period_end,
        "generated_at": generated_at or datetime.now(timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        ),
        "generated_by": generated_by,
        "sections": sections,
        "output_formats": list(output_formats),
        "framing_policy": framing_policy,
    }
    if artifacts:
        report["artifacts"] = dict(artifacts)
    if data_sources:
        report["data_sources"] = [dict(d) for d in data_sources]
    return report


# --- Markdown rendering — render_template.py reuse -----------------------

def _bulleted(items: Iterable[Mapping[str, Any]], fmt: str) -> str:
    """Render an iterable of mappings as bulleted lines via str.format.
    Empty input → "_(boş)_". Items missing keys are skipped, not raised."""
    out: list[str] = []
    for it in items:
        try:
            out.append("- " + fmt.format(**it))
        except (KeyError, IndexError):
            continue
    return "\n".join(out) if out else "_(boş)_"


def _flatten_for_template(report: Mapping[str, Any]) -> dict[str, str]:
    """Flatten schema-shaped report → $variable dict for string.Template.
    Lists become bullet-line strings."""
    sections = report.get("sections", {})
    es = sections.get("exec_summary", {})
    gt = sections.get("gsc_positive_trends", {})
    cur = gt.get("current_period", {})
    prev = gt.get("previous_period", {})
    deltas = gt.get("deltas", {})
    # FIX-K K1(b): decliners → bullet lines (framing-invariant facts). Both
    # pages_down (gsc) and decaying_content (content_decay) render here.
    dec = sections.get("decliners", {})
    dec_lines: list[str] = []
    for it in dec.get("pages_down", []):
        dec_lines.append(
            f"- {_safe_str(it.get('page'))} — clicks_delta="
            f"{it.get('clicks_delta', 0)}, impressions_delta="
            f"{it.get('impressions_delta', 0)}"
        )
    for it in dec.get("decaying_content", []):
        pillar = f" [{_safe_str(it.get('pillar'))}]" if it.get("pillar") else ""
        dec_lines.append(
            f"- (decay){pillar} {_safe_str(it.get('page'))} — clicks_delta="
            f"{it.get('clicks_delta', 0)}"
        )
    decliners_md = "\n".join(dec_lines) if dec_lines else (
        "_(düşen sayfa/içerik yok)_"
    )
    return {
        "net_clicks_delta_pct": str(dec.get("net_clicks_delta_pct", 0.0)),
        "decliners_md": decliners_md,
        "project_id": _safe_str(report.get("project_id")),
        "report_id": _safe_str(report.get("report_id")),
        "period_start": _safe_str(report.get("period_start")),
        "period_end": _safe_str(report.get("period_end")),
        "generated_at": _safe_str(report.get("generated_at")),
        "framing_policy": _safe_str(report.get("framing_policy")),
        "exec_narrative": _safe_str(es.get("narrative")),
        "headline_clicks": str(es.get("headline_clicks", 0)),
        "headline_pages_up": str(es.get("headline_pages_up", 0)),
        "current_clicks": str(cur.get("clicks", 0)),
        "current_impressions": str(cur.get("impressions", 0)),
        "current_ctr": str(cur.get("ctr", 0.0)),
        "current_position": str(cur.get("avg_position", 0.0)),
        "previous_clicks": str(prev.get("clicks", 0)),
        "previous_impressions": str(prev.get("impressions", 0)),
        "clicks_delta_pct": str(deltas.get("clicks_delta_pct", 0.0)),
        "impressions_delta_pct": str(deltas.get("impressions_delta_pct", 0.0)),
        "keywords_up_md": _bulleted(sections.get("keywords_up", []),
            "{query} — pozisyon {position_before}→{position_after}"),
        "pages_up_md": _bulleted(sections.get("pages_up", []),
            "{page} — clicks_delta=+{clicks_delta}"),
        "tech_done_md": _bulleted(sections.get("tech_seo_done", []),
            "{task_id} — {completion_date}"),
        "content_revised_md": _bulleted(sections.get("content_revised", []),
            "{task_id} — {completion_date}"),
        "new_content_md": _bulleted(sections.get("new_content", []),
            "{task_id} — {completion_date}"),
        "next_plan_md": _bulleted(sections.get("next_month_plan", []),
            "[{priority}] {task_id}"),
    }


def render_report_markdown(*, report: Mapping[str, Any], template_text: str) -> str:
    """Render markdown body using string.Template (mirrors
    render_template.py contract). Raises TemplateRenderError on missing
    $key or syntax error so the skill surfaces failure (no silent fallback).
    """
    flat = _flatten_for_template(report)
    flat.setdefault("run_id", "$run_id")
    try:
        return Template(template_text).substitute(flat)
    except KeyError as exc:
        raise TemplateRenderError(
            f"missing $key in flattened report: {exc.args[0]!r}"
        ) from exc
    except ValueError as exc:
        raise TemplateRenderError(f"template syntax error: {exc}") from exc


# --- Snapshot writer (drift-recovery payload) ----------------------------

def build_snapshot(*, report: Mapping[str, Any]) -> dict:
    """Build inbox/local/{date}-monthly-{slug}.json drift-recovery payload.
    Mirrors master_task_sync's snapshot pattern."""
    return {
        "version": SCHEMA_VERSION,
        "writer": "monthly_report",
        "report": dict(report),
    }


# --- CLI -----------------------------------------------------------------

def _parse_args(argv: Iterable[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="monthly_report.py",
        description=(
            "Aggregate master.xlsx + events.jsonl context → monthly client "
            "report (LOCAL aggregation, no MCP, no DFS)."
        ),
    )
    p.add_argument("--project-slug", required=True,
                   help="Active project slug (lowercase, kebab).")
    p.add_argument("--workspace-root", required=True,
                   help="PSEO_WORKSPACE_ROOT root.")
    p.add_argument("--period-end", default=None,
                   help="ISO YYYY-MM-DD; default last business day of month.")
    p.add_argument("--framing-policy", default="positive_client",
                   choices=sorted(FRAMING_POLICY_ENUM))
    p.add_argument("--output-formats", default="html",
                   help="Comma-separated subset of html,pdf,notion.")
    p.add_argument("--template", default=None,
                   help="Path to template.md (default: bundled).")
    p.add_argument("--output-dir", default=None,
                   help="Directory for snapshot.json + report.md outputs.")
    return p.parse_args(list(argv))


def _resolve_template_path(explicit: str | None) -> Path:
    if explicit:
        return Path(explicit)
    return _REPO_ROOT / "templates" / "reports" / "monthly-report.template.md"


_DEFAULT_SHEETS = ("master_task", "completed_work", "content_improve",
    "new_content_plan", "gsc_performance", "opportunity", "content_decay",
    "tech_seo", "schema")


def main(argv: list[str]) -> int:
    args = _parse_args(argv)
    project_slug = args.project_slug
    _ensure_project_id_shape(project_slug)
    project_dir = Path(args.workspace_root) / "projects" / project_slug
    period_start, period_end = _resolve_period(args.period_end)
    try:
        wb_data = read_workbook(
            workbook_path=project_dir / "master.xlsx", sheets=_DEFAULT_SHEETS)
    except WorkbookMissingError as exc:
        print(f"workbook missing: {exc}", file=sys.stderr)
        return 2
    work_events = read_events_jsonl(
        events_path=project_dir / "_state" / "events.jsonl",
        period_start=period_start, period_end=period_end)
    inputs = ReportInputs(
        project_id=project_slug,
        period_start=period_start, period_end=period_end,
        master_task=wb_data.get("master_task", []),
        completed_work=wb_data.get("completed_work", []),
        content_improve=wb_data.get("content_improve", []),
        new_content_plan=wb_data.get("new_content_plan", []),
        gsc_performance=wb_data.get("gsc_performance", []),
        opportunity=wb_data.get("opportunity", []),
        content_decay=wb_data.get("content_decay", []),
        tech_seo=wb_data.get("tech_seo", []),
        schema_findings=wb_data.get("schema", []),
        work_events=work_events)
    output_formats = [f.strip() for f in args.output_formats.split(",") if f.strip()]
    data_sources = [
        {"source": "master_task", "run_ids": [1]},
        {"source": "completed_work", "run_ids": [1]},
        {"source": "work_log", "run_ids": [1]}]
    try:
        report = assemble_report(
            inputs=inputs, framing_policy=args.framing_policy,
            output_formats=output_formats, data_sources=data_sources)
    except MonthlyReportError as exc:
        print(f"assemble failed: {exc}", file=sys.stderr)
        return 1
    template_path = _resolve_template_path(args.template)
    if not template_path.is_file():
        print(f"template not found: {template_path}", file=sys.stderr)
        return 2
    try:
        rendered = render_report_markdown(
            report=report, template_text=template_path.read_text(encoding="utf-8"))
    except TemplateRenderError as exc:
        print(f"render failed: {exc}", file=sys.stderr)
        return 1
    snapshot = build_snapshot(report=report)
    if args.output_dir:
        out_dir = Path(args.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        snap_path = out_dir / f"{period_end}-monthly-{project_slug}.json"
        snap_path.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8")
        report_path = out_dir / f"{period_end}-monthly.md"
        report_path.write_text(rendered, encoding="utf-8")
        print(json.dumps({
            "snapshot_path": str(snap_path.resolve()),
            "report_path": str(report_path.resolve()),
            "report_id": report["report_id"],
            "sections": list(report["sections"].keys()),
        }, ensure_ascii=False, indent=2))
    else:
        print(json.dumps({"snapshot": snapshot, "report_md": rendered},
                         ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))


__all__ = (
    "SCHEMA_VERSION", "REQUIRED_SECTIONS", "FRAMING_POLICY_ENUM",
    "OUTPUT_FORMATS_ENUM", "DATA_SOURCE_ENUM", "PRIORITY_RANK", "PERIOD_DAYS",
    "ReportInputs", "MonthlyReportError", "FramingPolicyEnumViolation",
    "OutputFormatEnumViolation", "DataSourceEnumViolation",
    "MissingSectionError", "ProjectIdShapeError", "WorkbookMissingError",
    "TemplateRenderError", "read_workbook", "read_events_jsonl",
    "assemble_report", "render_report_markdown", "build_snapshot", "main",
)
