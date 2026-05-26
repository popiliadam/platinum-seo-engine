#!/usr/bin/env python3
"""
whats_next.py — backing module for the `whats-next` meta/router skill.

Pure read + score + emit. NEVER invokes another skill. Four signal
sources (master_task TODO, content_decay RED, awaiting_approval
workflow runs, quick_wins pending), one fixed heuristic, one
events.jsonl row per run.

Public API
----------
SCORES                       — heuristic table (constant, for tests)
RECOMMENDED_SKILL            — kind → suggested downstream skill
SUGGESTED_NONE               — sentinel string for empty results
scan_master_task_todo(...)   — sheet scan, returns list[dict]
scan_content_decay_red(...)  — sheet scan, returns list[dict]
scan_pending_approvals(...)  — workflow_runner.list_runs filter
scan_quick_wins_pending(...) — sheet scan, returns list[dict]
score_candidate(candidate)   — int (heuristic)
rank_candidates(c, top_k)    — list[dict] sorted desc, truncated
render_summary(ranked, ...)  — markdown bullet list (str)
synthetic_task_id(run_id)    — derive `T-9NNNN` from workflow run hex
emit_router_event(...)       — events_writer.append_work wrapper
run(project_slug, ...)       — full 10-step orchestration

Cross-references
----------------
- Schemas: events.schema.json (event_kind=work, event_type=skill_whats_next
  canonical per v1.6-Phase-2 H-E; task_id pattern ^T-[0-9]{4,}$, note
  required), workflow-run.schema.json
  (outputs.* string-typed, F5).
- Cross-modules: scripts.state.workflow_runner,
  scripts.state.events_writer, openpyxl (read-only).
- ADRs: ADR-019, ADR-020, ADR-021, F5 (string-typed outputs).
"""

from __future__ import annotations

import json
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from scripts.state import events_writer, workflow_runner


# ---------------------------------------------------------------------------
# Constants — heuristic table (single source of truth, asserted by tests)
# ---------------------------------------------------------------------------

SCORES: dict[str, int] = {
    "pending_approval":      100,
    "content_decay_red":      80,
    "master_task_high":       60,   # CRITICAL or HIGH
    "quick_wins_pending":     50,
    "master_task_medium":     40,
    "sf_crawl_stale":         30,   # v1.8 Phase 4: SF MCP non-blocking
    "master_task_other":      20,   # LOW or unset
}

RECOMMENDED_SKILL: dict[str, str] = {
    "pending_approval":   "(resume awaiting_approval run)",
    "content_decay_red":  "content-decay",
    "master_task":        "done-protocol",
    "quick_wins_pending": "apply-quickwin",
    "sf_crawl_stale":     "sf-crawl-orchestrator",
}

# v1.8 Phase 4 SF MCP staleness threshold (operator-tunable via call).
_SF_CRAWL_STALE_DEFAULT_DAYS: int = 30

SUGGESTED_NONE: str = "(none)"

_TOP_K_MIN = 1
_TOP_K_MAX = 10
_MASTER_TASK_DATA_START_ROW = 4   # master-excel.schema sheets.master_task
_CONTENT_DECAY_DATA_START_ROW = 6  # master-excel.schema sheets.content_decay
_QUICK_WINS_DATA_START_ROW = 5     # master-excel.schema sheets.quick_wins


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class WhatsNextError(Exception):
    """Base class for whats-next router errors."""


class MasterExcelMissingError(WhatsNextError):
    """projects/{slug}/master.xlsx absent — DURUR #2."""


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _resolve_workspace_root(workspace_root: Path | None) -> Path:
    if workspace_root is not None:
        return Path(workspace_root)
    env = os.getenv("PSEO_WORKSPACE_ROOT")
    if env:
        return Path(env)
    raise WhatsNextError(
        "workspace_root not provided and PSEO_WORKSPACE_ROOT not set"
    )


def _master_excel_path(project_slug: str, workspace_root: Path | None) -> Path:
    ws = _resolve_workspace_root(workspace_root)
    return ws / "projects" / project_slug / "master.xlsx"


# ---------------------------------------------------------------------------
# Sheet scans (read-only; sheet-absent → empty list, NOT error — AMBER)
# ---------------------------------------------------------------------------

def _open_workbook(path: Path):
    if not path.exists():
        raise MasterExcelMissingError(
            f"master.xlsx missing at {path} — run init-project first (DURUR #2)"
        )
    return load_workbook(filename=str(path), read_only=True, data_only=True)


def scan_master_task_todo(
    project_slug: str,
    workspace_root: Path | None = None,
) -> list[dict[str, Any]]:
    """Return TODO rows from `master_task` sheet. Sheet-absent → []."""
    path = _master_excel_path(project_slug, workspace_root)
    wb = _open_workbook(path)
    try:
        if "master_task" not in wb.sheetnames:
            return []
        ws = wb["master_task"]
        out: list[dict[str, Any]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=_MASTER_TASK_DATA_START_ROW, values_only=True),
            start=_MASTER_TASK_DATA_START_ROW,
        ):
            if not row or row[0] is None:
                continue
            # Master-excel.schema cols: A=task_id, B=task, G=priority(idx 6), J=status(idx 9)
            task_id = row[0]
            task = row[1] if len(row) > 1 else None
            priority = row[6] if len(row) > 6 else None
            status = row[9] if len(row) > 9 else None
            if status != "TODO":
                continue
            out.append({
                "kind": "master_task",
                "row_id": row_idx,
                "task_id": task_id,
                "task": task,
                "priority": priority,
                "status": status,
            })
        return out
    finally:
        wb.close()


def scan_content_decay_red(
    project_slug: str,
    workspace_root: Path | None = None,
) -> list[dict[str, Any]]:
    """Return RED-trend rows from `content_decay`. Sheet-absent → [] (AMBER)."""
    path = _master_excel_path(project_slug, workspace_root)
    wb = _open_workbook(path)
    try:
        if "content_decay" not in wb.sheetnames:
            return []
        ws = wb["content_decay"]
        out: list[dict[str, Any]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=_CONTENT_DECAY_DATA_START_ROW, values_only=True),
            start=_CONTENT_DECAY_DATA_START_ROW,
        ):
            if not row or row[0] is None:
                continue
            # cols: A=url(0), E=delta_pct(4), F=trend(5), G=pillar(6)
            url = row[0]
            delta_pct = row[4] if len(row) > 4 else None
            trend = row[5] if len(row) > 5 else None
            if not isinstance(trend, str) or not trend.upper().startswith("RED"):
                continue
            out.append({
                "kind": "content_decay_red",
                "row_id": row_idx,
                "url": url,
                "delta_pct": delta_pct,
                "trend": trend,
            })
        return out
    finally:
        wb.close()


def scan_pending_approvals(
    project_slug: str,
    workspace_root: Path | None = None,
) -> list[dict[str, Any]]:
    """Return runs whose top-level status == awaiting_approval."""
    runs = workflow_runner.list_runs(
        project_slug,
        workspace_root=workspace_root,
        status_filter="awaiting_approval",
    )
    out: list[dict[str, Any]] = []
    for r in runs:
        meta = r.data.get("approval_meta") or {}
        out.append({
            "kind": "pending_approval",
            "row_id": r.run_id,
            "run_id": r.run_id,
            "skill": r.data.get("skill"),
            "subject": meta.get("subject"),
            "approver": meta.get("approver"),
        })
    return out


def scan_quick_wins_pending(
    project_slug: str,
    workspace_root: Path | None = None,
) -> list[dict[str, Any]]:
    """Return rows from `quick_wins` sheet (each row = pending approval)."""
    path = _master_excel_path(project_slug, workspace_root)
    wb = _open_workbook(path)
    try:
        if "quick_wins" not in wb.sheetnames:
            return []
        ws = wb["quick_wins"]
        out: list[dict[str, Any]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=_QUICK_WINS_DATA_START_ROW, values_only=True),
            start=_QUICK_WINS_DATA_START_ROW,
        ):
            if not row or row[0] is None:
                continue
            # cols: A=query, B=url, H=opportunity(7)
            out.append({
                "kind": "quick_wins_pending",
                "row_id": row_idx,
                "query": row[0],
                "url": row[1] if len(row) > 1 else None,
                "opportunity": row[7] if len(row) > 7 else None,
            })
        return out
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Heuristic — score + rank
# ---------------------------------------------------------------------------

def score_candidate(candidate: dict[str, Any]) -> int:
    """Map a candidate dict to its heuristic score (constant table)."""
    kind = candidate.get("kind")
    if kind == "pending_approval":
        return SCORES["pending_approval"]
    if kind == "content_decay_red":
        return SCORES["content_decay_red"]
    if kind == "quick_wins_pending":
        return SCORES["quick_wins_pending"]
    if kind == "sf_crawl_stale":
        return SCORES["sf_crawl_stale"]
    if kind == "master_task":
        prio = (candidate.get("priority") or "").upper()
        if prio in ("CRITICAL", "HIGH"):
            return SCORES["master_task_high"]
        if prio == "MEDIUM":
            return SCORES["master_task_medium"]
        return SCORES["master_task_other"]
    return 0


def _suggested_skill(candidate: dict[str, Any]) -> str:
    kind = candidate.get("kind")
    if kind == "master_task":
        return RECOMMENDED_SKILL["master_task"]
    return RECOMMENDED_SKILL.get(kind, SUGGESTED_NONE)


def _candidate_reason(c: dict[str, Any]) -> str:
    kind = c.get("kind")
    if kind == "pending_approval":
        return f"workflow {c.get('run_id')} ({c.get('skill')}) bekliyor; approve veya reject."
    if kind == "content_decay_red":
        return f"{c.get('url')} ({c.get('trend')}, delta={c.get('delta_pct')}); content-decay skill çalıştır."
    if kind == "master_task":
        return f"{c.get('task_id')} priority={c.get('priority')} \"{c.get('task')}\"; done-protocol başlat."
    if kind == "quick_wins_pending":
        return f"{c.get('query')} → {c.get('url')} (opportunity={c.get('opportunity')}); apply-quickwin."
    if kind == "sf_crawl_stale":
        return f"last SF crawl {c.get('last_seen') or '<none>'}; consider /pseo-sf-crawl {c.get('project_slug')}."
    return "(unknown)"


def _parse_iso_z(s: Any) -> datetime | None:
    """Parse an ISO 8601 string (possibly with Z) to an aware datetime.
    Returns None on parse failure."""
    if not isinstance(s, str) or not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return None


def suggest_sf_crawl_when_stale(
    project_slug: str,
    workspace_root: Path | None = None,
    threshold_days: int = _SF_CRAWL_STALE_DEFAULT_DAYS,
) -> dict[str, Any] | None:
    """v1.8 Phase 4 (D-SF-04 helper): if project has sf.mcp.enabled=true
    and the newest completed sf-crawl-orchestrator workflow run is older
    than ``threshold_days`` (or no run exists), return a non-blocking
    candidate suggestion for the whats-next router. Otherwise return None.

    The helper is read-only: it never invokes a skill nor mutates state.
    Caller (whats-next.run) decides whether to append it to the candidate
    list.
    """
    try:
        ws = _resolve_workspace_root(workspace_root)
    except WhatsNextError:
        return None
    cfg_path = ws / "projects" / project_slug / "project.config.json"
    if not cfg_path.exists():
        return None
    try:
        cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    sf_block = cfg.get("sf") or {}
    mcp_block = sf_block.get("mcp") or {}
    if not mcp_block.get("enabled"):
        # SF MCP not opted-in for this project — no suggestion surface.
        return None

    wf_dir = ws / "projects" / project_slug / "_state" / "workflows"
    newest: datetime | None = None
    if wf_dir.is_dir():
        for p in sorted(wf_dir.glob("*.json")):
            try:
                obj = json.loads(p.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            if obj.get("skill") != "sf-crawl-orchestrator":
                continue
            if obj.get("status") != "done":
                continue
            # Prefer completed_at; fall back to updated_at then created_at.
            ts = (
                _parse_iso_z(obj.get("completed_at"))
                or _parse_iso_z(obj.get("updated_at"))
                or _parse_iso_z(obj.get("created_at"))
            )
            if ts is None:
                continue
            if newest is None or ts > newest:
                newest = ts

    now = datetime.now(timezone.utc)
    if newest is None:
        last_seen_label: str | None = None
        stale = True
    else:
        delta = now - newest
        stale = delta > timedelta(days=threshold_days)
        last_seen_label = newest.strftime("%Y-%m-%d")

    if not stale:
        return None

    return {
        "kind": "sf_crawl_stale",
        "row_id": f"sf-mcp-{project_slug}",
        "project_slug": project_slug,
        "last_seen": last_seen_label,
        "threshold_days": threshold_days,
    }


def rank_candidates(
    candidates: Iterable[dict[str, Any]],
    top_k: int = 3,
) -> list[dict[str, Any]]:
    """Score, sort desc, break ties stably, truncate to top_k (clamped 1..10)."""
    if top_k < _TOP_K_MIN:
        top_k = _TOP_K_MIN
    if top_k > _TOP_K_MAX:
        top_k = _TOP_K_MAX
    def _row_sort_key(rid: Any) -> tuple[int, int, str]:
        # Prefer numeric ordering when row_id is an int; fall back to string.
        # Tuple shape (type_rank, int_value, str_value) keeps mixed types
        # comparable without TypeErrors and stable across runs.
        if isinstance(rid, int):
            return (0, rid, "")
        return (1, 0, str(rid))

    scored: list[tuple[int, str, tuple[int, int, str], dict[str, Any]]] = []
    for c in candidates:
        s = score_candidate(c)
        # Secondary keys: kind alphabetic, then row_id (numeric if possible).
        scored.append((s, str(c.get("kind") or ""), _row_sort_key(c.get("row_id")), c))
    # Sort: score DESC, then kind ASC, then row_id ASC.
    scored.sort(key=lambda t: (-t[0], t[1], t[2]))
    out: list[dict[str, Any]] = []
    for rank, (score, _kind, _rid, c) in enumerate(scored[:top_k], start=1):
        out.append({
            "rank": rank,
            "score": score,
            "kind": c.get("kind"),
            "suggested_skill": _suggested_skill(c),
            "reason": _candidate_reason(c),
            "candidate": c,
        })
    return out


# ---------------------------------------------------------------------------
# Markdown summary (the body of the events.jsonl router event note)
# ---------------------------------------------------------------------------

def render_summary(
    ranked: list[dict[str, Any]],
    project_slug: str,
    user_intent: str | None = None,
) -> str:
    """Markdown bullet list. Used both as return value and events note."""
    title = f"**whats-next — Top-{len(ranked)} recommendations** (project: {project_slug})"
    if user_intent:
        title += f" — intent: {user_intent}"
    if not ranked:
        return title + "\n\n_(no candidates surfaced — workspace is quiet.)_"
    lines = [title, ""]
    for r in ranked:
        lines.append(
            f"{r['rank']}. [score {r['score']:>3}] {r['kind']} → {r['suggested_skill']} — {r['reason']}"
        )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Synthetic task_id (events.schema requires ^T-[0-9]{4,}$)
# ---------------------------------------------------------------------------

def synthetic_task_id(run_id: str) -> str:
    """Derive a `T-9NNNN` router task_id from a workflow run_id hex token.

    Router ids land in the 90000..99998 band so they never collide with
    master_task `T-NNNN` ids (which use 1..9999).
    """
    parts = run_id.split("-")
    token = parts[-1] if parts else ""
    if not token:
        return "T-90000"
    try:
        n = int(token, 16)
    except ValueError:
        n = sum(ord(c) for c in token)
    return f"T-{90000 + (n % 9999)}"


# ---------------------------------------------------------------------------
# Router event emission (single events.jsonl row per run)
# ---------------------------------------------------------------------------

def emit_router_event(
    *,
    project_slug: str,
    run_id: str,
    note: str,
    workspace_root: Path | None = None,
) -> events_writer.EventResult:
    """Single event_kind=work, event_type=skill_whats_next, primary_source=manual.

    Canonical event_type per v1.6-Phase-2 H-E (events.schema event_type enum
    10→12 additive bump 2026-05-07; ADR-018 paterni reuse). Replaces legacy
    DSL workaround event_type=manual + note=[skill=whats-next].
    """
    return events_writer.append_work(
        project_id=project_slug,
        event_type="skill_whats_next",
        task_id=synthetic_task_id(run_id),
        note=note,
        primary_source="manual",
        workspace_root=workspace_root,
    )


# ---------------------------------------------------------------------------
# Orchestration — full 10-step run
# ---------------------------------------------------------------------------

def run(
    project_slug: str,
    *,
    user_intent: str | None = None,
    top_k: int = 3,
    workspace_root: Path | None = None,
) -> dict[str, Any]:
    """Full 10-step protocol — returns {recommendations, summary, run_id}."""
    # Step 1: create_run.
    handle = workflow_runner.create_run(
        skill="whats-next",
        project_slug=project_slug,
        steps=[
            {"name": "scan_master_task"},
            {"name": "scan_content_decay"},
            {"name": "scan_pending_approvals"},
            {"name": "scan_quick_wins_pending"},
            {"name": "score_and_rank"},
            {"name": "emit_router_event"},
        ],
        workspace_root=workspace_root,
    )

    # Steps 2-5: scan signal sources.
    candidates: list[dict[str, Any]] = []
    candidates += scan_master_task_todo(project_slug, workspace_root=workspace_root)
    candidates += scan_content_decay_red(project_slug, workspace_root=workspace_root)
    candidates += scan_pending_approvals(project_slug, workspace_root=workspace_root)
    candidates += scan_quick_wins_pending(project_slug, workspace_root=workspace_root)
    # Step 4.5 — v1.8 Phase 4 SF MCP staleness non-blocking suggestion.
    sf_stale = suggest_sf_crawl_when_stale(
        project_slug, workspace_root=workspace_root
    )
    if sf_stale is not None:
        candidates.append(sf_stale)

    # Steps 6-7: score, rank, build top-K.
    ranked = rank_candidates(candidates, top_k=top_k)

    # Step 8: emit single events.jsonl router event.
    summary = render_summary(ranked, project_slug=project_slug, user_intent=user_intent)
    emit_router_event(
        project_slug=project_slug,
        run_id=handle.run_id,
        note=summary,
        workspace_root=workspace_root,
    )

    # Step 9: build return value (caller prints / displays).
    out = {
        "run_id": handle.run_id,
        "recommendations": ranked,
        "summary": summary,
    }

    # Step 10: complete (F5: outputs.* STRING-TYPED — never int).
    workflow_runner.complete(
        handle.run_id,
        project_slug=project_slug,
        workspace_root=workspace_root,
        outputs={
            "recommendations_count": str(len(ranked)),
            "top_skill": ranked[0]["suggested_skill"] if ranked else SUGGESTED_NONE,
        },
    )
    return out


__all__: Iterable[str] = (
    "SCORES",
    "RECOMMENDED_SKILL",
    "SUGGESTED_NONE",
    "WhatsNextError",
    "MasterExcelMissingError",
    "scan_master_task_todo",
    "scan_content_decay_red",
    "scan_pending_approvals",
    "scan_quick_wins_pending",
    "suggest_sf_crawl_when_stale",
    "score_candidate",
    "rank_candidates",
    "render_summary",
    "synthetic_task_id",
    "emit_router_event",
    "run",
)
