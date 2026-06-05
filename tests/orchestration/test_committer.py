"""Idempotency proof for scripts/orchestration/committer.py (AMO batch 1b).

committer.commit wraps transaction.replace (NOT append): re-committing the same
step's rows must NOT duplicate. This test drives the REAL transaction.replace on
a throwaway master.xlsx (default master-excel schema, topical_map sheet — which
declares no allowed_writers, so writer='orchestrator' is accepted) and asserts
the rowcount is stable across two identical commits. The e2e stub harness covers
the rest with a fake commit_fn; this is the ONE focused real-wrap test.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from scripts.excel.transaction import WriteResult
from scripts.orchestration import committer

RUN_ID = "test-proj-2026-06-05-a1b2"
SLUG = "test-proj"


def _topical_row(**overrides) -> dict:
    base = {
        "pillar": "P01_sofa_sets",
        "cluster": "leather-sofas",
        "primary_keyword": "leather sofa",
        "monthly_volume": 4400,
        "data_source": "gsc",
        "assigned_url": "/leather-sofas",
        "page_type": "pillar",
        "status": "TODO",
        "priority": "HIGH",
        "note": "",
    }
    base.update(overrides)
    return base


def _setup(tmp_path: Path) -> Path:
    proj = tmp_path / "projects" / SLUG
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    return proj / "master.xlsx"


def test_commit_lands_rows(tmp_path: Path) -> None:
    wb_path = _setup(tmp_path)
    rows = [_topical_row(primary_keyword=f"kw {i}") for i in range(3)]

    result = committer.commit(wb_path, "topical_map", rows, run_id=RUN_ID, project_slug=SLUG)

    assert isinstance(result, WriteResult)
    assert result.rows_affected == 3
    ws = load_workbook(wb_path)["topical_map"]
    assert ws.max_row == 7  # header_row=4, data lands at 5..7
    assert ws["C5"].value == "kw 0"


def test_commit_is_idempotent_no_duplication(tmp_path: Path) -> None:
    wb_path = _setup(tmp_path)
    rows = [_topical_row(primary_keyword=f"kw {i}") for i in range(3)]

    r1 = committer.commit(wb_path, "topical_map", rows, run_id=RUN_ID, project_slug=SLUG)
    max1 = load_workbook(wb_path)["topical_map"].max_row

    # Re-commit the SAME run's SAME rows: replace clears + rewrites -> no growth.
    r2 = committer.commit(wb_path, "topical_map", rows, run_id=RUN_ID, project_slug=SLUG)
    max2 = load_workbook(wb_path)["topical_map"].max_row

    assert r1.rows_affected == r2.rows_affected == 3
    assert max1 == max2 == 7  # still 3 data rows — replace semantics, no dupes
