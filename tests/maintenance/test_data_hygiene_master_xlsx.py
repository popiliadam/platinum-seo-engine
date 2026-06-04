"""tests/maintenance/test_data_hygiene_master_xlsx.py — ADR-037 invariants.

Locks:
  - dry-run produces an audit-trail markdown but does NOT mutate the workbook;
  - apply on a clean workbook is idempotent (0 cells changed);
  - apply on a synthetic dirty workbook normalizes P1/P2/P3 → HIGH/MEDIUM/LOW
    and writes an audit trail listing each change;
  - re-running apply on the post-apply workbook yields a 0-cells audit (true
    idempotency).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO = Path(__file__).resolve().parents[2]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from scripts.maintenance import data_hygiene_master_xlsx as hygiene  # noqa: E402


def _make_master_task_workbook(
    tmp_path: Path,
    *,
    priorities: list[str],
) -> Path:
    """Build a minimal workspace-shape workbook with master_task at
    schema header_row=3 (matches demo-dental layout)."""
    project_dir = tmp_path / "projects" / "demo-dental"
    project_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = project_dir / "master.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    schema_path = REPO / "schemas" / "master-excel.schema.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    mt_def = schema["sheets"]["master_task"]
    columns = mt_def["required_columns"]
    header_row = mt_def["header_row"]
    data_start = mt_def["data_start_row"]

    # Build all sheets so the workbook validates structurally.
    for sheet_name, sdef in schema["sheets"].items():
        ws = wb.create_sheet(title=sheet_name)
        cols = sdef.get("required_columns", [])
        if not cols:
            continue
        sh_row = sdef.get("header_row", 1)
        for col in cols:
            ws[f"{col['col']}{sh_row}"] = col["name"]

    ws = wb["master_task"]
    col_map = {c["name"]: c["col"] for c in columns}
    for i, prio in enumerate(priorities):
        row = data_start + i
        # Schema-valid values (except the deliberately-dirty `priority`): the
        # --apply path runs transaction.update, which validates the WHOLE merged
        # row, so primary_source must be a real enum value and auto_generated a
        # real bool. created_date stays an ISO string (the validator types the
        # schema `date` as string|null, not a Python date object).
        ws[f"{col_map['task_id']}{row}"] = f"T-{20000 + i}"
        ws[f"{col_map['task']}{row}"] = f"task {i}"
        ws[f"{col_map['primary_source']}{row}"] = "manual"
        ws[f"{col_map['url']}{row}"] = "https://example.com/"
        ws[f"{col_map['category']}{row}"] = "test"
        ws[f"{col_map['priority']}{row}"] = prio
        ws[f"{col_map['impact']}{row}"] = "synthetic"
        ws[f"{col_map['duration_est_min']}{row}"] = 10
        ws[f"{col_map['status']}{row}"] = "TODO"
        ws[f"{col_map['created_date']}{row}"] = "2026-05-06"
        ws[f"{col_map['auto_generated']}{row}"] = False

    wb.save(workbook_path)
    return workbook_path


def test_scan_priority_drift_detects_pcodes(tmp_path: Path) -> None:
    workbook_path = _make_master_task_workbook(
        tmp_path, priorities=["P1", "P2", "HIGH", "P3", "MEDIUM"]
    )
    changes = hygiene._scan_priority_drift(workbook_path)
    olds = sorted(c.old for c in changes)
    assert olds == ["P1", "P2", "P3"]
    by_old = {c.old: c.new for c in changes}
    assert by_old == {"P1": "HIGH", "P2": "MEDIUM", "P3": "LOW"}


def test_scan_clean_workbook_is_no_op(tmp_path: Path) -> None:
    workbook_path = _make_master_task_workbook(
        tmp_path, priorities=["HIGH", "MEDIUM", "LOW", "CRITICAL"]
    )
    assert hygiene._scan_priority_drift(workbook_path) == []


def test_format_audit_includes_per_cell_log(tmp_path: Path) -> None:
    workbook_path = _make_master_task_workbook(
        tmp_path, priorities=["P1", "P2", "HIGH"]
    )
    changes = hygiene._scan_priority_drift(workbook_path)
    audit = hygiene._format_audit(changes, project_slug="demo-dental", apply_mode=False)
    assert "Mode:** DRY-RUN" in audit
    assert "Cells affected:** 2" in audit
    assert "P1 | HIGH" in audit
    assert "P2 | MEDIUM" in audit


def test_format_audit_clean_workbook_zero_cells(tmp_path: Path) -> None:
    audit = hygiene._format_audit([], project_slug="demo-dental", apply_mode=True)
    assert "Cells affected:** 0" in audit
    assert "0 cells outside severityEnum" in audit


def test_dry_run_does_not_mutate_workbook(tmp_path: Path) -> None:
    workbook_path = _make_master_task_workbook(tmp_path, priorities=["P1", "P2"])
    pre_bytes = workbook_path.read_bytes()
    rc = hygiene.main([
        "--workspace", str(tmp_path),
        "--project", "demo-dental",
        "--dry-run",
    ])
    assert rc == 0
    assert workbook_path.read_bytes() == pre_bytes  # exact byte parity


def test_apply_normalizes_and_then_idempotent(tmp_path: Path) -> None:
    """--apply normalizes P1/P2/P3 -> HIGH/MEDIUM/LOW via transaction.update,
    then a second --apply is a true no-op (idempotent).

    The prior skip claimed the apply path "requires transaction state dir + lock"
    — but that fixture already exists (tests/scripts/test_transaction.py). The
    real prerequisite is a schema-valid workbook: transaction.update validates
    the whole merged row, so the fixture rows had to become valid first."""
    workbook_path = _make_master_task_workbook(
        tmp_path, priorities=["P1", "P2", "P3", "HIGH"]
    )
    # transaction.update needs projects/{slug}/_state/ for its excel.lock,
    # backups, and provenance event (same setup as test_transaction.py).
    (workbook_path.parent / "_state").mkdir(parents=True, exist_ok=True)

    rc = hygiene.main(
        ["--workspace", str(tmp_path), "--project", "demo-dental", "--apply"]
    )
    assert rc == 0

    # No P-codes remain, and the cells now carry the mapped severities.
    assert hygiene._scan_priority_drift(workbook_path) == []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        priorities = sorted(
            r.get("priority")
            for r in hygiene._iter_rows_as_dicts(wb, "master_task")
        )
    finally:
        wb.close()
    assert priorities == ["HIGH", "HIGH", "LOW", "MEDIUM"]

    # Second apply on the now-clean workbook: 0 changes (true idempotency).
    rc2 = hygiene.main(
        ["--workspace", str(tmp_path), "--project", "demo-dental", "--apply"]
    )
    assert rc2 == 0
    assert hygiene._scan_priority_drift(workbook_path) == []
