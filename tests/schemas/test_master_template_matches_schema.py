"""FIX-J: the shipped Excel template MUST match master-excel.schema.json.

`templates/master-excel.xlsx` is the canonical workbook every project pack is
seeded from. If its sheet names or header rows drift from
`master-excel.schema.json` (e.g. a sheet gains a column in the schema but the
template is not regenerated), projects silently inherit a stale layout that the
transforms then fail to populate. This test locks the template ↔ schema
correspondence — including the columns GAP-M-1b added to `quick_wins` (K–N) and
`opportunity` (I–J).

Known, documented divergence (re-derived 2026-06-10): the schema defines a
`gbp_audit` sheet (added by the local/GBP gap work) that the binary template has
NOT yet been regenerated to include. Regenerating the .xlsx is out of FIX-J
scope (FIX-J touches schemas + tests + ARCHITECTURE only), so the gap is pinned
in `KNOWN_SCHEMA_ONLY_SHEETS` as an explicit tripwire: this test fires both when
a NEW schema sheet appears without a template, and when `gbp_audit` is finally
templated (forcing whoever does it to update this constant).
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
SCHEMA_PATH = REPO / "schemas" / "master-excel.schema.json"
TEMPLATE_PATH = REPO / "templates" / "master-excel.xlsx"

# Schema sheets not yet present in the shipped template. See module docstring.
KNOWN_SCHEMA_ONLY_SHEETS = {"gbp_audit"}


def _schema_sheets() -> dict:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))["sheets"]


def _expected_headers(sheet_def: dict) -> list | None:
    """Ordered column names a sheet's header row must carry, or None for KPI
    sheets (dashboard) that declare `required_cells` instead of columns."""
    cols = sheet_def.get("required_columns")
    if not cols or sheet_def.get("header_row") is None:
        return None
    return [c["name"] for c in cols]


def _actual_headers(ws, header_row: int) -> list:
    rows = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    values = list(rows[0]) if rows else []
    while values and values[-1] is None:  # trim trailing empty cells
        values.pop()
    return values


def _header_diff(expected: list, actual: list) -> list[str]:
    """Pure comparison: human-readable per-position differences (empty == match)."""
    diffs: list[str] = []
    if len(expected) != len(actual):
        diffs.append(f"column count {len(actual)} != expected {len(expected)}")
    for i in range(max(len(expected), len(actual))):
        exp = expected[i] if i < len(expected) else "<missing>"
        act = actual[i] if i < len(actual) else "<missing>"
        if exp != act:
            diffs.append(f"col[{i}] {act!r} != expected {exp!r}")
    return diffs


def test_template_sheets_are_all_schema_defined() -> None:
    """No orphan sheet in the template that the schema doesn't define."""
    schema_sheets = set(_schema_sheets())
    wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    try:
        orphans = set(wb.sheetnames) - schema_sheets
    finally:
        wb.close()
    assert not orphans, f"template has sheets not defined in master-excel.schema.json: {sorted(orphans)}"


def test_schema_only_sheets_match_known_gap() -> None:
    """Schema sheets absent from the template == the documented known gap.

    Fires if a NEW schema sheet ships without a template, or when the known gap
    (gbp_audit) is finally templated — both demand a deliberate update here."""
    schema_sheets = set(_schema_sheets())
    wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    try:
        schema_only = schema_sheets - set(wb.sheetnames)
    finally:
        wb.close()
    assert schema_only == KNOWN_SCHEMA_ONLY_SHEETS, (
        f"schema-only sheets {sorted(schema_only)} != documented "
        f"{sorted(KNOWN_SCHEMA_ONLY_SHEETS)}. A schema sheet was added/removed "
        "without regenerating templates/master-excel.xlsx — update the template "
        "or KNOWN_SCHEMA_ONLY_SHEETS with rationale."
    )


def test_template_header_rows_match_schema_columns() -> None:
    """Every sheet present in BOTH must carry the schema's exact header row
    (names, in column order) — locks GAP-M-1b's added columns."""
    schema_sheets = _schema_sheets()
    wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    failures: list[str] = []
    try:
        for name in wb.sheetnames:
            if name not in schema_sheets:
                continue
            expected = _expected_headers(schema_sheets[name])
            if expected is None:  # KPI sheet (dashboard) — no header row to match
                continue
            actual = _actual_headers(wb[name], schema_sheets[name]["header_row"])
            diff = _header_diff(expected, actual)
            if diff:
                failures.append(f"{name}: " + "; ".join(diff))
    finally:
        wb.close()
    assert not failures, "template header rows drift from schema:\n" + "\n".join(f"  - {f}" for f in failures)


def test_header_diff_detects_drift() -> None:
    """Teeth check: the comparison must flag a real divergence (so the
    locks above cannot silently pass on a mismatch)."""
    assert _header_diff(["query", "url"], ["query", "url"]) == []
    assert _header_diff(["query", "url"], ["query", "URL"])  # renamed column
    assert _header_diff(["a", "b", "c"], ["a", "b"])  # missing trailing column
