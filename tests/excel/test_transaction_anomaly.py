"""
Audit-atomicity tests for scripts/excel/transaction.py (hostile-audit #8).

After an atomic workbook write, transaction.py emits a provenance event into
events.jsonl. Operator decision D-B: if that emit FAILS, the workbook must be
KEPT (no roll-back), the caller must NOT be hard-blocked (no propagated
exception), and a DURABLE ANOMALY must be recorded into the _state/anomalies.jsonl
sidecar so the missing provenance event stays reconcilable.

These tests fault-inject events_writer.append_provenance and assert the D-B
contract on BOTH emit sites: the append/replace driver and update().
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.excel import transaction
from scripts.state import events_writer


def _setup_project(tmp_path: Path, slug: str = "test-proj") -> tuple[Path, Path]:
    proj = tmp_path / "projects" / slug
    state = proj / "_state"
    state.mkdir(parents=True, exist_ok=True)
    return proj, state


def _topical_row(**overrides) -> dict:
    base = {
        "pillar": "P01_sofa_sets",
        "cluster": "leather-sofas",
        "primary_keyword": "leather sofa",
        "monthly_volume": 4400,
        "data_source": "gsc",
        "assigned_url": "/leather-sofas",
        "page_type": "pillar",
        "status": "TODO",
        "priority": "HIGH",
        "note": "",
    }
    base.update(overrides)
    return base


def _anomaly_records(state: Path) -> list[dict]:
    path = state / "anomalies.jsonl"
    if not path.exists():
        return []
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


# ---------------------------------------------------------------------------
# Test 1 — append: emit failure records anomaly, keeps workbook, no exception
# ---------------------------------------------------------------------------

def test_append_provenance_emit_failure_records_anomaly(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    def _boom(**kwargs):
        raise events_writer.EventWriterError("simulated events.jsonl emit failure")

    monkeypatch.setattr(events_writer, "append_provenance", _boom)

    # Must NOT raise (D-B: emit hiccup never hard-blocks a legitimate write).
    result = transaction.append(
        wb_path, "topical_map", [_topical_row(primary_keyword="ghost kw")], "test-proj"
    )

    # Workbook was written and KEPT (no roll-back).
    assert wb_path.exists()
    wb = load_workbook(wb_path)
    ws = wb["topical_map"]
    values = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "ghost kw" in values

    # No provenance event id was produced → sentinel empty string.
    assert result.event_id == ""
    assert result.rows_affected == 1

    # A durable anomaly record was written to the SEPARATE sidecar.
    recs = _anomaly_records(state)
    assert len(recs) == 1
    rec = recs[0]
    assert rec["kind"] == "provenance_emit_failed"
    assert rec["detail"]["sheet"] == "topical_map"
    assert rec["detail"]["rows_written"] == 1
    assert rec["detail"]["operation"] == "project_excel"
    assert "simulated events.jsonl emit failure" in rec["error"]


# ---------------------------------------------------------------------------
# Test 2 — update(): the second emit site honors the same D-B contract
# ---------------------------------------------------------------------------

def test_update_provenance_emit_failure_records_anomaly(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    # Seed one row with a working emit (no anomaly yet).
    transaction.append(
        wb_path, "topical_map", [_topical_row(primary_keyword="seed-kw")], "test-proj"
    )
    assert _anomaly_records(state) == []

    def _boom(**kwargs):
        raise events_writer.EventWriterError("emit down during update")

    monkeypatch.setattr(events_writer, "append_provenance", _boom)

    result = transaction.update(
        wb_path, "topical_map",
        {"primary_keyword": "seed-kw"}, {"note": "touched"},
        "test-proj",
    )

    # The mutation is on disk (kept, not rolled back).
    wb = load_workbook(wb_path)
    ws = wb["topical_map"]
    notes = [ws.cell(row=r, column=10).value for r in range(1, ws.max_row + 1)]
    assert "touched" in notes

    assert result.event_id == ""
    assert result.rows_affected == 1

    recs = _anomaly_records(state)
    assert len(recs) == 1
    assert recs[0]["kind"] == "provenance_emit_failed"
    assert recs[0]["detail"]["sheet"] == "topical_map"
    assert "emit down during update" in recs[0]["error"]


# ---------------------------------------------------------------------------
# Test 3 — happy path writes NO anomaly (guard against always-recording)
# ---------------------------------------------------------------------------

def test_successful_emit_writes_no_anomaly(tmp_path: Path) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    result = transaction.append(
        wb_path, "topical_map", [_topical_row()], "test-proj"
    )

    # Real event id (uuid hex), and the sidecar was never created.
    assert result.event_id
    assert result.event_id != ""
    assert not (state / "anomalies.jsonl").exists()
