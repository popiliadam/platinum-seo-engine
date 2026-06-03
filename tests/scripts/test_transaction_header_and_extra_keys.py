"""tests/scripts/test_transaction_header_and_extra_keys.py — P1-09.

Two write-path integrity gaps in scripts/excel/transaction.py:

(a) _ensure_sheet_with_header always wrote the header to row 1, ignoring the
    schema's header_row. For the 18/19 sheets with header_row>=3 that disagrees
    with data_start_row (= header_row+1) used by replace() and every reader, so a
    freshly-created sheet had its header at row 1 but data expected at row 5.

(b) Per-row validation used additionalProperties:true and the write loop silently
    skipped any key without a column — so a typo'd column name vanished without a
    trace. Unknown keys now FAIL by default; allow_extra=True restores the skip.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.excel import transaction
from scripts.excel.transaction import RowSchemaError


def _setup_project(tmp_path: Path, slug: str = "test-proj") -> Path:
    proj = tmp_path / "projects" / slug
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    return proj


def _tech_seo_row(**overrides) -> dict:
    # tech_seo: header_row=3, data_start_row=4. Values match the known-valid row
    # used by the existing replace/update tests.
    base = {
        "issue_category": "Performance",
        "detail": "slow LCP",
        "affected_urls": "/page",
        "impact": "HIGH",
        "resolution": "lazy-load",
        "priority": "P1",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# (a) header lands at the schema header_row, not row 1
# ---------------------------------------------------------------------------

def test_create_sheet_writes_header_at_schema_header_row(tmp_path: Path) -> None:
    wb_path = _setup_project(tmp_path) / "master.xlsx"
    rows = [_tech_seo_row(detail=f"d{i}") for i in range(2)]
    transaction.append(wb_path, "tech_seo", rows, "test-proj")

    ws = load_workbook(wb_path)["tech_seo"]
    # P1-09(a): header at the declared header_row (3), and row 1 is NOT clobbered.
    assert ws.cell(row=3, column=1).value == "issue_category"
    assert ws.cell(row=1, column=1).value is None
    # ... so data lands at data_start_row (4), consistent with replace() + readers.
    assert ws.cell(row=4, column=2).value == "d0"
    assert ws.cell(row=5, column=2).value == "d1"


# ---------------------------------------------------------------------------
# (b) unknown row keys rejected by default; allow_extra=True restores skip
# ---------------------------------------------------------------------------

def test_unknown_row_key_rejected_by_default(tmp_path: Path) -> None:
    wb_path = _setup_project(tmp_path) / "master.xlsx"
    bad = _tech_seo_row()
    bad["typo_colum"] = "oops"  # not a schema column → likely a typo
    with pytest.raises(RowSchemaError):
        transaction.append(wb_path, "tech_seo", [bad], "test-proj")
    # The bad write must not have created the workbook (validation precedes lock).
    assert not wb_path.exists()


def test_unknown_row_key_allowed_with_allow_extra(tmp_path: Path) -> None:
    wb_path = _setup_project(tmp_path) / "master.xlsx"
    row = _tech_seo_row()
    row["extra_meta"] = "carry-along"  # deliberately not a column
    result = transaction.append(
        wb_path, "tech_seo", [row], "test-proj", allow_extra=True
    )
    assert result.rows_affected == 1
    ws = load_workbook(wb_path)["tech_seo"]
    # Real columns written at data_start_row; the extra key is skipped (no column).
    assert ws.cell(row=4, column=1).value == "Performance"
    header = [ws.cell(row=3, column=c).value for c in range(1, 7)]
    assert "extra_meta" not in header
