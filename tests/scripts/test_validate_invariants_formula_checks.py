"""tests/scripts/test_validate_invariants_formula_checks.py

F-02/F-03/F-04 ("dashboard holds no live =COUNTIF/=SUMIF/=AVERAGEIF formula")
had no test at all, and were reporting PASS over live formulas for as long as
they had existed. The shared workbook handle is opened ``data_only=True``, and
in that mode openpyxl never yields a formula string — it yields the cached
result, which is ``None`` for a freshly written formula. The old check sniffed
that handle for a leading ``'=' + token``, so it could match nothing.

The failure was invisible in the worst way: three of the five CRITICAL
invariants were green, and the only thing that ever made them fire was a
harmless TEXT cell that looked like a formula — the inverse of the stated rule.

These tests pin the fix in both directions. The one that matters most is
``test_skips_without_formula_view``: a check that cannot see its subject must
say SKIP, never PASS, because a PASS there is indistinguishable from a real
measurement and that is what let this survive.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from scripts.validation import validate_invariants as vi

_CASES = (
    ("F-02", vi.check_F_02, "=COUNTIF(A1:A2,1)"),
    ("F-03", vi.check_F_03, "=SUMIF(A1:A2,1)"),
    ("F-04", vi.check_F_04, "=AVERAGEIF(A1:A2,1)"),
)


def _dashboard(path: Path, cell_value: str | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "dashboard"
    ws["A1"] = 1
    ws["A2"] = 2
    if cell_value is not None:
        ws["A3"] = cell_value
    wb.save(str(path))
    wb.close()
    return path


def _verdict(path: Path, fn) -> dict:
    values = load_workbook(str(path), read_only=True, data_only=True)
    formulas = load_workbook(str(path), read_only=True, data_only=False)
    try:
        return fn(values, "testproj", formula_workbook=formulas)
    finally:
        values.close()
        formulas.close()


@pytest.mark.parametrize("rule_id,fn,formula", _CASES)
def test_catches_live_formula(tmp_path: Path, rule_id: str, fn, formula: str) -> None:
    """The regression itself: a live formula must turn the rule FAIL."""
    wb_path = _dashboard(tmp_path / f"{rule_id}.xlsx", formula)
    result = _verdict(wb_path, fn)
    assert result["verdict"] == "FAIL", (
        f"{rule_id} stayed {result['verdict']} over a live {formula} — the "
        f"formula view is not reaching the check: {result['evidence']}"
    )
    assert formula in result.get("sample_violations", []), (
        f"{rule_id} failed without naming the offending cell"
    )


@pytest.mark.parametrize("rule_id,fn,_formula", _CASES)
def test_passes_on_clean_dashboard(tmp_path: Path, rule_id: str, fn, _formula: str) -> None:
    """...and a clean sheet still passes, so the rule is not simply always red."""
    wb_path = _dashboard(tmp_path / f"{rule_id}-clean.xlsx")
    assert _verdict(wb_path, fn)["verdict"] == "PASS"


def test_each_rule_matches_only_its_own_token(tmp_path: Path) -> None:
    """Discrimination: one planted formula must trip exactly one rule.

    Without this, three rules that all fire on anything would look identical to
    three rules that each detect their own token.
    """
    for rule_id, _fn, formula in _CASES:
        wb_path = _dashboard(tmp_path / f"only-{rule_id}.xlsx", formula)
        failed = {
            rid for rid, fn, _ in _CASES
            if _verdict(wb_path, fn)["verdict"] == "FAIL"
        }
        assert failed == {rule_id}, (
            f"{formula} tripped {sorted(failed)}, expected only {rule_id}"
        )


@pytest.mark.parametrize("rule_id,fn,formula", _CASES)
def test_skips_without_formula_view(tmp_path: Path, rule_id: str, fn, formula: str) -> None:
    """No formula view ⇒ SKIP, never PASS.

    This is the test that would have caught the original defect. The rule cannot
    see a formula through a data_only=True handle, and reporting PASS there is a
    claim it has no basis for.
    """
    wb_path = _dashboard(tmp_path / f"{rule_id}-noview.xlsx", formula)
    values = load_workbook(str(wb_path), read_only=True, data_only=True)
    try:
        result = fn(values, "testproj")
    finally:
        values.close()
    assert result["verdict"] == "SKIP", (
        f"{rule_id} returned {result['verdict']} without a formula view — a "
        f"check that cannot see its subject has not measured it"
    )
    assert "NOT MEASURED" in result["evidence"]


def test_data_only_handle_really_hides_formulas(tmp_path: Path) -> None:
    """Pins the MECHANISM, so the fix cannot be undone by 'simplifying' it back.

    If someone later drops the second handle because 'one workbook is enough',
    this fails and says why.
    """
    wb_path = _dashboard(tmp_path / "mechanism.xlsx", "=AVERAGEIF(A1:A2,1)")
    values = load_workbook(str(wb_path), read_only=True, data_only=True)
    formulas = load_workbook(str(wb_path), read_only=True, data_only=False)
    try:
        cached = values["dashboard"]["A3"].value
        literal = formulas["dashboard"]["A3"].value
    finally:
        values.close()
        formulas.close()
    assert cached is None, (
        "data_only=True unexpectedly exposed a formula — if openpyxl changed "
        "this, F-02/03/04 may no longer need a second handle"
    )
    assert literal == "=AVERAGEIF(A1:A2,1)"
