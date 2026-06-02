#!/usr/bin/env python3
"""test_sf_import_wiring.py — scripts/ingestion/sf_import.py Step 6
(projection→replace) + Step 7 (sf_csv source-provenance) wiring (AC-10).

Distinct basename from tests/skills/test_sf_import.py (skill-level: frontmatter,
Tier validation, witness patterns) — pytest needs unique module basenames.

The headline guarantee (core AC): running ``project_and_write`` against a
real-layout master.xlsx is IDEMPOTENT — re-running refreshes the six SF sheets
in place rather than duplicating rows. This is the property ``transaction.replace``
provides and the reason ``sf_import`` projects via replace (not append).

Faithful integration: the workbook under test is a *copy* of the REAL workspace
master.xlsx
(``…/projects/aluminumstation-ca/master.xlsx``); the raw CSVs are the REAL
staged crawl (``…/sf-exports/2026-06-02/raw``). The real workbook is never
mutated — only the tmp copy is written.

Refs: scripts/ingestion/sf_import.py (project_and_write + main wiring),
scripts/ingestion/sf_projection.py (project_all), scripts/excel/transaction.py
(replace, idempotent), scripts/state/events_writer.py (append_provenance),
schemas/events.schema.json (source.kind=sf_csv), skills/ingestion/sf-import/
SKILL.md Step 6 + Step 7.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.excel import transaction
from scripts.ingestion import sf_import

# Real staged SF crawl + real workspace workbook (live fixtures).
_PROJECT_SLUG = "aluminumstation-ca"
_WS_PROJECT = Path(
    "/Users/apple/Documents/platinum-seo-workspace/projects/aluminumstation-ca"
)
_RAW_DIR = _WS_PROJECT / "sf-exports" / "2026-06-02" / "raw"
_REAL_MASTER = _WS_PROJECT / "master.xlsx"

# Live-proven per-sheet rowcounts (must match sf_projection.project_all output).
_EXPECTED_COUNTS = {
    "crawl_sitemap": 74,
    "redirect_404": 8,
    "schema": 2,
    "on_page_audit": 58,
    "tech_seo": 30,
    "robots_txt": 9,
}


# ---------------------------------------------------------------------------
# Fixtures — build a hermetic tmp workspace that mirrors the real layout, with
# a *copy* of the real master.xlsx at projects/{slug}/master.xlsx and a state
# dir under it. The real raw dir is reused read-only (projection never writes).
# ---------------------------------------------------------------------------

@pytest.fixture
def real_fixtures_available() -> bool:
    return _RAW_DIR.is_dir() and (_RAW_DIR / "internal_all.csv").exists() and _REAL_MASTER.exists()


@pytest.fixture
def tmp_ws(tmp_path: Path, real_fixtures_available: bool) -> dict:
    """Tmp workspace with a copy of the real master.xlsx and a symlinked raw dir.

    Returns ``{"workspace_root", "master_xlsx", "raw_dir", "events_path"}``.
    The raw dir is the REAL one (read-only consumption); the master.xlsx is a
    private tmp copy so the real workbook is never touched.
    """
    if not real_fixtures_available:
        pytest.skip(f"live SF fixtures not present at {_RAW_DIR} / {_REAL_MASTER}")
    ws = tmp_path / "workspace"
    proj = ws / "projects" / _PROJECT_SLUG
    proj.mkdir(parents=True, exist_ok=True)
    master = proj / "master.xlsx"
    shutil.copy2(_REAL_MASTER, master)
    # Symlink the real sf-exports dir under the tmp project so a
    # workspace-relative --sf-export-path resolves to a raw/ dir that lives
    # *under* the tmp workspace_root (mirrors production layout, no 17MB copy).
    sf_exports = proj / "sf-exports"
    sf_exports.symlink_to(_WS_PROJECT / "sf-exports")
    rel_export = Path("projects") / _PROJECT_SLUG / "sf-exports" / "2026-06-02"
    return {
        "workspace_root": ws,
        "master_xlsx": master,
        "raw_dir": _RAW_DIR,
        "rel_export_path": str(rel_export),
        "events_path": proj / "_state" / "events.jsonl",
    }


def _read_sheet_rowcount(master: Path, sheet: str) -> int:
    """Count non-empty data rows for a sheet, read from its schema data_start_row.

    Reads the schema's ``data_start_row`` via transaction's own resolver so the
    test honors the exact row model the writer uses (header block above data).
    """
    schema = transaction._load_schema(str(transaction._DEFAULT_SCHEMA_PATH))
    data_start = transaction._data_start_row(schema, sheet)
    wb = load_workbook(str(master))
    ws = wb[sheet]
    count = 0
    for r in range(data_start, (ws.max_row or 0) + 1):
        # A data row is "present" if its first column carries a value.
        if ws.cell(row=r, column=1).value not in (None, ""):
            count += 1
    return count


# ---------------------------------------------------------------------------
# Integration — project_and_write lands the live-proven rowcounts.
# ---------------------------------------------------------------------------

def test_project_and_write_lands_expected_rowcounts(tmp_ws: dict) -> None:
    counts = sf_import.project_and_write(
        tmp_ws["raw_dir"],
        tmp_ws["master_xlsx"],
        _PROJECT_SLUG,
        workspace_root=tmp_ws["workspace_root"],
    )
    # Return value matches the live-proven counts.
    assert counts == _EXPECTED_COUNTS

    # Reading the workbook back from each sheet's data_start_row matches too.
    for sheet, expected in _EXPECTED_COUNTS.items():
        observed = _read_sheet_rowcount(tmp_ws["master_xlsx"], sheet)
        assert observed == expected, f"{sheet}: read {observed}, expected {expected}"


# ---------------------------------------------------------------------------
# Idempotency — THE CORE AC. Two runs → identical rowcounts (no duplication).
# ---------------------------------------------------------------------------

def test_project_and_write_is_idempotent(tmp_ws: dict) -> None:
    counts1 = sf_import.project_and_write(
        tmp_ws["raw_dir"], tmp_ws["master_xlsx"], _PROJECT_SLUG,
        workspace_root=tmp_ws["workspace_root"],
    )
    rowcounts1 = {s: _read_sheet_rowcount(tmp_ws["master_xlsx"], s) for s in _EXPECTED_COUNTS}

    counts2 = sf_import.project_and_write(
        tmp_ws["raw_dir"], tmp_ws["master_xlsx"], _PROJECT_SLUG,
        workspace_root=tmp_ws["workspace_root"],
    )
    rowcounts2 = {s: _read_sheet_rowcount(tmp_ws["master_xlsx"], s) for s in _EXPECTED_COUNTS}

    assert counts1 == counts2 == _EXPECTED_COUNTS
    assert rowcounts1 == rowcounts2 == _EXPECTED_COUNTS


# ---------------------------------------------------------------------------
# Provenance — one sf_csv source event appended + schema-valid.
# ---------------------------------------------------------------------------

def test_project_and_write_emits_sf_csv_provenance(tmp_ws: dict) -> None:
    sf_import.project_and_write(
        tmp_ws["raw_dir"], tmp_ws["master_xlsx"], _PROJECT_SLUG,
        workspace_root=tmp_ws["workspace_root"],
    )
    events_path = tmp_ws["events_path"]
    assert events_path.exists(), "events.jsonl was not created"
    events = [
        json.loads(line)
        for line in events_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    sf_csv_events = [
        e for e in events
        if e.get("event_kind") == "provenance"
        and (e.get("source") or {}).get("kind") == "sf_csv"
    ]
    assert len(sf_csv_events) == 1, f"expected 1 sf_csv provenance event, got {len(sf_csv_events)}"

    ev = sf_csv_events[0]
    assert ev["operation"] == "ingest"
    assert ev["project_id"] == _PROJECT_SLUG
    assert ev["source"]["source_folder"]  # non-empty relative folder string
    assert ev["rows_written"] == sum(_EXPECTED_COUNTS.values())

    # Re-validate the event against events.schema.json (the writer validated it
    # on append; this confirms the emitted shape independently).
    from scripts.state import events_writer
    events_writer._validate_event(ev, events_writer._DEFAULT_SCHEMA_PATH)


def test_provenance_emitted_once_per_run_not_per_sheet(tmp_ws: dict) -> None:
    sf_import.project_and_write(
        tmp_ws["raw_dir"], tmp_ws["master_xlsx"], _PROJECT_SLUG,
        workspace_root=tmp_ws["workspace_root"],
    )
    events = [
        json.loads(line)
        for line in tmp_ws["events_path"].read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    sf_csv = [e for e in events if (e.get("source") or {}).get("kind") == "sf_csv"]
    # Exactly ONE sf_csv source event for the run (the 6 per-sheet replace()
    # calls each emit their own tool_computed event — those are separate).
    assert len(sf_csv) == 1
    tool_computed = [e for e in events if (e.get("source") or {}).get("kind") == "tool_computed"]
    assert len(tool_computed) == 6  # one per replaced sheet


# ---------------------------------------------------------------------------
# Error propagation — a RowSchemaError must surface, not be swallowed.
# ---------------------------------------------------------------------------

def test_project_and_write_propagates_schema_error(tmp_ws: dict, monkeypatch) -> None:
    """If a projection emits an invalid row, replace() raises RowSchemaError and
    project_and_write lets it propagate (DURUR — no swallow)."""
    def _bad_project_all(raw_dir, *, site_host=None):
        return {
            "crawl_sitemap": [{"category": "Crawl"}],  # missing required columns
        }
    monkeypatch.setattr(sf_import.sf_projection, "project_all", _bad_project_all)
    with pytest.raises(transaction.RowSchemaError):
        sf_import.project_and_write(
            tmp_ws["raw_dir"], tmp_ws["master_xlsx"], _PROJECT_SLUG,
            workspace_root=tmp_ws["workspace_root"],
        )


# ---------------------------------------------------------------------------
# main() CLI — dry-run returns 0 without writing; full run lands rows.
# ---------------------------------------------------------------------------

def test_main_dry_run_returns_zero_without_writing(tmp_ws: dict, capsys) -> None:
    # Before: every SF sheet's data block is whatever the real workbook shipped.
    before = {s: _read_sheet_rowcount(tmp_ws["master_xlsx"], s) for s in _EXPECTED_COUNTS}
    rc = sf_import.main([
        "--project", _PROJECT_SLUG,
        "--sf-export-path", tmp_ws["rel_export_path"],  # ws-relative; raw/ under it
        "--workspace-root", str(tmp_ws["workspace_root"]),
        "--dry-run",
    ])
    assert rc == 0
    out = capsys.readouterr().out
    assert "DRY-RUN" in out
    # No projection write happened — sheet rowcounts are unchanged.
    after = {s: _read_sheet_rowcount(tmp_ws["master_xlsx"], s) for s in _EXPECTED_COUNTS}
    assert after == before
    # And no events.jsonl provenance for this run.
    assert not tmp_ws["events_path"].exists() or not any(
        (json.loads(line).get("source") or {}).get("kind") == "sf_csv"
        for line in tmp_ws["events_path"].read_text(encoding="utf-8").splitlines()
        if line.strip()
    )


def test_main_full_run_projects_and_summarizes(tmp_ws: dict, capsys) -> None:
    rc = sf_import.main([
        "--project", _PROJECT_SLUG,
        "--sf-export-path", tmp_ws["rel_export_path"],
        "--workspace-root", str(tmp_ws["workspace_root"]),
    ])
    assert rc == 0
    out = capsys.readouterr().out
    # One OK summary line per sheet with its rowcount.
    for sheet, n in _EXPECTED_COUNTS.items():
        assert f"OK: {sheet}" in out
        assert str(n) in out
    # Sheets actually populated to the expected counts.
    for sheet, n in _EXPECTED_COUNTS.items():
        assert _read_sheet_rowcount(tmp_ws["master_xlsx"], sheet) == n
