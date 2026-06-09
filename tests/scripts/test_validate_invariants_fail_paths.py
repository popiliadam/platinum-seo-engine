"""tests/scripts/test_validate_invariants_fail_paths.py — FAIL-path regression lock.

Deterministic FAIL-verdict coverage for six consistency invariants whose
``check_F_NN`` functions previously had no dedicated test exercising their
FAIL branch. Each section builds the MINIMAL fixture that (a) satisfies the
precondition so the check does NOT short-circuit to SKIP, and (b) violates
the invariant so ``verdict == "FAIL"``. Where cheap, a paired clean fixture
proves the FAIL is meaningful (not always-FAIL).

Pattern mirrors:
- tests/scripts/test_validate_invariants_F16.py (openpyxl ``_build_*`` helper
  under tmp_path → ``check_F_NN(wb, project_slug=...)`` → plain asserts on
  verdict / affected_rows / sample_violations)
- tests/scripts/test_validate_invariants_F12.py (events.jsonl + _state state
  seeded under ``{ws}/projects/{slug}/_state/`` → ``check_F_NN(None, slug,
  workspace_root=ws)``)

Invariants covered (severity in parens):
- F-11 (HIGH)   — _state/workflows/*.json schema_version == "1.0"
- F-13 (HIGH)   — events.jsonl provenance.run_id is integer
- F-14 (HIGH)   — events.jsonl workflow.workflow_run_id pattern
- F-17 (HIGH)   — severity-typed cells ⊆ severityEnum {LOW,MEDIUM,HIGH,CRITICAL}
- F-20 (MEDIUM) — events.jsonl per-line size < 64 KiB
- F-21 (MEDIUM) — every cell value < 32767 chars (Excel hard limit)

Cross-references:
- scripts/validation/validate_invariants.py::check_F_11 / _13 / _14 / _17 / _20 / _21
- schemas/master-excel.schema.json (severityEnum-typed columns for F-17)
- schemas/events.schema.json (provenance run_id integer; workflow workflow_run_id)
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from scripts.validation.validate_invariants import (
    ConsistencyReportInvalidError,
    build_consistency_report,
    check_F_11,
    check_F_13,
    check_F_14,
    check_F_17,
    check_F_20,
    check_F_21,
)


# ---------------------------------------------------------------------------
# Shared workspace/state helpers ({ws}/projects/{slug}/_state/...)
# ---------------------------------------------------------------------------

def _state_dir(tmp_path: Path, slug: str) -> Path:
    """Create and return {ws}/projects/{slug}/_state/ (mkdir -p)."""
    d = tmp_path / "projects" / slug / "_state"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _seed_events(tmp_path: Path, slug: str, raw: str) -> Path:
    """Write verbatim ``raw`` text to {ws}/projects/{slug}/_state/events.jsonl."""
    events = _state_dir(tmp_path, slug) / "events.jsonl"
    events.write_text(raw, encoding="utf-8")
    return events


def _seed_events_lines(tmp_path: Path, slug: str, objs: list[dict]) -> Path:
    """Write one JSON object per line to events.jsonl."""
    raw = "".join(json.dumps(o) + "\n" for o in objs)
    return _seed_events(tmp_path, slug, raw)


def _load_ro(wb_path: Path) -> openpyxl.Workbook:
    """Re-open a saved workbook read-only + data_only (matches drift-check load)."""
    return openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)


# ===========================================================================
# F-11 — _state/workflows/*.json schema_version == "1.0"
#
# Precondition to evaluate (else SKIP): _state/workflows/ exists AND contains
# >= 1 *.json file. FAIL condition: at least one file has schema_version != "1.0".
# ===========================================================================

def _seed_workflow(tmp_path: Path, slug: str, name: str, schema_version: object) -> Path:
    wf_dir = _state_dir(tmp_path, slug) / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    p = wf_dir / name
    p.write_text(json.dumps({"schema_version": schema_version}), encoding="utf-8")
    return p


def test_F11_wrong_schema_version_fails(tmp_path: Path) -> None:
    """A workflow run file with schema_version != "1.0" → FAIL."""
    _seed_workflow(tmp_path, "f11", "run-a.json", "0.9")
    result = check_F_11(None, "f11", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert "run-a.json" in result["sample_violations"][0]
    assert "0.9" in result["sample_violations"][0]


def test_F11_unreadable_json_fails(tmp_path: Path) -> None:
    """A corrupt (non-JSON) workflow file → FAIL (counted as unreadable)."""
    wf_dir = _state_dir(tmp_path, "f11bad") / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    (wf_dir / "broken.json").write_text("{not valid json", encoding="utf-8")
    result = check_F_11(None, "f11bad", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert "broken.json: unreadable" in result["sample_violations"]


def test_F11_correct_schema_version_passes(tmp_path: Path) -> None:
    """All workflow files at schema_version "1.0" → PASS (FAIL is meaningful)."""
    _seed_workflow(tmp_path, "f11ok", "run-a.json", "1.0")
    _seed_workflow(tmp_path, "f11ok", "run-b.json", "1.0")
    result = check_F_11(None, "f11ok", workspace_root=tmp_path)
    assert result["verdict"] == "PASS", result


# ===========================================================================
# F-13 — events.jsonl provenance.run_id is integer
#
# Precondition to evaluate (else SKIP): events.jsonl exists. FAIL condition:
# a row with event_kind=="provenance" whose run_id is not an int (bool counts
# as non-int per the explicit isinstance(rid, bool) guard).
# ===========================================================================

def test_F13_string_run_id_fails(tmp_path: Path) -> None:
    """A provenance event with a string run_id → FAIL."""
    _seed_events_lines(
        tmp_path, "f13",
        [{"event_kind": "provenance", "event_id": "p1", "run_id": "NaN"}],
    )
    result = check_F_13(None, "f13", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert "p1" in result["sample_violations"][0]
    assert "'NaN'" in result["sample_violations"][0]


def test_F13_bool_run_id_fails(tmp_path: Path) -> None:
    """run_id=True is a bool (not a real int) → FAIL via the explicit guard."""
    _seed_events_lines(
        tmp_path, "f13bool",
        [{"event_kind": "provenance", "event_id": "p2", "run_id": True}],
    )
    result = check_F_13(None, "f13bool", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert "run_id=True" in result["sample_violations"][0]


def test_F13_integer_run_id_passes(tmp_path: Path) -> None:
    """A provenance event with an integer run_id → PASS (FAIL is meaningful)."""
    _seed_events_lines(
        tmp_path, "f13ok",
        [{"event_kind": "provenance", "event_id": "p1", "run_id": 7}],
    )
    result = check_F_13(None, "f13ok", workspace_root=tmp_path)
    assert result["verdict"] == "PASS", result


# ===========================================================================
# F-14 — events.jsonl workflow.workflow_run_id pattern
#
# Pattern: ^[a-z][a-z0-9-]*-\d{4}-\d{2}-\d{2}-[a-f0-9]{4}$
# Precondition to evaluate (else SKIP): events.jsonl exists. FAIL condition:
# a row with event_kind=="workflow" whose workflow_run_id is missing/non-str
# or does not match the {slug}-YYYY-MM-DD-{hash4} pattern.
# ===========================================================================

def test_F14_malformed_workflow_run_id_fails(tmp_path: Path) -> None:
    """A workflow event with a malformed workflow_run_id → FAIL."""
    _seed_events_lines(
        tmp_path, "f14",
        [{"event_kind": "workflow", "event_id": "w1",
          "workflow_run_id": "BAD ID with spaces"}],
    )
    result = check_F_14(None, "f14", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert "w1" in result["sample_violations"][0]


def test_F14_missing_workflow_run_id_fails(tmp_path: Path) -> None:
    """workflow_run_id absent (None) on a workflow event → FAIL (non-str)."""
    _seed_events_lines(
        tmp_path, "f14none",
        [{"event_kind": "workflow", "event_id": "w2"}],
    )
    result = check_F_14(None, "f14none", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert "workflow_run_id=None" in result["sample_violations"][0]


def test_F14_wellformed_workflow_run_id_passes(tmp_path: Path) -> None:
    """A canonical {slug}-YYYY-MM-DD-{hash4} workflow_run_id → PASS."""
    _seed_events_lines(
        tmp_path, "f14ok",
        [{"event_kind": "workflow", "event_id": "w1",
          "workflow_run_id": "myproj-2026-06-04-ab12"}],
    )
    result = check_F_14(None, "f14ok", workspace_root=tmp_path)
    assert result["verdict"] == "PASS", result


# ===========================================================================
# F-17 — severity-typed cells ⊆ severityEnum {LOW,MEDIUM,HIGH,CRITICAL}
#
# Severity-typed columns are discovered from master-excel.schema.json (any
# required_columns entry whose ref ends with "/severityEnum"); master_task.priority
# is one such column. Precondition to evaluate (else SKIP): >= 1 non-empty
# severity-typed cell present. FAIL condition: a value whose upper() is not in
# the 4-value enum.
#
# Header is placed at row 1: _resolve_header_row probes the schema header_row
# (master_task=3), finds it empty, and falls back to row 1 — so a 2-column
# row-1 header is read correctly by _iter_rows_as_dicts.
# ===========================================================================

def _build_master_task_priority_wb(tmp_path: Path, priority: object) -> openpyxl.Workbook:
    """Minimal workbook: master_task sheet, header at row 1, one data row whose
    priority cell carries ``priority``."""
    wb_path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    mt = wb.create_sheet("master_task")
    mt.append(["task_id", "priority"])
    mt.append(["T-1", priority])
    wb.save(str(wb_path))
    return _load_ro(wb_path)


def test_F17_invalid_severity_value_fails(tmp_path: Path) -> None:
    """A priority cell outside severityEnum → FAIL."""
    wb = _build_master_task_priority_wb(tmp_path, "SUPERHIGH")
    try:
        result = check_F_17(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert "master_task" in result["affected_sheets"]
    assert "master_task.priority='SUPERHIGH'" in result["sample_violations"]


def test_F17_valid_severity_value_passes(tmp_path: Path) -> None:
    """A priority cell within severityEnum → PASS (FAIL is meaningful).

    Note: the check upper()s before comparison, so "HIGH" is in-enum.
    """
    wb = _build_master_task_priority_wb(tmp_path, "HIGH")
    try:
        result = check_F_17(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "PASS", result


# ===========================================================================
# F-20 — events.jsonl per-line size < 64 KiB (_EVENTS_MAX_LINE_BYTES)
#
# Precondition to evaluate (else SKIP): events.jsonl exists. FAIL condition:
# a non-blank line whose raw byte length exceeds 64 KiB (65536).
# ===========================================================================

def test_F20_oversized_line_fails(tmp_path: Path) -> None:
    """An events.jsonl line over the 64 KiB cap → FAIL."""
    oversized = json.dumps(
        {"event_kind": "audit", "event_id": "a1", "blob": "x" * 70_000}
    )
    # Sanity: the fixture is genuinely over the cap this check enforces.
    assert len((oversized + "\n").encode("utf-8")) > 65_536
    _seed_events(tmp_path, "f20", oversized + "\n")
    result = check_F_20(None, "f20", workspace_root=tmp_path)
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert result["sample_violations"][0].startswith("line 1:")


def test_F20_within_cap_passes(tmp_path: Path) -> None:
    """A small events.jsonl line under the cap → PASS (FAIL is meaningful)."""
    small = json.dumps({"event_kind": "audit", "event_id": "a1"})
    _seed_events(tmp_path, "f20ok", small + "\n")
    result = check_F_20(None, "f20ok", workspace_root=tmp_path)
    assert result["verdict"] == "PASS", result


# ===========================================================================
# F-21 — every cell value < 32767 chars (_EXCEL_CELL_MAX_CHARS, Excel hard limit)
#
# No precondition gate (any workbook is scanned). FAIL condition: a string cell
# whose length is >= 32767. openpyxl permits writing exactly 32767 chars (the
# boundary), and the check uses ``>=`` — so a 32767-char cell triggers FAIL.
# ===========================================================================

def _build_single_cell_wb(tmp_path: Path, text: str) -> openpyxl.Workbook:
    wb_path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("sheet1")
    ws["A1"] = text
    wb.save(str(wb_path))
    return _load_ro(wb_path)


def test_F21_cell_at_excel_limit_fails(tmp_path: Path) -> None:
    """A cell exactly at the 32767-char Excel limit (>= check) → FAIL."""
    wb = _build_single_cell_wb(tmp_path, "x" * 32_767)
    try:
        result = check_F_21(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "FAIL", result
    assert result["affected_rows"] == 1
    assert "sheet1: 32767 chars" in result["sample_violations"]


def test_F21_short_cell_passes(tmp_path: Path) -> None:
    """A short cell well under the limit → PASS (FAIL is meaningful)."""
    wb = _build_single_cell_wb(tmp_path, "x" * 100)
    try:
        result = check_F_21(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "PASS", result


# ---------------------------------------------------------------------------
# #19 — consistency-report generated_at is strict UTC '…Z'
# ---------------------------------------------------------------------------

def test_build_consistency_report_rejects_naive_generated_at() -> None:
    """generated_at carries format:date-time in consistency-report.schema.json.
    Routing the report validator through the strict build_validator (finding #19)
    makes build_consistency_report reject a naive (tz-less) generated_at — the UTC
    '…Z' discipline is ENFORCED, not annotated. The rest of the report is valid,
    so the failure is unambiguously the timestamp."""
    results = [{"id": "CI-01", "verdict": "PASS"}]
    agg = {"overall": "GREEN", "total": 1, "pass_count": 1, "warn_count": 0, "fail_count": 0}
    # canonical '…Z' builds a valid report (control)…
    ok = build_consistency_report(
        rule_results=results, aggregation=agg, project_id="demo", report_id=1,
        generated_at="2026-06-08T12:00:00Z", run_id=1,
    )
    assert ok["generated_at"].endswith("Z")
    # …a naive stamp is rejected on its date-time format.
    with pytest.raises(ConsistencyReportInvalidError, match="generated_at"):
        build_consistency_report(
            rule_results=results, aggregation=agg, project_id="demo", report_id=1,
            generated_at="2026-06-08T12:00:00", run_id=1,  # naive — no 'Z'
        )
