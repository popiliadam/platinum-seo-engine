"""tests/scripts/test_validate_invariants_F16.py — Q-V1.2-OPP-COVERAGE-01 fix.

F-16 invariant defensive parsing regression-lock (Phase B post-closeout
schema-first override 17'inci uygulama doğum belgesi):

- Canonical format `"url | action"` (Phase 8 quickwins_transform.py
  paterni) → PASS expected (backward-compat with existing
  test_drift_check.py fixtures)
- Freeform format `"Optimize <url> for query 'X'"` (workspace gerçek
  master.xlsx 2026-05-06 state, manual/non-canonical drift) → PASS
  expected (Q-V1.2-OPP-COVERAGE-01 RESOLVED via _extract_url_from_action_field)
- Helper unit test ensures extraction priority (canonical > freeform
  fallback regex)

Cross-references:
- scripts/validation/validate_invariants.py::_extract_url_from_action_field
- scripts/validation/validate_invariants.py::check_F_16
- scripts/discovery/quickwins_transform.py:467 — canonical writer
- docs/OPEN_QUESTIONS.md Q-V1.2-OPP-COVERAGE-01
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from scripts.validation.validate_invariants import (
    _extract_url_from_action_field,
    check_F_16,
)


# ---------------------------------------------------------------------------
# Helper unit tests — _extract_url_from_action_field
# ---------------------------------------------------------------------------

def test_extract_url_canonical_format() -> None:
    """Canonical 'url | action' pattern → URL prefix returned."""
    text = "https://example.com/foo | optimize for query bar"
    assert _extract_url_from_action_field(text) == "https://example.com/foo"


def test_extract_url_freeform_format() -> None:
    """Freeform 'Optimize <url> for X' pattern → URL extracted via regex."""
    text = "Optimize https://dentnotion.com/gece-plagi-splint-nedir-nasil-kullanilir/ for query 'splint nedir'"
    assert _extract_url_from_action_field(text) == "https://dentnotion.com/gece-plagi-splint-nedir-nasil-kullanilir/"


def test_extract_url_freeform_with_pipe_in_action() -> None:
    """Pipe in action text but URL is freeform-embedded → falls back to regex."""
    text = "Note: redirect needed | https://example.com/target via canonical"
    # Canonical split returns "Note: redirect needed" (not URL prefix), so falls back to regex.
    assert _extract_url_from_action_field(text) == "https://example.com/target"


def test_extract_url_no_url_in_text() -> None:
    """No URL anywhere → empty string."""
    assert _extract_url_from_action_field("just a description without url") == ""


def test_extract_url_empty_or_none() -> None:
    """Empty/None → empty string (defensive)."""
    assert _extract_url_from_action_field("") == ""
    assert _extract_url_from_action_field(None) == ""  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# F-16 integration tests — both formats
# ---------------------------------------------------------------------------

def _build_workbook_with_opp_format(tmp_path: Path, action_text: str) -> openpyxl.Workbook:
    """Helper: minimal workbook with quick_wins + opportunity sheets."""
    wb_path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    qw = wb.create_sheet("quick_wins")
    qw.append(["query", "url", "current_position", "impressions_30d",
               "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
               "action", "priority"])
    qw.append(["foo", "https://example.com/page", 12, 100, 5, 5.0, 10,
               "High", "x", "MEDIUM"])
    op = wb.create_sheet("opportunity")
    op.append(["query", "opportunity_score", "current_position", "ctr_pct",
               "impressions_30d", "clicks_30d", "potential_clicks",
               "assigned_url_action"])
    op.append(["foo", 800, 12, 5.0, 100, 5, 10, action_text])
    wb.save(str(wb_path))
    return openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)


def test_F16_canonical_format_passes(tmp_path: Path) -> None:
    """Legacy canonical 'url | action' format → PASS (backward-compat)."""
    wb = _build_workbook_with_opp_format(
        tmp_path,
        "https://example.com/page | optimize title",
    )
    try:
        result = check_F_16(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "PASS", (
        f"Canonical format must PASS (backward-compat); evidence={result.get('evidence')}"
    )


def test_F16_freeform_format_passes(tmp_path: Path) -> None:
    """Freeform 'Optimize <url> for query X' format → PASS (Q-V1.2-OPP-COVERAGE-01)."""
    wb = _build_workbook_with_opp_format(
        tmp_path,
        "Optimize https://example.com/page for query 'foo'",
    )
    try:
        result = check_F_16(wb, project_slug="test")
    finally:
        wb.close()
    assert result["verdict"] == "PASS", (
        f"Freeform format must PASS post Q-V1.2-OPP-COVERAGE-01 fix; "
        f"evidence={result.get('evidence')}"
    )


def test_F16_genuine_orphan_still_fails(tmp_path: Path) -> None:
    """Genuine orphan (qw url not in opportunity at all) → FAIL (defensive parsing
    must not mask real drift)."""
    wb_path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    qw = wb.create_sheet("quick_wins")
    qw.append(["query", "url", "current_position", "impressions_30d",
               "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
               "action", "priority"])
    qw.append(["orphan_query", "https://example.com/orphan-page", 12, 100, 5, 5.0, 10,
               "High", "x", "MEDIUM"])
    op = wb.create_sheet("opportunity")
    op.append(["query", "opportunity_score", "current_position", "ctr_pct",
               "impressions_30d", "clicks_30d", "potential_clicks",
               "assigned_url_action"])
    # Different URL — genuine orphan
    op.append(["other_query", 800, 12, 5.0, 100, 5, 10,
               "Optimize https://example.com/other-page for query 'other_query'"])
    wb.save(str(wb_path))
    wb_ro = openpyxl.load_workbook(str(wb_path), read_only=True, data_only=True)
    try:
        result = check_F_16(wb_ro, project_slug="test")
    finally:
        wb_ro.close()
    assert result["verdict"] == "FAIL"
    assert result["affected_rows"] == 1
    assert "https://example.com/orphan-page" in result["sample_violations"]
