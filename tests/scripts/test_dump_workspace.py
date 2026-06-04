"""tests/scripts/test_dump_workspace.py — Y-07 v1.5-Phase-3 Tier 2
Manager session state summary tek komut (scripts/state/dump_workspace.py).

Aggregates the 5 data sources a manager session needs to triage state at a
glance:

  1. _state/events.jsonl son 5 satır (parsed JSON)
  2. outputs/master.xlsx#master_task TODO satır sayısı
  3. _state/workflows/<run_id>.json içinde status="awaiting_approval" olanlar
  4. _state/backups/master-excel-<TIMESTAMP>.xlsx son 7 (mtime-sorted)
  5. _state/consistency-report.json verdict (GREEN/AMBER/RED)

CLI semantics:
  - --project <slug> opsiyonel; eğer atlanırsa shared/active.json'daki bound
    slug kullanılır. Active.json yoksa raise.
  - All data sources are graceful: missing path -> None / empty list (manager
    session aware'liği için raise YERINE).

Authority: rules/append-only-state.md, ADR-021, ADR-035 PSEO_WORKSPACE_ROOT,
  schemas/master-excel.schema.json (master_task statusEnum),
  schemas/consistency-report.schema.json (verdict enum),
  scripts/state/workflow_runner.py (status="awaiting_approval" semantics).
Brief: docs/superpowers/plans/v1.5-audit-closure-brief.md Phase 3 Tier 2.
"""

from __future__ import annotations

import importlib.util
import json
import os
import time
from pathlib import Path

import pytest
from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[2]
SCRIPT_PATH = REPO / "scripts" / "state" / "dump_workspace.py"


@pytest.fixture(scope="module")
def dump_module():
    spec = importlib.util.spec_from_file_location("dump_workspace", SCRIPT_PATH)
    assert spec is not None and spec.loader is not None, (
        f"scripts/state/dump_workspace.py not loadable: {SCRIPT_PATH}"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _make_workspace(root: Path, slug: str = "demo") -> Path:
    """Create a minimal ADR-035 workspace tree at root and return project dir."""
    project = root / "projects" / slug
    (project / "_state" / "workflows").mkdir(parents=True)
    (project / "_state" / "backups").mkdir(parents=True)
    (project / "outputs").mkdir(parents=True)
    return project


def _write_events(project: Path, events: list[dict]) -> None:
    body = "\n".join(json.dumps(e) for e in events) + ("\n" if events else "")
    (project / "_state" / "events.jsonl").write_text(body, encoding="utf-8")


def _write_master_xlsx(project: Path, statuses: list[str]) -> None:
    """Create master.xlsx at the project ROOT (the live convention — every
    workspace project keeps master.xlsx at projects/<slug>/master.xlsx, and
    init-project writes it there) with a master_task sheet whose column J holds
    statusEnum values (header row = 1)."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("master_task")
    ws.cell(row=1, column=1, value="task_id")
    ws.cell(row=1, column=10, value="status")
    for i, status in enumerate(statuses, start=2):
        ws.cell(row=i, column=1, value=f"task-{i}")
        ws.cell(row=i, column=10, value=status)
    wb.save(project / "master.xlsx")


def _write_workflow(project: Path, run_id: str, status: str) -> None:
    (project / "_state" / "workflows" / f"{run_id}.json").write_text(
        json.dumps({"run_id": run_id, "status": status}), encoding="utf-8"
    )


def _write_backups(project: Path, n: int = 8) -> list[str]:
    """Create n backup files with monotonically increasing mtime; return
    filenames in chronological order (oldest first)."""
    backups_dir = project / "_state" / "backups"
    names: list[str] = []
    base_mtime = time.time() - 60 * 60 * 24 * n  # n days ago
    for i in range(n):
        ts = f"2026-04-{(i % 28) + 1:02d}T00-00-00Z"
        name = f"master-excel-{ts}.xlsx"
        path = backups_dir / name
        path.write_bytes(b"\x50\x4b\x03\x04")  # tiny PK ZIP magic
        mtime = base_mtime + i * 60
        os.utime(path, (mtime, mtime))
        names.append(name)
    return names


def _write_consistency_report(project: Path, verdict: str) -> None:
    (project / "_state" / "consistency-report.json").write_text(
        json.dumps(
            {
                "schema_version": "1.0",
                "report_id": 1,
                "generated_at": "2026-05-07T00:00:00Z",
                "project_id": project.name,
                "verdict": verdict,
                "checks": [{"check_id": "F-01", "category": "CSR", "verdict": "PASS"}],
            }
        ),
        encoding="utf-8",
    )


def _write_active_json(workspace_root: Path, slug: str) -> None:
    shared = workspace_root / "shared"
    shared.mkdir(parents=True, exist_ok=True)
    (shared / "active.json").write_text(
        json.dumps({"active_project": slug}), encoding="utf-8"
    )


# ---------- 1. explicit slug returns summary ----------


def test_dump_with_explicit_slug_returns_summary_dict(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["project"] == "demo"
    assert "events_tail" in result
    assert "master_task_todo_count" in result
    assert "workflow_pending_approvals" in result
    assert "backups_recent" in result
    assert "drift_verdict" in result


# ---------- 2. uses active.json when slug=None ----------


def test_dump_uses_active_json_when_slug_none(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")
    _write_active_json(tmp_path, "demo")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug=None)

    assert result["project"] == "demo"


# ---------- 2a. canonical active_project key resolves (P0-02) ----------


def test_dump_reads_active_project_canonical(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")
    (tmp_path / "shared").mkdir(parents=True, exist_ok=True)
    (tmp_path / "shared" / "active.json").write_text(
        json.dumps({"active_project": "demo", "updated_at": "2026-06-03T00:00:00Z"}),
        encoding="utf-8",
    )

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug=None)

    assert result["project"] == "demo"


# ---------- 2b. legacy slug key still resolves (P0-02 back-compat) ----------


def test_dump_legacy_slug_still_resolves(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")
    (tmp_path / "shared").mkdir(parents=True, exist_ok=True)
    (tmp_path / "shared" / "active.json").write_text(
        json.dumps({"slug": "demo"}), encoding="utf-8"
    )

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug=None)

    assert result["project"] == "demo"


# ---------- 3. raises when slug=None and no active.json ----------


def test_dump_raises_when_slug_none_and_active_json_missing(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")

    with pytest.raises(FileNotFoundError, match=r"active\.json|bound"):
        dump_module.dump_workspace(workspace_root=tmp_path, project_slug=None)


# ---------- 4. events.jsonl tail returns last 5 ----------


def test_events_tail_returns_last_5_lines(dump_module, tmp_path):
    project = _make_workspace(tmp_path, slug="demo")
    events = [{"event_id": f"evt-{i:02d}", "n": i} for i in range(10)]
    _write_events(project, events)

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    tail = result["events_tail"]
    assert len(tail) == 5
    assert tail[0]["n"] == 5
    assert tail[-1]["n"] == 9


# ---------- 5. master_task TODO count ----------


def test_master_task_todo_count_from_xlsx(dump_module, tmp_path):
    project = _make_workspace(tmp_path, slug="demo")
    _write_master_xlsx(
        project,
        statuses=["TODO", "DONE", "TODO", "ONGOING", "TODO", "BLOCKED"],
    )

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["master_task_todo_count"] == 3


# ---------- 6. master_task None when xlsx missing ----------


def test_master_task_todo_count_none_when_xlsx_missing(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")
    # outputs/master.xlsx deliberately not created

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["master_task_todo_count"] is None


# ---------- 7. pending approvals filtered ----------


def test_pending_approvals_only_awaiting_approval_status(dump_module, tmp_path):
    project = _make_workspace(tmp_path, slug="demo")
    _write_workflow(project, "wf-1", "running")
    _write_workflow(project, "wf-2", "awaiting_approval")
    _write_workflow(project, "wf-3", "done")
    _write_workflow(project, "wf-4", "awaiting_approval")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    pending = result["workflow_pending_approvals"]
    assert sorted(pending) == ["wf-2", "wf-4"]


# ---------- 8. pending empty when no workflows ----------


def test_pending_approvals_empty_when_no_workflows(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["workflow_pending_approvals"] == []


# ---------- 9. backups recent: last 7 by mtime ----------


def test_backups_recent_returns_last_7_by_mtime(dump_module, tmp_path):
    project = _make_workspace(tmp_path, slug="demo")
    names = _write_backups(project, n=8)  # 8 backups, last 7 should win

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    recent = result["backups_recent"]
    assert len(recent) == 7
    # Ordered most-recent first.
    assert recent[0] == names[-1]
    assert recent[-1] == names[1]  # oldest of the 8 (names[0]) is dropped


# ---------- 10. backups empty when dir missing ----------


def test_backups_recent_empty_when_dir_missing(dump_module, tmp_path):
    project = tmp_path / "projects" / "demo"
    (project / "_state").mkdir(parents=True)
    (project / "outputs").mkdir(parents=True)
    # backups dir deliberately not created

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["backups_recent"] == []


# ---------- 11. drift verdict from consistency-report ----------


def test_drift_verdict_from_consistency_report(dump_module, tmp_path):
    project = _make_workspace(tmp_path, slug="demo")
    _write_consistency_report(project, "AMBER")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["drift_verdict"] == "AMBER"


# ---------- 12. drift verdict None when report missing ----------


def test_drift_verdict_live_from_root_workbook_when_no_report(dump_module, tmp_path):
    """No consistency-report.json is ever persisted today (the live case), so
    drift_verdict must be computed LIVE from the root master.xlsx instead of
    returning null. A master_task with the wrong column count -> F-05 CRITICAL
    -> overall RED."""
    project = _make_workspace(tmp_path, slug="demo")
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("master_task")
    ws.append(["task_id", "status", "foo"])  # wrong column count -> F-05 RED
    ws.append(["t-1", "TODO", "x"])
    wb.save(project / "master.xlsx")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")
    assert result["drift_verdict"] == "RED"


def test_drift_verdict_none_when_report_missing(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")

    result = dump_module.dump_workspace(workspace_root=tmp_path, project_slug="demo")

    assert result["drift_verdict"] is None


# ---------- 13. invalid workspace raises ----------


def test_invalid_workspace_raises(dump_module, tmp_path):
    nonexistent = tmp_path / "not-a-workspace"

    with pytest.raises(FileNotFoundError, match=r"workspace|projects"):
        dump_module.dump_workspace(
            workspace_root=nonexistent, project_slug="demo"
        )


# ---------- 14. explicit slug not found raises ----------


def test_explicit_slug_not_found_raises(dump_module, tmp_path):
    _make_workspace(tmp_path, slug="demo")

    with pytest.raises(FileNotFoundError, match=r"unknown|project|slug"):
        dump_module.dump_workspace(
            workspace_root=tmp_path, project_slug="never-exists"
        )


# ---------- 15. main entry for CLI ----------


def test_module_exposes_main_for_cli(dump_module):
    assert hasattr(dump_module, "main")
    assert callable(dump_module.main)
