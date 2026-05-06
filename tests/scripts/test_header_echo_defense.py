"""tests/scripts/test_header_echo_defense.py — validator header echo defense.

ADR-037 codifies the existing Phase 14 W3-W2-C-a defense as a regression
gate.  When a workbook carries a duplicate-header artifact (transaction.py
writes the header at row 1 + bootstrap_excel.py writes it at the schema's
``header_row`` of 3 / 4 / 5), ``_resolve_header_row`` MUST resolve to the
schema-declared row by virtue of the 50%+ column-name overlap probe — it
MUST NOT fall back to row 1 and treat the schema header as data.

The synthetic fixtures below reproduce the dentnotion shape:

  - ``master_task``: header at row 1 (orphan), row 2 empty, row 3 header
    (canonical schema), data at row 4+.
  - ``quick_wins``: header at row 1 (orphan), rows 2-3 empty, row 4 header
    (canonical schema), data at row 5+.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[2]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from scripts.validation.validate_invariants import (  # noqa: E402
    _iter_rows_as_dicts,
    _resolve_header_row,
)


def _build_master_task_with_dup_header(tmp_path: Path) -> Path:
    """Build a master_task sheet that has a row-1 orphan header AND a
    row-3 schema-canonical header (mirrors dentnotion layout)."""
    schema_path = REPO / "schemas" / "master-excel.schema.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    mt_def = schema["sheets"]["master_task"]
    columns = mt_def["required_columns"]
    header_row = mt_def["header_row"]  # 3 for master_task
    data_start = mt_def["data_start_row"]  # 4

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("master_task")
    col_map = {c["name"]: c["col"] for c in columns}

    # Row 1: orphan header (transaction.py:_ensure_sheet_with_header).
    for col in columns:
        ws[f"{col['col']}1"] = col["name"]
    # Row 2: blank (None) — nothing to do.
    # Row 3: canonical header (bootstrap_excel.write_headers).
    for col in columns:
        ws[f"{col['col']}{header_row}"] = col["name"]
    # Row 4+: real data.
    ws[f"{col_map['task_id']}{data_start}"] = "T-30001"
    ws[f"{col_map['task']}{data_start}"] = "real task"
    ws[f"{col_map['priority']}{data_start}"] = "HIGH"
    ws[f"{col_map['status']}{data_start}"] = "TODO"
    ws[f"{col_map['url']}{data_start}"] = "https://example.com/"

    out = tmp_path / "master.xlsx"
    wb.save(out)
    return out


def test_resolve_header_row_picks_schema_row_under_dup_header(tmp_path: Path) -> None:
    workbook_path = _build_master_task_with_dup_header(tmp_path)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        resolved = _resolve_header_row(wb, "master_task")
    finally:
        wb.close()
    assert resolved == 3, (
        f"Expected schema header_row=3 (canonical) under a row-1 orphan-header "
        f"workbook; got {resolved}.  Header echo defense regression."
    )


def test_iter_rows_skips_dup_header_and_reads_real_data(tmp_path: Path) -> None:
    workbook_path = _build_master_task_with_dup_header(tmp_path)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = _iter_rows_as_dicts(wb, "master_task")
    finally:
        wb.close()
    # Validator iterates from header_row + 1 = 4.  Should get exactly 1 real
    # row and skip the row-1 orphan header.
    assert len(rows) == 1, f"Expected 1 data row, got {len(rows)}: {rows}"
    only = rows[0]
    assert only["task_id"] == "T-30001"
    assert only["priority"] == "HIGH"
    # Critically: the column-name strings ("task", "priority") MUST NOT
    # appear as data values.  This proves the dup-header was skipped.
    assert only.get("task") == "real task"
    assert only.get("priority") != "priority"


def test_dup_header_does_not_pollute_severity_check(tmp_path: Path) -> None:
    """If the dup-header leaked through, severity check would see literal
    'priority' in priority cells — proving the orphan header escaped.
    Lock the negative case explicitly."""
    workbook_path = _build_master_task_with_dup_header(tmp_path)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        rows = _iter_rows_as_dicts(wb, "master_task")
    finally:
        wb.close()
    bad = [r for r in rows if str(r.get("priority", "")).lower() == "priority"]
    assert not bad, f"Header echo defense breach — got {bad!r}"
