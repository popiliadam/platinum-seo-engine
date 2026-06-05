"""
Unit tests for scripts/excel/transaction.py.

Covers: happy-path writes, schema validation, formula policy, atomic-save
fault-injection, backup rotation, lock contention, status/severity enum
guards (ADR-018 hibrit), provenance event emission, plugin-agnostic
source check.
"""

from __future__ import annotations

import json
import multiprocessing as mp
import os
import time
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from scripts.excel import transaction
from scripts.excel.transaction import (
    CellValueTooLongError,
    FormulaPolicyViolation,
    LockHeldError,
    RowSchemaError,
    SchemaSheetMismatchError,
    WriteResult,
    WriterScopeError,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _setup_project(tmp_path: Path, slug: str = "test-proj") -> tuple[Path, Path]:
    """Create projects/{slug}/_state/ tree; return (project_dir, state_dir)."""
    proj = tmp_path / "projects" / slug
    state = proj / "_state"
    state.mkdir(parents=True, exist_ok=True)
    return proj, state


def _topical_row(**overrides) -> dict:
    base = {
        "pillar": "P01_sofa_sets",
        "cluster": "leather-sofas",
        "primary_keyword": "leather sofa",
        "monthly_volume": 4400,
        "data_source": "gsc",
        "assigned_url": "/leather-sofas",
        "page_type": "pillar",
        "status": "TODO",
        "priority": "HIGH",
        "note": "",
    }
    base.update(overrides)
    return base


def _master_task_row(**overrides) -> dict:
    base = {
        "task_id": "T-0001",
        "task": "fix slow page",
        "primary_source": "tech_fix",
        "related_sources": "tech_fix",
        "url": "/slow-page",
        "category": "performance",
        "priority": "HIGH",
        "impact": "speed",
        "duration_est_min": 30,
        "status": "TODO",
        "created_date": "2026-04-30",
        "done_date": None,
        "assignee": "ops",
        "note": "",
        "auto_generated": False,
        "dependencies": "",
        "effort_actual_min": None,
        "metric_impact_json": "",
        "work_log_ref": "",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# Test 1 — happy path: write topical_map rows, verify content + event
# ---------------------------------------------------------------------------

def test_happy_path_write_topical_map(tmp_path: Path) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    rows = [_topical_row(primary_keyword=f"kw {i}") for i in range(3)]
    result = transaction.append(wb_path, "topical_map", rows, "test-proj")

    assert isinstance(result, WriteResult)
    assert result.rows_affected == 3
    assert result.backup_path.exists()
    assert wb_path.exists()

    wb = load_workbook(wb_path)
    ws = wb["topical_map"]
    # P1-09: header lands at the schema header_row (topical_map=4); data follows
    # at data_start_row (5). Row 1 is no longer clobbered with a spurious header.
    assert ws.max_row == 7
    assert ws["A1"].value is None
    assert ws["A4"].value == "pillar"
    assert ws["C5"].value == "kw 0"
    assert ws["C7"].value == "kw 2"

    # Provenance event was emitted.
    events_path = state / "events.jsonl"
    assert events_path.exists()
    lines = events_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 1
    ev = json.loads(lines[0])
    assert ev["event_kind"] == "provenance"
    assert ev["target_excel_sheet"] == "topical_map"
    assert ev["rows_written"] == 3
    assert ev["operation"] == "project_excel"
    assert ev["event_id"] == result.event_id


# ---------------------------------------------------------------------------
# Test 2 — atomic save: monkeypatch os.replace to raise, original intact
# ---------------------------------------------------------------------------

def test_atomic_kill_mid_write_no_partial(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    # Seed an existing workbook with one row so we can verify it's untouched.
    transaction.append(wb_path, "topical_map", [_topical_row()], "test-proj")
    pre_mtime = wb_path.stat().st_mtime
    pre_size = wb_path.stat().st_size

    # Inject failure at os.replace inside transaction module.
    def _boom(src, dst):
        raise OSError(13, "simulated rename failure")

    monkeypatch.setattr(transaction.os, "replace", _boom)

    with pytest.raises(OSError, match="simulated rename failure"):
        transaction.append(wb_path, "topical_map",
                           [_topical_row(primary_keyword="ghost")], "test-proj")

    # Original workbook unchanged.
    assert wb_path.stat().st_size == pre_size
    # Tempfile must not leak in parent dir.
    leftovers = [p for p in proj.iterdir() if p.name.startswith(".master-") and p.name.endswith(".xlsx.tmp")]
    assert leftovers == [], f"leaked tempfile: {leftovers}"


# ---------------------------------------------------------------------------
# Test 3 — missing required column → RowSchemaError
# ---------------------------------------------------------------------------

def test_schema_validation_rejects_missing_required(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    bad = _topical_row()
    bad.pop("pillar")
    with pytest.raises(RowSchemaError):
        transaction.append(wb_path, "topical_map", [bad], "test-proj")


# ---------------------------------------------------------------------------
# Test 4 — formula prefix '=' rejected
# ---------------------------------------------------------------------------

def test_formula_policy_rejects_equals_prefix(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    row = _topical_row(note="=SUM(A:A)")
    with pytest.raises(FormulaPolicyViolation):
        transaction.append(wb_path, "topical_map", [row], "test-proj")


# ---------------------------------------------------------------------------
# Test 5 — backup rotation keeps 7
# ---------------------------------------------------------------------------

def test_backup_rotation_keeps_last_seven(tmp_path: Path) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    backup_dir = state / "backups" / "master"

    for i in range(10):
        transaction.append(
            wb_path, "topical_map",
            [_topical_row(primary_keyword=f"kw{i}")],
            "test-proj",
        )
        time.sleep(0.001)  # ensure ISO timestamps differ at microsecond grain

    files = sorted(backup_dir.iterdir())
    assert len(files) == 7, f"expected 7 backups, got {len(files)}"


# ---------------------------------------------------------------------------
# Test 6 — lock contention (multiprocessing): second writer fails fast
# ---------------------------------------------------------------------------

def _holder_proc(lock_path: str, hold_event_path: str, release_event_path: str) -> None:
    """Subprocess: acquire lock, signal ready, hold until told to release."""
    import fcntl as _fcntl
    import os as _os
    fd = _os.open(lock_path, _os.O_WRONLY | _os.O_CREAT, 0o644)
    _fcntl.flock(fd, _fcntl.LOCK_EX | _fcntl.LOCK_NB)
    Path(hold_event_path).touch()
    while not Path(release_event_path).exists():
        time.sleep(0.01)
    _fcntl.flock(fd, _fcntl.LOCK_UN)
    _os.close(fd)


def test_lock_contention_second_writer_fails_fast(tmp_path: Path) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    lock_path = state / "excel.lock"
    hold_event = tmp_path / "holder-up"
    release_event = tmp_path / "holder-release"

    # Pre-write the sentinel with a real PID + ts so stale-detection
    # treats it as live (not stale).
    holder = mp.Process(
        target=_holder_proc,
        args=(str(lock_path), str(hold_event), str(release_event)),
    )
    holder.start()
    try:
        # Wait until the holder has acquired flock.
        for _ in range(200):
            if hold_event.exists():
                break
            time.sleep(0.01)
        else:  # pragma: no cover
            holder.terminate()
            pytest.fail("holder process never acquired lock")

        # Pre-populate sentinel with the holder's pid so stale-check passes.
        lock_path.write_text(json.dumps({
            "pid": holder.pid,
            "ts": "2099-12-31T23:59:59.999999Z",  # future = not stale
        }), encoding="utf-8")

        with pytest.raises(LockHeldError):
            transaction.append(wb_path, "topical_map", [_topical_row()], "test-proj")
    finally:
        release_event.touch()
        holder.join(timeout=5)
        if holder.is_alive():  # pragma: no cover
            holder.terminate()


# ---------------------------------------------------------------------------
# Test 7 — statusEnum 7 values pass; "WIP" fails
# ---------------------------------------------------------------------------

def test_status_enum_validates_all_seven(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    statuses = ["TODO", "ONGOING", "EXISTS", "DONE", "BLOCKED", "DEFERRED", "CANCELED"]
    rows = [_topical_row(status=s, primary_keyword=f"kw-{s.lower()}") for s in statuses]
    result = transaction.append(wb_path, "topical_map", rows, "test-proj")
    assert result.rows_affected == 7

    # Now invalid status:
    bad = _topical_row(status="WIP", primary_keyword="kw-wip")
    with pytest.raises(RowSchemaError):
        transaction.append(wb_path, "topical_map", [bad], "test-proj")


# ---------------------------------------------------------------------------
# Test 8 — severityEnum on tech_seo: CRITICAL/HIGH/MEDIUM/LOW pass; URGENT fails
# ---------------------------------------------------------------------------

def test_severity_enum_validates(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"

    def _tech_row(impact: str) -> dict:
        return {
            "issue_category": "Performance",
            "detail": "slow LCP",
            "affected_urls": "/page",
            "impact": impact,
            "resolution": "lazy-load",
            "priority": "P1",
        }

    rows = [_tech_row(s) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]]
    transaction.append(wb_path, "tech_seo", rows, "test-proj")

    with pytest.raises(RowSchemaError):
        transaction.append(wb_path, "tech_seo", [_tech_row("URGENT")], "test-proj")


# ---------------------------------------------------------------------------
# Test 9 — provenance event has correct shape
# ---------------------------------------------------------------------------

def test_provenance_event_emitted_after_write(tmp_path: Path) -> None:
    proj, state = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    transaction.append(wb_path, "topical_map", [_topical_row()], "test-proj")

    events_path = state / "events.jsonl"
    lines = events_path.read_text(encoding="utf-8").strip().splitlines()
    assert len(lines) == 1
    ev = json.loads(lines[0])
    assert ev["event_kind"] == "provenance"
    assert ev["target_excel_sheet"] == "topical_map"
    assert ev["operation"] == "project_excel"
    assert ev["source"]["kind"] == "tool_computed"
    assert ev["rows_written"] == 1
    assert isinstance(ev["run_id"], int)


# ---------------------------------------------------------------------------
# Test 10 — plugin-agnostic source: no slug literals in module
# ---------------------------------------------------------------------------

def test_plugin_agnostic_no_slug_leak() -> None:
    src = Path(transaction.__file__).read_text(encoding="utf-8")
    forbidden = ("demo-dental", "demo-furniture", "demo-hvac", "demo-petcare")
    hits = [name for name in forbidden if name in src.lower()]
    assert hits == [], f"slug leak in transaction.py: {hits}"


# ---------------------------------------------------------------------------
# Test 11 — sheet missing from schema → SchemaSheetMismatchError
# ---------------------------------------------------------------------------

def test_unknown_sheet_rejected(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    with pytest.raises(SchemaSheetMismatchError):
        transaction.append(wb_path, "no_such_sheet", [], "test-proj")


# ---------------------------------------------------------------------------
# Test 12 — master_task writer scope guard
# ---------------------------------------------------------------------------

def test_master_task_requires_writer(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    row = _master_task_row()
    # No writer → WriterScopeError
    with pytest.raises(WriterScopeError):
        transaction.append(wb_path, "master_task", [row], "test-proj")
    # Disallowed writer → WriterScopeError
    with pytest.raises(WriterScopeError):
        transaction.append(wb_path, "master_task", [row], "test-proj", writer="rogue")
    # Allowed writer → OK
    result = transaction.append(
        wb_path, "master_task", [row], "test-proj", writer="human",
    )
    assert result.rows_affected == 1


# ---------------------------------------------------------------------------
# Test 13 — cell length cap
# ---------------------------------------------------------------------------

def test_cell_value_too_long(tmp_path: Path) -> None:
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    huge_note = "x" * 40_000
    with pytest.raises(CellValueTooLongError):
        transaction.append(
            wb_path, "topical_map",
            [_topical_row(note=huge_note)],
            "test-proj",
        )


# ---------------------------------------------------------------------------
# LC-1 — WRITER_REGISTRY codification (Q-W3W2B-WRITER-01; spec Risk R6)
# OPT-IN / advisory in v1.9; enforcement reserved v2.0. The write path is
# intentionally NOT gated — these tests lock that contract.
# ---------------------------------------------------------------------------

def _schema_sheet_names() -> set[str]:
    schema = json.loads(transaction._DEFAULT_SCHEMA_PATH.read_text(encoding="utf-8"))
    return set(schema.get("sheets", {}).keys())


def test_writer_registry_covers_every_schema_sheet() -> None:
    # Schema-first: the registry must enumerate ALL logical sheets, no more.
    assert set(transaction.WRITER_REGISTRY.keys()) == _schema_sheet_names()
    # Every value is a non-empty frozenset of non-empty writer identities.
    for sheet, writers in transaction.WRITER_REGISTRY.items():
        assert isinstance(writers, frozenset) and writers, sheet
        assert all(isinstance(w, str) and w for w in writers), sheet


def test_writer_registry_master_task_mirrors_schema_allowed_writers() -> None:
    # master_task is the one hard-gated sheet; registry must match the schema
    # SSoT exactly (existing writer pattern: human/dashboard_refresh/
    # done_protocol/master_task_sync).
    schema = json.loads(transaction._DEFAULT_SCHEMA_PATH.read_text(encoding="utf-8"))
    schema_writers = set(schema["sheets"]["master_task"]["allowed_writers"])
    assert transaction.WRITER_REGISTRY["master_task"] == frozenset(schema_writers)
    assert "master_task_sync" in transaction.WRITER_REGISTRY["master_task"]


def test_writer_registry_known_convention_mappings() -> None:
    # Prompt-specified convention anchors (LC-1): opportunity + quick_wins are
    # both produced by the quick-wins skill.
    assert transaction.WRITER_REGISTRY["opportunity"] == frozenset({"quick-wins"})
    assert transaction.WRITER_REGISTRY["quick_wins"] == frozenset({"quick-wins"})
    # B5-07: content_improve is PRODUCED by content-decay (its produces[] lists
    # "content-improve"; monthly-report's producer table agrees). The
    # content-remediation + faq-optimization skills declare content_improve
    # READ-ONLY (F-15 allowed_writers=null), so neither writes it.
    assert transaction.WRITER_REGISTRY["content_improve"] == frozenset({"content-decay"})


def test_writer_registry_status_ok_for_registered_writer() -> None:
    assert transaction.writer_registry_status("quick_wins", "quick-wins") == "ok"
    assert transaction.writer_registry_status("master_task", "master_task_sync") == "ok"


def test_writer_registry_status_unregistered_logs_warning(capsys: pytest.CaptureFixture) -> None:
    status = transaction.writer_registry_status("quick_wins", "rogue-skill")
    assert status == "unregistered"
    err = capsys.readouterr().err
    assert "off-convention" in err and "rogue-skill" in err and "advisory" in err


def test_writer_registry_status_vacuous_when_sheet_absent_or_writer_none() -> None:
    # Sheet not in the registry → vacuous OK (no convention to enforce).
    assert transaction.writer_registry_status("not_a_sheet", "anyone") == "ok"
    # writer None → vacuous OK (nothing claimed).
    assert transaction.writer_registry_status("quick_wins", None) == "ok"


def test_writer_registry_enforcement_off_by_default_r6() -> None:
    # R6: enforcement MUST be OFF in v1.9 so existing callers cannot break.
    assert transaction.WRITER_REGISTRY_ENFORCEMENT is False


def test_writer_registry_does_not_gate_writes_r6(tmp_path: Path) -> None:
    # R6 regression proof: a non-master_task write with an off-convention
    # writer STILL succeeds — WRITER_REGISTRY never touches the write path.
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    assert transaction.writer_registry_status("topical_map", "rogue-writer") == "unregistered"
    result = transaction.append(
        wb_path, "topical_map", [_topical_row()], "test-proj", writer="rogue-writer",
    )
    assert isinstance(result, WriteResult) and result.rows_affected == 1


# ---------------------------------------------------------------------------
# AC-10 — transaction.replace (idempotent sheet refresh)
#
# The template-seeded row model that existing tests miss: bootstrap_excel.py
# honors each sheet's header_row, so a header_row=3 sheet physically has
# row1=header, row2=blank, row3=header, and DATA starts at data_start_row (=4).
# Readers (weekly_summary, portfolio_heatmap, validate_invariants) all read
# from data_start_row. replace() MUST clear the data block and re-land data at
# data_start_row, leaving the header block (rows 1..data_start-1) verbatim.
# ---------------------------------------------------------------------------

def _seed_template_workbook(
    wb_path: Path,
    sheet: str,
    header_names: list[str],
    *,
    header_row: int,
    data_start_row: int,
    data_rows: list[list],
) -> None:
    """Build a workbook mimicking bootstrap_excel's template layout.

    Lays the schema header at row 1 AND at `header_row` (the bootstrap seeds a
    title/sub-header block then the real header), leaves any rows between blank,
    and seeds `data_rows` starting at `data_start_row`. This is the realistic
    layout the replace() writer must respect.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    # Header at row 1 (mirrors transaction._ensure_sheet_with_header) and at the
    # schema header_row. Rows in between (e.g. row 2) stay blank.
    for col_idx, name in enumerate(header_names, start=1):
        ws.cell(row=1, column=col_idx, value=name)
        ws.cell(row=header_row, column=col_idx, value=name)
    for offset, values in enumerate(data_rows):
        r = data_start_row + offset
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=r, column=col_idx, value=val)
    wb.save(str(wb_path))


# tech_seo: header_row=3, data_start_row=4 (the header_row=3 case).
_TECH_SEO_HEADERS = [
    "issue_category", "detail", "affected_urls", "impact", "resolution", "priority",
]


def _tech_seo_row(**overrides) -> dict:
    base = {
        "issue_category": "Performance",
        "detail": "slow LCP",
        "affected_urls": "/page",
        "impact": "HIGH",
        "resolution": "lazy-load",
        "priority": "P1",
    }
    base.update(overrides)
    return base


def _tech_seo_values(row: dict) -> list:
    return [row[h] for h in _TECH_SEO_HEADERS]


# robots_txt: header_row=4, data_start_row=5 (the header_row=4 case).
_ROBOTS_HEADERS = ["id", "level", "issue", "detail", "resolution"]


def _robots_row(**overrides) -> dict:
    base = {
        "id": "R-001",
        "level": "MEDIUM",
        "issue": "missing sitemap directive",
        "detail": "robots.txt lacks Sitemap:",
        "resolution": "add Sitemap line",
    }
    base.update(overrides)
    return base


def _robots_values(row: dict) -> list:
    return [row[h] for h in _ROBOTS_HEADERS]


def test_replace_clears_prior_data_then_writes(tmp_path: Path) -> None:
    # header_row=3 sheet (tech_seo): seed 5 data rows at row4-8, replace with 2.
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    seeded = [_tech_seo_row(detail=f"old-{i}") for i in range(5)]
    _seed_template_workbook(
        wb_path, "tech_seo", _TECH_SEO_HEADERS,
        header_row=3, data_start_row=4,
        data_rows=[_tech_seo_values(r) for r in seeded],
    )

    new_rows = [_tech_seo_row(detail="new-A"), _tech_seo_row(detail="new-B")]
    result = transaction.replace(wb_path, "tech_seo", new_rows, "test-proj")
    assert result.rows_affected == 2

    wb = load_workbook(wb_path)
    ws = wb["tech_seo"]
    # Data now lands at row4-5 only; max_row == 5 (old 5 rows gone).
    assert ws.max_row == 5
    # Header block (rows 1..3) UNCHANGED — header still at row1 and row3.
    assert ws.cell(row=1, column=1).value == "issue_category"
    assert ws.cell(row=3, column=1).value == "issue_category"
    assert ws.cell(row=2, column=1).value is None  # blank row preserved
    # The 2 new rows are exactly what we wrote, at data_start_row onward.
    assert ws.cell(row=4, column=2).value == "new-A"
    assert ws.cell(row=5, column=2).value == "new-B"
    # The old rows are gone — no "old-*" survives.
    survivors = [
        ws.cell(row=r, column=2).value
        for r in range(4, ws.max_row + 1)
    ]
    assert all(not str(v).startswith("old-") for v in survivors if v is not None)


def test_replace_is_idempotent(tmp_path: Path) -> None:
    # robots_txt: header_row=4, data_start_row=5. Replace same 3 rows twice →
    # identical rowcount + content (no duplication / drift).
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    _seed_template_workbook(
        wb_path, "robots_txt", _ROBOTS_HEADERS,
        header_row=4, data_start_row=5,
        data_rows=[],  # start with header block only
    )
    rows = [_robots_row(id=f"R-{i:03d}") for i in range(3)]

    r1 = transaction.replace(wb_path, "robots_txt", rows, "test-proj")
    wb1 = load_workbook(wb_path)
    ws1 = wb1["robots_txt"]
    snap1 = [
        [ws1.cell(row=r, column=c).value for c in range(1, len(_ROBOTS_HEADERS) + 1)]
        for r in range(5, ws1.max_row + 1)
    ]
    max1 = ws1.max_row

    r2 = transaction.replace(wb_path, "robots_txt", rows, "test-proj")
    wb2 = load_workbook(wb_path)
    ws2 = wb2["robots_txt"]
    snap2 = [
        [ws2.cell(row=r, column=c).value for c in range(1, len(_ROBOTS_HEADERS) + 1)]
        for r in range(5, ws2.max_row + 1)
    ]
    max2 = ws2.max_row

    assert r1.rows_affected == r2.rows_affected == 3
    assert max1 == max2 == 7  # rows 5,6,7 — no growth on second replace
    assert snap1 == snap2
    # Data starts exactly at row 5 (header_row=4 + 1).
    assert ws2.cell(row=5, column=1).value == "R-000"


def test_replace_validates_schema_and_does_not_clear_on_failure(tmp_path: Path) -> None:
    # Seed 4 rows; an enum-violating replace must RAISE and leave the 4 intact.
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    seeded = [_tech_seo_row(detail=f"keep-{i}") for i in range(4)]
    _seed_template_workbook(
        wb_path, "tech_seo", _TECH_SEO_HEADERS,
        header_row=3, data_start_row=4,
        data_rows=[_tech_seo_values(r) for r in seeded],
    )

    bad = [_tech_seo_row(impact="URGENT")]  # URGENT not in severityEnum
    with pytest.raises(RowSchemaError):
        transaction.replace(wb_path, "tech_seo", bad, "test-proj")

    # The original 4 rows must STILL be present — clear must NOT have happened.
    wb = load_workbook(wb_path)
    ws = wb["tech_seo"]
    assert ws.max_row == 7  # rows 4-7 = 4 data rows
    survivors = [ws.cell(row=r, column=2).value for r in range(4, 8)]
    assert survivors == ["keep-0", "keep-1", "keep-2", "keep-3"]


def test_replace_on_empty_data_block(tmp_path: Path) -> None:
    # Sheet with only the header block (rows 1-3), no data → replace writes
    # cleanly from row 4.
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    _seed_template_workbook(
        wb_path, "tech_seo", _TECH_SEO_HEADERS,
        header_row=3, data_start_row=4,
        data_rows=[],
    )

    rows = [_tech_seo_row(detail="first"), _tech_seo_row(detail="second")]
    result = transaction.replace(wb_path, "tech_seo", rows, "test-proj")
    assert result.rows_affected == 2

    wb = load_workbook(wb_path)
    ws = wb["tech_seo"]
    assert ws.cell(row=3, column=1).value == "issue_category"  # header intact
    assert ws.cell(row=4, column=2).value == "first"
    assert ws.cell(row=5, column=2).value == "second"
    assert ws.max_row == 5


def test_append_mode_unchanged(tmp_path: Path) -> None:
    # Regression guard: append still lands at _next_data_row (max_row+1), NOT a
    # recomputed data_start_row. P1-09: with the header now at the schema
    # header_row (topical_map=4), the first append lands at row 5 (= max_row+1)
    # and row 1 is no longer clobbered. The cursor is still physical: a second
    # append advances to row 6 rather than resetting.
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    transaction.append(wb_path, "topical_map", [_topical_row()], "test-proj")
    result = transaction.append(
        wb_path, "topical_map", [_topical_row(primary_keyword="second")], "test-proj"
    )
    assert result.rows_affected == 1

    wb = load_workbook(wb_path)
    ws = wb["topical_map"]
    assert ws["A1"].value is None            # row 1 not clobbered
    assert ws["A4"].value == "pillar"        # header at header_row=4
    assert ws["A5"].value == "P01_sofa_sets"  # first append at row 5
    assert ws["C6"].value == "second"        # second append advances to row 6
    assert ws.max_row == 6


def test_update_starts_at_data_start_row_not_row_2(tmp_path: Path) -> None:
    """update() must scan from the schema's data_start_row, never hardcoded row 2.

    A template-seeded sheet (tech_seo: header_row=3, data_start_row=4) keeps
    header/blank rows at 1-3. A schema-valid row planted ABOVE data_start_row
    must never be matched or mutated by update(); otherwise update() corrupts the
    header block and over-counts rows_affected. Regression guard for the
    `range(2, ...)` scan bug (the same data_start_row contract replace() honors).

    P1-09: the plant lives at row 2 (a blank row above data_start_row that is NOT
    the header row=3), since _ensure_sheet_with_header now correctly (re)writes the
    canonical header at row 3.
    """
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    # Real data at rows 4-5 (data_start_row=4); only row 4 carries the marker.
    seeded = [_tech_seo_row(detail="COLLIDE"), _tech_seo_row(detail="other")]
    _seed_template_workbook(
        wb_path, "tech_seo", _TECH_SEO_HEADERS,
        header_row=3, data_start_row=4,
        data_rows=[_tech_seo_values(r) for r in seeded],
    )
    # Adversary: plant a schema-valid row at row 2 (ABOVE data_start_row=4, and not
    # the header row=3) whose `detail` collides with the where clause. A correct
    # update() scans only from data_start_row and ignores it.
    wb = load_workbook(str(wb_path))
    ws = wb["tech_seo"]
    planted = _tech_seo_values(
        _tech_seo_row(detail="COLLIDE", resolution="ABOVE-UNTOUCHED")
    )
    for col_idx, val in enumerate(planted, start=1):
        ws.cell(row=2, column=col_idx, value=val)
    wb.save(str(wb_path))

    result = transaction.update(
        wb_path, "tech_seo",
        where={"detail": "COLLIDE"},
        set_={"resolution": "FIXED"},
        project_slug="test-proj",
    )

    res_col = _TECH_SEO_HEADERS.index("resolution") + 1
    check = load_workbook(wb_path)["tech_seo"]
    # Only the row-4 data row matches — the planted row-2 is above data_start_row.
    assert result.rows_affected == 1
    # Row 2 (above data_start_row) must be untouched by update().
    assert check.cell(row=2, column=res_col).value == "ABOVE-UNTOUCHED"
    # The real data row at row 4 IS updated (positive path still works).
    assert check.cell(row=4, column=res_col).value == "FIXED"


# ---------------------------------------------------------------------------
# Deep-audit HIGH (2026-06-04) — nullable enum/ref columns
#
# update() rebuilds the matched row from EVERY column and re-validates the merge.
# Enum/ref columns were non-nullable while typed/string columns were nullable, so
# a pre-existing BLANK enum cell (None, which the engine itself tolerates on
# write) made the full-row re-validation fail and silently rejected unrelated
# edits. Enum/ref columns must be nullable, consistent with typed/string columns.
# ---------------------------------------------------------------------------

def test_update_succeeds_with_blank_enum_cell_in_untouched_column(tmp_path: Path) -> None:
    """A blank enum/ref cell (impact) in a column update() does not touch must not
    block an edit to another column (resolution)."""
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    # Seed a tech_seo data row whose `impact` (severityEnum ref) cell is BLANK,
    # exactly as openpyxl reads an empty enum cell back (None).
    seeded = [_tech_seo_row(detail="TARGET", impact=None)]
    _seed_template_workbook(
        wb_path, "tech_seo", _TECH_SEO_HEADERS,
        header_row=3, data_start_row=4,
        data_rows=[_tech_seo_values(r) for r in seeded],
    )

    result = transaction.update(
        wb_path, "tech_seo",
        where={"detail": "TARGET"},
        set_={"resolution": "FIXED"},
        project_slug="test-proj",
    )
    assert result.rows_affected == 1

    check = load_workbook(wb_path)["tech_seo"]
    res_col = _TECH_SEO_HEADERS.index("resolution") + 1
    impact_col = _TECH_SEO_HEADERS.index("impact") + 1
    assert check.cell(row=4, column=res_col).value == "FIXED"
    # The untouched blank enum cell stays blank — not coerced, not blocking.
    assert check.cell(row=4, column=impact_col).value is None


def test_write_still_rejects_invalid_nonblank_enum(tmp_path: Path) -> None:
    """Nullability must NOT weaken the enum constraint: a non-blank value outside
    the enum is still rejected (only None becomes acceptable)."""
    proj, _ = _setup_project(tmp_path)
    wb_path = proj / "master.xlsx"
    with pytest.raises(RowSchemaError):
        transaction.append(wb_path, "tech_seo", [_tech_seo_row(impact="URGENT")], "test-proj")
