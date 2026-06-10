"""TDD lock for scripts/util/sheet_merge.py — idempotent prefix/keyed sheet merge.

sheet_merge survives sf-import's ``transaction.replace`` snapshot semantics: the
tech-SEO governance skills (hreflang-audit / facet-nav-audit / robots-policy-audit)
append ``HF-`` / ``FN-`` / ``RP-`` rows to the ``robots_txt`` sheet WITHOUT
clobbering sf_projection's ``R-NNN`` rows or each other, and migration-map keys
``redirect_404`` rows by ``url``. Both functions read the current sheet, drop only
their OWN id/key namespace, then re-land the union via the single idempotent
``committer.commit`` path (whole-block ``transaction.replace`` under the hood).

Contract (spec GAP-T2 §c + §d):
  1. idempotent double-run (stable row count, no dup ids)
  2. preserves foreign-prefix rows (R-/HF- survive an FN- merge)
  3. schema-validates output rows (RowSchemaError propagates on bad severityEnum)
  4. refuses unknown sheet (SchemaSheetMismatchError)
  + commit/lock passthrough (returns a WriteResult with the union rowcount)
  + keyed variant (redirect_404 by url): replace-by-key, preserve others, idempotent
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.excel.transaction import (
    RowSchemaError,
    SchemaSheetMismatchError,
    WriteResult,
)
from scripts.orchestration import committer
from scripts.util import sheet_merge

SLUG = "test-proj"
RUN_ID = "test-proj-2026-06-10-aa11"


def _wb(tmp_path: Path) -> Path:
    proj = tmp_path / "projects" / SLUG
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    return proj / "master.xlsx"


def _robots_row(id_, level="MEDIUM", issue="x", detail="d", resolution="r") -> dict:
    return {"id": id_, "level": level, "issue": issue, "detail": detail, "resolution": resolution}


def _redirect_row(url, action="301", target_url="/new", status="TODO", inlinks=0) -> dict:
    return {"url": url, "inlinks": inlinks, "action": action, "target_url": target_url, "status": status}


def _seed_robots(wb_path: Path, rows: list[dict]) -> None:
    committer.commit(wb_path, "robots_txt", rows, run_id=RUN_ID, project_slug=SLUG, writer="sf-import")


def _ids(wb_path: Path, sheet="robots_txt") -> list:
    ws = load_workbook(wb_path)[sheet]
    return [r[0] for r in ws.iter_rows(min_row=5, values_only=True) if r and r[0] is not None]


# --- merge_prefixed_rows -----------------------------------------------------

def test_merge_prefixed_preserves_foreign_rows(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    _seed_robots(wb, [_robots_row("R-001", "HIGH"), _robots_row("HF-001")])
    new = [_robots_row(None, "MEDIUM", issue="facet bloat"),
           _robots_row(None, "LOW", issue="unknown param")]
    res = sheet_merge.merge_prefixed_rows(
        wb, "robots_txt", new, id_prefix="FN-",
        run_id=RUN_ID, project_slug=SLUG, writer="facet-nav-audit",
    )
    assert isinstance(res, WriteResult)
    assert res.rows_affected == 4
    ids = _ids(wb)
    assert {"R-001", "HF-001", "FN-001", "FN-002"} == set(ids)


def test_merge_prefixed_fresh_workbook_generates_sequential_ids(tmp_path: Path) -> None:
    wb = _wb(tmp_path)  # does not exist yet — merge must create via committer
    new = [_robots_row(None, "HIGH"), _robots_row(None, "MEDIUM"), _robots_row(None, "LOW")]
    sheet_merge.merge_prefixed_rows(
        wb, "robots_txt", new, id_prefix="FN-",
        run_id=RUN_ID, project_slug=SLUG, writer="facet-nav-audit",
    )
    assert _ids(wb) == ["FN-001", "FN-002", "FN-003"]


def test_merge_prefixed_idempotent_no_dupes(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    _seed_robots(wb, [_robots_row("R-001", "HIGH")])
    new = [_robots_row(None, "MEDIUM"), _robots_row(None, "LOW")]
    kw = dict(id_prefix="FN-", run_id=RUN_ID, project_slug=SLUG, writer="facet-nav-audit")
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", new, **kw)
    ids1 = sorted(_ids(wb))
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", new, **kw)
    ids2 = sorted(_ids(wb))
    assert ids1 == ids2 == ["FN-001", "FN-002", "R-001"]


def test_merge_prefixed_does_not_mutate_caller_rows(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    new = [_robots_row(None, "HIGH")]
    sheet_merge.merge_prefixed_rows(
        wb, "robots_txt", new, id_prefix="FN-",
        run_id=RUN_ID, project_slug=SLUG, writer="facet-nav-audit",
    )
    assert new[0]["id"] is None  # caller dict untouched (immutability)


def test_merge_prefixed_schema_error_propagates(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    bad = [_robots_row(None, level="BOGUS")]  # not a severityEnum value
    with pytest.raises(RowSchemaError):
        sheet_merge.merge_prefixed_rows(
            wb, "robots_txt", bad, id_prefix="FN-",
            run_id=RUN_ID, project_slug=SLUG, writer="facet-nav-audit",
        )


def test_merge_prefixed_unknown_sheet(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    with pytest.raises(SchemaSheetMismatchError):
        sheet_merge.merge_prefixed_rows(
            wb, "not_a_sheet", [], id_prefix="FN-",
            run_id=RUN_ID, project_slug=SLUG, writer="x",
        )


# --- merge_keyed_rows (redirect_404, no id column) ---------------------------

def test_merge_keyed_replaces_by_key_preserves_others(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    committer.commit(
        wb, "redirect_404",
        [_redirect_row("/a", status="TODO"), _redirect_row("/b", status="TODO")],
        run_id=RUN_ID, project_slug=SLUG, writer="sf-import",
    )
    new = [_redirect_row("/a", status="DONE"), _redirect_row("/c", status="TODO")]
    res = sheet_merge.merge_keyed_rows(
        wb, "redirect_404", new, key_column="url",
        run_id=RUN_ID, project_slug=SLUG, writer="migration-map",
    )
    ws = load_workbook(wb)["redirect_404"]
    status_by_url = {r[0]: r[4] for r in ws.iter_rows(min_row=5, values_only=True) if r and r[0]}
    assert status_by_url == {"/a": "DONE", "/b": "TODO", "/c": "TODO"}
    assert res.rows_affected == 3


def test_merge_keyed_idempotent(tmp_path: Path) -> None:
    wb = _wb(tmp_path)
    committer.commit(
        wb, "redirect_404", [_redirect_row("/b", status="TODO")],
        run_id=RUN_ID, project_slug=SLUG, writer="sf-import",
    )
    new = [_redirect_row("/a", status="TODO")]
    kw = dict(key_column="url", run_id=RUN_ID, project_slug=SLUG, writer="migration-map")
    sheet_merge.merge_keyed_rows(wb, "redirect_404", new, **kw)
    urls1 = sorted(_ids(wb, "redirect_404"))
    sheet_merge.merge_keyed_rows(wb, "redirect_404", new, **kw)
    urls2 = sorted(_ids(wb, "redirect_404"))
    assert urls1 == urls2 == ["/a", "/b"]
