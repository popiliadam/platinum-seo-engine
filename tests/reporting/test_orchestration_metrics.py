"""tests/reporting/test_orchestration_metrics.py — AMO independent oracle (batch 2d).

TDD lock for scripts/reporting/orchestration_metrics.py: the READ-ONLY oracle that
re-derives the structured-error rate from on-disk artifacts (raw provenance drop →
transform output → COMMITTED master.xlsx) WITHOUT trusting coverage.verdict.

Reconcile contract (one code_verified step of one run):
  R1 identity    raw.provenance.run_id/slug == this run's run_id/slug
  R2 untruncated raw.provenance.declared_count == len(raw.rows)
  R3 output      transform output exists and is a list (or {"rows": [...]})
  R4 workbook    committed-row-count(sheet) == len(transform output)   [workbook given]
  R5 self-report coverage scored_count == committed-row-count           [workbook given]

Workbooks are seeded both directly with openpyxl (header at header_row, data from
data_start_row — the layout transaction.replace writes) and, in one integration
test, via the REAL committer.commit so the reader is proven against the writer.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from scripts.orchestration.coverage import coverage_path
from scripts.orchestration.workflows import audit_suite
from scripts.orchestration.workflows.monthly_maintenance import inbox_path, output_path
from scripts.reporting import orchestration_metrics as om

# The three code_verified STEPS (name -> committed sheet), per monthly_maintenance.
STEP_SHEETS = {
    "gsc_pull": "gsc_performance",
    "quick_wins": "quick_wins",
    "content_decay": "content_decay",
}


# ---------------------------------------------------------------------------
# Seeding helpers — place artifacts at the REAL paths via the workflow helpers.
# ---------------------------------------------------------------------------

def _write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj), encoding="utf-8")


def _raw_drop(run_id: str, slug: str, count: int, *, declared=None, rows=None) -> dict:
    rows = rows if rows is not None else [{"i": i} for i in range(count)]
    return {
        "provenance": {
            "run_id": run_id,
            "slug": slug,
            "site_url": "https://example.test",
            "window": "recent",
            "tool": "mcp__gsc__search_analytics",
            "fetched_at": "2026-06-06T00:00:00Z",
            "declared_count": count if declared is None else declared,
        },
        "rows": rows,
    }


def _seed_raw(ws_root, run_id, slug, step, drop) -> None:
    _write_json(inbox_path(ws_root, run_id, slug, step), drop)


def _seed_output(ws_root, run_id, slug, sheet, n) -> None:
    _write_json(output_path(ws_root, run_id, slug, sheet), [{"i": i} for i in range(n)])


def _cov_step(name, *, scored_count, vclass="code_verified", status="satisfied") -> dict:
    return {
        "name": name,
        "verification_class": vclass,
        "status": status,
        "scored_count": scored_count,
    }


def _cov_record(run_id, slug, steps, verdict) -> dict:
    return {
        "run_id": run_id,
        "steps": steps,
        "required_satisfied": True,
        "verdict": verdict,
    }


def _seed_workbook(path, sheet, data_start_row, n_rows, *, header_row=None) -> Path:
    """One-sheet workbook: header at header_row, n contiguous data rows from
    data_start_row (the layout transaction.replace lands)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    hr = data_start_row - 1 if header_row is None else header_row
    ws.cell(row=hr, column=1, value="col_a")
    for r in range(n_rows):
        ws.cell(row=data_start_row + r, column=1, value=f"v{r}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _seed_multi_workbook(path, sheet_counts: dict) -> Path:
    """Workbook with one sheet per key; each sheet's data_start_row is read from
    the master-excel schema so the layout matches the committer per-sheet."""
    wb = Workbook()
    default = wb.active
    first = True
    for sheet, n in sheet_counts.items():
        dsr = om.data_start_row_for(sheet)
        if first:
            ws, ws.title, first = default, sheet, False
        else:
            ws = wb.create_sheet(sheet)
        ws.cell(row=dsr - 1, column=1, value="col_a")
        for r in range(n):
            ws.cell(row=dsr + r, column=1, value=f"v{r}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def _seed_run_artifacts(ws_root, run_id, slug, spec: dict, verdict: str) -> dict:
    """spec: {step: {"raw": n, "output": n?, "scored": n, "declared": n?}}.

    Seeds the raw drop + (optional) transform output for each step and returns the
    coverage record (NOT written to disk — caller writes it if it wants on-disk).
    """
    cov_steps = []
    for step, vals in spec.items():
        sheet = STEP_SHEETS[step]
        _seed_raw(
            ws_root, run_id, slug, step,
            _raw_drop(run_id, slug, vals.get("raw", 0), declared=vals.get("declared")),
        )
        if "output" in vals:
            _seed_output(ws_root, run_id, slug, sheet, vals["output"])
        cov_steps.append(_cov_step(step, scored_count=vals.get("scored", 0)))
    return _cov_record(run_id, slug, cov_steps, verdict)


def _setup_step(tmp_path, *, raw_slug=None, declared=None, output_n=4, omit_output=False,
                committed_n=4, scored=4, raw_n=4) -> dict:
    """Build a single happy gsc_pull step, optionally perturbing ONE input, and
    return its reconcile_step result. Defaults are all-consistent (happy)."""
    ws_root = tmp_path / "ws"
    run_id, slug, step, sheet = "r1", "demo", "gsc_pull", "gsc_performance"
    _seed_raw(
        ws_root, run_id, slug, step,
        _raw_drop(run_id, slug if raw_slug is None else raw_slug, raw_n, declared=declared),
    )
    if not omit_output:
        _seed_output(ws_root, run_id, slug, sheet, output_n)
    wb = _seed_workbook(tmp_path / "m.xlsx", sheet, 5, committed_n)
    return om.reconcile_step(
        workspace_root=ws_root, slug=slug, run_id=run_id, step_name=step, sheet=sheet,
        coverage_step=_cov_step(step, scored_count=scored), workbook_path=wb,
    )


# ---------------------------------------------------------------------------
# Audit-workflow seeding — same conventions, but the transform output is keyed by
# each step's EXPLICIT output_file (schema_audit -> schema_audit.json while its
# sheet is `schema`) and three of the four steps are model_attested.
# ---------------------------------------------------------------------------

_AUDIT = {s["name"]: s for s in audit_suite.STEPS}


def _seed_output_file(ws_root, run_id, slug, output_file, n) -> None:
    """Seed an audit transform output at its EXPLICIT output_file (the real path
    audit_suite writes), NOT {sheet}.json."""
    _write_json(
        audit_suite.output_path(ws_root, run_id, slug, output_file),
        [{"i": i} for i in range(n)],
    )


def _seed_audit_run(ws_root, run_id, slug, spec: dict, verdict: str) -> dict:
    """spec: {audit_step: {"raw","output"?,"scored"}}. Seeds the raw drop +
    (optional) transform output (keyed by output_file) per step; returns the
    coverage record (verification_class taken from audit_suite.STEPS)."""
    cov_steps = []
    for step, vals in spec.items():
        entry = _AUDIT[step]
        _seed_raw(ws_root, run_id, slug, step, _raw_drop(run_id, slug, vals.get("raw", 0)))
        if "output" in vals:
            _seed_output_file(ws_root, run_id, slug, entry["output_file"], vals["output"])
        cov_steps.append(_cov_step(
            step, scored_count=vals.get("scored", 0),
            vclass=entry["verification_class"],
        ))
    return _cov_record(run_id, slug, cov_steps, verdict)


def _audit_workbook(path) -> Path:
    """A committed workbook with all four audit sheets at the {3,1,5,2} layout the
    happy audit spec targets (committed counts; tests vary OUTPUT to break R4)."""
    return _seed_multi_workbook(path, {
        "tech_seo": 3, "schema": 1, "on_page_audit": 5, "cannibalization": 2,
    })


def _audit_happy_spec(**overrides) -> dict:
    """The all-reconciling audit spec (committed == output per step). The attested
    steps legitimately drop >50% of raw (advisory high_silent_skip). Override a
    step immutably, e.g. _audit_happy_spec(tech_audit={"output": 9})."""
    spec = {
        "tech_audit": {"raw": 50, "output": 3, "scored": 3},
        "schema_audit": {"raw": 20, "output": 1, "scored": 1},
        "on_page_audit": {"raw": 5, "output": 5, "scored": 5},
        "cannibalization": {"raw": 40, "output": 2, "scored": 2},
    }
    return {k: {**spec[k], **overrides.get(k, {})} for k in spec}


# ---------------------------------------------------------------------------
# 1. committed_row_count
# ---------------------------------------------------------------------------

def test_committed_row_count_counts_contiguous_rows(tmp_path):
    wb = _seed_workbook(tmp_path / "m.xlsx", "gsc_performance", 5, 3)
    assert om.committed_row_count(wb, "gsc_performance", 5) == 3


def test_committed_row_count_empty_sheet_is_zero(tmp_path):
    wb = _seed_workbook(tmp_path / "m.xlsx", "gsc_performance", 5, 0)
    assert om.committed_row_count(wb, "gsc_performance", 5) == 0


def test_committed_row_count_missing_workbook_is_none(tmp_path):
    assert om.committed_row_count(tmp_path / "nope.xlsx", "gsc_performance", 5) is None


def test_committed_row_count_missing_sheet_is_none(tmp_path):
    wb = _seed_workbook(tmp_path / "m.xlsx", "gsc_performance", 5, 2)
    assert om.committed_row_count(wb, "quick_wins", 5) is None


def test_committed_row_count_stops_at_first_empty_row(tmp_path):
    """A gap below data_start_row ends the contiguous block (matches the writer,
    which lands a single contiguous block — a gap would be genuine corruption)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "gsc_performance"
    ws.cell(row=4, column=1, value="col_a")
    ws.cell(row=5, column=1, value="a")
    ws.cell(row=6, column=1, value="b")
    ws.cell(row=8, column=1, value="d")  # row 7 is empty -> block ends at 2
    p = tmp_path / "m.xlsx"
    wb.save(str(p))
    assert om.committed_row_count(p, "gsc_performance", 5) == 2


def test_committed_row_count_counts_row_with_any_nonempty_cell(tmp_path):
    """A row counts if ANY cell is non-None, even when column 1 is blank."""
    wb = Workbook()
    ws = wb.active
    ws.title = "gsc_performance"
    ws.cell(row=4, column=1, value="col_a")
    ws.cell(row=5, column=3, value="only-col-3")  # col 1/2 blank
    p = tmp_path / "m.xlsx"
    wb.save(str(p))
    assert om.committed_row_count(p, "gsc_performance", 5) == 1


# ---------------------------------------------------------------------------
# data_start_row_for
# ---------------------------------------------------------------------------

def test_data_start_row_for_known_sheets():
    assert om.data_start_row_for("gsc_performance") == 5
    assert om.data_start_row_for("quick_wins") == 5
    assert om.data_start_row_for("content_decay") == 6


def test_data_start_row_for_unknown_sheet_raises():
    with pytest.raises(KeyError):
        om.data_start_row_for("not_a_real_sheet")


# ---------------------------------------------------------------------------
# Workflow-agnostic resolution: _step_output_file + _resolve_workflow_steps
# ---------------------------------------------------------------------------

def test_step_output_file_prefers_explicit_then_sheet():
    # monthly entry (no output_file) -> {sheet}.json
    assert om._step_output_file(
        {"name": "gsc_pull", "sheet": "gsc_performance"}
    ) == "gsc_performance.json"
    # audit entry (explicit output_file) -> output_file, NOT {sheet}.json
    assert om._step_output_file(
        {"name": "schema_audit", "sheet": "schema", "output_file": "schema_audit.json"}
    ) == "schema_audit.json"


def test_resolve_workflow_steps_monthly():
    rec = _cov_record("r1", "demo", [
        _cov_step("gsc_pull", scored_count=1),
        _cov_step("quick_wins", scored_count=1),
        _cov_step("content_decay", scored_count=1),
        _cov_step("monthly_report", scored_count=0, vclass="model_attested"),
    ], verdict="pass")
    name, steps = om._resolve_workflow_steps(rec)
    assert name == "monthly"
    assert {s["name"] for s in steps} == {"gsc_pull", "quick_wins", "content_decay"}


def test_resolve_workflow_steps_audit():
    rec = _cov_record("r1", "demo", [
        _cov_step(s["name"], scored_count=1, vclass=s["verification_class"])
        for s in audit_suite.STEPS
    ], verdict="pass")
    name, steps = om._resolve_workflow_steps(rec)
    assert name == "audit"
    assert {s["name"] for s in steps} == {
        "tech_audit", "schema_audit", "on_page_audit", "cannibalization"}


def test_resolve_workflow_steps_unknown_is_none():
    rec = _cov_record("r1", "demo", [_cov_step("mystery_step", scored_count=1)], "pass")
    assert om._resolve_workflow_steps(rec) is None


# ---------------------------------------------------------------------------
# 2. reconcile_step — HAPPY
# ---------------------------------------------------------------------------

def test_reconcile_step_happy(tmp_path):
    res = _setup_step(tmp_path)  # all consistent at 4
    assert res["independent_ok"] is True
    assert res["failed_checks"] == []
    assert res["raw_count"] == 4
    assert res["output_count"] == 4
    assert res["committed_count"] == 4
    assert res["scored_count"] == 4


# ---------------------------------------------------------------------------
# 3. reconcile_step — each perturbation maps to the RIGHT failed_check
# ---------------------------------------------------------------------------

def test_reconcile_step_identity_mismatch(tmp_path):
    res = _setup_step(tmp_path, raw_slug="someone-else")
    assert res["independent_ok"] is False
    assert res["failed_checks"] == [om.IDENTITY_MISMATCH]


def test_reconcile_step_truncated(tmp_path):
    res = _setup_step(tmp_path, declared=99)  # declared_count != len(rows)
    assert res["independent_ok"] is False
    assert res["failed_checks"] == [om.TRUNCATED]


def test_reconcile_step_output_missing(tmp_path):
    res = _setup_step(tmp_path, omit_output=True)
    assert res["independent_ok"] is False
    assert res["failed_checks"] == [om.OUTPUT_MISSING]


def test_reconcile_step_workbook_mismatch(tmp_path):
    res = _setup_step(tmp_path, committed_n=2, scored=2)  # committed(2) != output(4)
    assert res["independent_ok"] is False
    assert res["failed_checks"] == [om.WORKBOOK_MISMATCH]


def test_reconcile_step_scored_count_mismatch(tmp_path):
    res = _setup_step(tmp_path, scored=99)  # scored(99) != committed(4)
    assert res["independent_ok"] is False
    assert res["failed_checks"] == [om.SCORED_COUNT_MISMATCH]


def test_reconcile_step_raw_missing(tmp_path):
    ws_root = tmp_path / "ws"
    _seed_output(ws_root, "r1", "demo", "gsc_performance", 4)
    wb = _seed_workbook(tmp_path / "m.xlsx", "gsc_performance", 5, 4)
    res = om.reconcile_step(
        workspace_root=ws_root, slug="demo", run_id="r1", step_name="gsc_pull",
        sheet="gsc_performance", coverage_step=_cov_step("gsc_pull", scored_count=4),
        workbook_path=wb,
    )
    assert res["independent_ok"] is False
    assert om.RAW_MISSING in res["failed_checks"]
    assert res["raw_count"] is None


def test_reconcile_step_no_workbook_skips_r4_r5(tmp_path):
    """No workbook -> R4/R5 skipped (committed_count None); R1-R3 still judged."""
    ws_root = tmp_path / "ws"
    _seed_raw(ws_root, "r1", "demo", "gsc_pull", _raw_drop("r1", "demo", 4))
    _seed_output(ws_root, "r1", "demo", "gsc_performance", 4)
    res = om.reconcile_step(
        workspace_root=ws_root, slug="demo", run_id="r1", step_name="gsc_pull",
        sheet="gsc_performance", coverage_step=_cov_step("gsc_pull", scored_count=4),
        workbook_path=None,
    )
    assert res["independent_ok"] is True
    assert res["committed_count"] is None
    assert om.WORKBOOK_ABSENT not in res["failed_checks"]


def test_reconcile_step_filtering_does_not_false_flag(tmp_path):
    """quick_wins legitimately FILTERS: raw=10, output=3, committed=3, scored=3.
    Reconcile compares committed-vs-OUTPUT (not raw) -> reconciles cleanly."""
    ws_root = tmp_path / "ws"
    _seed_raw(ws_root, "r1", "demo", "quick_wins", _raw_drop("r1", "demo", 10))
    _seed_output(ws_root, "r1", "demo", "quick_wins", 3)
    wb = _seed_workbook(tmp_path / "m.xlsx", "quick_wins", 5, 3)
    res = om.reconcile_step(
        workspace_root=ws_root, slug="demo", run_id="r1", step_name="quick_wins",
        sheet="quick_wins", coverage_step=_cov_step("quick_wins", scored_count=3),
        workbook_path=wb,
    )
    assert res["independent_ok"] is True
    assert res["failed_checks"] == []
    assert res["raw_count"] == 10 and res["output_count"] == 3
    assert res["high_silent_skip"] is True  # advisory: dropped >50% of raw input


# ---------------------------------------------------------------------------
# 4. reconcile_run — reconciled / mismatch / fake_green
# ---------------------------------------------------------------------------

def _full_run(tmp_path, run_id="r1", slug="demo", *, verdict="pass",
              quick_output=2, decay_scored=1):
    """Seed a 3-step run against a committed workbook of {3,2,1}. Defaults all
    reconcile; override quick_output / decay_scored to break R4 / R5."""
    ws_root = tmp_path / "ws"
    wb = _seed_multi_workbook(
        tmp_path / f"{run_id}.xlsx",
        {"gsc_performance": 3, "quick_wins": 2, "content_decay": 1},
    )
    rec = _seed_run_artifacts(ws_root, run_id, slug, {
        "gsc_pull": {"raw": 3, "output": 3, "scored": 3},
        "quick_wins": {"raw": 2, "output": quick_output, "scored": 2},
        "content_decay": {"raw": 1, "output": 1, "scored": decay_scored},
    }, verdict=verdict)
    return ws_root, wb, rec, slug


def test_reconcile_run_all_reconciled(tmp_path):
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="pass")
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "reconciled"
    assert out["fake_green"] is False
    assert len(out["steps"]) == 3


def test_reconcile_run_one_step_mismatch_not_fake_green(tmp_path):
    # quick_wins output(9) != committed(2); verdict honestly NOT pass.
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="incomplete", quick_output=9)
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "mismatch"
    assert out["fake_green"] is False  # self-report did not claim pass


def test_reconcile_run_fake_green_detected(tmp_path):
    # content_decay scored(9) != committed(1) -> mismatch, but verdict CLAIMS pass.
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="pass", decay_scored=9)
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "mismatch"
    assert out["fake_green"] is True
    assert out["self_reported_verdict"] == "pass"


def test_reconcile_run_loads_coverage_from_disk(tmp_path):
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="pass")
    _write_json(coverage_path(ws_root, slug, "r1"), rec)
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb)  # no coverage_record passed
    assert out["independent_verdict"] == "reconciled"
    assert out["self_reported_verdict"] == "pass"


def test_reconcile_run_reconciles_attested_steps_in_steps_table(tmp_path):
    """MIGRATED from test_reconcile_run_skips_model_attested_steps to the NEW
    contract: a step that IS in the workflow's STEPS is reconciled regardless of
    verification_class — so audit's THREE model_attested steps ARE reconciled (they
    each commit a sheet). A coverage step NOT in any STEPS (a synthetic report
    step) is still skipped."""
    ws_root = tmp_path / "ws"
    slug = "demo"
    wb = _audit_workbook(tmp_path / "m.xlsx")
    rec = _seed_audit_run(ws_root, "r1", slug, _audit_happy_spec(), "pass")
    # a synthetic report step NOT in any workflow's STEPS -> must be skipped
    rec = {**rec, "steps": [
        *rec["steps"],
        _cov_step("audit_report", scored_count=0, vclass="model_attested"),
    ]}
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    reconciled_steps = {s["step"] for s in out["steps"]}
    # all 4 STEPS reconciled, INCLUDING the 3 model_attested ones (not skipped)
    assert reconciled_steps == {
        "tech_audit", "schema_audit", "on_page_audit", "cannibalization"}
    assert "audit_report" not in reconciled_steps  # synthetic report step skipped
    assert out["independent_verdict"] == "reconciled"


def test_reconcile_run_audit_all_reconciled(tmp_path):
    """A full audit run where every committed sheet == its transform output
    reconciles; the attested steps appear in the result (not skipped)."""
    ws_root = tmp_path / "ws"
    wb = _audit_workbook(tmp_path / "m.xlsx")
    rec = _seed_audit_run(ws_root, "r1", "demo", _audit_happy_spec(), "pass")
    out = om.reconcile_run(workspace_root=ws_root, slug="demo", run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "reconciled"
    assert out["fake_green"] is False
    assert {s["step"] for s in out["steps"]} == {
        "tech_audit", "schema_audit", "on_page_audit", "cannibalization"}
    # advisory only: an attested step drops >50% of raw yet still reconciles
    tech = next(s for s in out["steps"] if s["step"] == "tech_audit")
    assert tech["high_silent_skip"] is True
    assert tech["independent_ok"] is True


def test_reconcile_run_audit_attested_step_fake_green(tmp_path):
    """The backstop FIRES on an ATTESTED step: tech_audit's committed(3) != its
    transform output(9) while the coverage verdict CLAIMS pass -> mismatch +
    fake_green. This is the whole point of generalizing the oracle to attested
    sheet-writers."""
    ws_root = tmp_path / "ws"
    wb = _audit_workbook(tmp_path / "m.xlsx")  # committed tech_seo == 3
    rec = _seed_audit_run(
        ws_root, "r1", "demo",
        _audit_happy_spec(tech_audit={"output": 9}),  # output(9) != committed(3)
        "pass",
    )
    out = om.reconcile_run(workspace_root=ws_root, slug="demo", run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "mismatch"
    assert out["fake_green"] is True
    assert out["self_reported_verdict"] == "pass"
    tech = next(s for s in out["steps"] if s["step"] == "tech_audit")
    assert om.WORKBOOK_MISMATCH in tech["failed_checks"]


def test_reconcile_run_audit_reads_output_file_not_sheet(tmp_path):
    """schema_audit's CLI writes schema_audit.json though its sheet is `schema`.
    The oracle must read the OUTPUT_FILE (schema_audit.json), NOT {sheet}.json
    (schema.json). A correct schema_audit.json + a DECOY schema.json (that would
    mismatch if wrongly read) still reconciles."""
    ws_root = tmp_path / "ws"
    wb = _audit_workbook(tmp_path / "m.xlsx")  # committed schema == 1
    rec = _seed_audit_run(ws_root, "r1", "demo", _audit_happy_spec(), "pass")
    # Decoy at the sheet-keyed path: 3 rows. If the oracle (wrongly) read
    # schema.json it would see 3 != committed(1) -> mismatch. It must NOT.
    _write_json(
        audit_suite.output_path(ws_root, "r1", "demo", "schema.json"),
        [{"i": i} for i in range(3)],
    )
    out = om.reconcile_run(workspace_root=ws_root, slug="demo", run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "reconciled"
    schema_step = next(s for s in out["steps"] if s["step"] == "schema_audit")
    assert schema_step["independent_ok"] is True
    assert schema_step["output_count"] == 1  # schema_audit.json (1), NOT schema.json (3)


def test_reconcile_run_unresolved_workflow(tmp_path):
    """A record whose steps match NO workflow -> unresolved_workflow: reported,
    excluded from the error-rate denominator, never silently passed. A PRESENT
    workbook proves the resolution check fires BEFORE the workbook check (so the
    verdict is not the unrelated workbook_absent)."""
    ws_root = tmp_path / "ws"
    wb = _seed_multi_workbook(tmp_path / "m.xlsx", {"gsc_performance": 1})
    rec = _cov_record("r1", "demo", [_cov_step("mystery_step", scored_count=1)], "pass")
    out = om.reconcile_run(workspace_root=ws_root, slug="demo", run_id="r1",
                           workbook_path=wb, coverage_record=rec)
    assert out["independent_verdict"] == "unresolved_workflow"
    assert out["fake_green"] is False  # unresolved is NOT a mismatch
    assert out["steps"] == []
    m = om.structured_error_rate([out])
    assert m["unresolved_workflow"] == 1
    assert m["reconcilable"] == 0
    assert m["error_rate"] == 0.0


# ---------------------------------------------------------------------------
# 5. reconcile_run — workbook absent (reported, never scored as pass)
# ---------------------------------------------------------------------------

def test_reconcile_run_workbook_none_is_absent(tmp_path):
    ws_root, _wb, rec, slug = _full_run(tmp_path, verdict="pass")
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=None, coverage_record=rec)
    assert out["independent_verdict"] == "workbook_absent"
    assert all(s["committed_count"] is None for s in out["steps"])
    assert out["fake_green"] is False  # absent is NOT a mismatch


def test_reconcile_run_unreadable_workbook_is_absent(tmp_path):
    ws_root, _wb, rec, slug = _full_run(tmp_path, verdict="pass")
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=tmp_path / "does_not_exist.xlsx",
                           coverage_record=rec)
    assert out["independent_verdict"] == "workbook_absent"


# ---------------------------------------------------------------------------
# 6. structured_error_rate
# ---------------------------------------------------------------------------

def test_structured_error_rate_mix():
    runs = [
        {"run_id": "a", "independent_verdict": "reconciled", "fake_green": False},
        {"run_id": "b", "independent_verdict": "reconciled", "fake_green": False},
        {"run_id": "c", "independent_verdict": "reconciled", "fake_green": False},
        {"run_id": "d", "independent_verdict": "mismatch", "fake_green": True},
        {"run_id": "e", "independent_verdict": "workbook_absent", "fake_green": False},
    ]
    m = om.structured_error_rate(runs)
    assert m["total"] == 5
    assert m["reconcilable"] == 4
    assert m["reconciled"] == 3
    assert m["mismatched"] == 1
    assert m["workbook_absent"] == 1
    assert m["error_rate"] == 0.25
    assert m["fake_green_count"] == 1
    assert m["mismatched_run_ids"] == ["d"]


def test_structured_error_rate_no_reconcilable_is_zero():
    runs = [{"run_id": "a", "independent_verdict": "workbook_absent", "fake_green": False}]
    m = om.structured_error_rate(runs)
    assert m["total"] == 1
    assert m["reconcilable"] == 0
    assert m["error_rate"] == 0.0
    assert m["workbook_absent"] == 1


def test_structured_error_rate_empty():
    m = om.structured_error_rate([])
    assert m["total"] == 0 and m["reconcilable"] == 0 and m["error_rate"] == 0.0


# ---------------------------------------------------------------------------
# 7. oracle_report — end-to-end over a coverage dir with 2 runs
# ---------------------------------------------------------------------------

def test_oracle_report_two_runs(tmp_path):
    ws_root = tmp_path / "ws"
    slug = "demo"
    wb = _seed_multi_workbook(
        tmp_path / "master.xlsx",
        {"gsc_performance": 3, "quick_wins": 2, "content_decay": 1},
    )
    rec1 = _seed_run_artifacts(ws_root, "run1", slug, {
        "gsc_pull": {"raw": 3, "output": 3, "scored": 3},
        "quick_wins": {"raw": 2, "output": 2, "scored": 2},
        "content_decay": {"raw": 1, "output": 1, "scored": 1},
    }, verdict="pass")
    _write_json(coverage_path(ws_root, slug, "run1"), rec1)
    rec2 = _seed_run_artifacts(ws_root, "run2", slug, {
        "gsc_pull": {"raw": 3, "output": 3, "scored": 3},
        "quick_wins": {"raw": 2, "output": 5, "scored": 2},  # output(5) != committed(2)
        "content_decay": {"raw": 1, "output": 1, "scored": 1},
    }, verdict="pass")  # verdict CLAIMS pass while committed data disagrees
    _write_json(coverage_path(ws_root, slug, "run2"), rec2)

    report = om.oracle_report(workspace_root=ws_root, slug=slug, workbook_path=wb)
    verdicts = {r["run_id"]: r["independent_verdict"] for r in report["runs"]}
    assert verdicts == {"run1": "reconciled", "run2": "mismatch"}
    m = report["metrics"]
    assert m["total"] == 2
    assert m["reconcilable"] == 2
    assert m["reconciled"] == 1
    assert m["mismatched"] == 1
    assert m["error_rate"] == 0.5
    assert m["fake_green_count"] == 1
    assert m["mismatched_run_ids"] == ["run2"]


def test_oracle_report_no_runs_is_empty(tmp_path):
    report = om.oracle_report(workspace_root=tmp_path / "ws", slug="demo")
    assert report["runs"] == []
    assert report["metrics"]["total"] == 0


def test_oracle_report_mixed_workflows(tmp_path):
    """A slug holding BOTH a monthly run and an audit run: each auto-resolves to
    its OWN workflow (monthly STEPS / audit STEPS) and reconciles correctly under
    one oracle_report call; the metrics aggregate both."""
    ws_root = tmp_path / "ws"
    slug = "demo"
    wb = _seed_multi_workbook(tmp_path / "master.xlsx", {
        "gsc_performance": 3, "quick_wins": 2, "content_decay": 1,   # monthly sheets
        "tech_seo": 3, "schema": 1, "on_page_audit": 5, "cannibalization": 2,  # audit
    })
    mrec = _seed_run_artifacts(ws_root, "monthly1", slug, {
        "gsc_pull": {"raw": 3, "output": 3, "scored": 3},
        "quick_wins": {"raw": 2, "output": 2, "scored": 2},
        "content_decay": {"raw": 1, "output": 1, "scored": 1},
    }, verdict="pass")
    _write_json(coverage_path(ws_root, slug, "monthly1"), mrec)
    arec = _seed_audit_run(ws_root, "audit1", slug, _audit_happy_spec(), "pass")
    _write_json(coverage_path(ws_root, slug, "audit1"), arec)

    report = om.oracle_report(workspace_root=ws_root, slug=slug, workbook_path=wb)
    verdicts = {r["run_id"]: r["independent_verdict"] for r in report["runs"]}
    assert verdicts == {"monthly1": "reconciled", "audit1": "reconciled"}
    m = report["metrics"]
    assert m["total"] == 2
    assert m["reconciled"] == 2
    assert m["reconcilable"] == 2
    assert m["error_rate"] == 0.0
    assert m["fake_green_count"] == 0


# ---------------------------------------------------------------------------
# 8. CLI smoke
# ---------------------------------------------------------------------------

def test_cli_smoke_exits_zero_and_prints_rate(tmp_path, capsys):
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="pass")
    _write_json(coverage_path(ws_root, slug, "r1"), rec)
    rc = om.main(["--workspace-root", str(ws_root), "--slug", slug,
                  "--workbook", str(wb)])
    assert rc == 0
    out = capsys.readouterr().out.lower()
    assert "error rate" in out
    assert "0.0%" in out  # all reconciled


def test_cli_smoke_no_workbook_notes_degradation(tmp_path, capsys):
    ws_root = tmp_path / "ws"
    slug = "demo"
    rec = _seed_run_artifacts(ws_root, "r1", slug, {
        "gsc_pull": {"raw": 3, "output": 3, "scored": 3},
    }, verdict="pass")
    _write_json(coverage_path(ws_root, slug, "r1"), rec)
    rc = om.main(["--workspace-root", str(ws_root), "--slug", slug])
    assert rc == 0
    out = capsys.readouterr().out
    assert "workbook_absent" in out
    assert "R4/R5" in out  # explicit no-silent-degradation note


def test_cli_smoke_surfaces_fake_green_loudly(tmp_path, capsys):
    ws_root, wb, rec, slug = _full_run(tmp_path, verdict="pass", decay_scored=9)
    _write_json(coverage_path(ws_root, slug, "r1"), rec)
    rc = om.main(["--workspace-root", str(ws_root), "--slug", slug,
                  "--workbook", str(wb)])
    assert rc == 0
    out = capsys.readouterr().out
    assert "FAKE-GREEN" in out
    assert "r1" in out


# ---------------------------------------------------------------------------
# Integration — prove the reader against the REAL committer (transaction.replace)
# ---------------------------------------------------------------------------

def test_integration_committer_replace_reconciles(tmp_path):
    """Seed the committed workbook via the REAL committer.commit and confirm the
    oracle's independent row-count reads exactly what transaction.replace wrote."""
    from scripts.orchestration import committer

    slug = "demo"
    ws_root = tmp_path / "ws"
    project_dir = ws_root / "projects" / slug
    project_dir.mkdir(parents=True)
    wb_path = project_dir / "master.xlsx"

    def _gsc_row(i):
        return {
            "url": f"https://example.test/p{i}",
            "clicks_recent": i, "clicks_previous": 0, "clicks_delta": i,
            "clicks_delta_pct": 0.0, "impressions_recent": i * 10,
            "impressions_previous": 0, "impressions_delta": i * 10,
            "ctr_recent": 0.1, "position_recent": 5.0, "position_previous": 6.0,
            "note": "",
        }

    rows = [_gsc_row(i) for i in range(1, 4)]  # 3 committed rows
    committer.commit(wb_path, "gsc_performance", rows,
                     run_id="r1", project_slug=slug)

    _seed_raw(ws_root, "r1", slug, "gsc_pull", _raw_drop("r1", slug, 3))
    _seed_output(ws_root, "r1", slug, "gsc_performance", 3)
    rec = _cov_record("r1", slug, [_cov_step("gsc_pull", scored_count=3)], "pass")

    assert om.committed_row_count(wb_path, "gsc_performance", 5) == 3
    # This test's contract is the ROW-COUNTER vs the real committer, not workflow
    # resolution. The single-step record is a minimal probe, not a real monthly run
    # (which always carries all 3 STEPS), so it would not strict-subset-resolve to a
    # workflow. Pass `steps=` explicitly — the back-compat path the oracle preserves
    # for exactly this — to isolate the row-counter assertion from resolution.
    out = om.reconcile_run(workspace_root=ws_root, slug=slug, run_id="r1",
                           workbook_path=wb_path, coverage_record=rec,
                           steps=[{"name": "gsc_pull", "sheet": "gsc_performance"}])
    assert out["independent_verdict"] == "reconciled"
    assert out["steps"][0]["committed_count"] == 3
