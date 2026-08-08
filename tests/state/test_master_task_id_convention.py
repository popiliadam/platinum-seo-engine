"""tests/state/test_master_task_id_convention.py — master_task.task_id ratchet.

`rules/master-task-id.md` is marked `status: enforced` and fixes the canonical
form `^T-[0-9]{4,}$`, with ONE named legacy exception: the `MT-W3W2B-NNN` batch
created during Phase 14 W3-W2-B, which the rule says cannot be corrected
retroactively because the state is append-only.

Nothing checked it. The rule's own closing line admits as much — "drift-check
F-XX aday: master_task.task_id pattern cross-check (Phase 16+ scope)" — and in
its absence a THIRD, FOURTH and FIFTH format appeared: a 2026-08-08 sweep of the
bound workspace found 125 non-conforming ids across six projects in six shapes
(`MT-001`, `QW-001`, `T-301-01`, `T-PIL-PEK-01`, `T-0026-V`, `T-PIVOT-15G`).

This file is a RATCHET, not a clean-up. Renaming 125 live ids is an operator
decision with real consequences — archived `events.jsonl` rows and
`completed_work` entries reference them, and the ledgers are append-only, so a
rename creates dangling references that cannot be repaired in place. So the
existing debt is PINNED per project, by count, and the gate fails when it GROWS.
It also fails when it SHRINKS, which is the only way a ratchet stays honest: the
pin has to be lowered deliberately, with the fix, rather than drifting loose.

Widening the pattern to admit the shapes already on disk was considered and
rejected: that is loosening the gate to reach green, and it would bless
ad-hoc id invention as a convention.
"""
from __future__ import annotations

import os
import re
import warnings
from pathlib import Path

import pytest

# Mirrors schemas/master-excel.schema.json#/definitions/taskIdPattern and the
# events.schema.json legacy bypass. Kept as literals rather than imported so a
# schema edit cannot silently relax this gate too.
_CANONICAL = re.compile(r"^T-[0-9]{4,}$")
_LEGACY_W3W2B = re.compile(r"^MT-W[0-9]+W[0-9]+[A-Z]?-[0-9]+$")

# Non-conforming ids ALREADY on disk, per project, measured 2026-08-08.
# Lower a number when the ids are actually fixed; never raise one to make the
# suite green — a rise is new drift and is exactly what this gate exists for.
KNOWN_DRIFT: dict[str, int] = {
    "bayder": 54,        # MT-NNN
    "bigcat-tr": 5,      # QW-NNN
    "dentnotion": 64,    # T-NNN-NN, T-AAA-AAA-NN, MT-FIYAT-NN (19 of them)
    "katrenur-tr": 1,    # T-0026-V
    "rkturizm-tr": 1,    # T-PIVOT-15G
}


def conforms(task_id: str) -> bool:
    """True when ``task_id`` is canonical or the ONE named legacy batch."""
    if not isinstance(task_id, str):
        return False
    s = task_id.strip()
    return bool(_CANONICAL.match(s) or _LEGACY_W3W2B.match(s))


def nonconforming(task_ids) -> list[str]:
    """The subset of ``task_ids`` that violates the convention, order preserved."""
    return [t for t in task_ids if not conforms(t)]


# --- the classifier itself (pure, runs everywhere including CI) --------------

@pytest.mark.parametrize("task_id", ["T-0001", "T-10001", "T-999999", "MT-W3W2B-001", "MT-W12W4A-77"])
def test_conforming_ids_are_accepted(task_id: str) -> None:
    assert conforms(task_id)


@pytest.mark.parametrize("task_id", [
    "MT-001",         # bayder
    "QW-001",         # bigcat-tr
    "T-301-01",       # dentnotion
    "T-PIL-PEK-01",   # dentnotion
    "MT-FIYAT-01",    # dentnotion — NOT the W3W2B batch the rule exempts
    "T-0026-V",       # katrenur-tr
    "T-PIVOT-15G",    # rkturizm-tr
    "T-001",          # too few digits for the canonical form
    "t-0001",         # wrong case
    "",
])
def test_ad_hoc_ids_are_rejected(task_id: str) -> None:
    assert not conforms(task_id)


def test_nonconforming_preserves_only_the_offenders() -> None:
    assert nonconforming(["T-0001", "MT-001", "MT-W3W2B-002", "QW-001"]) == ["MT-001", "QW-001"]


# --- the ratchet over the bound workspace ------------------------------------

def _workspace_root() -> Path | None:
    raw = os.environ.get("PSEO_WORKSPACE_ROOT")
    if not raw:
        return None
    p = Path(raw).expanduser()
    return p if p.exists() else None


def _master_task_ids(workbook: Path) -> list[str]:
    """Every task_id in the workbook's master_task sheet (header row skipped)."""
    import openpyxl

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        if "master_task" not in wb.sheetnames:
            return []
        rows = list(wb["master_task"].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = header.index("task_id") if "task_id" in header else 0
    out: list[str] = []
    for row in rows[1:]:
        value = row[idx] if idx < len(row) else None
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "task_id":
            out.append(text)
    return out


def _workbooks() -> list[Path]:
    root = _workspace_root()
    if root is None:
        return []
    return sorted((root / "projects").glob("*/master.xlsx"))


@pytest.mark.parametrize("workbook", _workbooks(), ids=lambda p: p.parent.name if p else "<none>")
def test_task_id_drift_does_not_grow(workbook: Path) -> None:
    """Per-project non-conforming count must equal its pin — exactly."""
    slug = workbook.parent.name
    offenders = nonconforming(_master_task_ids(workbook))
    pinned = KNOWN_DRIFT.get(slug, 0)
    if len(offenders) > pinned:
        fresh = offenders[pinned:] if pinned else offenders
        pytest.fail(
            f"{slug}: NEW task_id drift — {len(offenders)} non-conforming ids, "
            f"pinned at {pinned}. rules/master-task-id.md fixes the canonical "
            f"form ^T-[0-9]{{4,}}$; the only exempt legacy batch is MT-W3W2B-NNN. "
            f"Do NOT raise the pin to go green. Examples: {fresh[:5]}"
        )
    assert len(offenders) == pinned, (
        f"{slug}: {len(offenders)} non-conforming ids but pinned at {pinned} — "
        f"drift was fixed without lowering the pin. Set KNOWN_DRIFT['{slug}'] = "
        f"{len(offenders)} so the ratchet keeps measuring."
    )


def test_unbound_workspace_is_reported_as_zero_coverage() -> None:
    """Unbound means this ratchet checked NOTHING — say so rather than pass quietly."""
    if _workspace_root() is not None:
        return
    assert _workbooks() == []
    warnings.warn(
        "master_task.task_id convention NOT CHECKED — PSEO_WORKSPACE_ROOT is "
        "unbound, so the ratchet contributed zero cases. Coverage gap, not a pass.",
        UserWarning,
        stacklevel=2,
    )
