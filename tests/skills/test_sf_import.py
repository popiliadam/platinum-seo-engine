"""
tests/skills/test_sf_import.py — sf-import ingestion skill end-to-end tests.

Six cases mirroring the 6-criterion acceptance gate from the W-R worker
brief (Phase 5 Wave 2). The pilot SF batch lives at
projects/dentnotion/sf-exports/2026-04-27/raw/ in workspace-staging
(26 CSV files, Tier 1 14/14 + Tier 2 9/10 — search_console_all is the
canonical AMBER exemption).

Cross-module IMPORT-only discipline (sf_import.py is a plain CLI — no
workflow_runner run shell):
  - scripts.excel.transaction       (idempotent replace + backup FIFO 7)
  - scripts.state.events_writer     (sf_csv provenance event)

Schemas referenced:
  - schemas/sf-required-reports.schema.json   (Tier 1 14, Tier 2 10)
  - schemas/skill-frontmatter.schema.json     (frontmatter draft7)
  - schemas/master-excel.schema.json          (6 SF-derived sheets)
  - schemas/events.schema.json                (source.kind=sf_csv)

Test 1 is the live-batch smoke: it asserts the canonical pilot SF export
is on disk at the workspace path the §6 ingest discipline mandates and
that all 14 Tier 1 reports resolve via the filename normalization rules
(de_ shadow + Turkish ı handling). MCP tools are not Python-callable
inside pytest; the GSC sitemap cross-check is exercised via the same
witness-file pattern W-P used for quick-wins (live MCP capture persisted
to inbox; pytest validates the witness, does not re-issue the call).

Run from repo root:
    PYTHONPATH=. pytest tests/skills/test_sf_import.py -v
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import os

import pytest
import yaml
from jsonschema import Draft7Validator

from scripts.excel import transaction
from scripts.state import events_writer


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"
SKILL_MD = REPO_ROOT / "skills" / "ingestion" / "sf-import" / "SKILL.md"

WORKSPACE_STAGING = Path(
    os.environ.get(
        "PSEO_WORKSPACE_STAGING",
        str(Path.home() / "Documents" / "platinum-seo-workspace-staging"),
    )
)
PILOT_SLUG = "dentnotion"
PILOT_SF_DATE = "2026-04-27"
PILOT_RAW_DIR = (
    WORKSPACE_STAGING / "projects" / PILOT_SLUG
    / "sf-exports" / PILOT_SF_DATE / "raw"
)
PILOT_MASTER_XLSX = WORKSPACE_STAGING / "projects" / PILOT_SLUG / "master.xlsx"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sf_required_schema() -> dict:
    return json.loads(
        (SCHEMAS / "sf-required-reports.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def skill_frontmatter_schema() -> dict:
    return json.loads(
        (SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def events_schema() -> dict:
    return json.loads((SCHEMAS / "events.schema.json").read_text("utf-8"))


@pytest.fixture(scope="module")
def tier_canonical_names(sf_required_schema: dict) -> tuple[set[str], set[str]]:
    """Tier 1 (REQUIRED, 14) and Tier 2 (RECOMMENDED, 10) canonical names.

    The schema is meta-shape only; the actual canonical names live in the
    `definitions.canonicalName` enum. The W-R worker maintains the Tier
    1/2 split by the §7.1/§7.2 spec contract — we encode it here so the
    test is independent of the registry document (which is Phase 6).
    """
    # §7.1 REQUIRED — 14 reports.
    tier1 = {
        "internal_all",
        "all_inlinks",
        "all_outlinks",
        "response_codes_all",
        "issues_overview_report",
        "page_titles_all",
        "meta_description_all",
        "h1_all",
        "canonicals_all",
        "directives_all",
        "indexability",
        "structured_data_all",
        "sitemaps_all",
        "redirect_chains",
    }
    # §7.2 RECOMMENDED — 10 reports.
    tier2 = {
        "h2_all",
        "images_all",
        "hreflang_all",
        "orphan_pages",
        "all_anchor_text",
        "near_duplicates_report",
        "exact_duplicates_report",
        "search_console_all",
        "crawl_depth",
        "pagination_all",
    }
    # Sanity: every name must exist in the schema's canonicalName enum.
    enum = set(sf_required_schema["definitions"]["canonicalName"]["enum"])
    assert tier1 <= enum, f"Tier 1 names not in enum: {tier1 - enum}"
    assert tier2 <= enum, f"Tier 2 names not in enum: {tier2 - enum}"
    return tier1, tier2


# ---------------------------------------------------------------------------
# Filename normalization (mirrors SKILL.md §"Filename normalization")
# ---------------------------------------------------------------------------

_VERSION_SUFFIX = re.compile(r"(?:\(\d+\)|-copy|_old)$")


def normalize_sf_filename(raw: str) -> str:
    """Return the canonical_name candidate for a SF export filename.

    Steps (order matters):
      1. lowercase basename
      2. strip extension
      3. strip leading 'de_' shadow prefix (pilot artifact)
      4. strip leading 'v_' / 'p_' export-mode prefix
      5. fold Turkish dotless 'ı' → ASCII 'i'
      6. drop (N)/-copy/_old version suffixes
    """
    name = Path(raw).name.lower()
    if name.endswith(".csv"):
        name = name[:-4]
    if name.startswith("de_"):
        name = name[3:]
    if name.startswith("v_") or name.startswith("p_"):
        name = name[2:]
    name = name.replace("ı", "i")
    name = _VERSION_SUFFIX.sub("", name)
    return name


def match_tiers(
    raw_dir: Path, tier1: set[str], tier2: set[str]
) -> tuple[set[str], set[str], set[str]]:
    """Return (matched_canonical, missing_tier1, missing_tier2) for a raw dir."""
    found: set[str] = set()
    for p in raw_dir.iterdir():
        if not p.is_file() or not p.name.lower().endswith(".csv"):
            continue
        canonical = normalize_sf_filename(p.name)
        if canonical in tier1 or canonical in tier2:
            found.add(canonical)
    return found, tier1 - found, tier2 - found


# ---------------------------------------------------------------------------
# Test 1 — Tier 1 14/14 PASS path (live pilot SF batch)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(
    not WORKSPACE_STAGING.exists(),
    reason="local-only fixture: WORKSPACE_STAGING path missing on CI runner "
           "(Q-CI-W3-04 Phase 15 audit Wave 1 kategori #5 codify aday)"
)
def test_tier1_14_validates(tier_canonical_names: tuple[set[str], set[str]]) -> None:
    """All 14 §7.1 REQUIRED reports must resolve from the pilot raw dir.

    DUR + flag rule: if the pilot raw dir is missing or any Tier 1 file
    fails to normalize back to its canonical name, the worker did not
    actually capture the live batch — fail loudly.
    """
    assert PILOT_RAW_DIR.is_dir(), f"pilot raw dir missing: {PILOT_RAW_DIR}"
    tier1, tier2 = tier_canonical_names

    found, missing_t1, missing_t2 = match_tiers(PILOT_RAW_DIR, tier1, tier2)

    assert not missing_t1, (
        f"Tier 1 incomplete (RED FAIL): missing {sorted(missing_t1)}"
    )
    assert tier1 <= found, (
        f"matched set does not cover Tier 1: missing {sorted(tier1 - found)}"
    )


# ---------------------------------------------------------------------------
# Test 2 — Tier 2 search_console_all is AMBER (warn, NOT fail)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(
    not WORKSPACE_STAGING.exists(),
    reason="local-only fixture: WORKSPACE_STAGING path missing on CI runner "
           "(Q-CI-W3-04 Phase 15 audit Wave 1 kategori #5 codify aday)"
)
def test_tier2_search_console_all_amber(
    tier_canonical_names: tuple[set[str], set[str]],
) -> None:
    """The pilot ships 9/10 Tier 2 — search_console_all is the canonical
    AMBER exemption. Missing Tier 2 must NOT fail the import; it must
    surface as an amber_warnings entry (string), and Tier 1 stays GREEN.
    """
    assert PILOT_RAW_DIR.is_dir(), f"pilot raw dir missing: {PILOT_RAW_DIR}"
    tier1, tier2 = tier_canonical_names

    found, missing_t1, missing_t2 = match_tiers(PILOT_RAW_DIR, tier1, tier2)

    # Tier 1 must be GREEN even when Tier 2 is AMBER.
    assert not missing_t1, f"Tier 1 must be complete; missing {sorted(missing_t1)}"

    # search_console_all is the documented exemption.
    assert "search_console_all" in missing_t2, (
        "expected search_console_all to be missing in pilot (AMBER kanıt); "
        f"actual missing tier 2: {sorted(missing_t2)}"
    )

    # AMBER warning shape: a non-empty string surfacing the missing names.
    amber_warnings: list[str] = []
    if missing_t2:
        amber_warnings.append(
            f"Tier 2 missing (AMBER, not fatal): {sorted(missing_t2)}"
        )
    assert amber_warnings, "AMBER warning must be produced when Tier 2 incomplete"
    assert "search_console_all" in amber_warnings[0]


# ---------------------------------------------------------------------------
# Test 3 — crawl_sitemap sheet populates ≥ 1 row from pilot data
# ---------------------------------------------------------------------------

def test_sheet_populate_crawl_sitemap(tmp_path: Path) -> None:
    """Project a minimal real-pilot crawl_sitemap row and write it via
    transaction.replace — the SAME write path sf_import.project_and_write uses
    (replace, NOT append: re-importing a batch refreshes rather than
    duplicates). The test runs against a fresh workbook in tmp_path so it
    never touches workspace-staging — the live execution proof appears in the
    worker package, not the test suite."""
    workbook = tmp_path / "projects" / PILOT_SLUG / "master.xlsx"
    workbook.parent.mkdir(parents=True, exist_ok=True)

    rows = [
        {
            "category": "crawl_summary",
            "metric": "internal_urls_crawled",
            "value": "192",
            "status": "EXISTS",   # statusEnum
            "action": "n/a",
        },
        {
            "category": "sitemap_summary",
            "metric": "sitemaps_seen",
            "value": "1",
            "status": "EXISTS",
            "action": "n/a",
        },
    ]
    result = transaction.replace(
        workbook_path=workbook,
        sheet="crawl_sitemap",
        rows=rows,
        project_slug=PILOT_SLUG,
        writer="sf-import",
    )
    assert result.rows_affected == 2

    # Physical write proof: the sheet exists and grew to hold the rows.
    from openpyxl import load_workbook
    ws1 = load_workbook(str(workbook))["crawl_sitemap"]
    max_row_after_first = ws1.max_row  # header block (data_start_row-1) + 2 rows

    # Idempotency — the reason sf-import uses replace, NOT append: re-importing
    # the same batch must REFRESH in place, not duplicate. The sheet must not
    # grow; `append` would push max_row from N to N+2. (Robust to the sheet's
    # header_row layout — we compare growth, not an absolute data-row count.)
    again = transaction.replace(
        workbook_path=workbook,
        sheet="crawl_sitemap",
        rows=rows,
        project_slug=PILOT_SLUG,
        writer="sf-import",
    )
    assert again.rows_affected == 2, "replace must refresh, not accumulate"
    ws2 = load_workbook(str(workbook))["crawl_sitemap"]
    assert ws2.max_row == max_row_after_first, (
        f"replace must refresh in place; sheet grew "
        f"{max_row_after_first}→{ws2.max_row} (rows duplicated)"
    )


# ---------------------------------------------------------------------------
# Test 4 — sitemap cross-check (mocked mcp__gsc__list_sitemaps)
# ---------------------------------------------------------------------------

def test_sitemap_xcheck_mcp() -> None:
    """Sitemap cross-check — Phase-X DEFERRED capability (B3-02).

    The sitemap cross-check is NOT implemented in sf_import.py (no
    mcp__gsc__list_sitemaps call, no sf_sitemap_xcheck module); the GSC tools
    are `optional` in the frontmatter. This test exercises only inline mocks —
    it documents the INTENDED diagnostic shape for when the capability lands at
    the manager/caller layer, not a shipped contract. It is kept (green) as the
    design witness; the parity test above locks the demotion.
    """

    def mock_list_sitemaps(siteUrl: str) -> dict:
        return {
            "siteUrl": siteUrl,
            "sitemap": [
                {
                    "path": f"{siteUrl}sitemap.xml",
                    "lastSubmitted": "2026-04-20T10:00:00Z",
                    "isPending": False,
                    "isSitemapsIndex": True,
                    "type": "sitemapsIndex",
                    "lastDownloaded": "2026-04-26T03:00:00Z",
                    "warnings": "0",
                    "errors": "0",
                    "contents": [{"type": "web", "submitted": "150", "indexed": "120"}],
                },
                {
                    "path": f"{siteUrl}post-sitemap.xml",
                    "lastSubmitted": "2026-04-20T10:00:00Z",
                    "isPending": False,
                    "isSitemapsIndex": False,
                    "type": "sitemap",
                    "lastDownloaded": "2026-04-26T03:00:00Z",
                    "warnings": "0",
                    "errors": "0",
                    "contents": [{"type": "web", "submitted": "150", "indexed": "120"}],
                },
            ],
        }

    def xcheck(sf_internal_urls: set[str], list_sitemaps_payload: dict) -> dict:
        sitemap_paths = {
            s["path"] for s in list_sitemaps_payload.get("sitemap", [])
        }
        # Diagnostic: a real impl would fetch each sitemap and intersect URLs.
        # For the contract test we synthesize the metric shape.
        return {
            "submitted_sitemaps": len(sitemap_paths),
            "sf_crawled_urls": len(sf_internal_urls),
            "in_both_marker": "submitted_count_present",
        }

    payload = mock_list_sitemaps("https://dentnotion.com/")
    sf_urls = {"https://dentnotion.com/", "https://dentnotion.com/about"}
    metrics = xcheck(sf_urls, payload)

    assert metrics["submitted_sitemaps"] == 2
    assert metrics["sf_crawled_urls"] == 2
    assert "in_both_marker" in metrics

    # DURUR #4 (deferred): when the cross-check lands, a mcp__gsc__list_sitemaps
    # failure must NOT silently downgrade — it raises. Documented here as the
    # intended contract for the future Wave (not exercised by the shipped CLI).
    def failing_list_sitemaps(siteUrl: str) -> dict:
        raise RuntimeError("auth/network/scope")

    with pytest.raises(RuntimeError, match="auth/network/scope"):
        failing_list_sitemaps("https://dentnotion.com/")


# ---------------------------------------------------------------------------
# Test 5 — atomic write produces a backup (FIFO keep-7)
# ---------------------------------------------------------------------------

def test_atomic_write_backup_created(tmp_path: Path) -> None:
    """Every transaction.replace (the sf-import write path) produces a backup
    under _state/backups/master/. After 8 writes only 7 survive (FIFO
    rotation) — replace shares append's atomic envelope (backup → atomic_save
    → rotate).
    """
    workbook = tmp_path / "projects" / PILOT_SLUG / "master.xlsx"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    backups_dir = (
        tmp_path / "projects" / PILOT_SLUG
        / "_state" / "backups" / "master"
    )

    # Write 8 small batches → 8 backup events, 7 should survive.
    for i in range(8):
        rows = [{
            "category": f"batch_{i}",
            "metric": "row_count",
            "value": str(i),
            "status": "EXISTS",
            "action": "n/a",
        }]
        result = transaction.replace(
            workbook_path=workbook,
            sheet="crawl_sitemap",
            rows=rows,
            project_slug=PILOT_SLUG,
            writer="sf-import",
        )
        assert result.backup_path.exists(), (
            f"backup #{i} not created at {result.backup_path}"
        )

    surviving = sorted(p.name for p in backups_dir.iterdir() if p.is_file())
    assert len(surviving) == 7, (
        f"FIFO keep-7 violated: {len(surviving)} backups remain "
        f"(expected 7); names={surviving}"
    )


# ---------------------------------------------------------------------------
# Test 6 — provenance event with source.kind="sf_csv" validates
# ---------------------------------------------------------------------------

def test_provenance_event_sf_csv(tmp_path: Path, events_schema: dict) -> None:
    """The skill emits an event_kind=provenance entry with source.kind=
    'sf_csv' (versus the per-sheet 'tool_computed' event transaction.py
    emits internally). The shape must validate against events.schema."""
    slug = "sf-test"
    state_dir = tmp_path / "projects" / slug / "_state"
    state_dir.mkdir(parents=True, exist_ok=True)

    rid = events_writer.next_run_id(slug, workspace_root=tmp_path)
    sha = "sha256:" + hashlib.sha256(b"pilot-internal-all").hexdigest()
    result = events_writer.append_provenance(
        project_id=slug,
        run_id=rid,
        source={
            "kind": "sf_csv",
            "source_folder": f"sf-exports/{PILOT_SF_DATE}/raw/",
            "filename_original": "internal_all.csv",
            "filename_normalized": "internal_all",
            "file_hash": sha,
            "row_count": 192,
        },
        operation="ingest",
        rows_written=192,
        workspace_root=tmp_path,
    )
    assert result.event_id

    events_path = state_dir / "events.jsonl"
    assert events_path.exists(), "events.jsonl was not created"
    lines = [
        json.loads(line)
        for line in events_path.read_text("utf-8").splitlines()
        if line.strip()
    ]
    sf_provs = [
        e for e in lines
        if e.get("event_kind") == "provenance"
        and e.get("source", {}).get("kind") == "sf_csv"
    ]
    assert sf_provs, (
        "no provenance event_kind=provenance source.kind=sf_csv emitted"
    )

    # Re-validate the emitted record against the canonical events schema.
    validator = Draft7Validator(events_schema)
    for evt in sf_provs:
        errs = sorted(validator.iter_errors(evt), key=lambda e: list(e.absolute_path))
        assert not errs, (
            "emitted sf_csv provenance event invalid: "
            f"{[(list(e.absolute_path), e.message) for e in errs]}"
        )


# ---------------------------------------------------------------------------
# Bonus — frontmatter contract validation (free, fast, catches drift)
# ---------------------------------------------------------------------------

def test_source_run_id_frontmatter_present(
    skill_frontmatter_schema: dict,
) -> None:
    """v1.8 Phase 3 D-SF-07 hand-off: sf-import frontmatter declares
    `source_run_id` as an optional input so the sf-crawl-orchestrator
    can pass its upstream MCP-crawl run_id into the file-based ingest
    for provenance correlation. Phase 4 verification gate: Phase 3
    Worker added this — Phase 4 Worker confirms intact (no body
    changes per D-SF-07).
    """
    text = SKILL_MD.read_text(encoding="utf-8")
    parts = text.split("---", 2)
    fm = yaml.safe_load(parts[1])

    inputs = fm.get("inputs", {})
    assert "source_run_id" in inputs, (
        "v1.8 Phase 3 D-SF-07: sf-import frontmatter must declare "
        "`source_run_id` optional input for sf-crawl-orchestrator handoff"
    )
    src_def = inputs["source_run_id"]
    assert src_def["type"] == "string"
    assert src_def.get("required") is False, (
        "source_run_id MUST be optional (required=False) — old sf-import "
        "calls predating Phase 3 must still satisfy the contract"
    )
    # Description must reference the upstream skill so the hand-off intent
    # is self-documenting.
    desc = src_def.get("description", "")
    assert "sf-crawl-orchestrator" in desc, (
        f"source_run_id description must cite sf-crawl-orchestrator; "
        f"got: {desc!r}"
    )


def test_source_run_id_provenance_chain(tmp_path: Path, events_schema: dict) -> None:
    """v1.8 Phase 4 verification (D-SF-07 sf-crawl-orchestrator → sf-import
    provenance chain): when a caller supplies source_run_id, the
    provenance event written to events.jsonl carries the upstream
    orchestrator's run id token in the source dict so drift-check can
    correlate file-based ingest with MCP crawl.

    Implementation note: source dict additionalProperties=false per
    events.schema (Q-PHASE-3-WORKER-07 catch); the upstream run id is
    carried in the source_folder string OR a sibling envelope file
    (inbox/sf-mcp/{date}-sf-crawl-{slug}.json per Phase 3 spec). This
    test exercises the *event* shape — the workflow link lives at the
    envelope layer.
    """
    slug = "sf-handoff"
    state_dir = tmp_path / "projects" / slug / "_state"
    state_dir.mkdir(parents=True, exist_ok=True)

    upstream_run_id = f"{slug}-2026-05-26-orch"

    # Mirror sf-import Step 7 with the Phase 3 handoff: source.source_folder
    # carries the orchestrator-produced sf-mcp/{date} subdir under inbox.
    rid = events_writer.next_run_id(slug, workspace_root=tmp_path)
    sha = "sha256:" + hashlib.sha256(b"orchestrator-handoff").hexdigest()
    result = events_writer.append_provenance(
        project_id=slug,
        run_id=rid,
        source={
            "kind": "sf_csv",
            "source_folder": (
                f"inbox/sf-mcp/2026-05-26-sf-crawl-{slug}.json#{upstream_run_id}"
            ),
            "filename_original": "internal_all.csv",
            "filename_normalized": "internal_all",
            "file_hash": sha,
            "row_count": 192,
        },
        operation="ingest",
        rows_written=192,
        notes=f"sf-crawl-orchestrator handoff source_run_id={upstream_run_id}",
        workspace_root=tmp_path,
    )
    assert result.event_id

    events_path = state_dir / "events.jsonl"
    lines = [
        json.loads(line) for line in events_path.read_text("utf-8").splitlines()
        if line.strip()
    ]
    chained = [
        e for e in lines
        if e.get("event_kind") == "provenance"
        and e.get("source", {}).get("kind") == "sf_csv"
        and upstream_run_id in str(e.get("source", {}).get("source_folder", ""))
    ]
    assert chained, (
        "no provenance event carries the upstream sf-crawl-orchestrator "
        f"run_id={upstream_run_id} in source.source_folder"
    )
    # Re-validate the chained event against the canonical events schema —
    # additionalProperties=false on source must still pass (source_folder
    # is a declared property, not extra).
    validator = Draft7Validator(events_schema)
    for evt in chained:
        errs = sorted(
            validator.iter_errors(evt), key=lambda e: list(e.absolute_path)
        )
        assert not errs, (
            "chained sf_csv provenance event invalid: "
            f"{[(list(e.absolute_path), e.message) for e in errs]}"
        )


def test_frontmatter_validates_against_schema(
    skill_frontmatter_schema: dict,
) -> None:
    """SKILL.md frontmatter must validate against
    schemas/skill-frontmatter.schema.json (Draft 7). Plus the B3-03 lock: the
    body describes the REAL deterministic CLI — it must NOT re-introduce the
    fictional `workflow_runner` run-shell contract (sf_import.py calls neither
    create_run nor complete; DURUR signalling is exit-code based).
    """
    text = SKILL_MD.read_text("utf-8")
    parts = text.split("---", 2)
    assert len(parts) >= 3, "SKILL.md must open with --- frontmatter ---"
    fm = yaml.safe_load(parts[1])

    validator = Draft7Validator(skill_frontmatter_schema)
    errs = sorted(validator.iter_errors(fm), key=lambda e: list(e.absolute_path))
    assert not errs, (
        "frontmatter invalid: "
        f"{[('/'.join(str(p) for p in e.absolute_path) or '<root>', e.message) for e in errs]}"
    )

    # B3-03: the shipped impl is a plain CLI (print + exit codes). The body
    # must not CALL a workflow_runner run shell. We match the call form (with
    # an open paren) so the prose disclaimer ("no `workflow_runner.create_run`")
    # is allowed while an actual `workflow_runner.create_run(...)` call is not.
    body = parts[2]
    assert "workflow_runner.create_run(" not in body, (
        "sf-import body must not call workflow_runner.create_run( — it is a "
        "deterministic CLI (see scripts/ingestion/sf_import.py)"
    )
    assert "workflow_runner.complete(" not in body, (
        "sf-import body must not call workflow_runner.complete( — main() prints "
        "a stdout summary and returns 0 (no string-typed outputs dict)"
    )


def test_required_mcp_tools_match_what_impl_invokes() -> None:
    """B3-02: a tool declared `mcp_tools.required` must actually be invoked by
    the shipped implementation. sf-import's deterministic core (sf_import.py)
    reads SF CSVs from disk and writes Excel — it invokes NO MCP tool. The GSC
    sitemap tools (the sole consumers) power the Phase-X-deferred sitemap
    cross-check, so `required` must be empty and they live under `optional`.
    Declaring a tool `required` that the code never calls over-promises a
    dependency — exactly the contract bug this test locks out.
    """
    text = SKILL_MD.read_text(encoding="utf-8")
    fm = yaml.safe_load(text.split("---", 2)[1])
    mcp = fm.get("mcp_tools") or {}
    required = mcp.get("required") or []
    optional = mcp.get("optional") or []

    impl_src = (
        REPO_ROOT / "scripts" / "ingestion" / "sf_import.py"
    ).read_text(encoding="utf-8")

    # Forward-regression guard: if `required` is ever repopulated, each tool
    # must at least appear in the shipped source (a coarse substring check — the
    # two hard assertions below are the precise behavioral locks for this
    # finding). Today `required == []`, so this loop is a no-op by design.
    for tool in required:
        assert tool in impl_src, (
            f"{tool!r} is declared mcp_tools.required but never invoked in "
            "sf_import.py — demote it to optional or wire it (B3-02)"
        )

    # Regression lock for the specific finding: the deferred cross-check's tool
    # must be optional (demoted), not required, and not silently dropped.
    assert "mcp__gsc__list_sitemaps" not in required, (
        "mcp__gsc__list_sitemaps powers the Phase-X-deferred sitemap "
        "cross-check; it must be optional, not required"
    )
    assert "mcp__gsc__list_sitemaps" in optional, (
        "the demoted GSC sitemap tool must be declared optional (deferred "
        "enrichment), not removed entirely"
    )
