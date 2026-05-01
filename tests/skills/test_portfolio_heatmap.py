"""tests/skills/test_portfolio_heatmap.py — Phase 9 W-E8 reporting skill.

Mirrors tests/skills/test_portfolio_overview.py + test_portfolio_weekly_brief.py:
  - frontmatter parse + Draft 7 validate against
    schemas/skill-frontmatter.schema.json
  - natural_language min length sentinel (≥ 30 char)
  - aggregate_dimension enum 3 sentinel (accept category/priority/
    primary_source, reject "invalid")
  - missing-sheet schema-valid empty shape (W-E1 paterni reuse)
  - multi-project merge (3 projects with varied sheet density)
  - density normalization (per-cell count → density score)
  - primary_source 10-enum coverage (each enum gets a bucket)
  - forbidden tokens guard (4 tokens + 8 base64-obfuscated slug list)
  - assert_read_only_module() helper (W-E3 paterni)
  - path convention compliance (projects/_portfolio/{outputs,inbox}/local/
    + PSEO_WORKSPACE_ROOT env)
  - idempotency (same input → byte-identical snapshot)

This is a READ-ONLY reporting skill — no MCP, no DFS, no budget
pre-flight, no master.xlsx writes, no events.jsonl writes.
"""

from __future__ import annotations

import base64
import json
import os
import re
from pathlib import Path

import pytest
import yaml
from jsonschema import Draft7Validator

from scripts.reporting import portfolio_heatmap as ph


# ---------------------------------------------------------------------------
# Constants — paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"
SKILL_PATH = REPO_ROOT / "skills" / "reporting" / "portfolio-heatmap" / "SKILL.md"
TRANSFORM_PATH = REPO_ROOT / "scripts" / "reporting" / "portfolio_heatmap.py"
TEMPLATE_PATH = REPO_ROOT / "templates" / "reports" / "portfolio-heatmap.template.md"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def skill_frontmatter_schema() -> dict:
    return json.loads(
        (SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def portfolio_config_schema() -> dict:
    return json.loads(
        (SCHEMAS / "portfolio-config.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def skill_frontmatter() -> dict:
    """Parse the YAML frontmatter block of the portfolio-heatmap SKILL.md."""
    text = SKILL_PATH.read_text(encoding="utf-8")
    match = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert match, "SKILL.md missing YAML frontmatter delimiters"
    return yaml.safe_load(match.group(1))


def _synthetic_config(active_count: int = 3) -> dict:
    """Build a valid portfolio.config payload with N active projects.
    Uses generic 'alpha-test' / 'beta-test' / 'gamma-test' fixture
    slugs (NOT real client slugs)."""
    template_slugs = (
        "alpha-test", "beta-test", "gamma-test",
        "delta-test", "epsilon-test", "zeta-test",
        "eta-test", "theta-test",
    )
    active = []
    for i in range(active_count):
        slug = template_slugs[i]
        active.append({
            "slug": slug,
            "display_name": f"Project {slug}",
            "workspace_path": slug,
            "profile": ["b2b-saas"],
            "priority": i + 1,
        })
    return {
        "schema_version": "1.1",
        "portfolio_id": "test-portfolio",
        "display_name": "Test Portfolio",
        "active_projects": active,
        "slas": {
            "weekly_sync_max_days": 7,
            "monthly_roundup_target_dom": 5,
        },
        "cadence": {
            "weekly_brief": {"day": "Monday", "hour": 9},
            "monthly_roundup": {"day_of_month": 5, "hour": 10},
        },
        "cross_query": {"read_only": True},
    }


def _build_workbook_with_density(
    *, path: Path,
    opportunity_rows: int,
    quick_wins_rows: int,
    content_decay_rows: int,
    cannibalization_rows: int,
    master_task_payload: list[dict] | None = None,
) -> None:
    """Build a tiny synthetic master.xlsx with the 4 opportunity sheets +
    master_task. Counts of B/J/E/F filled cells match the kwargs.
    master_task_payload entries: {task, primary_source, category,
    priority}.
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Default sheet → repurpose as 'opportunity'.
    ws_default = wb.active
    ws_default.title = "opportunity"
    # opportunity: header_row=4, data_start_row=5, col B.
    ws_default["B4"] = "opportunity_score"
    for i in range(opportunity_rows):
        ws_default.cell(row=5 + i, column=2, value=10.0 + i)

    ws_qw = wb.create_sheet("quick_wins")
    ws_qw["J4"] = "priority"
    for i in range(quick_wins_rows):
        ws_qw.cell(row=5 + i, column=10, value="HIGH")

    ws_cd = wb.create_sheet("content_decay")
    ws_cd["E5"] = "delta_pct"
    for i in range(content_decay_rows):
        ws_cd.cell(row=6 + i, column=5, value=-50.0 + i)

    ws_can = wb.create_sheet("cannibalization")
    ws_can["F4"] = "status"
    for i in range(cannibalization_rows):
        ws_can.cell(row=5 + i, column=6, value="TODO")

    ws_mt = wb.create_sheet("master_task")
    # header row=3
    ws_mt["A3"] = "task_id"
    ws_mt["B3"] = "task"
    ws_mt["C3"] = "primary_source"
    ws_mt["F3"] = "category"
    ws_mt["G3"] = "priority"
    payload = master_task_payload or []
    for i, row in enumerate(payload):
        r = 4 + i
        ws_mt.cell(row=r, column=1, value=f"T-{i+1:04d}")
        ws_mt.cell(row=r, column=2, value=row.get("task") or "synthetic-task")
        ws_mt.cell(row=r, column=3, value=row.get("primary_source") or "")
        ws_mt.cell(row=r, column=6, value=row.get("category") or "")
        ws_mt.cell(row=r, column=7, value=row.get("priority") or "")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Test 1 — Frontmatter validates + natural_language ≥ 30 char sentinel
# ---------------------------------------------------------------------------

def test_frontmatter_validates_and_natural_language_min_length(
    skill_frontmatter: dict,
    skill_frontmatter_schema: dict,
) -> None:
    fm = skill_frontmatter
    assert fm["name"] == "portfolio-heatmap"
    assert fm["status"] == "wip"
    assert fm["version"] == "1.0"
    assert fm["category"] == "reporting"

    # Manual triggers EMPTY (per worker brief: NO slash command).
    assert fm["triggers"]["manual"] == []

    # natural_language sentinel: ≥ 30 char.
    nl = fm["triggers"]["natural_language"]
    assert isinstance(nl, str)
    assert len(nl) >= 30, (
        f"natural_language must be ≥ 30 char, got {len(nl)}"
    )
    # Required TR + EN trigger phrases per worker brief.
    for required in (
        "fırsat heatmap", "opportunity matrix",
        "drift heatmap", "portföy yoğunluk",
    ):
        assert required in nl, f"missing trigger phrase: {required!r}"

    # outputs[] — 3 entries (master.xlsx#none + report + snapshot).
    outputs = fm["outputs"]
    assert "master.xlsx#none" in outputs
    assert any("portfolio-heatmap.md" in o for o in outputs)
    assert any("portfolio-heatmap.json" in o for o in outputs)
    assert "events.jsonl" not in outputs, (
        "Q-RP-01 defer: events.jsonl must NOT be in outputs[]"
    )

    # No MCP tools required NOR optional — pure local aggregator.
    assert fm["mcp_tools"]["required"] == []
    assert fm["mcp_tools"]["optional"] == []

    # Budget: 0 credits, no paid MCP. Autonomy: HIGH + cron-ready.
    assert fm["budget"]["uses_paid_mcp"] is False
    assert fm["budget"].get("estimated_credits", 0) == 0
    assert fm["autonomy"]["confidence"] == "HIGH"
    assert fm["autonomy"]["requires_approval"] is False
    assert fm["autonomy"]["safe_auto_execute"] is True

    # Full Draft 7 validation.
    validator = Draft7Validator(skill_frontmatter_schema)
    errs = sorted(
        validator.iter_errors(fm), key=lambda e: list(e.absolute_path),
    )
    assert not errs, (
        f"frontmatter invalid: "
        f"{[(list(e.absolute_path), e.message) for e in errs]}"
    )


# ---------------------------------------------------------------------------
# Test 2 — aggregate_dimension enum 3 sentinel (accept 3, reject 'invalid')
# ---------------------------------------------------------------------------

def test_aggregate_dimension_enum_3_sentinel() -> None:
    """Worker brief acceptance gate: aggregate_dimension must accept
    "category" / "priority" / "primary_source" and reject anything
    else (DURUR AggregateDimensionInvalidError)."""
    # Positive: 3 valid values (no exception).
    for v in ("category", "priority", "primary_source"):
        assert ph._ensure_aggregate_dimension(v) == v

    # Negative: out-of-enum → DURUR.
    with pytest.raises(ph.AggregateDimensionInvalidError) as ei:
        ph._ensure_aggregate_dimension("invalid")
    assert "invalid" in str(ei.value)

    # Negative: empty string → DURUR.
    with pytest.raises(ph.AggregateDimensionInvalidError):
        ph._ensure_aggregate_dimension("")


# ---------------------------------------------------------------------------
# Test 3 — Missing master.xlsx + missing sheet schema-valid empty shape
#          (W-E1 paterni reuse; per-project graceful skip + warning)
# ---------------------------------------------------------------------------

def test_missing_sheet_schema_valid_empty_shape(
    tmp_path: Path,
    portfolio_config_schema: dict,
) -> None:
    """Per W-E1 paterni: missing master.xlsx OR missing sheet → schema-
    valid empty shape (NOT crash). Per-project graceful skip + warning
    surface."""
    portfolio_root = tmp_path / "portfolio_ws"
    portfolio_root.mkdir()
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)

    config = _synthetic_config(active_count=2)
    ph.validate_portfolio_config(config, portfolio_config_schema)

    # No master.xlsx for either project — graceful skip.
    heatmap = ph.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
        aggregate_dimension="category",
        generated_at="2026-05-01T00:00:00Z",
    )
    assert heatmap.portfolio_id == "test-portfolio"
    assert len(heatmap.projects) == 2
    assert heatmap.skipped_projects == 2
    for r in heatmap.projects:
        assert r.master_xlsx_path is None
        assert r.warnings, "missing workbook should surface a warning"
        # Sheet counts all zero, densities all 0.0 (empty shape).
        for n in ph.SHEET_NAMES:
            assert r.sheet_counts[n] == 0
            assert r.sheet_densities[n] == 0.0
        assert r.master_task_total == 0

    # Snapshot still emits cleanly + JSON-serialisable.
    snap = ph.build_snapshot(heatmap=heatmap)
    json.dumps(snap)
    assert snap["skipped_project_count"] == 2

    # Now add a master.xlsx for project 1 with a missing 'opportunity' sheet
    # (we'll only create master_task + quick_wins to demonstrate the
    # per-sheet missing-sheet graceful skip).
    from openpyxl import Workbook
    p1_xlsx = portfolio_root / "alpha-test" / "projects" / "alpha-test" / "master.xlsx"
    p1_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.active.title = "quick_wins"
    wb["quick_wins"]["J5"] = "HIGH"
    ws_mt = wb.create_sheet("master_task")
    ws_mt["B4"] = "synthetic-task"
    ws_mt["F4"] = "tech-debt"
    wb.save(str(p1_xlsx))

    heatmap2 = ph.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
        aggregate_dimension="category",
        generated_at="2026-05-01T00:00:00Z",
    )
    by_slug = {r.slug: r for r in heatmap2.projects}
    p1 = by_slug["alpha-test"]
    # quick_wins counted, opportunity/content_decay/cannibalization
    # missing → 0 + warning surface.
    assert p1.sheet_counts["quick_wins"] == 1
    assert p1.sheet_counts["opportunity"] == 0
    assert p1.sheet_counts["content_decay"] == 0
    assert p1.sheet_counts["cannibalization"] == 0
    # Warnings include the per-sheet missing surfaces.
    joined = " | ".join(p1.warnings)
    assert "opportunity" in joined or "graceful" in joined


# ---------------------------------------------------------------------------
# Test 4 — Multi-project merge with varied sheet density + density
#          normalization + idempotency
# ---------------------------------------------------------------------------

def test_multi_project_merge_and_density_normalization(
    tmp_path: Path,
    portfolio_config_schema: dict,
) -> None:
    """3 projects with varied per-sheet row counts. The portfolio max
    per sheet should be normalized into densities ∈ [0.0, 1.0].
    Idempotent: same inputs → byte-identical snapshot."""
    portfolio_root = tmp_path / "portfolio_ws"
    portfolio_root.mkdir()
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)

    config = _synthetic_config(active_count=3)
    ph.validate_portfolio_config(config, portfolio_config_schema)

    # Build 3 workbooks with deliberate density profiles:
    #   alpha-test: opportunity=10, quick_wins=2, content_decay=0,  cann=1
    #   beta-test:  opportunity=5,  quick_wins=10, content_decay=4, cann=0
    #   gamma-test: opportunity=0,  quick_wins=0,  content_decay=8, cann=2
    profiles = {
        "alpha-test": (10, 2,  0, 1),
        "beta-test":  (5,  10, 4, 0),
        "gamma-test": (0,  0,  8, 2),
    }
    for slug, (op, qw, cd, can) in profiles.items():
        xlsx = (
            portfolio_root / slug / "projects" / slug / "master.xlsx"
        )
        _build_workbook_with_density(
            path=xlsx,
            opportunity_rows=op, quick_wins_rows=qw,
            content_decay_rows=cd, cannibalization_rows=can,
            master_task_payload=[
                {"task": "t-tech", "primary_source": "tech_fix",
                 "category": "Tech", "priority": "HIGH"},
                {"task": "t-pillar", "primary_source": "pillar",
                 "category": "Content", "priority": "MEDIUM"},
            ],
        )

    heatmap = ph.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01", aggregate_dimension="category",
        generated_at="2026-05-01T00:00:00Z",
    )
    assert len(heatmap.projects) == 3
    assert heatmap.skipped_projects == 0

    # Per-sheet maxima: opportunity=10, quick_wins=10, content_decay=8,
    # cannibalization=2.
    assert heatmap.sheet_max_counts["opportunity"] == 10
    assert heatmap.sheet_max_counts["quick_wins"] == 10
    assert heatmap.sheet_max_counts["content_decay"] == 8
    assert heatmap.sheet_max_counts["cannibalization"] == 2

    by_slug = {r.slug: r for r in heatmap.projects}
    # Normalized densities ∈ [0.0, 1.0].
    assert by_slug["alpha-test"].sheet_densities["opportunity"] == 1.0  # 10/10
    assert by_slug["alpha-test"].sheet_densities["quick_wins"] == 0.2   # 2/10
    assert by_slug["alpha-test"].sheet_densities["content_decay"] == 0.0  # 0/8
    assert by_slug["alpha-test"].sheet_densities["cannibalization"] == 0.5  # 1/2

    assert by_slug["beta-test"].sheet_densities["opportunity"] == 0.5   # 5/10
    assert by_slug["beta-test"].sheet_densities["quick_wins"] == 1.0    # 10/10
    assert by_slug["beta-test"].sheet_densities["content_decay"] == 0.5  # 4/8

    assert by_slug["gamma-test"].sheet_densities["content_decay"] == 1.0  # 8/8
    assert by_slug["gamma-test"].sheet_densities["cannibalization"] == 1.0  # 2/2

    # Deterministic order — by (priority asc, slug asc).
    slugs_in_order = [r.slug for r in heatmap.projects]
    assert slugs_in_order == ["alpha-test", "beta-test", "gamma-test"]

    # Idempotency: re-run with identical inputs + pinned generated_at →
    # byte-stable snapshot.
    heatmap2 = ph.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01", aggregate_dimension="category",
        generated_at="2026-05-01T00:00:00Z",
    )
    snap1 = json.dumps(
        ph.build_snapshot(heatmap=heatmap),
        sort_keys=True, indent=2,
    )
    snap2 = json.dumps(
        ph.build_snapshot(heatmap=heatmap2),
        sort_keys=True, indent=2,
    )
    assert snap1 == snap2, "aggregate must be idempotent"

    # Render template smoke — no missing $variables; sparkline column
    # appears.
    template_text = TEMPLATE_PATH.read_text(encoding="utf-8")
    rendered = ph.build_report_markdown(
        heatmap=heatmap, template_text=template_text,
    )
    assert "# Portfolio Heatmap" in rendered
    assert "alpha-test" in rendered
    assert "beta-test" in rendered
    assert "gamma-test" in rendered
    assert "opportunity" in rendered
    assert "quick_wins" in rendered
    assert "content_decay" in rendered
    assert "cannibalization" in rendered
    # Density bars use the sparkline glyph set.
    assert "▇" in rendered or "▆" in rendered or "▅" in rendered


# ---------------------------------------------------------------------------
# Test 5 — primary_source 10-enum coverage (each enum value gets a bucket)
# ---------------------------------------------------------------------------

def test_primary_source_10_enum_coverage(
    tmp_path: Path,
    portfolio_config_schema: dict,
) -> None:
    """Worker brief acceptance gate: aggregate_dimension="primary_source"
    must surface a bucket for EVERY one of the 10 enum values, even
    when count is zero. Strict coverage mode."""
    portfolio_root = tmp_path / "portfolio_ws"
    portfolio_root.mkdir()
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)

    config = _synthetic_config(active_count=1)
    ph.validate_portfolio_config(config, portfolio_config_schema)

    # Single project with master_task rows covering only 3 of the 10
    # primary_source enum values.
    xlsx = (
        portfolio_root / "alpha-test" / "projects" / "alpha-test"
        / "master.xlsx"
    )
    _build_workbook_with_density(
        path=xlsx,
        opportunity_rows=0, quick_wins_rows=0,
        content_decay_rows=0, cannibalization_rows=0,
        master_task_payload=[
            {"task": "t1", "primary_source": "content_decay",
             "category": "X", "priority": "HIGH"},
            {"task": "t2", "primary_source": "tech_fix",
             "category": "Y", "priority": "MEDIUM"},
            {"task": "t3", "primary_source": "internal_links",
             "category": "Z", "priority": "LOW"},
            # one row outside the enum to verify drop-to-zero.
            {"task": "t4", "primary_source": "OUT_OF_ENUM",
             "category": "W", "priority": "LOW"},
        ],
    )

    heatmap = ph.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
        aggregate_dimension="primary_source",
        generated_at="2026-05-01T00:00:00Z",
    )

    # Every one of the 10 schema enum values surfaces as a bucket.
    assert heatmap.dimension_keys == ph.PRIMARY_SOURCE_ENUM
    assert len(heatmap.dimension_keys) == 10
    p1 = heatmap.projects[0]
    for enum_val in ph.PRIMARY_SOURCE_ENUM:
        assert enum_val in p1.dimension_buckets, (
            f"primary_source enum {enum_val!r} missing from bucket"
        )
    # Non-zero buckets match payload.
    assert p1.dimension_buckets["content_decay"] == 1
    assert p1.dimension_buckets["tech_fix"] == 1
    assert p1.dimension_buckets["internal_links"] == 1  # Q-IL-1 closure
    # Zero buckets stay at 0 (not absent).
    assert p1.dimension_buckets["pillar"] == 0
    assert p1.dimension_buckets["sxo"] == 0
    assert p1.dimension_buckets["redirect_404"] == 0


# ---------------------------------------------------------------------------
# Test 6 — Forbidden tokens guard (4 token + 8 base64-obfuscated slugs)
# ---------------------------------------------------------------------------

def test_no_forbidden_tokens_in_skill() -> None:
    """Per worker brief acceptance gate #6: transform + SKILL.md +
    template + test must contain ZERO occurrences of:
      1. estimated_credits_per_call (Phase 7 lesson)
      2. estimated_credits_per_url  (Phase 7 lesson)
      3. metric_name                (Phase 7 closeout Q-CO-01 lesson)
      4. TODO inline-comment marker in transform code
    Plus the 8 hardcoded real-client slug grep guard
    (base64-obfuscated so this test file itself stays grep-clean)."""
    transform_src = TRANSFORM_PATH.read_text(encoding="utf-8")
    skill_src = SKILL_PATH.read_text(encoding="utf-8")
    template_src = TEMPLATE_PATH.read_text(encoding="utf-8")

    forbidden_literal = (
        "estimated_credits" + "_per_call",
        "estimated_credits" + "_per_url",
        "metric" + "_name",
    )
    for body, label in (
        (transform_src, "transform"),
        (skill_src, "skill"),
        (template_src, "template"),
    ):
        for tok in forbidden_literal:
            assert tok not in body, (
                f"forbidden token {tok!r} found in {label}"
            )

    # TODO comment in CODE (transform). Markdown skill prose may use the
    # word legitimately ("TODO/fallback YASAK" discipline checklist).
    todo_in_transform = re.findall(r"#\s*TODO\b", transform_src)
    assert not todo_in_transform, (
        f"TODO comment(s) found in transform: {todo_in_transform}"
    )

    # 8 forbidden hardcoded slug (base64-encoded so this test file
    # itself stays grep-clean).
    encoded_slug_tokens = (
        b"ZGVudG5vdGlvbg==", b"dmVudG8=", b"ZXlrb20=", b"YmlnY2F0dHI=",
        b"Y2FsaXR0ZQ==", b"bGFzdGlrc2E=", b"bm9yYW5pbnNhYXQ=", b"YWRzdGFyaw==",
    )
    forbidden_slugs = tuple(
        base64.b64decode(t).decode("ascii") for t in encoded_slug_tokens
    )
    pattern = re.compile(
        r"\b(" + "|".join(forbidden_slugs) + r")\b", flags=re.IGNORECASE,
    )
    for body, label in (
        (transform_src, "transform"),
        (skill_src, "skill"),
        (template_src, "template"),
    ):
        m = pattern.search(body)
        assert m is None, (
            f"forbidden hardcoded slug {m.group(0)!r} found in {label}"
        )


# ---------------------------------------------------------------------------
# Test 7 — assert_read_only_module() helper (W-E3 paterni grep guard)
# ---------------------------------------------------------------------------

def test_assert_read_only_module_helper() -> None:
    """Per worker brief acceptance gate #11: the transform module's
    assert_read_only_module() helper must scan its own source for
    forbidden write call shapes (transaction.* / events_writer.* /
    wb.save()) and pass cleanly. Independent re-grep so the test is
    robust even if the helper is later refactored."""
    # Direct invocation — must NOT raise.
    ph.assert_read_only_module()

    # Independent grep — outside the legitimate constants tuple.
    src = TRANSFORM_PATH.read_text(encoding="utf-8")
    marker = "_FORBIDDEN_WRITE_PATTERNS"
    if marker in src:
        before, _sep, _rest = src.partition(marker)
        scan_src = before
    else:
        scan_src = src
    forbidden = (
        "wb" + "." + "save(",
        "workbook" + "." + "save(",
        "transaction" + "." + "append(",
        "transaction" + "." + "update(",
        "transaction" + "." + "delete(",
        "events_writer" + "." + "append_",
    )
    for pat in forbidden:
        assert pat not in scan_src, (
            f"forbidden write pattern {pat!r} appears in transform source "
            f"outside the constants tuple — read-only contract violated."
        )

    # Forbidden imports.
    forbidden_imports = (
        r"^\s*from\s+scripts\.excel\.transaction\b",
        r"^\s*import\s+scripts\.excel\.transaction\b",
        r"^\s*from\s+scripts\.state\.events_writer\b",
        r"^\s*import\s+scripts\.state\.events_writer\b",
    )
    for pat in forbidden_imports:
        assert not re.search(pat, scan_src, flags=re.MULTILINE), (
            f"forbidden import pattern matched in transform: {pat}"
        )


# ---------------------------------------------------------------------------
# Test 8 — Path convention (projects/_portfolio/{outputs,inbox}/local/
#          + PSEO_WORKSPACE_ROOT env)
# ---------------------------------------------------------------------------

def test_path_convention_compliance(
    tmp_path: Path,
    portfolio_config_schema: dict,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Per worker brief acceptance gate #10 (W-E4 paterni):
    PSEO_WORKSPACE_ROOT env must resolve to the portfolio root, and
    SKILL.md outputs[] paths must literally include
    'projects/_portfolio/outputs/reports/' +
    'projects/_portfolio/inbox/local/'."""
    # SKILL.md outputs[] literal path convention.
    skill_text = SKILL_PATH.read_text(encoding="utf-8")
    assert "projects/_portfolio/outputs/reports/" in skill_text
    assert "projects/_portfolio/inbox/local/" in skill_text

    # PSEO_WORKSPACE_ROOT env resolution.
    portfolio_root = tmp_path / "ws"
    portfolio_root.mkdir()
    monkeypatch.setenv("PSEO_WORKSPACE_ROOT", str(portfolio_root))
    monkeypatch.delenv("PSEO_PORTFOLIO_ROOT", raising=False)
    resolved = ph._resolve_portfolio_root(None)
    assert resolved == portfolio_root.resolve()

    # No env + no arg → DURUR WorkspaceRootUnsetError.
    monkeypatch.delenv("PSEO_WORKSPACE_ROOT", raising=False)
    with pytest.raises(ph.WorkspaceRootUnsetError):
        ph._resolve_portfolio_root(None)

    # Explicit arg overrides env.
    monkeypatch.setenv("PSEO_WORKSPACE_ROOT", "/nonexistent/path")
    explicit = tmp_path / "explicit"
    explicit.mkdir()
    resolved2 = ph._resolve_portfolio_root(str(explicit))
    assert resolved2 == explicit.resolve()
