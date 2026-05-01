"""
tests/skills/test_dfs_pull.py — dfs-pull ingestion skill end-to-end tests.

Seven cases mirroring the W-V worker brief acceptance gate (Phase 6
Wave 1, paid-MCP authority). The TR forwarding workaround (live test
1835229: dataforseo-mcp-server@2.8.9 returns location_code=2840 even
when 2792 is requested) is tested via mock fixtures — no live API
needed.

Cross-module IMPORT-only discipline:
  - scripts.excel.transaction       (atomic per-sheet write + backup FIFO 7)
  - scripts.state.workflow_runner   (10-step run shell)
  - scripts.state.events_writer     (dataforseo_mcp provenance event)
  - scripts.budget.check_budget     (§16.8 pre-flight; FIRST paid skill)
  - scripts.ingestion.dfs_pull      (transform + TR workaround helpers)

Schemas referenced:
  - schemas/skill-frontmatter.schema.json    (frontmatter draft7)
  - schemas/master-excel.schema.json         (cluster_keywords, opportunity)
  - schemas/events.schema.json               (source.kind=dataforseo_mcp)

Run from repo root:
    PYTHONPATH=. pytest tests/skills/test_dfs_pull.py -v
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
import yaml
from jsonschema import Draft7Validator

from scripts.excel import transaction
from scripts.ingestion import dfs_pull
from scripts.state import events_writer


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"
SKILL_MD = REPO_ROOT / "skills" / "ingestion" / "dfs-pull" / "SKILL.md"


# ---------------------------------------------------------------------------
# Fixtures — schemas
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def skill_frontmatter_schema() -> dict:
    return json.loads(
        (SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def master_excel_schema() -> dict:
    return json.loads((SCHEMAS / "master-excel.schema.json").read_text("utf-8"))


@pytest.fixture(scope="module")
def events_schema() -> dict:
    return json.loads((SCHEMAS / "events.schema.json").read_text("utf-8"))


# ---------------------------------------------------------------------------
# Fixtures — mock DFS payloads (TR forwarding workaround test data)
# ---------------------------------------------------------------------------

@pytest.fixture
def overview_payload_us_drift() -> dict:
    """Mock keyword_overview response from the wrapper bug (1835229):
    location_code=2840 (US) returned even though 2792 (TR) was requested.
    """
    return {
        "version": "0.1.20240101",
        "status_code": 20000,
        "tasks": [{
            "id": "06241000-1535-0095-0000-7e9b3aab3bcc",
            "status_code": 20000,
            "data": {
                "location_code": 2792,    # what we *requested*
                "language_code": "tr",
            },
            "result": [{
                # what the wrapper actually served (THE BUG)
                "location_code": 2840,    # US
                "language_code": "en",
                "items": [
                    {
                        "keyword": "diş kliniği",
                        "keyword_info": {"search_volume": 9100, "competition": "MEDIUM"},
                        "search_intent_info": {"main_intent": "commercial"},
                        "language_code": "en",
                    }
                ],
            }],
        }],
    }


@pytest.fixture
def overview_payload_tr_correct() -> dict:
    """Mock keyword_overview response after workaround C (HTTP bypass):
    location_code=2792 / language_code='tr' honoured.
    """
    return {
        "version": "0.1.20240101",
        "status_code": 20000,
        "tasks": [{
            "id": "06241000-1535-0095-0000-7e9b3aab3bcc",
            "status_code": 20000,
            "data": {"location_code": 2792, "language_code": "tr"},
            "result": [{
                "location_code": 2792,
                "language_code": "tr",
                "items": [
                    {
                        "keyword": "diş kliniği",
                        "keyword_info": {"search_volume": 12100, "competition": "MEDIUM"},
                        "search_intent_info": {"main_intent": "commercial"},
                        "language_code": "tr",
                    },
                    {
                        "keyword": "implant fiyatları",
                        "keyword_info": {"search_volume": 8100, "competition": "HIGH"},
                        "search_intent_info": {"main_intent": "transactional"},
                        "language_code": "tr",
                    },
                    {
                        "keyword": "diş beyazlatma",
                        "keyword_info": {"search_volume": 5400, "competition": "LOW"},
                        "search_intent_info": {"main_intent": "informational"},
                        "language_code": "tr",
                    },
                ],
            }],
        }],
    }


@pytest.fixture
def volume_payload_tr_correct() -> dict:
    """Mock Google Ads search_volume response (workaround B alt endpoint),
    locale honoured. Keys match what dfs_pull.load_volume_index expects.
    """
    return {
        "tasks": [{
            "data": {"location_code": 2792, "language_code": "tr"},
            "result": [{
                "location_code": 2792,
                "language_code": "tr",
                "items": [
                    {"keyword": "diş kliniği",
                     "search_volume": 12100, "competition": "MEDIUM", "cpc": 1.05},
                    {"keyword": "implant fiyatları",
                     "search_volume": 8100, "competition": "HIGH", "cpc": 2.30},
                    {"keyword": "diş beyazlatma",
                     "search_volume": 5400, "competition": "LOW", "cpc": 0.65},
                ],
            }],
        }],
    }


# ---------------------------------------------------------------------------
# Test (a) — Frontmatter parse + schema validate
# ---------------------------------------------------------------------------

def test_frontmatter_validates_against_schema(
    skill_frontmatter_schema: dict,
) -> None:
    """SKILL.md frontmatter must validate against
    schemas/skill-frontmatter.schema.json (Draft 7).
    """
    text = SKILL_MD.read_text("utf-8")
    parts = text.split("---", 2)
    assert len(parts) >= 3, "SKILL.md must open with --- frontmatter ---"
    fm = yaml.safe_load(parts[1])

    validator = Draft7Validator(skill_frontmatter_schema)
    errs = sorted(validator.iter_errors(fm),
                  key=lambda e: list(e.absolute_path))
    assert not errs, (
        "frontmatter invalid: "
        f"{[('/'.join(str(p) for p in e.absolute_path) or '<root>', e.message) for e in errs]}"
    )

    # Brief-mandated values.
    assert fm["category"] == "ingestion"
    assert fm["status"] == "wip"
    assert fm["budget"]["uses_paid_mcp"] is True
    assert fm["budget"]["estimated_credits"] > 0
    # Required MCP tools present.
    req = fm["mcp_tools"]["required"]
    assert "mcp__dataforseo__keywords_data_google_ads_search_volume" in req
    assert "mcp__dataforseo__dataforseo_labs_google_keyword_overview" in req


# ---------------------------------------------------------------------------
# Test (b) — Transform unit: TR-correct fixture → expected rows
# ---------------------------------------------------------------------------

def test_transform_tr_correct_payload(
    overview_payload_tr_correct: dict,
    volume_payload_tr_correct: dict,
) -> None:
    """Happy path: payload honors TR, transform emits 3 cluster_keywords
    rows + 3 opportunity rows, sorted by volume desc / score desc."""
    out = dfs_pull.transform(
        overview_payload_tr_correct,
        raw_volume=volume_payload_tr_correct,
        cluster_default="dental",
        location_code=2792,
        language_code="tr",
    )
    ck = out["cluster_keywords"]
    op = out["opportunity"]

    assert len(ck) == 3
    assert len(op) == 3

    # Sort: monthly_volume desc → diş kliniği(12100) > implant fiyatları(8100) > diş beyazlatma(5400)
    assert ck[0]["keyword"] == "diş kliniği"
    assert ck[0]["monthly_volume"] == 12100
    assert ck[0]["data_source"] == "dfs_keyword_overview"
    assert ck[0]["cluster"] == "dental"
    assert ck[0]["intent"] == "Commercial"  # search_intent main_intent="commercial"

    # Schema-shape: exactly the 11 cluster_keywords columns.
    assert set(ck[0].keys()) == set(dfs_pull.CLUSTER_KEYWORDS_COLUMNS)

    # Opportunity scoring: diş beyazlatma (5400 × 1.0 LOW) = 5400 should
    # outrank implant fiyatları (8100 × 0.3 HIGH) = 2430 and diş kliniği
    # (12100 × 0.6 MEDIUM) = 7260 — actually diş kliniği wins.
    assert op[0]["query"] == "diş kliniği"
    # Schema-shape: 8 opportunity columns.
    assert set(op[0].keys()) == set(dfs_pull.OPPORTUNITY_COLUMNS)


# ---------------------------------------------------------------------------
# Test (c) — Schema column match (cluster_keywords + opportunity)
# ---------------------------------------------------------------------------

def test_transform_columns_match_master_schema(
    overview_payload_tr_correct: dict,
    master_excel_schema: dict,
) -> None:
    """Transform output must match master-excel.schema column names
    exactly — no extra keys, no missing keys."""
    out = dfs_pull.transform(
        overview_payload_tr_correct,
        cluster_default="dental",
        location_code=2792,
        language_code="tr",
    )

    sheets = master_excel_schema["sheets"]
    ck_cols = {c["name"] for c in sheets["cluster_keywords"]["required_columns"]}
    op_cols = {c["name"] for c in sheets["opportunity"]["required_columns"]}

    # Module constants must equal the schema (not a superset, not a subset).
    assert set(dfs_pull.CLUSTER_KEYWORDS_COLUMNS) == ck_cols, (
        f"cluster_keywords drift: module={set(dfs_pull.CLUSTER_KEYWORDS_COLUMNS)} "
        f"schema={ck_cols}"
    )
    assert set(dfs_pull.OPPORTUNITY_COLUMNS) == op_cols, (
        f"opportunity drift: module={set(dfs_pull.OPPORTUNITY_COLUMNS)} "
        f"schema={op_cols}"
    )

    # Each emitted row carries exactly the schema's columns, nothing more.
    for row in out["cluster_keywords"]:
        assert set(row.keys()) == ck_cols
    for row in out["opportunity"]:
        assert set(row.keys()) == op_cols


# ---------------------------------------------------------------------------
# Test (d) — TR workaround method tests (the 1835229 bug fixture)
# ---------------------------------------------------------------------------

def test_tr_workaround_detects_us_drift(
    overview_payload_us_drift: dict,
    overview_payload_tr_correct: dict,
) -> None:
    """The wrapper bug returns location_code=2840/lang='en' even when
    2792/'tr' was requested. response_honors_tr() must:
      - return False for the drifted (US) payload,
      - return True for the workaround-C HTTP-bypass payload.
    """
    # Sanity: the env-detector pulls the SERVED locale, not the echoed
    # request — that's the whole point of the bug.
    served_loc, served_lang = dfs_pull.detect_response_locale(overview_payload_us_drift)
    assert served_loc == 2840
    assert served_lang == "en"

    # response_honors_tr respects served, not echoed → False.
    assert dfs_pull.response_honors_tr(
        overview_payload_us_drift,
        expected_location=2792, expected_language="tr",
    ) is False

    # The TR-correct fixture passes.
    assert dfs_pull.response_honors_tr(
        overview_payload_tr_correct,
        expected_location=2792, expected_language="tr",
    ) is True

    # Workaround C HTTP body builder: location_code/language_code carried
    # in the JSON body (wrapper bypass — REST honours these).
    body = dfs_pull.build_http_payload_tr(
        keywords=["diş kliniği"], location_code=2792, language_code="tr",
    )
    assert isinstance(body, list) and len(body) == 1
    assert body[0]["location_code"] == 2792
    assert body[0]["language_code"] == "tr"

    # Workaround A heuristic filter: drops obviously-non-TR rows.
    rows = [
        {"keyword": "kw-tr",  "language_code": "tr"},
        {"keyword": "kw-en",  "language_code": "en"},   # filtered out
        {"keyword": "kw-no-evidence"},                  # kept (no neg evidence)
    ]
    kept = dfs_pull.filter_to_tr_heuristic(rows)
    kept_kw = {r["keyword"] for r in kept}
    assert "kw-tr" in kept_kw
    assert "kw-no-evidence" in kept_kw
    assert "kw-en" not in kept_kw


# ---------------------------------------------------------------------------
# Test (e) — Budget pre-flight integration (FIRST paid skill)
# ---------------------------------------------------------------------------

def test_budget_preflight_integration(tmp_path: Path) -> None:
    """estimate_credits() > 0 triggers preflight_budget(). When projected
    usage stays under budget → returns envelope. When projected > budget →
    BudgetError DURUR.
    """
    # Estimate: 10 keywords × (1.0 + 0.5) = 15 credits.
    est = dfs_pull.estimate_credits(10)
    assert est == 15.0

    # PASS path: budget=500, no prior usage, estimate=15 → projected=15.
    cfg_path = tmp_path / "project-config.json"
    cfg_path.write_text(json.dumps({
        "dataforseo": {"budget_credits_per_day": 500}
    }))
    events_path = tmp_path / "events.jsonl"
    # No events file = treated as zero usage.
    envelope = dfs_pull.preflight_budget(
        estimated_credits=est,
        project_config_path=cfg_path,
        events_path=events_path,
    )
    assert envelope["exceeded"] is False
    assert envelope["budget_per_day"] == 500
    assert envelope["projected_used"] == 15.0
    assert envelope["remaining_after"] == 485.0

    # FAIL path: budget=10, estimate=15 → projected=15 > 10.
    cfg_low = tmp_path / "project-config-low.json"
    cfg_low.write_text(json.dumps({
        "dataforseo": {"budget_credits_per_day": 10}
    }))
    with pytest.raises(dfs_pull.BudgetError):
        dfs_pull.preflight_budget(
            estimated_credits=est,
            project_config_path=cfg_low,
            events_path=events_path,
        )


# ---------------------------------------------------------------------------
# Test (f) — DURUR fired: TR workaround all-fail + missing creds + drift
# ---------------------------------------------------------------------------

def test_durur_conditions_fire(
    overview_payload_us_drift: dict,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """DURUR coverage:
      #4: TR workaround all-fail → TrWorkaroundFailed when transform sees
          a non-TR payload and skip_tr_check=False.
      #8: DATAFORSEO_USERNAME / DATAFORSEO_PASSWORD missing →
          CredentialError on http_credentials_from_env().
    """
    # #4 — transform refuses to silently emit US data.
    with pytest.raises(dfs_pull.TrWorkaroundFailed):
        dfs_pull.transform(
            overview_payload_us_drift,
            location_code=2792, language_code="tr",
            skip_tr_check=False,
        )

    # Caller can opt in via skip_tr_check=True after running A/B/C — then
    # transform proceeds (this is the normal happy path AFTER workaround).
    out = dfs_pull.transform(
        overview_payload_us_drift,
        location_code=2792, language_code="tr",
        skip_tr_check=True,
    )
    assert len(out["cluster_keywords"]) == 1   # one item in the fixture

    # #8 — credentials missing.
    monkeypatch.delenv("DATAFORSEO_USERNAME", raising=False)
    monkeypatch.delenv("DATAFORSEO_PASSWORD", raising=False)
    with pytest.raises(dfs_pull.CredentialError):
        dfs_pull.http_credentials_from_env()

    # #8 happy: both set.
    monkeypatch.setenv("DATAFORSEO_USERNAME", "test@example.com")
    monkeypatch.setenv("DATAFORSEO_PASSWORD", "redacted")
    user, pwd = dfs_pull.http_credentials_from_env()
    assert user == "test@example.com"
    assert pwd == "redacted"


# ---------------------------------------------------------------------------
# Test (g) — Smoke E2E: provenance event emitted + opportunity sheet shared
#                       writer (F-09 invariant) verified
# ---------------------------------------------------------------------------

def test_smoke_e2e_with_mock(
    tmp_path: Path,
    overview_payload_tr_correct: dict,
    volume_payload_tr_correct: dict,
    events_schema: dict,
) -> None:
    """End-to-end smoke:
      1. Transform mock TR payload → 3 cluster_keywords + 3 opportunity rows.
      2. transaction.append both sheets to a fresh workbook in tmp_path.
      3. Emit a dataforseo_mcp provenance event via events_writer.
      4. Re-validate the event against schemas/events.schema.json.
      5. Assert opportunity is shared-writer safe: append a quick-wins
         row AFTER dfs-pull and ensure both survive (F-09 invariant).
    """
    slug = "dfs-test"
    workbook = tmp_path / "projects" / slug / "master.xlsx"
    workbook.parent.mkdir(parents=True, exist_ok=True)
    state_dir = tmp_path / "projects" / slug / "_state"
    state_dir.mkdir(parents=True, exist_ok=True)

    out = dfs_pull.transform(
        overview_payload_tr_correct,
        raw_volume=volume_payload_tr_correct,
        cluster_default="dental",
        location_code=2792, language_code="tr",
    )

    # Step 8 — atomic per-sheet append.
    res_ck = transaction.append(
        workbook_path=workbook,
        sheet="cluster_keywords",
        rows=out["cluster_keywords"],
        project_slug=slug,
        writer="dfs-pull",
    )
    assert res_ck.rows_affected == 3

    res_op = transaction.append(
        workbook_path=workbook,
        sheet="opportunity",
        rows=out["opportunity"],
        project_slug=slug,
        writer="dfs-pull",
    )
    assert res_op.rows_affected == 3

    # F-09 invariant: opportunity is shared with quick-wins. Append a
    # quick-wins-shaped row and confirm BOTH survive (no overwrite).
    qw_op_row = {
        "query": "kw-from-quickwins",
        "opportunity_score": 9999.0,
        "current_position": 12.5,
        "ctr_pct": 1.2,
        "impressions_30d": 850,
        "clicks_30d": 5,
        "potential_clicks": 80,
        "assigned_url_action": "https://example.com/page | refresh meta",
    }
    res_qw = transaction.append(
        workbook_path=workbook,
        sheet="opportunity",
        rows=[qw_op_row],
        project_slug=slug,
        writer="quick-wins",
    )
    assert res_qw.rows_affected == 1

    # Verify physical row count: 3 dfs + 1 qw = 4 data rows + 1 header.
    from openpyxl import load_workbook
    wb = load_workbook(str(workbook))
    ws = wb["opportunity"]
    assert ws.max_row == 5, (
        f"shared-writer F-09 invariant broken: opportunity has "
        f"{ws.max_row - 1} data rows, expected 4"
    )

    # Step 9 — provenance event with source.kind=dataforseo_mcp.
    rid = events_writer.next_run_id(slug, workspace_root=tmp_path)
    result = events_writer.append_provenance(
        project_id=slug,
        run_id=rid,
        source={
            "kind": "dataforseo_mcp",
            "mcp_server": "dataforseo",
            "mcp_tool": "dataforseo__dataforseo_labs_google_keyword_overview",
            "response_bytes": len(json.dumps(overview_payload_tr_correct)),
        },
        operation="project_excel",
        target_excel_sheet="cluster_keywords",
        rows_written=3,
        cost={
            "provider": "dataforseo",
            "credits": 4.5,   # 3 keywords × 1.5
            "budget_key": "project.config.dataforseo.budget_credits_per_day",
        },
        workspace_root=tmp_path,
    )
    assert result.event_id

    events_path = state_dir / "events.jsonl"
    lines = [
        json.loads(line)
        for line in events_path.read_text("utf-8").splitlines()
        if line.strip()
    ]
    dfs_provs = [
        e for e in lines
        if e.get("event_kind") == "provenance"
        and e.get("source", {}).get("kind") == "dataforseo_mcp"
    ]
    assert dfs_provs, (
        "no provenance event_kind=provenance source.kind=dataforseo_mcp emitted"
    )

    # Re-validate against the canonical events schema.
    validator = Draft7Validator(events_schema)
    for evt in dfs_provs:
        errs = sorted(validator.iter_errors(evt),
                      key=lambda e: list(e.absolute_path))
        assert not errs, (
            "emitted dataforseo_mcp provenance event invalid: "
            f"{[(list(e.absolute_path), e.message) for e in errs]}"
        )
