"""TDD lock for skills/discovery/robots-policy-audit/SKILL.md + wiring (GAP-T3).

Contract checks (spec GAP-T3 §d):
  - frontmatter validates against skill-frontmatter.schema.json, category=discovery,
    status=wip, name matches dir;
  - BOTH output artifacts declared (master.xlsx#robots_txt + the proposed-txt path);
  - template + slash command exist; produces edges present;
  - body documents AMBER (not fail) on live-fetch down + DURUR + recommendation-only;
  - the transform -> sheet_merge.merge_prefixed_rows(RP-) write is idempotent and
    preserves foreign rows.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import yaml
from jsonschema import Draft7Validator
from openpyxl import load_workbook

from scripts.discovery import robots_policy_transform as rpt
from scripts.orchestration import committer
from scripts.util import sheet_merge

ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = ROOT / "schemas"
SKILL_PATH = ROOT / "skills" / "discovery" / "robots-policy-audit" / "SKILL.md"
TEMPLATE = ROOT / "templates" / "reports" / "robots-policy-audit.template.md"
COMMAND = ROOT / "commands" / "pseo-robots-policy.md"


def _fm() -> dict:
    text = SKILL_PATH.read_text(encoding="utf-8")
    m = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert m, "SKILL.md missing YAML frontmatter delimiters"
    return yaml.safe_load(m.group(1))


def _body() -> str:
    text = SKILL_PATH.read_text(encoding="utf-8")
    return text[re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL).end():]


def test_frontmatter_validates() -> None:
    fm = _fm()
    schema = json.loads((SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8"))
    errors = sorted(Draft7Validator(schema).iter_errors(fm), key=lambda e: e.path)
    assert not errors, f"frontmatter invalid: {[e.message for e in errors]}"
    assert fm["name"] == "robots-policy-audit"
    assert fm["category"] == "discovery"
    assert fm["status"] == "wip"
    assert fm["budget"]["uses_paid_mcp"] is False


def test_outputs_declare_robots_sheet_and_proposed_artifact() -> None:
    fm = _fm()
    outs = fm["outputs"]
    master = json.loads((SCHEMAS / "master-excel.schema.json").read_text("utf-8"))
    anchors = [o.split("#", 1)[1] for o in outs if o.startswith("master.xlsx#")]
    assert "robots_txt" in anchors and anchors[0] in master["sheets"]
    assert any("robots.proposed.txt" in o for o in outs), "proposed-txt artifact must be declared"


def test_produces_edges_present() -> None:
    assert set(_fm()["produces"]) >= {"drift-check", "content-remediation"}


def test_template_and_command_exist() -> None:
    assert TEMPLATE.exists()
    assert COMMAND.exists()


def test_body_documents_amber_durur_recommendation() -> None:
    body = _body().lower()
    assert "amber" in body, "must document the AMBER fallback on live-fetch failure"
    assert "durur" in body
    assert "recommendation" in body or "operator" in body


def test_transform_to_merge_write_idempotent(tmp_path: Path) -> None:
    slug, run_id = "test-proj", "test-proj-2026-06-10-rp01"
    proj = tmp_path / "projects" / slug
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    wb = proj / "master.xlsx"
    committer.commit(
        wb, "robots_txt",
        [{"id": "R-001", "level": "LOW", "issue": "x", "detail": "d", "resolution": "r"},
         {"id": "FN-001", "level": "MEDIUM", "issue": "facet", "detail": "d", "resolution": "r"}],
        run_id=run_id, project_slug=slug, writer="sf-import",
    )
    out = rpt.transform(
        "User-agent: *\nnoindex: /x/\n", directives_rows=[], internal_rows=[],
        lifecycle_rows=[], platform="custom", domain="x.test",
    )
    kw = dict(id_prefix="RP-", run_id=run_id, project_slug=slug, writer="robots-policy-audit")
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", out["robots_txt_rows"], **kw)
    ids1 = {r[0] for r in load_workbook(wb)["robots_txt"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", out["robots_txt_rows"], **kw)
    ids2 = {r[0] for r in load_workbook(wb)["robots_txt"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    assert ids1 == ids2
    assert {"R-001", "FN-001"} <= ids2, "foreign R-/FN- rows must be preserved"
    assert any(str(i).startswith("RP-") for i in ids2)
