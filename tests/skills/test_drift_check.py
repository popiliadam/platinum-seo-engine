"""
tests/skills/test_drift_check.py — drift-check governance skill tests.

Eleven cases mirroring the W-S worker brief acceptance gate:

  1.  test_all_pass_green_verdict        — synthetic clean wb, 20/20 PASS → GREEN
  2.  test_critical_fail_red_verdict     — F-01 fail (statusEnum drift) → RED
  3.  test_high_fail_red_verdict         — F-08 sparse-pilot routing → AMBER (manual triage)
  4.  test_medium_fail_amber_verdict     — F-18 fail (ISO 8601 invalid) → AMBER
  5.  test_skip_missing_sheet_amber      — sheet missing → SKIP + AMBER (not RED)
  6.  test_consistency_report_schema_valid — JSON schema validate
  7.  test_unknown_rule_id_raises        — aggregator rejects synthetic ids
  8.  test_provenance_event_emitted      — events.jsonl event_kind=audit
  9.  test_workspace_read_only_no_writes — master.xlsx SHA-256 unchanged
  10. test_f15_manual_review_amber       — F-15 → manual_review_required[] populated
  11. test_auto_repair_available_flag    — F-22 (FIFO 7 backups) → auto_repair_available=true

Schemas referenced:
  - schemas/master-excel.schema.json (statusEnum 7, severityEnum 4, columns)
  - schemas/events.schema.json (event_kind=audit branch)
  - schemas/consistency-report.schema.json (output report shape)
"""

from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

import pytest
from jsonschema import Draft7Validator
from openpyxl import Workbook, load_workbook

from scripts.state import events_writer
from scripts.validation import validate_invariants as vi


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"


# ---------------------------------------------------------------------------
# Fixtures — synthetic workbook builder
# ---------------------------------------------------------------------------

def _build_clean_wb(path: Path) -> None:
    """Write a synthetic master.xlsx that ALL 20 rules pass on (within
    the SKIP semantics for sheets not present)."""
    wb = Workbook()
    # quick_wins (D-03 idempotent URLs)
    qw = wb.active
    qw.title = "quick_wins"
    qw.append(["query", "url", "current_position", "impressions_30d",
               "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
               "action", "priority"])
    qw.append(["foo", "https://example.com/a", 12, 100, 5, 5.0, 10,
               "High", "do x", "MEDIUM"])
    # opportunity (F-16 fk: every qw url ⊆ opp.url)
    op = wb.create_sheet("opportunity")
    op.append(["query", "opportunity_score", "current_position", "ctr_pct",
               "impressions_30d", "clicks_30d", "potential_clicks",
               "assigned_url_action"])
    op.append(["foo", 800, 12, 5.0, 100, 5, 10,
               "https://example.com/a | do x"])
    # crawl_sitemap (F-08: provide URL coverage)
    cs = wb.create_sheet("crawl_sitemap")
    cs.append(["category", "metric", "value", "status", "action"])
    cs.append(["url", "page", "https://example.com/a", "EXISTS", ""])
    # gsc_performance (also F-08 source)
    gp = wb.create_sheet("gsc_performance")
    gp.append(["url", "clicks_recent", "clicks_previous", "clicks_delta",
               "clicks_delta_pct", "impressions_recent",
               "impressions_previous", "impressions_delta",
               "ctr_recent", "position_recent", "position_previous", "note"])
    gp.append(["https://example.com/a", 5, 4, 1, 25.0, 100, 90, 10, 5.0,
               12.0, 13.0, ""])
    # dashboard (F-02/F-03/F-04 only check absence of formula cells)
    db = wb.create_sheet("dashboard")
    db.append(["KPI", "value"])
    db.append(["total_pillars", 0])
    # master_task — empty (header only) so F-01/F-09/F-18 PASS trivially
    mt = wb.create_sheet("master_task")
    mt.append(["task_id", "task", "primary_source", "related_sources", "url",
               "category", "priority", "impact", "duration_est_min", "status",
               "created_date", "done_date", "assignee", "note", "auto_generated",
               "dependencies", "effort_actual_min", "metric_impact_json",
               "work_log_ref"])
    wb.save(str(path))


def _setup_workspace(tmp_path: Path, slug: str) -> Path:
    """Create projects/{slug}/_state/ skeleton + return master.xlsx path."""
    pdir = tmp_path / "projects" / slug
    (pdir / "_state" / "workflows").mkdir(parents=True, exist_ok=True)
    (pdir / "_state" / "backups" / "master").mkdir(parents=True, exist_ok=True)
    (pdir / "outputs" / "reports").mkdir(parents=True, exist_ok=True)
    cfg = {"project_id": slug, "language": {"content_locale": "tr-TR"}, "market": "TR"}
    (pdir / "project.config.json").write_text(json.dumps(cfg), "utf-8")
    return pdir / "master.xlsx"


@pytest.fixture
def clean_wb_path(tmp_path: Path) -> Path:
    slug = "drifttest"
    wb_path = _setup_workspace(tmp_path, slug)
    _build_clean_wb(wb_path)
    return wb_path


@pytest.fixture
def consistency_schema() -> dict:
    return json.loads(
        (SCHEMAS / "consistency-report.schema.json").read_text("utf-8")
    )


@pytest.fixture
def events_schema() -> dict:
    return json.loads((SCHEMAS / "events.schema.json").read_text("utf-8"))


# ---------------------------------------------------------------------------
# Test 1 — synthetic clean wb → GREEN
# ---------------------------------------------------------------------------

def test_all_pass_green_verdict(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    agg = vi.aggregate_verdicts(results)
    # All-clean expectation: no FAIL anywhere; SKIPs are allowed (which
    # would land us on AMBER) — assert no RED at minimum, plus that the
    # 5 CRITICAL rules all PASS (no SKIP among them either).
    by_id = {r["id"]: r for r in results}
    for rid in ("F-01", "F-02", "F-03", "F-04", "F-05"):
        assert by_id[rid]["verdict"] == "PASS", (
            f"{rid} expected PASS, got {by_id[rid]['verdict']}: "
            f"{by_id[rid]['evidence']}"
        )
    # F-08, F-10, F-16 must PASS (URLs intersect, idempotent, FK satisfied).
    for rid in ("F-08", "F-10", "F-16"):
        assert by_id[rid]["verdict"] == "PASS", (
            f"{rid} expected PASS, got {by_id[rid]['verdict']}: "
            f"{by_id[rid]['evidence']}"
        )
    assert agg["overall"] in ("GREEN", "AMBER")  # SKIPs may push to AMBER
    assert agg["fail_count"] == 0


# ---------------------------------------------------------------------------
# Test 2 — F-01 fail (statusEnum drift) → RED
# ---------------------------------------------------------------------------

def test_critical_fail_red_verdict(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    # Mutate the wb: append a master_task row with bogus status.
    wb = load_workbook(str(clean_wb_path))
    mt = wb["master_task"]
    bogus_row = ["T-9999", "bad task", "manual", "", "", "tech", "MEDIUM",
                 "low", 30, "FAKE_STATUS",  # ← outside 7-value enum
                 "2026-04-30T00:00:00Z", "", "", "", False, "", 0, "", ""]
    mt.append(bogus_row)
    wb.save(str(clean_wb_path))
    wb.close()

    wb_ro = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb_ro, slug, workspace_root=tmp_path)
    finally:
        wb_ro.close()
    agg = vi.aggregate_verdicts(results)
    by_id = {r["id"]: r for r in results}
    assert by_id["F-01"]["verdict"] == "FAIL"
    assert "FAKE_STATUS" in by_id["F-01"]["sample_violations"][0]
    assert agg["overall"] == "RED"


# ---------------------------------------------------------------------------
# Test 3 — F-08 sparse-pilot routing → AMBER (manual triage)
# ---------------------------------------------------------------------------

def test_high_fail_red_verdict(tmp_path: Path) -> None:
    """Sparse pilot wb (only quick_wins, no crawl_sitemap, no gsc_performance).
    F-08 returns FAIL with manual_triage=True, which the aggregator routes
    to AMBER (not RED) — F2 flag from the W-S worker brief."""
    slug = "sparse"
    wb_path = _setup_workspace(tmp_path, slug)
    wb = Workbook()
    qw = wb.active
    qw.title = "quick_wins"
    qw.append(["query", "url", "current_position", "impressions_30d",
               "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
               "action", "priority"])
    qw.append(["foo", "https://example.com/orphan", 12, 100, 5, 5.0, 10,
               "High", "x", "MEDIUM"])
    op = wb.create_sheet("opportunity")
    op.append(["query", "opportunity_score", "current_position", "ctr_pct",
               "impressions_30d", "clicks_30d", "potential_clicks",
               "assigned_url_action"])
    op.append(["foo", 800, 12, 5.0, 100, 5, 10,
               "https://example.com/orphan | x"])
    wb.save(str(wb_path))

    wb_ro = load_workbook(str(wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb_ro, slug, workspace_root=tmp_path)
    finally:
        wb_ro.close()
    agg = vi.aggregate_verdicts(results)
    by_id = {r["id"]: r for r in results}
    assert by_id["F-08"]["verdict"] == "FAIL"
    assert by_id["F-08"]["manual_triage"] is True
    # Manual triage routes the FAIL to AMBER (not RED).
    assert "F-08" in agg["manual_review_required"]
    assert agg["overall"] in ("AMBER",)  # explicit: NOT RED


# ---------------------------------------------------------------------------
# Test 4 — F-18 fail (ISO 8601 invalid) → AMBER
# ---------------------------------------------------------------------------

def test_medium_fail_amber_verdict(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    wb = load_workbook(str(clean_wb_path))
    mt = wb["master_task"]
    bogus_row = ["T-1000", "bad date", "manual", "", "", "tech", "MEDIUM",
                 "low", 30, "TODO",
                 "not-a-date",  # ← unparseable
                 "", "", "", False, "", 0, "", ""]
    mt.append(bogus_row)
    wb.save(str(clean_wb_path))
    wb.close()

    wb_ro = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb_ro, slug, workspace_root=tmp_path)
    finally:
        wb_ro.close()
    agg = vi.aggregate_verdicts(results)
    by_id = {r["id"]: r for r in results}
    assert by_id["F-18"]["verdict"] == "FAIL"
    # MEDIUM FAIL → AMBER (no CRITICAL/HIGH FAIL present).
    assert agg["overall"] == "AMBER"


# ---------------------------------------------------------------------------
# Test 5 — sheet missing → SKIP + AMBER (not RED)
# ---------------------------------------------------------------------------

def test_skip_missing_sheet_amber(tmp_path: Path) -> None:
    slug = "minimal"
    wb_path = _setup_workspace(tmp_path, slug)
    wb = Workbook()
    qw = wb.active
    qw.title = "quick_wins"
    qw.append(["query", "url", "current_position", "impressions_30d",
               "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
               "action", "priority"])
    # Add an opportunity sheet so F-16 doesn't SKIP+AMBER but PASSes.
    op = wb.create_sheet("opportunity")
    op.append(["query", "opportunity_score", "current_position", "ctr_pct",
               "impressions_30d", "clicks_30d", "potential_clicks",
               "assigned_url_action"])
    # Add crawl_sitemap so F-08 doesn't trigger sparse-pilot path
    cs = wb.create_sheet("crawl_sitemap")
    cs.append(["category", "metric", "value", "status", "action"])
    wb.save(str(wb_path))

    wb_ro = load_workbook(str(wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb_ro, slug, workspace_root=tmp_path)
    finally:
        wb_ro.close()
    agg = vi.aggregate_verdicts(results)
    by_id = {r["id"]: r for r in results}
    # master_task missing → F-01 SKIP, F-09 SKIP, F-18 SKIP
    assert by_id["F-01"]["verdict"] == "SKIP"
    assert by_id["F-09"]["verdict"] == "SKIP"
    assert by_id["F-18"]["verdict"] == "SKIP"
    # SKIP must NOT promote to RED — must be AMBER.
    assert agg["overall"] == "AMBER"
    assert agg["fail_count"] == 0


# ---------------------------------------------------------------------------
# Test 6 — consistency-report.json schema validate
# ---------------------------------------------------------------------------

def test_consistency_report_schema_valid(clean_wb_path: Path, tmp_path: Path,
                                          consistency_schema: dict) -> None:
    slug = "drifttest"
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    agg = vi.aggregate_verdicts(results)
    report = vi.build_consistency_report(
        rule_results=results,
        aggregation=agg,
        project_id=slug,
        report_id=1,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        run_id=1,
    )
    # Re-validate from outside (build_consistency_report does internal validate
    # but we double-check end-to-end).
    Draft7Validator(consistency_schema).validate(report)
    assert report["schema_version"] == "1.0"
    # 20 baseline + F-23 (v1.8 Phase 4) + F-24 (v1.9 Phase 2) +
    # F-25 (v1.9 Phase 3) + F-26 (v1.9 Phase 4) = 24.
    assert len(report["checks"]) == 24
    assert report["verdict"] in ("GREEN", "AMBER", "RED")


# ---------------------------------------------------------------------------
# Test 7 — aggregator rejects unknown rule id
# ---------------------------------------------------------------------------

def test_unknown_rule_id_raises() -> None:
    bad = [{
        "id": "F-99",
        "severity": "HIGH",
        "verdict": "PASS",
        "evidence": "synthetic",
        "rule": "synthetic",
        "category": "csr_foundation",
    }]
    with pytest.raises(vi.UnknownRuleError):
        vi.aggregate_verdicts(bad)


# ---------------------------------------------------------------------------
# Test 8 — events.jsonl audit event emitted
# ---------------------------------------------------------------------------

def test_provenance_event_emitted(clean_wb_path: Path, tmp_path: Path,
                                   events_schema: dict) -> None:
    slug = "drifttest"
    # Mimic the skill's step 7: emit an event_kind=audit row.
    events_writer.append_audit(
        project_id=slug,
        audit_action="accessed",
        audit_target="invariants:21",
        actor="drift-check",
        notes="verdict=GREEN fails=0",
        workspace_root=tmp_path,
    )
    events_path = tmp_path / "projects" / slug / "_state" / "events.jsonl"
    assert events_path.exists()
    lines = [json.loads(l) for l in events_path.read_text("utf-8").splitlines()
             if l.strip()]
    audits = [e for e in lines if e.get("event_kind") == "audit"]
    assert audits, "no event_kind=audit emitted"
    rec = audits[-1]
    assert rec["audit_action"] == "accessed"
    assert rec["audit_target"] == "invariants:21"
    # Re-validate with the canonical events schema.
    Draft7Validator(events_schema).validate(rec)


# ---------------------------------------------------------------------------
# Test 9 — workspace read-only: master.xlsx SHA-256 unchanged
# ---------------------------------------------------------------------------

def test_workspace_read_only_no_writes(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    sha_before = hashlib.sha256(clean_wb_path.read_bytes()).hexdigest()
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
        agg = vi.aggregate_verdicts(results)
        report = vi.build_consistency_report(
            rule_results=results,
            aggregation=agg,
            project_id=slug,
            report_id=1,
            generated_at="2026-04-30T00:00:00Z",
        )
        assert report["verdict"] in ("GREEN", "AMBER", "RED")
    finally:
        wb.close()
    sha_after = hashlib.sha256(clean_wb_path.read_bytes()).hexdigest()
    assert sha_after == sha_before, (
        "master.xlsx mutated by drift-check evaluation — "
        "read-only governance violated"
    )


# ---------------------------------------------------------------------------
# Test 10 — F-15 manual review populated → AMBER
# ---------------------------------------------------------------------------

def test_f15_manual_review_amber(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    # Add cannibalization rows so F-15 fires (manual triage path).
    wb = load_workbook(str(clean_wb_path))
    cn = wb.create_sheet("cannibalization")
    cn.append(["conflict_pair", "overlapping_queries_est", "total_impact",
               "resolution", "note", "status", "priority"])
    cn.append(["page-a vs page-b", 5, "high", "redirect", "review",
               "TODO", "HIGH"])
    wb.save(str(clean_wb_path))
    wb.close()

    wb_ro = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb_ro, slug, workspace_root=tmp_path)
    finally:
        wb_ro.close()
    agg = vi.aggregate_verdicts(results)
    by_id = {r["id"]: r for r in results}
    assert by_id["F-15"]["verdict"] == "FAIL"
    assert by_id["F-15"]["manual_triage"] is True
    assert "F-15" in agg["manual_review_required"]
    # Aggregator must NOT route this to RED.
    assert agg["overall"] == "AMBER"


# ---------------------------------------------------------------------------
# Test 11 — F-22 auto_repair_available flag set
# ---------------------------------------------------------------------------

def test_auto_repair_available_flag(clean_wb_path: Path, tmp_path: Path) -> None:
    slug = "drifttest"
    backups = tmp_path / "projects" / slug / "_state" / "backups" / "master"
    backups.mkdir(parents=True, exist_ok=True)
    # Create 8 backup files (one over the FIFO cap of 7).
    for i in range(8):
        fname = f"master-2026-04-30T00-00-0{i}-000000Z.xlsx"
        (backups / fname).write_bytes(b"dummy")

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-22"]["verdict"] == "FAIL"
    assert by_id["F-22"]["auto_repair_available"] is True
    assert by_id["F-22"]["affected_rows"] == 1  # 8 - 7
    # And the aggregator surfaces it as an auto_repair_available rule id.
    agg = vi.aggregate_verdicts(results)
    assert "F-22" in agg["auto_repair_available"]


# ---------------------------------------------------------------------------
# Test 12 — F-23 (v1.8 Phase 4): sf-crawl-orchestrator workflow + registry sync
# ---------------------------------------------------------------------------

def test_f23_sf_crawl_orchestrator_registry_mismatch_fail(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-23 violation: a project has a completed sf-crawl-orchestrator
    workflow run, BUT the engine repo-root mcp-tool-registry.json is
    swapped to a synthetic copy that does NOT list 'sf' under servers.

    Expected: check_F_23 → FAIL HIGH → aggregator RED.
    """
    slug = "drifttest"
    # Write a synthetic completed sf-crawl-orchestrator workflow run.
    wf_dir = tmp_path / "projects" / slug / "_state" / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    run_payload = {
        "schema_version": "1.0",
        "run_id": f"{slug}-2026-05-26-deaf",
        "skill": "sf-crawl-orchestrator",
        "project_slug": slug,
        "status": "done",
        "created_at": "2026-05-26T10:00:00Z",
        "updated_at": "2026-05-26T11:00:00Z",
        "completed_at": "2026-05-26T11:00:00Z",
        "steps": [{"name": "complete", "status": "done"}],
    }
    (wf_dir / f"{run_payload['run_id']}.json").write_text(
        json.dumps(run_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Monkey-patch _REPO_ROOT to a tmp_path/engine fixture that ships a
    # registry WITHOUT the 'sf' server.
    engine_root = tmp_path / "engine"
    engine_root.mkdir(parents=True, exist_ok=True)
    bogus_registry = {
        "schema_version": "1.0",
        "servers": {
            "gsc": {"display_name": "GSC", "tools": []},
            "dataforseo": {"display_name": "DFS", "tools": []},
            "scrapling": {"display_name": "Scrapling", "tools": []},
        },
    }
    (engine_root / "mcp-tool-registry.json").write_text(
        json.dumps(bogus_registry, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-23"]["verdict"] == "FAIL", (
        f"F-23 must FAIL when sf-crawl-orchestrator run exists "
        f"but registry missing 'sf'; got {by_id['F-23']}"
    )
    assert by_id["F-23"]["severity"] == "HIGH"
    # FAIL HIGH (no manual_triage) → aggregator RED.
    agg = vi.aggregate_verdicts(results)
    assert agg["overall"] == "RED", (
        f"F-23 FAIL HIGH must drive aggregator RED; got {agg['overall']}"
    )


def test_f23_no_orchestrator_runs_pass(
    clean_wb_path: Path, tmp_path: Path,
) -> None:
    """F-23 vacuously PASS when no sf-crawl-orchestrator workflow runs
    exist — the invariant is conditional, no evidence means no demand
    on the registry."""
    slug = "drifttest"
    # No workflow runs exist for this slug (clean fixture).
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-23"]["verdict"] == "PASS"
    assert "vacuous" in by_id["F-23"]["evidence"].lower() or \
           "no completed" in by_id["F-23"]["evidence"].lower()


def test_f23_orchestrator_runs_with_sf_registry_pass(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-23 PASS when sf-crawl-orchestrator runs exist AND registry has
    'sf' under servers (the current engine baseline post-Phase 1)."""
    slug = "drifttest"
    wf_dir = tmp_path / "projects" / slug / "_state" / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    run_payload = {
        "schema_version": "1.0",
        "run_id": f"{slug}-2026-05-26-cafe",
        "skill": "sf-crawl-orchestrator",
        "project_slug": slug,
        "status": "done",
        "created_at": "2026-05-26T10:00:00Z",
        "updated_at": "2026-05-26T11:00:00Z",
        "completed_at": "2026-05-26T11:00:00Z",
        "steps": [{"name": "complete", "status": "done"}],
    }
    (wf_dir / f"{run_payload['run_id']}.json").write_text(
        json.dumps(run_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Synthesize a registry that DOES list 'sf' (mirrors live engine).
    engine_root = tmp_path / "engine"
    engine_root.mkdir(parents=True, exist_ok=True)
    good_registry = {
        "schema_version": "1.0",
        "servers": {
            "gsc": {"display_name": "GSC", "tools": []},
            "sf": {"display_name": "SF 24 MCP", "tools": []},
        },
    }
    (engine_root / "mcp-tool-registry.json").write_text(
        json.dumps(good_registry, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-23"]["verdict"] == "PASS"
    assert "sf-crawl-orchestrator" in by_id["F-23"]["evidence"]


# ---------------------------------------------------------------------------
# Test 13 — F-24 (v1.9 Phase 2): .mcp.json ↔ mcp-tool-registry.json key sync
# ---------------------------------------------------------------------------

def _write_engine_mcp_pair(
    engine_root: Path,
    mcp_servers: list[str] | None,
    registry_servers: list[str] | None,
) -> None:
    """Write synthetic .mcp.json (mcpServers) + mcp-tool-registry.json
    (servers) into a fixture engine root. F-16 safe: only ever touches
    tmp_path, NEVER the real repo-root .mcp.json. Pass None to omit a file
    (exercises the SKIP branch)."""
    engine_root.mkdir(parents=True, exist_ok=True)
    if mcp_servers is not None:
        (engine_root / ".mcp.json").write_text(
            json.dumps({"mcpServers": {s: {} for s in mcp_servers}}, indent=2),
            encoding="utf-8",
        )
    if registry_servers is not None:
        (engine_root / "mcp-tool-registry.json").write_text(
            json.dumps(
                {"schema_version": "1.0",
                 "servers": {s: {"display_name": s, "tools": []}
                             for s in registry_servers}},
                indent=2,
            ),
            encoding="utf-8",
        )


def test_f24_match_pass(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-24 PASS when the .mcp.json mcpServers key set equals the registry
    servers key set AFTER the D-SF-02 ScraplingServer→scrapling alias.

    .mcp.json carries 'ScraplingServer' (capital), the registry carries
    'scrapling' (lowercase). The EXPLICIT alias normalizes them to the same
    key, so {gsc, dataforseo, scrapling, sf} == {gsc, dataforseo, scrapling,
    sf} → PASS. A naive .lower() would yield 'scraplingserver' != 'scrapling'
    and false-FAIL — this case is the regression guard for the D-SF-02
    case-fold (spec risk R3).
    """
    slug = "drifttest"
    engine_root = tmp_path / "engine"
    _write_engine_mcp_pair(
        engine_root,
        mcp_servers=["gsc", "dataforseo", "ScraplingServer", "sf"],
        registry_servers=["gsc", "dataforseo", "scrapling", "sf"],
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-24"]["verdict"] == "PASS", (
        f"F-24 must PASS when key sets match post ScraplingServer→scrapling "
        f"normalization; got {by_id['F-24']}"
    )
    assert by_id["F-24"]["severity"] == "HIGH"


def test_f24_orphan_transport_fail(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-24 FAIL when .mcp.json declares a server absent from the registry
    (orphan transport): .mcp.json +'newserver' → FAIL HIGH → aggregator RED."""
    slug = "drifttest"
    engine_root = tmp_path / "engine"
    _write_engine_mcp_pair(
        engine_root,
        mcp_servers=["gsc", "dataforseo", "scrapling", "sf", "newserver"],
        registry_servers=["gsc", "dataforseo", "scrapling", "sf"],
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-24"]["verdict"] == "FAIL", (
        f"F-24 must FAIL on orphan transport 'newserver'; got {by_id['F-24']}"
    )
    assert by_id["F-24"]["severity"] == "HIGH"
    assert any(
        "newserver" in v for v in by_id["F-24"].get("sample_violations", [])
    ), f"orphan_transport:newserver expected in violations; got {by_id['F-24']}"
    # FAIL HIGH (no manual_triage) → aggregator RED.
    agg = vi.aggregate_verdicts(results)
    assert agg["overall"] == "RED", (
        f"F-24 FAIL HIGH must drive aggregator RED; got {agg['overall']}"
    )


def test_f24_orphan_inventory_fail(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-24 FAIL when the registry inventories a server absent from .mcp.json
    (orphan inventory): registry +'legacy' → FAIL HIGH."""
    slug = "drifttest"
    engine_root = tmp_path / "engine"
    _write_engine_mcp_pair(
        engine_root,
        mcp_servers=["gsc", "dataforseo", "scrapling", "sf"],
        registry_servers=["gsc", "dataforseo", "scrapling", "sf", "legacy"],
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-24"]["verdict"] == "FAIL", (
        f"F-24 must FAIL on orphan inventory 'legacy'; got {by_id['F-24']}"
    )
    assert by_id["F-24"]["severity"] == "HIGH"
    assert any(
        "legacy" in v for v in by_id["F-24"].get("sample_violations", [])
    ), f"orphan_inventory:legacy expected in violations; got {by_id['F-24']}"


def test_f24_either_file_missing_skip(
    clean_wb_path: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    """F-24 SKIP when either engine file is missing (engine state ambiguous).

    Here the registry exists but .mcp.json is absent → SKIP (NOT FAIL):
    drift-check must not hard-fail on an incomplete engine checkout. Mirrors
    the spec FE-1 'SKIP cases: Either file missing' contract.
    """
    slug = "drifttest"
    engine_root = tmp_path / "engine"
    # Write ONLY the registry; .mcp.json deliberately omitted.
    _write_engine_mcp_pair(
        engine_root, mcp_servers=None, registry_servers=["gsc", "sf"],
    )
    monkeypatch.setattr(vi, "_REPO_ROOT", engine_root)

    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-24"]["verdict"] == "SKIP", (
        f"F-24 must SKIP when .mcp.json is missing; got {by_id['F-24']}"
    )
    assert ".mcp.json" in by_id["F-24"]["evidence"]


# ---------------------------------------------------------------------------
# F-25 (v1.9 Phase 3) — sf.mcp.enabled=true ⇒ schema_version >= 1.5
#
# F-25 is a PER-PROJECT invariant: it reads projects/{slug}/project.config.json
# only. It NEVER touches the engine repo-root .mcp.json (F-16 safe) — so these
# tests rewrite the tmp-workspace config built by `clean_wb_path` and never go
# near a real project.config.json.
# ---------------------------------------------------------------------------

def _write_project_config(tmp_path: Path, slug: str, config: dict) -> None:
    """Overwrite projects/{slug}/project.config.json in the tmp workspace.

    F-25 reads ONLY this per-project file; the engine repo-root .mcp.json and
    mcp-tool-registry.json are untouched (F-16 discipline)."""
    cfg_path = tmp_path / "projects" / slug / "project.config.json"
    cfg_path.write_text(json.dumps(config), encoding="utf-8")


def test_f25_enabled_true_v1_5_pass(clean_wb_path: Path, tmp_path: Path) -> None:
    """sf.mcp.enabled=true + schema_version='1.5' → PASS (coupling satisfied)."""
    slug = "drifttest"
    _write_project_config(tmp_path, slug, {
        "project_id": slug,
        "schema_version": "1.5",
        "language": {"content_locale": "tr-TR"}, "market": "TR",
        "sf": {"mcp": {"enabled": True}},
    })
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-25"]["verdict"] == "PASS", (
        f"F-25 must PASS when sf.mcp.enabled=true on schema_version 1.5; "
        f"got {by_id['F-25']}"
    )
    assert by_id["F-25"]["severity"] == "HIGH"


def test_f25_enabled_true_v1_4_fail(clean_wb_path: Path, tmp_path: Path) -> None:
    """sf.mcp.enabled=true + schema_version='1.4' → FAIL HIGH → aggregator RED.

    Migration 0005 makes this combination impossible in normal operation (the
    sf block only lands on v1.5), but the defensive check catches a config that
    was hand-edited past the migration lock."""
    slug = "drifttest"
    _write_project_config(tmp_path, slug, {
        "project_id": slug,
        "schema_version": "1.4",
        "language": {"content_locale": "tr-TR"}, "market": "TR",
        "sf": {"mcp": {"enabled": True}},
    })
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-25"]["verdict"] == "FAIL", (
        f"F-25 must FAIL when sf.mcp.enabled=true on schema_version 1.4; "
        f"got {by_id['F-25']}"
    )
    assert by_id["F-25"]["severity"] == "HIGH"
    # FAIL HIGH (no manual_triage) → aggregator RED.
    agg = vi.aggregate_verdicts(results)
    assert agg["overall"] == "RED", (
        f"F-25 FAIL HIGH must drive aggregator RED; got {agg['overall']}"
    )


def test_f25_enabled_false_v1_4_pass(clean_wb_path: Path, tmp_path: Path) -> None:
    """sf.mcp.enabled=false + schema_version='1.4' → PASS.

    The version gate only applies when SF MCP is turned ON; the default
    file-drop behavior (enabled=false) is valid on any schema version."""
    slug = "drifttest"
    _write_project_config(tmp_path, slug, {
        "project_id": slug,
        "schema_version": "1.4",
        "language": {"content_locale": "tr-TR"}, "market": "TR",
        "sf": {"mcp": {"enabled": False}},
    })
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-25"]["verdict"] == "PASS", (
        f"F-25 must PASS when sf.mcp.enabled=false regardless of version; "
        f"got {by_id['F-25']}"
    )
    assert by_id["F-25"]["severity"] == "HIGH"


def test_f25_sf_block_absent_pass(clean_wb_path: Path, tmp_path: Path) -> None:
    """No sf block at all + schema_version='1.4' → PASS (default behavior).

    A config with no `sf` key is the common case; it must not trip the gate
    on any schema version."""
    slug = "drifttest"
    _write_project_config(tmp_path, slug, {
        "project_id": slug,
        "schema_version": "1.4",
        "language": {"content_locale": "tr-TR"}, "market": "TR",
    })
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-25"]["verdict"] == "PASS", (
        f"F-25 must PASS when no sf block is present; got {by_id['F-25']}"
    )
    assert by_id["F-25"]["severity"] == "HIGH"


def test_f25_config_missing_skip(clean_wb_path: Path, tmp_path: Path) -> None:
    """project.config.json absent → SKIP (state ambiguous, surfaces AMBER).

    Mirrors the spec FE-2 'SKIP cases: project.config.json missing OR
    schema_version field absent' contract. drift-check must not hard-fail when
    a project has no config yet."""
    slug = "drifttest"
    (tmp_path / "projects" / slug / "project.config.json").unlink()
    wb = load_workbook(str(clean_wb_path), data_only=True, read_only=True)
    try:
        results = vi.evaluate_all(wb, slug, workspace_root=tmp_path)
    finally:
        wb.close()
    by_id = {r["id"]: r for r in results}
    assert by_id["F-25"]["verdict"] == "SKIP", (
        f"F-25 must SKIP when project.config.json is missing; "
        f"got {by_id['F-25']}"
    )
    assert "missing" in by_id["F-25"]["evidence"]


# ---------------------------------------------------------------------------
# F-26 (v1.9 Phase 4) — orphan SF GUI crawl detection (MCP-aware → AMBER)
#
# F-26 is MEDIUM severity (the FIRST MEDIUM v1.9 invariant): a FAIL surfaces
# AMBER, never RED — an orphan crawl is an operator cleanup hint, not a data
# break (D-V1.9-11 / spec v2.2 line 210). The check is MCP-aware: it does a 1s
# SfMcpClient.health() probe BEFORE any sf_crawl_progress call and SKIPs (never
# hangs drift-check) when the probe fails (Risk R2).
#
# These tests INJECT a fake client (health + call_tool) — no real SF MCP socket
# is ever opened (mirrors tests/scripts/test_sf_mcp_client.py's injection
# ethos). check_F_26 is called DIRECTLY (not via evaluate_all, which does not
# thread an mcp_client through); the fake is the seam that keeps the network —
# and the repo-root .mcp.json — entirely out of the test path (F-16 safe;
# fixtures live only in tmp_path).
# ---------------------------------------------------------------------------

class _FakeSfMcpClient:
    """Stand-in for scripts.util.sf_mcp_client.SfMcpClient (health + call_tool).

    Mirrors the real API surface F-26 depends on: ``health() -> bool`` and
    ``call_tool(tool_name, **kwargs) -> dict``. Records calls so tests can
    assert the probe-before-progress ordering and the exact tool invoked.
    Never touches a socket."""

    def __init__(self, *, healthy: bool = True,
                 progress_status: str = "IN_PROGRESS") -> None:
        self._healthy = healthy
        self._progress_status = progress_status
        self.health_calls = 0
        self.tool_calls: list[tuple[str, dict]] = []

    def health(self) -> bool:
        self.health_calls += 1
        return self._healthy

    def call_tool(self, tool_name: str, **kwargs) -> dict:
        self.tool_calls.append((tool_name, dict(kwargs)))
        return {"status": self._progress_status}


def _write_sf_workflow(
    tmp_path: Path, slug: str, *, run_id: str, status: str,
    crawl_id: str | None = None, skill: str = "sf-crawl-orchestrator",
) -> None:
    """Write a synthetic sf-crawl-orchestrator workflow-run JSON in the tmp
    workspace. crawl_id is persisted the way the orchestrator persists it — in a
    step's ``output_ref`` as ``crawl_id=<value>`` (sf-crawl-orchestrator
    SKILL.md trigger step), because workflow-run.schema.json is
    additionalProperties:false at root. F-16 safe: only ever tmp_path."""
    wf_dir = tmp_path / "projects" / slug / "_state" / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    steps: list[dict] = [{"name": "preflight", "status": "done"}]
    if crawl_id is not None:
        steps.append({"name": "trigger_crawl", "status": "done",
                      "output_ref": f"crawl_id={crawl_id}"})
    payload: dict = {
        "schema_version": "1.0",
        "run_id": run_id,
        "skill": skill,
        "project_slug": slug,
        "status": status,
        "started_at": "2026-05-26T10:00:00Z",
        "updated_at": "2026-05-26T11:00:00Z",
        "steps": steps,
    }
    if status == "paused":
        payload["paused_at"] = "2026-05-26T11:00:00Z"
    if status == "failed":
        payload["ended_at"] = "2026-05-26T11:00:00Z"
        payload["failure_reason"] = {"code": "timeout", "message": "stalled"}
    (wf_dir / f"{run_id}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def test_f26_no_orphan_pass(tmp_path: Path) -> None:
    """PASS — no orphan. Covers spec FE-3 PASS cases (a) AND (b):
    (a) a ``done`` sf-crawl-orchestrator run is ignored (not paused/failed); and
    (b) a ``paused`` run whose SF GUI reports a terminal status (DONE, NOT
    IN_PROGRESS) → workflow state and GUI agree the crawl ended → no orphan.

    Also pins the status filter (only paused/failed runs are probed) and the
    health-probe-then-progress ordering."""
    slug = "drifttest"
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-dead",
                       status="done", crawl_id="crawl-done")
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-cafe",
                       status="paused", crawl_id="crawl-paused")
    client = _FakeSfMcpClient(healthy=True, progress_status="DONE")
    # check_F_26 ignores the workbook (reads _state/workflows + SF MCP only).
    result = vi.check_F_26(None, slug, workspace_root=tmp_path, mcp_client=client)
    assert result["verdict"] == "PASS", result
    assert result["severity"] == "MEDIUM"
    assert client.health_calls == 1, "health() probe must run once"
    # Only the paused run is probed; the done run is filtered out.
    assert client.tool_calls == [
        ("sf_crawl_progress", {"crawl_id": "crawl-paused"})
    ], client.tool_calls


def test_f26_orphan_detected_amber(tmp_path: Path) -> None:
    """AMBER — a ``failed`` workflow whose SF GUI still reports IN_PROGRESS is an
    orphan (workflow state + GUI disagree). FAIL MEDIUM → aggregator AMBER, NOT
    RED (D-V1.9-11)."""
    slug = "drifttest"
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-beef",
                       status="failed", crawl_id="crawl-fail")
    client = _FakeSfMcpClient(healthy=True, progress_status="IN_PROGRESS")
    result = vi.check_F_26(None, slug, workspace_root=tmp_path, mcp_client=client)
    assert result["verdict"] == "FAIL", result
    assert result["severity"] == "MEDIUM"
    assert result["affected_rows"] == 1
    assert any("crawl-fail" in v for v in result["sample_violations"]), result
    # MEDIUM FAIL must surface AMBER, never RED.
    agg = vi.aggregate_verdicts([result])
    assert agg["overall"] == "AMBER", (
        f"F-26 MEDIUM FAIL must be AMBER (operator hint), not RED; "
        f"got {agg['overall']}"
    )


def test_f26_mcp_unreachable_skip(tmp_path: Path) -> None:
    """SKIP — a ``paused`` run exists but the 1s health probe fails (MCP down).
    drift-check must NOT hang or hard-fail: F-26 SKIPs (→ AMBER) and never calls
    sf_crawl_progress (Risk R2 mitigation)."""
    slug = "drifttest"
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-cafe",
                       status="paused", crawl_id="crawl-paused")
    client = _FakeSfMcpClient(healthy=False)  # health probe fails (MCP down)
    result = vi.check_F_26(None, slug, workspace_root=tmp_path, mcp_client=client)
    assert result["verdict"] == "SKIP", result
    assert result["severity"] == "MEDIUM"
    assert client.health_calls == 1, "health() probe must be attempted"
    assert client.tool_calls == [], (
        "no sf_crawl_progress call may fire after a failed health probe (R2)"
    )
    # SKIP surfaces AMBER (never RED).
    agg = vi.aggregate_verdicts([result])
    assert agg["overall"] == "AMBER"


def test_f26_multiple_orphans_amber(tmp_path: Path) -> None:
    """AMBER — two paused runs both still IN_PROGRESS per the SF GUI → two
    orphans, two sample_violations, affected_rows == 2."""
    slug = "drifttest"
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-aaaa",
                       status="paused", crawl_id="crawl-aaaa")
    _write_sf_workflow(tmp_path, slug, run_id=f"{slug}-2026-05-26-bbbb",
                       status="paused", crawl_id="crawl-bbbb")
    client = _FakeSfMcpClient(healthy=True, progress_status="IN_PROGRESS")
    result = vi.check_F_26(None, slug, workspace_root=tmp_path, mcp_client=client)
    assert result["verdict"] == "FAIL", result
    assert result["severity"] == "MEDIUM"
    assert result["affected_rows"] == 2, result
    assert len(result["sample_violations"]) == 2, result
    joined = " ".join(result["sample_violations"])
    assert "crawl-aaaa" in joined and "crawl-bbbb" in joined, result
    # Both stalled runs were probed (1 health probe, 2 progress calls).
    assert client.health_calls == 1
    assert len(client.tool_calls) == 2
    agg = vi.aggregate_verdicts([result])
    assert agg["overall"] == "AMBER"
