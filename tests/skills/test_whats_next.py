"""
tests/skills/test_whats_next.py — whats-next meta/router skill tests.

Five cases exercising the router pattern end-to-end on tmp_path
fixtures (no workspace-staging side effects). Whats-next is read-only
+ one events.jsonl row per run; no Excel writes, no skill invocations.

Tests
-----
1. test_master_task_todo_scan
   Synthetic master_task sheet (TODO + non-TODO rows) → only TODO
   surfaces as a candidate, with priority/task_id captured.

2. test_content_decay_red_scan_amber_when_absent
   Sheet absent → empty list, NOT error (AMBER). Sheet present with
   RED rows → those rows surface; non-RED filtered out.

3. test_pending_approvals_scan
   Real workflow_runner round-trip: create_run + request_approval
   produces an awaiting_approval run that the scan picks up.

4. test_top_k_ranked_output
   Default k=3, custom k=5; rank order matches the heuristic table
   (pending_approval=100 > content_decay_red=80 > master_task_high=60
   > quick_wins_pending=50 > master_task_medium=40).

5. test_router_event_emitted
   `run()` writes exactly one event_kind=work, event_type=skill_whats_next
   row into events.jsonl with task_id matching ^T-9[0-9]{4}$ and the
   markdown summary in `note`. Workflow run completes with
   outputs.recommendations_count + outputs.top_skill STRING-TYPED
   (F5 invariant).

Schemas referenced
------------------
- schemas/skill-frontmatter.schema.json
- schemas/events.schema.json (event_kind=work, event_type=skill_whats_next
  canonical per v1.6-Phase-2 H-E, task_id pattern, note required,
  primary_source=manual)
- schemas/workflow-run.schema.json (outputs.* string-typed)
- schemas/master-excel.schema.json (sheet column conventions)
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pytest
from jsonschema import Draft7Validator
from openpyxl import Workbook

from scripts.meta import whats_next
from scripts.state import events_writer, workflow_runner


# ---------------------------------------------------------------------------
# Constants — paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def events_schema() -> dict:
    return json.loads((SCHEMAS / "events.schema.json").read_text("utf-8"))


@pytest.fixture(scope="module")
def workflow_schema() -> dict:
    return json.loads((SCHEMAS / "workflow-run.schema.json").read_text("utf-8"))


@pytest.fixture
def slug() -> str:
    return "wn-test"


def _setup_ws(tmp_path: Path, slug: str) -> Path:
    """Create projects/{slug}/_state/workflows/ + projects/{slug}/_state/."""
    proj = tmp_path / "projects" / slug
    (proj / "_state" / "workflows").mkdir(parents=True, exist_ok=True)
    return proj


def _make_master_xlsx(
    proj_dir: Path,
    *,
    master_task_rows: list[tuple[Any, ...]] | None = None,
    content_decay_rows: list[tuple[Any, ...]] | None = None,
    quick_wins_rows: list[tuple[Any, ...]] | None = None,
) -> Path:
    """Build a synthetic master.xlsx matching master-excel.schema layout.

    Each sheet uses its schema-mandated header_row + data_start_row (4/5/6).
    """
    wb = Workbook()
    # openpyxl gives a default sheet — remove it so we control what exists.
    default = wb.active
    wb.remove(default)

    if master_task_rows is not None:
        ws = wb.create_sheet("master_task")
        # Header at row 3 (data_start_row=4).
        headers = ["task_id", "task", "primary_source", "related_sources",
                   "url", "category", "priority", "impact",
                   "duration_est_min", "status", "created_date", "done_date",
                   "assignee", "note", "auto_generated", "dependencies",
                   "effort_actual_min", "metric_impact_json", "work_log_ref"]
        ws.append([])  # row 1
        ws.append([])  # row 2
        ws.append(headers)  # row 3
        for r in master_task_rows:
            # Pad to 19 cols.
            row = list(r) + [None] * (19 - len(r))
            ws.append(row)

    if content_decay_rows is not None:
        ws = wb.create_sheet("content_decay")
        # Header at row 5 (data_start_row=6).
        headers = ["url", "clicks_previous", "clicks_recent", "clicks_delta",
                   "delta_pct", "trend", "pillar", "action"]
        for _ in range(4):
            ws.append([])  # rows 1-4
        ws.append(headers)  # row 5
        for r in content_decay_rows:
            row = list(r) + [None] * (8 - len(r))
            ws.append(row)

    if quick_wins_rows is not None:
        ws = wb.create_sheet("quick_wins")
        # Header at row 4 (data_start_row=5).
        headers = ["query", "url", "current_position", "impressions_30d",
                   "clicks_30d", "ctr_pct", "potential_clicks", "opportunity",
                   "action", "priority"]
        for _ in range(3):
            ws.append([])  # rows 1-3
        ws.append(headers)  # row 4
        for r in quick_wins_rows:
            row = list(r) + [None] * (10 - len(r))
            ws.append(row)

    if not wb.sheetnames:
        # openpyxl requires at least one sheet to save.
        wb.create_sheet("_placeholder")

    target = proj_dir / "master.xlsx"
    wb.save(str(target))
    return target


# ---------------------------------------------------------------------------
# Test 1 — master_task TODO scan
# ---------------------------------------------------------------------------

def test_master_task_todo_scan(tmp_path: Path, slug: str) -> None:
    proj = _setup_ws(tmp_path, slug)
    _make_master_xlsx(
        proj,
        master_task_rows=[
            # task_id, task, primary_source, related_sources, url, category,
            # priority, impact, duration_est_min, status, ...
            ("T-0001", "Refresh meta tags", "manual", None, None,
             None, "HIGH", None, 30, "TODO"),
            ("T-0002", "Done already", "manual", None, None,
             None, "MEDIUM", None, 15, "DONE"),
            ("T-0003", "Add FAQ schema", "schema", None, None,
             None, "MEDIUM", None, 45, "TODO"),
            ("T-0004", "Low priority cleanup", "manual", None, None,
             None, "LOW", None, 10, "TODO"),
        ],
    )

    out = whats_next.scan_master_task_todo(slug, workspace_root=tmp_path)
    # Three TODO rows survive; the DONE row is filtered.
    assert len(out) == 3
    assert all(c["kind"] == "master_task" for c in out)
    assert {c["task_id"] for c in out} == {"T-0001", "T-0003", "T-0004"}
    # Priority capture intact.
    by_id = {c["task_id"]: c for c in out}
    assert by_id["T-0001"]["priority"] == "HIGH"
    assert by_id["T-0003"]["priority"] == "MEDIUM"
    assert by_id["T-0004"]["priority"] == "LOW"
    # Status preserved as the literal stored value (no decoration).
    assert all(c["status"] == "TODO" for c in out)


# ---------------------------------------------------------------------------
# Test 2 — content_decay sheet absent → AMBER (empty list, not error)
# ---------------------------------------------------------------------------

def test_content_decay_red_scan_amber_when_absent(
    tmp_path: Path, slug: str,
) -> None:
    proj = _setup_ws(tmp_path, slug)
    # Build a workbook WITHOUT a content_decay sheet (only master_task).
    _make_master_xlsx(
        proj,
        master_task_rows=[
            ("T-0001", "x", "manual", None, None, None, "HIGH",
             None, 5, "TODO"),
        ],
    )

    # AMBER: absent sheet returns [], does NOT raise.
    out = whats_next.scan_content_decay_red(slug, workspace_root=tmp_path)
    assert out == [], "absent content_decay sheet must return [], not raise"

    # Now rebuild with a content_decay sheet containing RED + non-RED rows.
    _make_master_xlsx(
        proj,
        master_task_rows=[],
        content_decay_rows=[
            # url, clicks_previous, clicks_recent, clicks_delta, delta_pct,
            # trend, pillar, action
            ("https://x.test/a", 100, 50, -50, -50.0, "RED",       "P01_x", "fix"),
            ("https://x.test/b", 100, 110,  10,  10.0, "GREEN",    "P01_x", "ok"),
            ("https://x.test/c", 100, 60,  -40, -40.0, "RED-fast", "P01_x", "fix"),
        ],
    )
    out = whats_next.scan_content_decay_red(slug, workspace_root=tmp_path)
    assert len(out) == 2, f"only RED-trend rows surface, got {out}"
    urls = {c["url"] for c in out}
    assert urls == {"https://x.test/a", "https://x.test/c"}
    assert all(c["kind"] == "content_decay_red" for c in out)


# ---------------------------------------------------------------------------
# Test 3 — pending approvals scan picks up awaiting_approval workflow runs
# ---------------------------------------------------------------------------

def test_pending_approvals_scan(tmp_path: Path, slug: str) -> None:
    _setup_ws(tmp_path, slug)
    # Create a run, immediately gate it on approval.
    handle = workflow_runner.create_run(
        skill="quick-wins",
        project_slug=slug,
        steps=[{"name": "transform"}, {"name": "request_approval"}],
        workspace_root=tmp_path,
    )
    workflow_runner.request_approval(
        handle.run_id,
        project_slug=slug,
        approver="user",
        subject="Approve top-N quick-wins write?",
        step_index=1,
        workspace_root=tmp_path,
    )
    # Create a SECOND run that stays "running" — should NOT be picked up.
    other = workflow_runner.create_run(
        skill="sf-import",
        project_slug=slug,
        steps=[{"name": "ingest"}],
        workspace_root=tmp_path,
    )
    assert other.status == "running"

    out = whats_next.scan_pending_approvals(slug, workspace_root=tmp_path)
    assert len(out) == 1, f"only the awaiting_approval run surfaces, got {out}"
    cand = out[0]
    assert cand["kind"] == "pending_approval"
    assert cand["run_id"] == handle.run_id
    assert cand["skill"] == "quick-wins"
    assert cand["approver"] == "user"
    assert "Approve" in cand["subject"]


# ---------------------------------------------------------------------------
# Test 4 — top-K ranking respects the heuristic table (and tie-breaks stably)
# ---------------------------------------------------------------------------

def test_top_k_ranked_output() -> None:
    candidates = [
        {"kind": "master_task",        "row_id": 4,
         "task_id": "T-0001", "priority": "HIGH",   "task": "x"},
        {"kind": "master_task",        "row_id": 5,
         "task_id": "T-0002", "priority": "MEDIUM", "task": "y"},
        {"kind": "master_task",        "row_id": 6,
         "task_id": "T-0003", "priority": "LOW",    "task": "z"},
        {"kind": "content_decay_red",  "row_id": 6,
         "url": "https://x.test/a", "trend": "RED", "delta_pct": -50.0},
        {"kind": "pending_approval",   "row_id": "wn-test-2026-04-30-aaaa",
         "run_id": "wn-test-2026-04-30-aaaa", "skill": "quick-wins",
         "subject": "go?", "approver": "user"},
        {"kind": "quick_wins_pending", "row_id": 5,
         "query": "kw", "url": "https://x.test/q", "opportunity": "Medium"},
    ]

    # Default top_k=3 — pending_approval (100) > content_decay_red (80) >
    # master_task HIGH (60).
    top3 = whats_next.rank_candidates(candidates, top_k=3)
    assert len(top3) == 3
    assert [r["score"] for r in top3] == [100, 80, 60]
    assert [r["kind"] for r in top3] == [
        "pending_approval", "content_decay_red", "master_task",
    ]
    # Rank values monotonic.
    assert [r["rank"] for r in top3] == [1, 2, 3]
    # Top suggested skill resolves via RECOMMENDED_SKILL.
    assert top3[0]["suggested_skill"] == whats_next.RECOMMENDED_SKILL["pending_approval"]

    # top_k=5 surfaces the full ladder.
    top5 = whats_next.rank_candidates(candidates, top_k=5)
    assert [r["score"] for r in top5] == [100, 80, 60, 50, 40]
    # top_k=10 includes the LOW master_task at the tail (score=20).
    top10 = whats_next.rank_candidates(candidates, top_k=10)
    assert top10[-1]["score"] == 20

    # Heuristic table itself is the SSoT — assert it matches the spec.
    # v1.8 Phase 4: sf_crawl_stale=30 additive (slots below master_task_medium
    # so it never displaces a higher-priority signal in top-K).
    assert whats_next.SCORES == {
        "pending_approval":   100,
        "content_decay_red":   80,
        "master_task_high":    60,
        "quick_wins_pending":  50,
        "master_task_medium":  40,
        "sf_crawl_stale":      30,
        "master_task_other":   20,
    }


# ---------------------------------------------------------------------------
# Test 5 — full router run emits a schema-valid events.jsonl row + F5 outputs
# ---------------------------------------------------------------------------

def test_suggest_sf_crawl_when_stale(tmp_path: Path, slug: str) -> None:
    """v1.8 Phase 4 (D-SF-04): the helper surfaces a non-blocking
    sf_crawl_stale candidate when:
      (a) project.config.sf.mcp.enabled == True, AND
      (b) no completed sf-crawl-orchestrator workflow run exists, OR
          the newest completed run is older than `threshold_days`.

    When sf.mcp.enabled is False / absent → helper returns None
    regardless of workflow state. When a fresh run exists (today) →
    helper returns None.
    """
    proj = _setup_ws(tmp_path, slug)

    # Case 1 — sf.mcp not opted-in: helper returns None.
    cfg_path = proj / "project.config.json"
    cfg_path.write_text(
        json.dumps({
            "schema_version": "1.5",
            "project_id": slug,
            "domain": "https://example.com/",
            "market": "TR",
            "language": {"content_locale": "tr-TR"},
            "sf": {"mcp": {"enabled": False}},
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    assert whats_next.suggest_sf_crawl_when_stale(
        slug, workspace_root=tmp_path,
    ) is None

    # Case 2 — sf.mcp opted-in, no workflow runs → STALE suggestion.
    cfg_path.write_text(
        json.dumps({
            "schema_version": "1.5",
            "project_id": slug,
            "domain": "https://example.com/",
            "market": "TR",
            "language": {"content_locale": "tr-TR"},
            "sf": {"mcp": {"enabled": True, "url": "http://127.0.0.1:11435/mcp"}},
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    sug = whats_next.suggest_sf_crawl_when_stale(
        slug, workspace_root=tmp_path,
    )
    assert sug is not None, (
        "sf.mcp.enabled=True + no workflow runs must surface a stale suggestion"
    )
    assert sug["kind"] == "sf_crawl_stale"
    assert sug["project_slug"] == slug
    assert sug["last_seen"] is None
    assert sug["threshold_days"] == 30

    # Case 3 — sf.mcp opted-in, FRESH run (today) → None.
    wf_dir = proj / "_state" / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    fresh_run = {
        "schema_version": "1.0",
        "run_id": f"{slug}-2026-05-26-fre1",
        "skill": "sf-crawl-orchestrator",
        "project_slug": slug,
        "status": "done",
        "created_at": "2026-05-26T10:00:00Z",
        "updated_at": "2026-05-26T11:00:00Z",
        "ended_at": "2026-05-26T11:00:00Z",
        "steps": [{"name": "complete", "status": "done"}],
    }
    (wf_dir / f"{fresh_run['run_id']}.json").write_text(
        json.dumps(fresh_run, ensure_ascii=False),
        encoding="utf-8",
    )
    # Use a very generous threshold so the test isn't time-fragile.
    sug = whats_next.suggest_sf_crawl_when_stale(
        slug, workspace_root=tmp_path, threshold_days=365_000,
    )
    assert sug is None, (
        f"fresh sf-crawl-orchestrator run + huge threshold must suppress "
        f"the suggestion; got {sug}"
    )

    # Case 4 — sf.mcp opted-in, STALE run (>30d ago) → STALE suggestion.
    stale_run = {
        "schema_version": "1.0",
        "run_id": f"{slug}-2024-01-01-stal",
        "skill": "sf-crawl-orchestrator",
        "project_slug": slug,
        "status": "done",
        "created_at": "2024-01-01T00:00:00Z",
        "updated_at": "2024-01-01T01:00:00Z",
        "ended_at": "2024-01-01T01:00:00Z",
        "steps": [{"name": "complete", "status": "done"}],
    }
    # Wipe fresh fixture, replace with stale-only state.
    (wf_dir / f"{fresh_run['run_id']}.json").unlink()
    (wf_dir / f"{stale_run['run_id']}.json").write_text(
        json.dumps(stale_run, ensure_ascii=False),
        encoding="utf-8",
    )
    sug = whats_next.suggest_sf_crawl_when_stale(
        slug, workspace_root=tmp_path,
    )
    assert sug is not None
    assert sug["last_seen"] == "2024-01-01"

    # Case 5 — score for the sf_crawl_stale kind = 30 (below master_task_medium
    # but above master_task_other), as documented in the SCORES table.
    assert whats_next.SCORES["sf_crawl_stale"] == 30
    assert whats_next.score_candidate({"kind": "sf_crawl_stale"}) == 30


def test_sf_crawl_staleness_uses_ended_at_not_updated_at(tmp_path: Path, slug: str) -> None:
    """Staleness must be measured from the run's terminal timestamp (ended_at —
    the field workflow_runner actually writes), NOT from updated_at. The old code
    read a non-existent `completed_at` and fell through to `updated_at`, so a
    post-completion touch of updated_at would mask a long-stale crawl (deep-audit)."""
    proj = _setup_ws(tmp_path, slug)
    (proj / "project.config.json").write_text(
        json.dumps({
            "schema_version": "1.5",
            "project_id": slug,
            "domain": "https://example.com/",
            "market": "TR",
            "language": {"content_locale": "tr-TR"},
            "sf": {"mcp": {"enabled": True, "url": "http://127.0.0.1:11435/mcp"}},
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    wf_dir = proj / "_state" / "workflows"
    wf_dir.mkdir(parents=True, exist_ok=True)
    # Crawl actually ENDED in 2024 (long stale), but updated_at was touched far in
    # the future. Staleness must follow ended_at (2024), not updated_at (2099).
    run = {
        "schema_version": "1.0",
        "run_id": f"{slug}-2024-01-01-endd",
        "skill": "sf-crawl-orchestrator",
        "project_slug": slug,
        "status": "done",
        "created_at": "2024-01-01T00:00:00Z",
        "ended_at": "2024-01-01T01:00:00Z",
        "updated_at": "2099-12-31T00:00:00Z",
        "steps": [{"name": "complete", "status": "done"}],
    }
    (wf_dir / f"{run['run_id']}.json").write_text(
        json.dumps(run, ensure_ascii=False), encoding="utf-8",
    )

    sug = whats_next.suggest_sf_crawl_when_stale(slug, workspace_root=tmp_path)
    assert sug is not None, (
        "a crawl that ENDED in 2024 is stale regardless of a later updated_at touch"
    )
    assert sug["last_seen"] == "2024-01-01"


def test_router_event_emitted(
    tmp_path: Path,
    slug: str,
    events_schema: dict,
    workflow_schema: dict,
) -> None:
    proj = _setup_ws(tmp_path, slug)
    # Workspace contains: 1 TODO master_task (HIGH), 1 RED decay row,
    # 1 quick_wins row pending, plus an awaiting_approval workflow.
    _make_master_xlsx(
        proj,
        master_task_rows=[
            ("T-0001", "Refresh meta", "manual", None, None,
             None, "HIGH", None, 30, "TODO"),
        ],
        content_decay_rows=[
            ("https://x.test/a", 100, 40, -60, -60.0, "RED",
             "P01_x", "rewrite"),
        ],
        quick_wins_rows=[
            ("kw-a", "https://x.test/q", 12.5, 500, 5, 1.0, 30,
             "Medium", "fix meta", "MEDIUM"),
        ],
    )
    pending = workflow_runner.create_run(
        skill="quick-wins",
        project_slug=slug,
        steps=[{"name": "transform"}, {"name": "request_approval"}],
        workspace_root=tmp_path,
    )
    workflow_runner.request_approval(
        pending.run_id, project_slug=slug,
        approver="user", subject="Approve top-N?",
        step_index=1, workspace_root=tmp_path,
    )

    result = whats_next.run(
        slug,
        user_intent="ne yapayım",
        top_k=3,
        workspace_root=tmp_path,
    )

    # Top-3 in correct order.
    ranked = result["recommendations"]
    assert len(ranked) == 3
    assert [r["kind"] for r in ranked] == [
        "pending_approval", "content_decay_red", "master_task",
    ]

    # Summary string carries the project slug + intent.
    assert slug in result["summary"]
    assert "ne yapayım" in result["summary"]

    # ---- events.jsonl: exactly ONE work/manual row, schema-valid -----------
    events_path = tmp_path / "projects" / slug / "_state" / "events.jsonl"
    lines = [
        json.loads(l) for l in events_path.read_text("utf-8").splitlines()
        if l.strip()
    ]
    work_router = [
        e for e in lines
        if e.get("event_kind") == "work" and e.get("event_type") == "skill_whats_next"
    ]
    assert len(work_router) == 1, (
        f"expected exactly one work/skill_whats_next router event, got {len(work_router)}: "
        f"{work_router}"
    )
    evt = work_router[0]
    # task_id pattern (events.schema): ^T-[0-9]{4,}$
    assert re.match(r"^T-[0-9]{4,}$", evt["task_id"]), evt["task_id"]
    # Synthetic router id band: 90000..99998 (mod 9999 + 90000).
    n = int(evt["task_id"].split("-")[1])
    assert 90000 <= n <= 99999, f"synthetic id outside router band: {evt['task_id']}"
    # note required for skill_X canonical entries (events.schema v1.6-Phase-2 H-E).
    assert evt["note"] and "whats-next" in evt["note"]
    assert evt["primary_source"] == "manual"
    # Re-validate against the canonical events schema.
    errs = sorted(Draft7Validator(events_schema).iter_errors(evt),
                  key=lambda e: list(e.absolute_path))
    assert not errs, f"router event invalid: {[(list(e.absolute_path), e.message) for e in errs]}"

    # ---- F5: outputs.* string-typed in workflow-run.json --------------------
    run_path = (tmp_path / "projects" / slug / "_state" / "workflows"
                / f"{result['run_id']}.json")
    run_data = json.loads(run_path.read_text("utf-8"))
    assert run_data["status"] == "done"
    outputs = run_data.get("outputs") or {}
    assert "recommendations_count" in outputs
    assert "top_skill" in outputs
    # Both MUST be strings (F5 invariant — int outputs break workflow-run.schema).
    assert isinstance(outputs["recommendations_count"], str)
    assert isinstance(outputs["top_skill"], str)
    # Re-validate the run JSON against workflow-run.schema.
    errs = sorted(Draft7Validator(workflow_schema).iter_errors(run_data),
                  key=lambda e: list(e.absolute_path))
    assert not errs, f"workflow run invalid: {[(list(e.absolute_path), e.message) for e in errs]}"
