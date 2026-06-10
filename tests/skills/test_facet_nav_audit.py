"""TDD lock for skills/discovery/facet-nav-audit/SKILL.md + its wiring (GAP-T2).

Contract checks (spec GAP-T2 §d):
  - frontmatter parses + validates against skill-frontmatter.schema.json,
    category=discovery, status=wip, name matches dir;
  - outputs '#robots_txt' anchor matches a real master-excel sheet;
  - referenced report template exists; slash command exists;
  - produces edges (drift-check, robots-policy-audit) are present;
  - the transform -> sheet_merge.merge_prefixed_rows write path is idempotent and
    preserves foreign (sf_projection R-) rows.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import pytest
import yaml
from jsonschema import Draft7Validator
from openpyxl import load_workbook

from scripts.discovery import facet_nav_audit_transform as fnt
from scripts.orchestration import committer
from scripts.util import sheet_merge

ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = ROOT / "schemas"
SKILL_PATH = ROOT / "skills" / "discovery" / "facet-nav-audit" / "SKILL.md"
TEMPLATE = ROOT / "templates" / "reports" / "facet-nav-audit.template.md"
COMMAND = ROOT / "commands" / "pseo-facet-audit.md"


def _frontmatter() -> dict:
    text = SKILL_PATH.read_text(encoding="utf-8")
    m = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert m, "SKILL.md missing YAML frontmatter delimiters"
    return yaml.safe_load(m.group(1))


def _body() -> str:
    text = SKILL_PATH.read_text(encoding="utf-8")
    return text[re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL).end():]


def test_frontmatter_validates() -> None:
    fm = _frontmatter()
    schema = json.loads((SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8"))
    errors = sorted(Draft7Validator(schema).iter_errors(fm), key=lambda e: e.path)
    assert not errors, f"frontmatter invalid: {[e.message for e in errors]}"
    assert fm["name"] == "facet-nav-audit"
    assert fm["category"] == "discovery"
    assert fm["status"] == "wip"
    assert fm["budget"]["uses_paid_mcp"] is False


def test_outputs_anchor_matches_master_sheet() -> None:
    fm = _frontmatter()
    master = json.loads((SCHEMAS / "master-excel.schema.json").read_text("utf-8"))
    anchors = [o.split("#", 1)[1] for o in fm["outputs"] if o.startswith("master.xlsx#")]
    assert "robots_txt" in anchors
    for a in anchors:
        assert a in master["sheets"], f"output anchor #{a} not a real master sheet"


def test_produces_edges_present() -> None:
    fm = _frontmatter()
    assert set(fm["produces"]) >= {"drift-check", "robots-policy-audit"}


def test_template_and_command_exist() -> None:
    assert TEMPLATE.exists(), "facet-nav-audit report template missing"
    assert COMMAND.exists(), "/pseo-facet-audit command missing"


def test_body_documents_durur_and_recommendation_only() -> None:
    body = _body().lower()
    assert "durur" in body
    # R-130: engine never deploys to client infra
    assert "recommendation" in body or "operator" in body


def test_transform_to_merge_write_is_idempotent(tmp_path: Path) -> None:
    slug, run_id = "test-proj", "test-proj-2026-06-10-fn01"
    proj = tmp_path / "projects" / slug
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    wb = proj / "master.xlsx"
    # foreign sf_projection row must survive
    committer.commit(
        wb, "robots_txt",
        [{"id": "R-001", "level": "LOW", "issue": "x", "detail": "d", "resolution": "r"}],
        run_id=run_id, project_slug=slug, writer="sf-import",
    )
    internal = [
        {"Address": "https://x.test/?s=chair", "Indexability": "Indexable"},
        {"Address": "https://x.test/p?utm_source=x", "Indexability": "Indexable"},
    ]
    out = fnt.transform(internal, platform="custom")
    kw = dict(id_prefix="FN-", run_id=run_id, project_slug=slug, writer="facet-nav-audit")
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", out["robots_txt_rows"], **kw)
    ids1 = {r[0] for r in load_workbook(wb)["robots_txt"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    sheet_merge.merge_prefixed_rows(wb, "robots_txt", out["robots_txt_rows"], **kw)
    ids2 = {r[0] for r in load_workbook(wb)["robots_txt"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    assert ids1 == ids2
    assert "R-001" in ids2, "foreign sf_projection R- row must be preserved"
    assert any(str(i).startswith("FN-") for i in ids2), "FN- findings must be written"
