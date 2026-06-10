"""TDD lock for skills/planning/migration-map/SKILL.md + its wiring (GAP-T4).

Contract checks (spec GAP-T4 §d):
  - frontmatter validates against skill-frontmatter.schema.json, category=planning,
    status=wip, name matches dir;
  - outputs '#redirect_404' anchor matches a real master-excel sheet;
  - referenced report template + slash command exist; produces edges present;
  - body documents the two modes (plan/verify), the mode-enum DURUR, the
    approval-before-write gate, and that the redirect_deployed work event is
    emitted ONLY on the operator-confirm path (never autonomous);
  - redirect_deployed is enum-legal in events.schema.json (no schema change);
  - the build_map -> sheet_merge.merge_keyed_rows write path is idempotent and
    preserves foreign (sf_projection) redirect_404 rows.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import yaml
from jsonschema import Draft7Validator
from openpyxl import load_workbook

from scripts.orchestration import committer
from scripts.planning import migration_map_transform as mmt
from scripts.util import sheet_merge

ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = ROOT / "schemas"
SKILL_PATH = ROOT / "skills" / "planning" / "migration-map" / "SKILL.md"
TEMPLATE = ROOT / "templates" / "reports" / "migration-map.template.md"
COMMAND = ROOT / "commands" / "pseo-migration-map.md"


def _fm() -> dict:
    text = SKILL_PATH.read_text(encoding="utf-8")
    m = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert m, "SKILL.md missing YAML frontmatter delimiters"
    return yaml.safe_load(m.group(1))


def _body() -> str:
    text = SKILL_PATH.read_text(encoding="utf-8")
    return text[re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL).end():]


def test_frontmatter_validates() -> None:
    fm = _fm()
    schema = json.loads((SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8"))
    errors = sorted(Draft7Validator(schema).iter_errors(fm), key=lambda e: e.path)
    assert not errors, f"frontmatter invalid: {[e.message for e in errors]}"
    assert fm["name"] == "migration-map"
    assert fm["category"] == "planning"
    assert fm["status"] == "wip"
    assert fm["budget"]["uses_paid_mcp"] is False


def test_outputs_anchor_matches_master_sheet() -> None:
    fm = _fm()
    master = json.loads((SCHEMAS / "master-excel.schema.json").read_text("utf-8"))
    anchors = [o.split("#", 1)[1] for o in fm["outputs"] if o.startswith("master.xlsx#")]
    assert "redirect_404" in anchors
    for a in anchors:
        assert a in master["sheets"], f"output anchor #{a} not a real master sheet"


def test_produces_edges_present() -> None:
    fm = _fm()
    assert set(fm["produces"]) >= {"drift-check", "master-task-sync", "indexing-ping"}


def test_template_and_command_exist() -> None:
    assert TEMPLATE.exists(), "migration-map report template missing"
    assert COMMAND.exists(), "/pseo-migration-map command missing"


def test_body_documents_modes_durur_and_approval() -> None:
    body = _body().lower()
    assert "plan" in body and "verify" in body, "both modes must be documented"
    assert "durur" in body
    # approval-before-write
    assert "request_approval" in body or "onay" in body
    # recommendation-only / operator-deployed (R-135)
    assert "operator" in body or "recommendation" in body


def test_body_gates_redirect_deployed_on_confirm() -> None:
    body = _body()
    assert "redirect_deployed" in body, "the confirm-path work event must be named"
    low = body.lower()
    # emitted only on the operator-confirm path, never autonomous
    assert "confirm" in low or "onay" in low
    assert "never autonomous" in low or "gated" in low or "approval" in low


def test_redirect_deployed_event_is_enum_legal() -> None:
    events = json.loads((SCHEMAS / "events.schema.json").read_text("utf-8"))
    work_enum = events["properties"]["event_type"]["enum"]
    assert "redirect_deployed" in work_enum, (
        "redirect_deployed must be an existing work event_type (no schema change)"
    )


def test_build_map_to_merge_keyed_write_is_idempotent(tmp_path: Path) -> None:
    slug, run_id = "test-proj", "test-proj-2026-06-10-mm01"
    proj = tmp_path / "projects" / slug
    (proj / "_state").mkdir(parents=True, exist_ok=True)
    wb = proj / "master.xlsx"
    # foreign sf_projection redirect_404 row (different url) must survive
    committer.commit(
        wb, "redirect_404",
        [{"url": "https://x.test/legacy", "inlinks": 1, "action": "301",
          "target_url": "https://x.test/kept", "status": "DONE"}],
        run_id=run_id, project_slug=slug, writer="sf-import",
    )
    out = mmt.build_map(
        [{"Address": "https://x.test/old-a", "Inlinks": 3}],
        [{"old_url": "https://x.test/old-a", "new_url": "https://x.test/new-a", "action": "301"}],
        [], [],
    )
    kw = dict(key_column="url", run_id=run_id, project_slug=slug, writer="migration-map")
    sheet_merge.merge_keyed_rows(wb, "redirect_404", out["redirect_rows"], **kw)
    urls1 = {r[0] for r in load_workbook(wb)["redirect_404"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    sheet_merge.merge_keyed_rows(wb, "redirect_404", out["redirect_rows"], **kw)
    urls2 = {r[0] for r in load_workbook(wb)["redirect_404"].iter_rows(min_row=5, values_only=True) if r and r[0]}
    assert urls1 == urls2, "re-running must not duplicate keyed rows"
    assert "https://x.test/legacy" in urls2, "foreign sf_projection redirect_404 row must be preserved"
    assert "https://x.test/old-a" in urls2, "migration row must be written"
