"""tests/skills/test_portfolio_task_heatmap.py — Phase 9 W-E6.

portfolio-task-heatmap reporting skill — multi-project READ-ONLY
project x category x priority density aggregator.

Mirrors tests/skills/test_portfolio_overview.py + test_portfolio_weekly_brief.py:
  - frontmatter parse + Draft 7 validate against
    schemas/skill-frontmatter.schema.json
  - statusEnum 7 sentinel (all 7 statusEnum values recognized)
  - severityEnum 4 sentinel (CRITICAL/HIGH/MEDIUM/LOW recognized)
  - missing master_task tolerance (graceful skip, not crash)
  - heatmap density empty cell (zero-count cell renders correctly)
  - status_check_drift advisory non-DURUR (mismatch surfaces WARNING,
    transform proceeds)
  - natural_language min length (>= 30 char) sentinel
  - forbidden tokens guard (4 grep tokens + 8 plugin slugs, CLEAN)
  - assert_read_only_module() helper grep guard (W-E3 reuse)
  - path convention sentinel (projects/_portfolio/ literal +
    PSEO_WORKSPACE_ROOT env)

This is a READ-ONLY reporting skill — no MCP, no DFS, no budget
pre-flight, no master.xlsx writes, no events.jsonl writes.
"""

from __future__ import annotations

import base64
import json
import os
import re
from pathlib import Path

import pytest
import yaml
from jsonschema import Draft7Validator

from scripts.reporting import portfolio_task_heatmap as pth


REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMAS = REPO_ROOT / "schemas"
SKILL_PATH = (
    REPO_ROOT / "skills" / "reporting" / "portfolio-task-heatmap"
    / "SKILL.md"
)
TRANSFORM_PATH = (
    REPO_ROOT / "scripts" / "reporting" / "portfolio_task_heatmap.py"
)
TEMPLATE_PATH = (
    REPO_ROOT / "templates" / "reports"
    / "portfolio-task-heatmap.template.md"
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def skill_frontmatter_schema() -> dict:
    return json.loads(
        (SCHEMAS / "skill-frontmatter.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def portfolio_config_schema() -> dict:
    return json.loads(
        (SCHEMAS / "portfolio-config.schema.json").read_text("utf-8")
    )


@pytest.fixture(scope="module")
def skill_frontmatter() -> dict:
    """Parse the YAML frontmatter block of the portfolio-task-heatmap SKILL.md."""
    text = SKILL_PATH.read_text(encoding="utf-8")
    match = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert match, "SKILL.md missing YAML frontmatter delimiters"
    return yaml.safe_load(match.group(1))


def _synthetic_config(active_count: int = 2) -> dict:
    """Build a valid portfolio.config payload with N active projects.
    Slugs are generic test fixtures (NOT real client slugs)."""
    active = []
    for i in range(active_count):
        active.append({
            "slug": f"alpha-test-{i+1}",
            "display_name": f"Alpha Test {i+1}",
            "workspace_path": f"./workspace-{i+1}",
            "profile": ["e-commerce"],
            "priority": i + 1,
        })
    return {
        "schema_version": "1.1",
        "portfolio_id": "test-portfolio",
        "display_name": "Test Portfolio",
        "active_projects": active,
        "slas": {
            "weekly_sync_max_days": 7,
            "monthly_roundup_target_dom": 5,
        },
        "cadence": {
            "weekly_brief": {"day": "Monday", "hour": 9},
            "monthly_roundup": {"day_of_month": 5, "hour": 10},
        },
        "cross_query": {"read_only": True},
    }


def _build_master_task_xlsx(
    path: Path,
    rows: list[tuple[str, str, str, str]],
) -> None:
    """Write a minimal master.xlsx with master_task sheet for tests.
    Each row = (task_id, category, priority, status).
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "master_task"
    # Header row at row 3 (master-excel.schema.json header_row=3).
    ws.cell(row=3, column=1, value="task_id")
    ws.cell(row=3, column=6, value="category")
    ws.cell(row=3, column=7, value="priority")
    ws.cell(row=3, column=10, value="status")
    # Data starts at row 4.
    for i, (tid, cat, pri, sta) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=tid)
        ws.cell(row=r, column=6, value=cat)
        ws.cell(row=r, column=7, value=pri)
        ws.cell(row=r, column=10, value=sta)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Test 1 — Frontmatter validates + natural_language >= 30 char sentinel
# ---------------------------------------------------------------------------

def test_frontmatter_validates_and_natural_language_min_length(
    skill_frontmatter: dict,
    skill_frontmatter_schema: dict,
) -> None:
    """SKILL.md frontmatter must:
      (a) parse + name=portfolio-task-heatmap, status=wip, version="1.0",
          category=reporting,
      (b) declare 0 required AND 0 optional MCP tools,
      (c) outputs include master.xlsx#none + report path + inbox snapshot,
      (d) triggers.manual = [] (NO slash command per brief),
      (e) natural_language >= 30 char (Phase 9 W2 brief gate #7),
      (f) budget + autonomy contracts,
      (g) Draft 7 validation against the canonical schema.
    """
    fm = skill_frontmatter
    assert fm["name"] == "portfolio-task-heatmap"
    assert fm["status"] == "wip"
    assert fm["version"] == "1.0"
    assert fm["category"] == "reporting"

    # No required NOR optional MCP tools.
    assert fm.get("mcp_tools", {}).get("required", []) == []
    assert fm.get("mcp_tools", {}).get("optional", []) == []

    # Outputs: master.xlsx#none + report + inbox snapshot.
    outputs = fm["outputs"]
    assert "master.xlsx#none" in outputs
    assert any(
        "portfolio-task-heatmap.md" in o
        and "outputs/reports/" in o
        for o in outputs
    )
    assert any(
        "inbox/local/" in o
        and "portfolio-task-heatmap.json" in o
        for o in outputs
    )
    # NO events.jsonl write (Q-RP-01 deferred).
    assert "events.jsonl" not in outputs

    # Manual triggers empty (brief: NO slash command).
    assert fm["triggers"]["manual"] == []

    # natural_language sentinel (>= 30 char).
    nl = fm["triggers"]["natural_language"]
    assert isinstance(nl, str), "natural_language must be a string block"
    assert len(nl) >= 30, (
        f"natural_language too short ({len(nl)} chars); >= 30 required."
    )
    # Required brief phrases.
    for required in (
        "task heatmap", "task yoğunluk", "task density",
        "category priority dağılım",
    ):
        assert required in nl, f"missing required trigger phrase: {required!r}"

    # Budget: 0 credits, no paid MCP.
    assert fm["budget"]["uses_paid_mcp"] is False
    assert fm["budget"].get("estimated_credits", 0) == 0

    # Autonomy: HIGH + cron-ready.
    assert fm["autonomy"]["confidence"] == "HIGH"
    assert fm["autonomy"]["requires_approval"] is False
    assert fm["autonomy"]["safe_auto_execute"] is True

    # Full Draft 7 validation.
    validator = Draft7Validator(skill_frontmatter_schema)
    errs = sorted(
        validator.iter_errors(fm), key=lambda e: list(e.absolute_path),
    )
    assert not errs, (
        f"frontmatter invalid: "
        f"{[(list(e.absolute_path), e.message) for e in errs]}"
    )


# ---------------------------------------------------------------------------
# Test 2 — statusEnum 7 + severityEnum 4 sentinels
# ---------------------------------------------------------------------------

def test_status_and_severity_enum_sentinels() -> None:
    """master-excel.schema.json defines statusEnum (7 values) and
    severityEnum (4 values). The transform must surface both as
    constants and recognize all enum members in density / status_dist.
    """
    # statusEnum 7 values per schema #/definitions/statusEnum.
    assert pth.STATUS_ENUM == (
        "TODO", "ONGOING", "EXISTS", "DONE",
        "BLOCKED", "DEFERRED", "CANCELED",
    )
    assert len(pth.STATUS_ENUM) == 7

    # severityEnum 4 values per schema #/definitions/severityEnum.
    assert pth.SEVERITY_ENUM == ("CRITICAL", "HIGH", "MEDIUM", "LOW")
    assert len(pth.SEVERITY_ENUM) == 4

    # Closed statuses (excluded from open-task density).
    assert pth.CLOSED_STATUSES == frozenset({"DONE", "CANCELED"})


# ---------------------------------------------------------------------------
# Test 3 — Missing master.xlsx tolerated (graceful skip, no crash)
# ---------------------------------------------------------------------------

def test_missing_master_task_tolerated(
    tmp_path: Path, portfolio_config_schema: dict,
) -> None:
    """Per the SKILL.md gate: per-project missing master.xlsx is NOT
    a DURUR. The aggregator skips gracefully — empty density + warning."""
    portfolio_root = tmp_path / "portfolio_ws"
    portfolio_root.mkdir()
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)
    config = _synthetic_config(active_count=2)
    (portfolio_root / "projects" / "_portfolio" / "portfolio.config.json"
     ).write_text(json.dumps(config), encoding="utf-8")

    # Validate first (schema clean).
    pth.validate_portfolio_config(config, portfolio_config_schema)

    # Aggregate — must NOT raise.
    batch = pth.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
    )
    assert batch.portfolio_id == "test-portfolio"
    assert len(batch.projects) == 2
    assert batch.skipped_projects == 2
    for ph in batch.projects:
        assert ph.master_xlsx_path is None
        assert ph.warnings, "missing workbook should surface warning"
        assert ph.density == {}
        assert ph.open_total == 0


# ---------------------------------------------------------------------------
# Test 4 — Heatmap density empty cell + statusEnum/severityEnum coverage
# ---------------------------------------------------------------------------

def test_heatmap_density_with_open_and_closed_tasks(
    tmp_path: Path, portfolio_config_schema: dict,
) -> None:
    """Build a real master.xlsx fixture + assert density correctness.
    Closed statuses (DONE/CANCELED) must NOT appear in density;
    open statuses must appear at correct (category, severity) cells;
    out-of-enum priority -> OTHER bucket; empty cells render as 0."""
    portfolio_root = tmp_path / "portfolio_ws"
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)
    workspace = tmp_path / "workspace-1"
    workspace.mkdir()
    config = {
        "schema_version": "1.1",
        "portfolio_id": "test-portfolio",
        "display_name": "Test Portfolio",
        "active_projects": [{
            "slug": "alpha-test-1",
            "display_name": "Alpha Test 1",
            "workspace_path": str(workspace),
            "profile": ["e-commerce"],
            "priority": 1,
        }],
        "slas": {
            "weekly_sync_max_days": 7,
            "monthly_roundup_target_dom": 5,
        },
        "cadence": {
            "weekly_brief": {"day": "Monday", "hour": 9},
            "monthly_roundup": {"day_of_month": 5, "hour": 10},
        },
        "cross_query": {"read_only": True},
    }
    (portfolio_root / "projects" / "_portfolio" / "portfolio.config.json"
     ).write_text(json.dumps(config), encoding="utf-8")

    master_xlsx = workspace / "projects" / "alpha-test-1" / "master.xlsx"
    _build_master_task_xlsx(master_xlsx, [
        # category, priority, status
        ("T-1", "Performance", "CRITICAL", "TODO"),
        ("T-2", "Performance", "HIGH",     "ONGOING"),
        ("T-3", "Performance", "HIGH",     "DONE"),     # closed -> excluded
        ("T-4", "SEO",         "MEDIUM",   "TODO"),
        ("T-5", "SEO",         "LOW",      "BLOCKED"),
        ("T-6", "SEO",         "LOW",      "CANCELED"), # closed -> excluded
        ("T-7", "SEO",         "URGENT",   "EXISTS"),   # bad severity -> OTHER
        ("T-8", "Content",     "CRITICAL", "DEFERRED"),
    ])
    pth.validate_portfolio_config(config, portfolio_config_schema)
    batch = pth.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
    )
    assert len(batch.projects) == 1
    p = batch.projects[0]
    assert p.master_xlsx_path is not None
    # Open tasks: T-1 (Perf/CRITICAL), T-2 (Perf/HIGH), T-4 (SEO/MED),
    # T-5 (SEO/LOW), T-7 (SEO/OTHER), T-8 (Content/CRITICAL) = 6 open.
    assert p.open_total == 6
    # density spot-checks.
    assert p.density["Performance"]["CRITICAL"] == 1
    assert p.density["Performance"]["HIGH"] == 1
    # DONE T-3 must NOT be counted in density.
    assert p.density["Performance"].get("DONE", 0) == 0
    # Empty cell sentinel: Performance / MEDIUM = 0.
    assert p.density["Performance"]["MEDIUM"] == 0
    assert p.density["Performance"]["LOW"] == 0
    assert p.density["SEO"]["MEDIUM"] == 1
    assert p.density["SEO"]["LOW"] == 1
    # Out-of-enum severity -> OTHER bucket.
    assert p.density["SEO"].get(pth.OTHER_BUCKET, 0) == 1
    assert p.density["Content"]["CRITICAL"] == 1
    # Status distribution: 8 rows, 7 statusEnum + 0 OTHER.
    assert sum(p.status_distribution.values()) == 8
    assert p.status_distribution.get(pth.OTHER_BUCKET, 0) == 0
    # consistency-report absent -> no drift.
    assert p.consistency_verdict is None
    assert p.status_check_drift is False

    # Render snapshot + report — both must succeed.
    snapshot = pth.build_snapshot(batch=batch)
    json.dumps(snapshot)  # JSON-serialisable round-trip.
    template_text = TEMPLATE_PATH.read_text(encoding="utf-8")
    report = pth.build_report_markdown(
        batch=batch, template_text=template_text,
    )
    assert "Portfolio Task Heatmap" in report
    assert "alpha-test-1" in report
    assert "Performance" in report
    assert "SEO" in report

    # Idempotency: same inputs + generated_at -> byte-stable snapshot.
    batch2 = pth.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01", generated_at=batch.generated_at,
    )
    snap1 = json.dumps(
        pth.build_snapshot(batch=batch), sort_keys=True,
    )
    snap2 = json.dumps(
        pth.build_snapshot(batch=batch2), sort_keys=True,
    )
    assert snap1 == snap2, "aggregate must be idempotent given same now"


# ---------------------------------------------------------------------------
# Test 5 — status_check_drift advisory non-DURUR (W-E3 paterni reuse)
# ---------------------------------------------------------------------------

def test_status_check_drift_advisory_non_durur(
    tmp_path: Path, portfolio_config_schema: dict,
) -> None:
    """When per-project consistency-report verdict != GREEN, the project
    surfaces status_check_drift = True + warning, BUT the transform
    PROCEEDS — heatmap density remains valid + report renders. This is
    the W-E3 advisory paterni reuse: WARNING NOT DURUR."""
    portfolio_root = tmp_path / "portfolio_ws"
    (portfolio_root / "projects" / "_portfolio").mkdir(parents=True)
    workspace = tmp_path / "workspace-1"
    workspace.mkdir()
    config = {
        "schema_version": "1.1",
        "portfolio_id": "test-portfolio",
        "display_name": "Test Portfolio",
        "active_projects": [{
            "slug": "alpha-test-1",
            "display_name": "Alpha Test 1",
            "workspace_path": str(workspace),
            "profile": ["e-commerce"],
            "priority": 1,
        }],
        "slas": {
            "weekly_sync_max_days": 7,
            "monthly_roundup_target_dom": 5,
        },
        "cadence": {
            "weekly_brief": {"day": "Monday", "hour": 9},
            "monthly_roundup": {"day_of_month": 5, "hour": 10},
        },
        "cross_query": {"read_only": True},
    }
    pth.validate_portfolio_config(config, portfolio_config_schema)
    master_xlsx = workspace / "projects" / "alpha-test-1" / "master.xlsx"
    _build_master_task_xlsx(master_xlsx, [
        ("T-1", "Performance", "CRITICAL", "TODO"),
        ("T-2", "SEO",         "HIGH",     "ONGOING"),
    ])
    # Plant a consistency-report with verdict=AMBER -> drift advisory.
    state_dir = workspace / "projects" / "alpha-test-1" / "_state"
    state_dir.mkdir(parents=True, exist_ok=True)
    (state_dir / "consistency-report.json").write_text(json.dumps({
        "schema_version": "1.0",
        "report_id": 1,
        "generated_at": "2026-05-01T10:00:00Z",
        "project_id": "alpha-test-1",
        "verdict": "AMBER",
        "checks": [{
            "check_id": "CSR-F-01",
            "category": "csr_foundation",
            "verdict": "WARN",
        }],
    }), encoding="utf-8")

    batch = pth.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
    )
    assert len(batch.projects) == 1
    p = batch.projects[0]
    # Transform PROCEEDED — density present.
    assert p.open_total == 2
    # Drift surfaces as advisory (non-DURUR).
    assert p.consistency_verdict == "AMBER"
    assert p.status_check_drift is True
    assert any("status_check_drift" in w for w in p.warnings)
    # Report renders the advisory block.
    report = pth.build_report_markdown(
        batch=batch,
        template_text=TEMPLATE_PATH.read_text(encoding="utf-8"),
    )
    assert "status_check_drift" in report
    assert "AMBER" in report

    # Counter-case: GREEN verdict -> NO drift.
    (state_dir / "consistency-report.json").write_text(json.dumps({
        "schema_version": "1.0",
        "report_id": 2,
        "generated_at": "2026-05-01T11:00:00Z",
        "project_id": "alpha-test-1",
        "verdict": "GREEN",
        "checks": [{
            "check_id": "CSR-F-01",
            "category": "csr_foundation",
            "verdict": "PASS",
        }],
    }), encoding="utf-8")
    batch_green = pth.aggregate(
        portfolio_root=portfolio_root, config=config,
        run_date="2026-05-01",
    )
    p_green = batch_green.projects[0]
    assert p_green.consistency_verdict == "GREEN"
    assert p_green.status_check_drift is False


# ---------------------------------------------------------------------------
# Test 6 — assert_read_only_module() helper grep guard (W-E3 reuse)
# ---------------------------------------------------------------------------

def test_assert_read_only_module_helper() -> None:
    """Per portfolio-config.schema.json#cross_query.read_only=true (const),
    portfolio_task_heatmap.py must contain ZERO openpyxl write call sites.
    Verified by W-E3 helper reuse: assert_read_only_module()."""
    # Direct invocation — must NOT raise.
    pth.assert_read_only_module()

    # Independent re-grep so the test is robust even if the helper itself
    # is later refactored. Patterns assembled at runtime so this test
    # file ALSO does not contain the literal write-call substrings.
    src = TRANSFORM_PATH.read_text(encoding="utf-8")
    forbidden = (
        "." + "save(",
        "transaction" + "." + "append(",
        "transaction" + "." + "update(",
        "transaction" + "." + "delete(",
    )
    for pat in forbidden:
        assert pat not in src, (
            f"forbidden write pattern {pat!r} appears verbatim in "
            "portfolio_task_heatmap.py source — read-only contract violated."
        )

    # No transaction.* / events_writer.* imports.
    forbidden_imports = (
        r"^\s*from\s+scripts\.excel\.transaction\b",
        r"^\s*import\s+scripts\.excel\.transaction\b",
        r"^\s*from\s+scripts\.state\.events_writer\b",
        r"^\s*import\s+scripts\.state\.events_writer\b",
    )
    for pattern in forbidden_imports:
        assert not re.search(pattern, src, flags=re.MULTILINE), (
            f"forbidden import pattern matched: {pattern}"
        )


# ---------------------------------------------------------------------------
# Test 7 — Forbidden tokens guard (4 grep tokens + 8 plugin slug guards)
# ---------------------------------------------------------------------------

def test_no_forbidden_tokens_in_skill() -> None:
    """Per worker brief acceptance gate #6: transform + SKILL.md +
    template must contain ZERO occurrences of:
      1. estimated_credits_per_call (Phase 7 lesson)
      2. estimated_credits_per_url  (Phase 7 lesson)
      3. metric_name                (Q-CO-01 closeout lesson)
      4. TODO comment in transform code
      5. 8 hardcoded real-client slugs (plugin-agnostik discipline)

    Slug tokens are obfuscated as base64-decoded strings so this test
    file itself doesn't trip the grep guard (mirrors W-D1 + W-E4 pattern).
    """
    transform_src = TRANSFORM_PATH.read_text(encoding="utf-8")
    skill_src = SKILL_PATH.read_text(encoding="utf-8")
    template_src = TEMPLATE_PATH.read_text(encoding="utf-8")

    # Tokens 1-3: literal grep CLEAN across transform/SKILL/template.
    forbidden_literal = (
        "estimated_credits" + "_per_call",
        "estimated_credits" + "_per_url",
        "metric" + "_name",
    )
    for body, label in (
        (transform_src, "transform"),
        (skill_src, "skill"),
        (template_src, "template"),
    ):
        for tok in forbidden_literal:
            assert tok not in body, (
                f"forbidden token {tok!r} found in {label}"
            )

    # Token 4: TODO comment in transform CODE only. Markdown skill prose
    # may use the word legitimately. Allow STATUS_ENUM "TODO" entries.
    todo_comment_pattern = re.compile(r"#\s*TODO\b")
    assert not todo_comment_pattern.search(transform_src), (
        "TODO inline-comment marker found in transform code "
        "(STATUS_ENUM tuple entries are allowed)."
    )

    # Token 5: hardcoded plugin slugs. Base64-encoded so the grep
    # doesn't trip on this test file.
    encoded_slug_tokens = (
        b"ZGVudG5vdGlvbg==", b"dmVudG8=", b"ZXlrb20=", b"YmlnY2F0dHI=",
        b"Y2FsaXR0ZQ==", b"bGFzdGlrc2E=", b"bm9yYW5pbnNhYXQ=",
        b"YWRzdGFyaw==",
    )
    forbidden_slugs = tuple(
        base64.b64decode(t).decode("ascii") for t in encoded_slug_tokens
    )
    pattern = re.compile(
        r"\b(" + "|".join(forbidden_slugs) + r")\b",
        flags=re.IGNORECASE,
    )
    for body, label in (
        (transform_src, "transform"),
        (skill_src, "skill"),
        (template_src, "template"),
    ):
        m = pattern.search(body)
        assert m is None, (
            f"forbidden hardcoded slug {m.group(0)!r} found in {label}"
        )


# ---------------------------------------------------------------------------
# Test 8 — Path convention sentinel (W-E4 surface compliance)
# ---------------------------------------------------------------------------

def test_path_convention_compliance(monkeypatch) -> None:
    """Worker brief acceptance gate #10: path convention compliance.
      - projects/_portfolio/{outputs,inbox}/local/ literal in transform
      - PSEO_WORKSPACE_ROOT env var honored as fallback resolver
      - frontmatter outputs[] reflects the same paths.
    """
    # Transform exposes the path constants.
    assert pth.OUTPUTS_REPORTS_REL == "projects/_portfolio/outputs/reports"
    assert pth.INBOX_LOCAL_REL == "projects/_portfolio/inbox/local"

    # Transform source contains the literal path strings (anchored).
    src = TRANSFORM_PATH.read_text(encoding="utf-8")
    assert "projects/_portfolio/outputs/reports" in src, (
        "transform must contain projects/_portfolio/outputs/reports literal"
    )
    assert "projects/_portfolio/inbox/local" in src, (
        "transform must contain projects/_portfolio/inbox/local literal"
    )
    assert "PSEO_WORKSPACE_ROOT" in src, (
        "transform must reference PSEO_WORKSPACE_ROOT env var"
    )

    # Frontmatter outputs[] reflects the path convention.
    text = SKILL_PATH.read_text(encoding="utf-8")
    match = re.match(r"^---\n(.*?)\n---\n", text, flags=re.DOTALL)
    assert match
    fm = yaml.safe_load(match.group(1))
    outputs = fm["outputs"]
    assert any(
        "projects/_portfolio/outputs/reports/" in o
        and "portfolio-task-heatmap.md" in o
        for o in outputs
    )
    assert any(
        "projects/_portfolio/inbox/local/" in o
        and "portfolio-task-heatmap.json" in o
        for o in outputs
    )

    # PSEO_WORKSPACE_ROOT env var resolution (W-E4 paterni).
    # Set the env var to a real dir; _resolve_portfolio_root must return it.
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td).resolve()
        monkeypatch.setenv("PSEO_WORKSPACE_ROOT", str(td_path))
        resolved = pth._resolve_portfolio_root(None)
        assert resolved == td_path, (
            f"PSEO_WORKSPACE_ROOT env not honored; got {resolved}"
        )

    # Negative: env var unset + no arg + no .pse-workspace -> DURUR.
    monkeypatch.delenv("PSEO_WORKSPACE_ROOT", raising=False)
    # Run from a temp dir that has no .pse-workspace marker.
    with tempfile.TemporaryDirectory() as td:
        cwd_before = os.getcwd()
        try:
            os.chdir(td)
            with pytest.raises(pth.WorkspaceRootUnsetError):
                pth._resolve_portfolio_root(None)
        finally:
            os.chdir(cwd_before)
